
import re
from contextlib import contextmanager
from collections import defaultdict
from flask import request, render_template, send_file, flash, url_for, abort
from flask import jsonify
from flask import request, redirect, url_for, flash, current_app, session
from flask import send_file
from flask import request, render_template, redirect, url_for, flash
from werkzeug.security import generate_password_hash, check_password_hash
from flask import session, redirect, url_for, flash, render_template, request
from urllib.parse import urljoin, urlparse
from werkzeug.security import check_password_hash
from flask import Response
from flask import (
    request, render_template, redirect, url_for, flash, send_file, session
)
from openpyxl.utils import get_column_letter
import csv
from werkzeug.security import generate_password_hash
from typing import Optional, Callable
from contextlib import closing
from flask import Flask, request, redirect, url_for, flash, render_template, jsonify, Response, session, send_file
from io import StringIO
# import bcrypt
import io
import secrets
from functools import wraps
from datetime import datetime, date
import xlsxwriter
from io import BytesIO
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
import secrets
import hashlib
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path
from dotenv import load_dotenv
import inspect
from flask import Flask, current_app
from urllib.parse import urlparse, urljoin
from collections import Counter
from config import DevConfig, ProdConfig, TestConfig
from authz import require_role
# import win32print
import mysql.connector
from mysql.connector import IntegrityError
from mysql.connector import Error
from mysql.connector.errors import InterfaceError
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm, mm
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, LongTable
from reportlab.platypus import Image as RLImage
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.enums import TA_LEFT, TA_CENTER
import math
import csv
import os, time
from werkzeug.utils import secure_filename


# 1) Load_dotenv before anything else
load_dotenv()


# Constants

TERMS = ["Term 1", "Term 2", "Term 3"]
ALLOWED_EXTS = {"csv", "xlsx", "xls"}
ALLOWED_EXTS_HP = ALLOWED_EXTS
SCHOOL_FEE_TYPES = ("fees", "schoolfees", "school_fees")
HOLIDAY_NAME = "Holiday Package"
OTHER_NAME = "Other Assessments"
CORE_CODES = ["ENG", "MATH", "SCI", "SST"]
AGG_MAP = {"D1":1,"D2":2,"C3":3,"C4":4,"C5":5,"C6":6,"P7":7,"P8":8,"F9":9}
ASSET_EXCEL_FILENAME = "Asset_Register.xlsx"
ASSET_PDF_FILENAME = "Asset_Register.pdf"
COMPONENT_FIELDS = ("other_mark", "holiday_mark", "bot_mark", "midterm_mark", "eot_mark", "ca_mark")
TERM2NO = {"term 1": 1, "term 2": 2, "term 3": 3}
KG_CLASSES = ["Baby","Middle","Top"]


# Kindergarten Report
# --- School branding (adjust freely or load from DB/config) ---
SCHOOL_NAME = "DEMO DAY & BOARDING"
SCHOOL_SUB = "PRIMARY SCHOOL – KAMPALA"
SCHOOL_PHONE_LINES = ["+256778878411, +256759685640, +256773589232, +256750347624"]
SCHOOL_ADDRESS = "P.O Box 1X1X1 Kampala"
SCHOOL_EMAIL = ""
SCHOOL_LOGO_PATH = "static/icons/logo.jpg" # put your crest here
# palette sampled from the banner
COL_NAVY = colors.HexColor("#0b2a5b")
COL_BLUE = colors.HexColor("#1f66c5")
COL_BLUE2 = colors.HexColor("#1556a8")
BBAND_PALETTE = [COL_NAVY, COL_BLUE]
USE_TWO_COLOR_PALETTE = True

# photo
#PHOTO_UPLOAD_FOLDER = os.path.join(app.root_path, "static", "uploads", "students")
#os.makedirs(PHOTO_UPLOAD_FOLDER, exist_ok=True)



BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# All uploads go under the Flask static folder
STATIC_DIR = os.path.join(BASE_DIR, "static")
PHOTO_SUBDIR = os.path.join("uploads", "students") # relative inside static
PHOTO_UPLOAD_FOLDER = os.path.join(STATIC_DIR, PHOTO_SUBDIR)
os.makedirs(PHOTO_UPLOAD_FOLDER, exist_ok=True)

try:
    from PIL import Image
    import pillow_heif
    pillow_heif.register_heif_opener()
except ImportError:
    Image = None


# Photo validation config
ALLOWED_PHOTO_EXTS = {"jpg", "jpeg", "png", "webp", "heic", "heif"}
ALLOWED_PHOTO_MIMES = {"image/jpeg", "image/png", "image/webp", "image/heic", "image/heif"}
MAX_PHOTO_SIZE = 3 * 1024 * 1024 # 5 MB (you can reduce if you want)


REMARK_OPTIONS = [ 
    "Exceeds the level of development expected",
    "At the level of development Expected",
    "Towards the level of Development Expected",
    "Exceeding the level Expected",
    "Trying to get to the level of development expected",
    "Not attempted",
    "Not yet covered or assessed",
]

RECOMMENDATION_OPTIONS = [
    "We need to keep supporting the learner for best results.",
    "The learner has shown great improvement; continue encouraging practice at home.",
    "The learner needs close supervision and extra support in key areas.",
]







AREA_STYLES = {
    # Colours borrowed from the sample PDF palette, but named with YOUR areas
    "KNOWLEDGE": {
        "bg": colors.HexColor("#3F6CB5"), # blue
        "label": "GENERAL KNOWLEDGE",
    },
    "LANGUAGE": {
        "bg": colors.HexColor("#6F2F8F"), # purple
        "label": "LANGUAGE",
    },
    "NUMBERS": {
        "bg": colors.HexColor("#E31E24"), # red / orange
        "label": "NUMBERS",
    },
    "PHYSICAL DEVELOPMENT": {
        "bg": colors.HexColor("#00A651"), # green
        "label": "PHYSICAL DEVELOPMENT",
    },
}


# Ensure each area has a label for vertical band
for _k, _v in AREA_STYLES.items():
    _v.setdefault("label", _k)
    
    
    
CLASS_ENUMS = ("Baby", "Middle", "Top", "P1", "P2", "P3", "P4", "P5", "P6", "P7")
CLASS_ENUM_SET = set(CLASS_ENUMS)


WEEKS = [f"Week {i}" for i in range(1, 15)] # Week 1 .. Week 14


# ------------ TEACHER WORK PLANS ------------

SUPERVISOR_ROLES = {
    "dos",
    "headteacher",
    "deputyheadteacher",
    "director",
    "admin",
    "classmanager",
    "classteacher",
}
# ---------- DB helpers ----------


def get_db_connection():
    conn = mysql.connector.connect(
        host=os.getenv("MYSQLHOST"),
        port=int(os.getenv("MYSQLPORT", "3306")),
        user=os.getenv("MYSQLUSER"),
        password=os.getenv("MYSQLPASSWORD"),
        database=os.getenv("MYSQLDATABASE"),
        autocommit=False,
        connection_timeout=10,
        charset="utf8mb4",
        collation="utf8mb4_general_ci",
        buffered=True
    )
    return conn


def safe_ping(conn):
    """Ensure the connection is alive; reconnect if needed."""
    try:
        if conn:
            conn.ping(reconnect=True, attempts=1, delay=0)
            return True
    except Exception:
        return False
    return False


def safe_rollback(conn):
    """Rollback without crashing if the server is gone."""
    try:
        if conn and conn.is_connected():
            conn.rollback()
    except Exception:
        pass


def safe_close(cur=None, conn=None):
    """Close cursor/connection without raising secondary errors."""
    try:
        if cur:
            cur.close()
    except Exception:
        pass
    try:
        if conn:
            conn.close()
    except Exception:
        pass

# ---------- usage pattern (example route) ----------

# @app.route("/some_action", methods=["POST"])
# def some_action():
# conn, cur = None, None
# try:
# conn = get_db_connection()
# safe_ping(conn)
# cur = conn.cursor(dictionary=True)
# # ... do your queries here ...
# conn.commit()
# flash("Done.", "success")
# except Error as e:
# safe_rollback(conn)
# flash(f"MySQL error: {e}", "danger")
# except Exception as e:
# safe_rollback(conn)
# flash(f"Error: {e}", "danger")
# finally:
# safe_close(cur, conn)
# return redirect(url_for("dashboard"))


def configure_logging(app):
    if getattr(app, "_logging_configured", False):
        return
    app.logger.setLevel(app.config.get("LOG_LEVEL", "INFO"))
    console = logging.StreamHandler()
    console.setLevel(app.config.get("LOG_LEVEL", "INFO"))
    console.setFormatter(logging.Formatter(
        "[%(asctime)s] %(levelname)s in %(module)s: %(message)s"))
    app.logger.addHandler(console)

    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    file_handler = RotatingFileHandler(
        log_dir / "app.log", maxBytes=2_000_000, backupCount=5)
    file_handler.setLevel(app.config.get("LOG_LEVEL", "INFO"))
    file_handler.setFormatter(logging.Formatter(
        "%(asctime)s %(levelname)s %(name)s %(funcName)s: %(message)s"))
    app.logger.addHandler(file_handler)

    app._logging_configured = True


def bootstrap():
    """
    Initialize database schema and seed defaults.
    Tries to call your project's ensure_* and seed_* helpers if present.
    Safe to run multiple times (idempotent).
    """

    import inspect

    def _call_if_exists(name, conn=None):
        """Helper to safely call ensure_*/seed_* if it exists."""
        fn = globals().get(name)
        if not callable(fn):
            return False, f"{name} (missing)"
        try:
            sig = inspect.signature(fn)
            if len(sig.parameters) >= 1 and conn is not None:
                fn(conn)  # call with connection
            else:
                fn()  # call without connection
            return True, f"{name} (ok)"
        except TypeError:
            try:
                fn() if conn is not None else fn()
                return True, f"{name} (ok: alt)"
            except Exception as e:
                return False, f"{name} (error: {e})"
        except Exception as e:
            return False, f"{name} (error: {e})"

    # --- Open a DB connection ---
    try:
        conn = get_db_connection()
    except NameError as e:
        raise RuntimeError(
            "get_db_connection() not found. Import it before calling bootstrap()."
        ) from e

    ran = []

    try:
        # --- Core schema creators ---
        for fname in [
            # academics
            "ensure_students_table",
            "ensure_subjects_table",
            "ensure_record_score_table",
            "ensure_results_table",
            "ensure_midterms_table",
            "ensure_reports_table",

            # finance & ops
            "ensure_expense_schema",
            "ensure_payroll_schema",
            "ensure_class_fees_schema",
            "ensure_classes_schema",
            "ensure_assets_schema",
            "ensure_fees_schema",
            "ensure_bursaries_schema",
            "ensure_requirements_schema",
            "ensure_other_income_schema",
            # "ensure_expense_categories",

            # people
            "ensure_users_table",
            "ensure_teachers_employees_schema",
            # "ensure_employees_table",
            # "ensure_teachers_table",

            # calendar & structure
            "ensure_academic_years_schema",
            "ensure_term_dates_schema",
            "ensure_streams_schema",
            "ensure_classes_schema",

            # misc
            "ensure_archived_students_table",
            "ensure_promotions_log_schema",
            "ensure_promotion_lock",
            "ensure_transport_schema",
            "ensure_report_comments_table",
            "ensure_grading_scale_schema",
            "ensure_comment_rules_schema",
            "ensure_join_columns",
            # "ensure_subject_papers_schema",
            # "ensure_teacher_subjects_table",
            # "ensure_requirements_has_year",
            # "ensure_fees_has_comment",
            # "ensure_fees_has_receipt_no",
        ]:
            ok, msg = _call_if_exists(fname, conn)
            ran.append(msg)

        # --- Seeders (run if present) ---
        for fname in [
            "seed_default_admin",
            "seed_default_classes",
            "seed_expense_categories",
        ]:
            ok, msg = _call_if_exists(fname, conn)
            ran.append(msg)

        # --- Optional guards / migrations ---
        for fname in [
            # "apply_schema_guards",
            "run_migrations",
        ]:
            ok, msg = _call_if_exists(fname, conn)
            ran.append(msg)

        # Commit pending changes
        try:
            conn.commit()
        except Exception:
            pass

    finally:
        try:
            conn.close()
        except Exception:
            pass

    # --- Log or print summary ---
    try:
        from flask import current_app as _app
        if _app and _app.logger:
            for line in ran:
                _app.logger.info(f"[bootstrap] {line}")
    except Exception:
        for line in ran:
            print(f"[bootstrap] {line}")
            
# Fallback TERMS if not defined elsewhere
try:
    TERMS  # keep your global TERMS if it exists
except NameError:
    TERMS = ('Term 1', 'Term 2', 'Term 3')

TERM_NO = {'Term 1': 1, 'Term 2': 2, 'Term 3': 3}

def _class_options():
    """Return distinct class names for the filter dropdown."""
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT class_name
            FROM students
            WHERE class_name IS NOT NULL
            ORDER BY class_name
        """)
        rows = cur.fetchall() or []
        # rows is list of tuples when dictionary=False (default)
        if rows and isinstance(rows[0], dict):
            return [r.get("class_name") for r in rows if r.get("class_name")]
        return [r[0] for r in rows if r and r[0]]
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()

def _active_year_term():
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)  # << important
        cur.execute("SELECT year, current_term FROM academic_years WHERE is_active=1 LIMIT 1")
        row = cur.fetchone()
        return (int(row["year"]), row["current_term"]) if row else (None, None)
    finally:
        try: cur.close()
        except: pass
        conn.close()


def norm_class(s: str | None) -> str | None:
    if not s:
        return None

    raw = str(s).strip()
    u = raw.upper().replace(" ", "")

    # canonical map
    canon = {
        "BABY": "Baby",
        "MIDDLE": "Middle",
        "TOP": "Top",
        "P1": "P1",
        "P2": "P2",
        "P3": "P3",
        "P4": "P4",
        "P5": "P5",
        "P6": "P6",
        "P7": "P7",
    }

    # direct match
    if u in canon:
        return canon[u]

    # heuristics: P3A, P3-WEST, etc.
    if u.startswith("P") and len(u) >= 2 and u[1].isdigit():
        n = int(u[1])
        if 1 <= n <= 7:
            return f"P{n}"

    # heuristics: BABYCLASS, MIDDLECLASS, TOPCLASS
    for k in ("BABY", "MIDDLE", "TOP"):
        if u.startswith(k):
            return canon[k]

    return None
    
def normalize_class_enum(value):
    """
    Take a raw class label from promotions_log and
    return a safe ENUM value for students.class_name,
    or None if we can't recognise it.
    """
    if not value:
        return None

    v = str(value).strip().upper()

    # already a clean enum like P3, Baby, etc.
    if v in CLASS_ENUM_SET:
        return v

    # Heuristics for 'P3 A', 'P3 WEST', etc.
    if v.startswith("P") and len(v) >= 2 and v[1].isdigit():
        cand = v[:2] # 'P3'
        if cand in CLASS_ENUM_SET:
            return cand

    # Heuristics for 'Baby A'
    if v.startswith("KG") and len(v) >= 3 and v[2].isdigit():
        cand = v[:3] # 'Baby'
        if cand in CLASS_ENUM_SET:
            return cand

    # Not recognised
    return None


def norm_section(value: str | None) -> str | None:
    if not value:
        return None
    s = value.strip().lower()
    if s in ("day", "d"):
        return "Day"
    if s in ("boarding", "board", "b"):
        return "Boarding"
    return None  # unknowns rejected


def class_expected_amount(student, conn):
    """
    Return the configured fee for a student based STRICTLY on
    (class_name, section). No fallback to NULL section.
    """
    if not student:
        return 0.0
    cls = (student.get("class_name") or "").strip()
    sec = (student.get("section") or "").strip()
    if not cls or not sec:
        return 0.0

    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT amount
          FROM class_fees
         WHERE class_name = %s
           AND LOWER(section) = LOWER(%s)
         LIMIT 1
        """,
        (cls, sec)
    )
    row = cur.fetchone()
    cur.close()
    return float(row[0]) if row and row[0] is not None else 0.0
#============================WORKPLANS==================================

def current_employee_id() -> int | None:
    """
    Returns employees.id of the logged-in user, or None if not linked.
    Works with either session['employee_id'] or the older session['staff_id'].
    """
    return session.get("employee_id") or session.get("staff_id")


# helper for simple word wrapping in PDF
def _wrap_text(text: str, max_len: int) -> list[str]:
    """
    Very small helper to wrap long text for PDF export.
    Splits on spaces; not perfect but good enough.
    """
    words = text.split()
    if not words:
        return []
    lines = []
    line = words[0]
    for w in words[1:]:
        if len(line) + 1 + len(w) > max_len:
            lines.append(line)
            line = w
        else:
            line += " " + w
    lines.append(line)
    return lines 
    
#======================TIMETABLE=================================
def extract_upload_blob(file_storage):
    """
    Validate a single uploaded document (PDF/image) and return:
       (filename, mime, blob_bytes)

    Raises ValueError for invalid uploads.
    """
    if not file_storage or file_storage.filename == "":
        raise ValueError("No file selected.")

    filename = secure_filename(file_storage.filename)
    if "." not in filename:
        raise ValueError("Invalid file (no extension).")

    ext = filename.rsplit(".", 1)[1].lower()
    if ext not in ALLOWED_DOC_EXTS:
        raise ValueError("Only PDF or image files are allowed.")

    mime = file_storage.mimetype or "application/octet-stream"
    blob = file_storage.read()
    if not blob:
        raise ValueError("Uploaded file is empty.")

    return filename, mime, blob

# ====Students' Expected fee displayed in students finance report======


def _table_exists(conn, table_name: str) -> bool:
    """
    Returns True if the given table exists in the current database.
    Works safely with dictionary=True cursors.
    """
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COUNT(1) AS cnt
        FROM information_schema.tables
        WHERE table_schema = DATABASE()
          AND table_name = %s
    """, (table_name,))
    row = cur.fetchone()
    cur.close()
    return bool(row and (row.get("cnt") or 0) > 0)


def _index_exists(conn, table: str, index_name: str) -> bool:
    """
    Returns True if the given index exists on `table`.
    Works with dictionary=True cursors.
    """
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COUNT(1) AS cnt
        FROM information_schema.statistics
        WHERE table_schema = DATABASE()
          AND table_name = %s
          AND index_name = %s
    """, (table, index_name))  # <-- use `table`, not `table_name`
    row = cur.fetchone()
    cur.close()
    return bool(row and (row.get("cnt") or 0) > 0)


def column_exists(conn, table_name: str, column_name: str) -> bool:
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT 1
        FROM information_schema.columns
        WHERE table_schema = DATABASE()
          AND table_name = %s
          AND column_name = %s
        LIMIT 1
        """,
        (table_name, column_name)
    )
    exists = cur.fetchone() is not None
    cur.close()
    return exists


def _norm_section(val: str) -> str:
    s = (val or "").strip().lower()
    if s in ("day", "d"):
        return "Day"
    if s in ("boarding", "board", "b"):
        return "Boarding"
    return ""


def norm_sex(v: str | None) -> str | None:
    s = (v or "").strip().upper()
    if s in ("M", "MALE"):
        return "M"
    if s in ("F", "FEMALE"):
        return "F"
    return None


def norm_stream(v: str | None) -> str | None:
    s = (v or "").strip().upper()
    return s or "A"


def get_active_academic_year():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT year, current_term FROM academic_years WHERE is_active = 1 LIMIT 1"
    )
    row = cur.fetchone()
    cur.close()
    conn.close

    if not row:
        # Fallback if no active row exists
        from datetime import datetime
        y = int(datetime.now().strftime("%Y"))
        return {"year": y, "current_term": "Term 1", "term": "Term 1"}

    ct = row["current_term"]
    return {"year": int(row["year"]), "current_term": ct, "term": ct}


def upsert_admin_user():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("SHOW COLUMNS FROM users")
    cols = [r[0].lower() for r in cur.fetchall()]

    if "password_hash" not in cols:
        cur.execute("ALTER TABLE users ADD COLUMN password_hash VARCHAR(150)")
    if "status" not in cols:
        cur.execute(
            "ALTER TABLE users ADD COLUMN status VARCHAR(50) DEFAULT 'active'")
    if "role" not in cols:
        cur.execute(
            "ALTER TABLE users ADD COLUMN role VARCHAR(20) DEFAULT 'admin'")

    cur.execute("SELECT id FROM users WHERE username = %s", ("admin",))
    row = cur.fetchone()
    if row:
        cur.execute("""
            UPDATE users
               SET password_hash = %s, role = 'admin', status = 'active'
             WHERE id = %s
        """, (generate_password_hash("admin123"), row[0]))
    else:
        cur.execute("""
            INSERT INTO users (username, password_hash, role, status)
            VALUES (%s, %s, 'admin', 'active')
        """, ("admin", generate_password_hash("admin123")))

    conn.commit()
    cur.close()
    conn.close()


def hp_resolve_student_id(conn, student_number):
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id FROM students WHERE student_number = %s",
                (student_number,))
    row = cur.fetchone()
    cur.close
    return row["id"] if row else None


# --- Role helpers ---
def _norm_role(val):
    """Return a canonical, lowercase role string."""
    return (str(val or "").strip().lower())


def require_login(f):
    @wraps(f)
    def _inner(*args, **kwargs):
        if "user_id" not in session:
            flash("Please login.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return _inner


def require_role(*roles):
    # normalize decorator inputs once
    wanted = tuple(_norm_role(r) for r in roles if r)

    def wrapper(f):
        @wraps(f)
        def inner(*args, **kwargs):
            if "user_id" not in session or "role" not in session:
                flash("Please login.", "warning")
                return redirect(url_for("login"))

            srole = _norm_role(session.get("role"))
            if wanted and srole not in wanted:
                flash("Access denied.", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return inner
    return wrapper


# allowed roles
ALLOWED_ROLES = (
    "admin", "bursar", "teacher", "headteacher",
    "director", "clerk", "dos", "deputyheadteacher", "classmanager"
)

#  FIX EXISTING FEES==============================================================

# utils/fees_fix.py

# ---- helpers ---------------------------------------------------------------

TERM_ORDER = {"Term 1": 1, "Term 2": 2, "Term 3": 3,
              "1": 1, "2": 2, "3": 3}  # tolerate numeric terms in DB


def _term_rank(term: str) -> int:
    t = (term or '').strip().lower()
    return 1 if t == 'term 1' else 2 if t == 'term 2' else 3 if t == 'term 3' else 0


def _fetchone(cur, sql, params=()):
    row = cur.execute(sql, params).fetchone()
    return dict(row) if row else None


def _fetchval(cur, sql, params=(), default=None):
    row = _fetchone(cur, sql, params)
    if not row:
        return default
    # return first column’s value
    return next(iter(row.values()), default)

# ---- main fix --------------------------------------------------------------


# ======================================
# register student & photo
# ======================================

# ======================================
# register student & photo
# ======================================



def save_student_photo(
    file_storage,
    existing_path: str | None = None,
    existing_blob: bytes | None = None,
    existing_mime: str | None = None,
) -> tuple[str | None, bytes | None, str | None]:
    """
    Save new student photo with validation, resize and compression.

    - If no new file is sent -> keep existing (path, blob, mime).
    - If a new file is sent -> delete old file (if any), then save new one.
    - Always store a resized/compressed copy in the DB as BLOB.

    Returns (relative_path, blob_bytes, mime).
    Raises ValueError for invalid files.
    """

    # --- no new upload: keep everything as is ---
    if not file_storage or file_storage.filename == "":
        return existing_path, existing_blob, existing_mime

    # ---------- extension check ----------
    filename = secure_filename(file_storage.filename)
    if "." not in filename:
        raise ValueError("Invalid photo file (no extension).")

    ext = filename.rsplit(".", 1)[1].lower()
    allowed_exts = set(ALLOWED_PHOTO_EXTS) | {"heic"}
    if ext not in allowed_exts:
        raise ValueError("Invalid photo type. Allowed: JPG, JPEG, HEIC, PNG.")

    # ---------- mimetype check ----------
    mimetype = (file_storage.mimetype or "").lower()

    if mimetype and ext in ("jpg", "jpeg", "png"):
        if mimetype not in ALLOWED_PHOTO_MIMES:
            raise ValueError("Invalid photo mimetype. Allowed: JPEG, PNG.")
    # For HEIC we are lenient – different browsers send different mimetypes

    # ---------- size check ----------
    size = getattr(file_storage, "content_length", None)
    if size is not None and size > MAX_PHOTO_SIZE:
        raise ValueError("Photo is too large. Max size is 5 MB.")

    # ---------- delete old file, if any ----------
    if existing_path:
        old_full = os.path.join(app.root_path, existing_path)
        if os.path.exists(old_full):
            try:
                os.remove(old_full)
            except Exception:
                pass

    # ---------- build new filename on disk ----------
    unique_name = f"{int(time.time())}_{filename}"
    full_path = os.path.join(PHOTO_UPLOAD_FOLDER, unique_name)

    blob_bytes: bytes

    # If Pillow is missing OR it's HEIC, just save raw bytes, then read them back
    if Image is None or ext == "heic":
        file_storage.stream.seek(0)
        data = file_storage.read()
        blob_bytes = data
        # also keep a copy on disk
        with open(full_path, "wb") as f:
            f.write(data)
        # best guess for mime
        if not mimetype:
            mimetype = "image/heic" if ext == "heic" else "application/octet-stream"
    else:
        # Pillow path: resize + compress, use resulting bytes both on disk and in DB
        try:
            file_storage.stream.seek(0)
            img = Image.open(file_storage.stream)

            if img.mode not in ("RGB", "RGBA"):
                img = img.convert("RGB")

            img.thumbnail((600, 600))

            buf = BytesIO()
            if ext in ("jpg", "jpeg"):
                img.save(buf, format="JPEG", quality=75, optimize=True)
                mimetype = "image/jpeg"
            else:
                img.save(buf, format="PNG", optimize=True)
                mimetype = "image/png"

            blob_bytes = buf.getvalue()

            with open(full_path, "wb") as f:
                f.write(blob_bytes)

        except Exception:
            # As a last resort: save whatever came in
            file_storage.stream.seek(0)
            data = file_storage.read()
            blob_bytes = data
            with open(full_path, "wb") as f:
                f.write(data)
            if not mimetype:
                mimetype = "application/octet-stream"

    rel_path = PHOTO_SUBDIR +"/"+ unique_name
    return rel_path, blob_bytes, mimetype



#================================================================================================#

def fix_existing_fees(get_conn=None) -> int:
    """
    Recomputes expected_amount, bursary_amount, carried_forward for all fees.
    Can use get_conn() (factory) if provided; otherwise opens its own connection.
    """
    conn = get_conn() if get_conn is not None else get_db_connection()
    close_after = True  # we always close what we open here

    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM fees WHERE payment_type = 'school_fees'")
    fees = cur.fetchall()
    cur.close()

    updated = 0
    for fee in fees:
        sid, term, year = fee['student_id'], fee['term'], fee['year']

        # student section (day/boarding enforced)
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT class_name, level, section FROM students WHERE id = %s",
            (sid,)
        )
        student = cur.fetchone()
        cur.close()
        if not student:
            continue

        # expected class fee (prefer exact section)
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT amount FROM class_fees
            WHERE class_name = %s AND (section = %s OR section IS NULL)
            LIMIT 1
        """, (student['class_name'], student['section']))
        class_fee = cur.fetchone()
        expected = class_fee['amount'] if class_fee else 0.0

        # bursary for that period
        cur.execute("""
            SELECT COALESCE(SUM(amount),0) AS total
            FROM bursaries
            WHERE student_id = %s AND term = %s AND year = %s
        """, (sid, term, year))
        bursary = cur.fetchone()
        bursary_amount = float(
            bursary['total'] if bursary and bursary['total'] else 0.0)

        # previous period balance (closest prior)
        cur.execute("""
            SELECT expected_amount, bursary_amount, amount_paid
            FROM fees
            WHERE student_id = %s
              AND (year < %s OR (year = %s AND term <> %s))
              AND payment_type = 'school_fees'
            ORDER BY year DESC,
                     CASE LOWER(term)
                       WHEN 'term 3' THEN 3
                       WHEN 'term 2' THEN 2
                       WHEN 'term 1' THEN 1
                       ELSE 0
                     END DESC
            LIMIT 1
        """, (sid, year, year, term))
        prev = cur.fetchone()
        cur.close()

        carried = 0.0
        if prev:
            carried = (float(prev['expected_amount'] or 0.0)
                       - float(prev['bursary_amount'] or 0.0)
                       - float(prev['amount_paid'] or 0.0))
            carried = max(carried, 0.0)

        cur = conn.cursor(dictionary=True)
        cur.execute("""
            UPDATE fees
               SET expected_amount=%s,
                   bursary_amount=%s,
                   carried_forward=%s
             WHERE id=%s
        """, (expected, bursary_amount, carried, fee['id']))
        cur.close()
        updated += 1

    conn.commit()
    if close_after:
        conn.close()
    return updated


def generate_student_number(conn, *_) -> str:
    """
    Format: STD-YYYY-NNN (e.g. STD-2025-001)
    Uses the 4-digit current year and a per-year running sequence.
    """
    year = datetime.now().year

    # Use a plain cursor so fetchone()[0] always works
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT COUNT(*) FROM students WHERE SUBSTRING(student_number, 5, 4) = %s",
            (str(year),)
        )
        count = cur.fetchone()[0] or 0
        seq = int(count) + 1
        return f"STD-{year}-{seq:03d}"
    finally:
        cur.close()


def generate_fees_code(conn) -> str:
    """
    Format: FC-XXXXXXXX (8 hex chars). Loops until unique in students.fees_code.
    """
    cur = conn.cursor()
    try:
        while True:
            cand = f"FC-{secrets.token_hex(4).upper()}"  # 8 hex chars
            cur.execute(
                "SELECT 1 FROM students WHERE fees_code = %s LIMIT 1", (cand,))
            if cur.fetchone() is None:
                return cand
    finally:
        cur.close()


def log_action(user_id, action):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "INSERT INTO audit_trail (user_id, action, timestamp) VALUES (%s, %s, NOW())",
        (user_id, action),
    )
    conn.commit()
    cur.close()
    conn.close()


# ---USB PRINTERS WITHOUT DEVICE ID---


# --- USERS schema + default admin -----------------------------------------


def ensure_students_table(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS students (
            id INT AUTO_INCREMENT PRIMARY KEY,
            first_name VARCHAR(50) NOT NULL,
            Middle_name VARCHAR(50),
            last_name VARCHAR(50) NOT NULL,
            sex VARCHAR(10),
            year_completed INT,
            class_name ENUM('Baby','Middle','Top','P1','P2','P3','P4','P5','P6','P7') NOT NULL DEFAULT 'Baby',
            stream VARCHAR(15),
            section ENUM('Day','Boarding') DEFAULT 'Day',
            combination VARCHAR(20),
            fees_amount DOUBLE,
            student_number VARCHAR(20) UNIQUE,
            academic_year_id INT,
            year_of_joining VARCHAR(10),
            term_joined VARCHAR(15),
            date_joined VARCHAR(15),
            cumulative_average VARCHAR(10),
            cumulative_grade VARCHAR(10),
            cumulative_comment VARCHAR(100),
            residence VARCHAR(50),
            house VARCHAR(15),
            parent_name VARCHAR(50),
            parent2_name VARCHAR(50),
            parent_contact VARCHAR(30),
            parent2_contact VARCHAR(30),
            fees_code VARCHAR(20),
            parent_email VARCHAR(50),
            archived TINYINT NOT NULL DEFAULT 0,
            current_class VARCHAR(15),
            status ENUM('active','dropped','left','completed') NOT NULL DEFAULT 'active'
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    for stmt in (
        "CREATE INDEX ix_students_class ON students(class_name)",
        "CREATE INDEX ix_students_sn ON students(student_number)",
        "CREATE INDEX ix_students_order ON students(last_name, first_name, Middle_name)",
    ):
        try:
            cur.execute(stmt)
        except Exception:
            pass
    conn.commit()
    cur.close()


def ensure_subjects_table(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS subjects (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(30) NOT NULL,
            code VARCHAR(10),
            UNIQUE KEY uq_subjects_name (name)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()

# put this ONCE (near your other ensure_* fns) and ABOVE marks_hub()

# Kindergarten Ensures
# ------------------ SCHEMA ENSURE KINDERGARTEN ------------------ #

def ensure_robotics_checklist_schema(conn):
    """Ensure robotics_checklist table exists (safe to call often)."""
    ddl = """
    CREATE TABLE IF NOT EXISTS robotics_checklist (
      id INT NOT NULL AUTO_INCREMENT,
      student_id INT NOT NULL,
      term VARCHAR(16) NOT NULL,
      year INT NOT NULL,
      area VARCHAR(64) NOT NULL,
      area_code VARCHAR(8) NOT NULL,
      section VARCHAR(64) NOT NULL,
      label VARCHAR(64) NOT NULL,
      competence TEXT NOT NULL,
      tick TINYINT(1) NOT NULL DEFAULT 0,
      remark VARCHAR(150) DEFAULT NULL,
      created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
      updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
      PRIMARY KEY (id),
      UNIQUE KEY uq_rc_student_item (
          student_id, term, year,
          area, section, label,
          competence(120)
      ),
      KEY idx_rc_student_period (student_id, term, year),
      CONSTRAINT fk_rc_student
        FOREIGN KEY (student_id) REFERENCES students(id)
        ON DELETE CASCADE
    ) ENGINE=InnoDB
      DEFAULT CHARSET=utf8mb4
      COLLATE=utf8mb4_0900_ai_ci;
    """
    cur = conn.cursor()
    cur.execute(ddl)
    conn.commit()
    cur.close()
    
def ensure_robotics_checklist_meta_schema(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS robotics_checklist_meta (
            id INT NOT NULL AUTO_INCREMENT,
            student_id INT NOT NULL,
            term VARCHAR(16) NOT NULL,
            year INT NOT NULL,
            overall_remark TEXT NULL,
            special_communication TEXT NULL,
            next_term_begin DATE NULL,
            next_term_end DATE NULL,
            school_fees VARCHAR(64) NULL,
            school_fees_daycare VARCHAR(64) NULL,
            updated_at DATETIME NOT NULL
             DEFAULT CURRENT_TIMESTAMP
            ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            UNIQUE KEY uq_meta (student_id, term, year),
            CONSTRAINT fk_meta_student
            FOREIGN KEY (student_id) REFERENCES students(id)
            ON DELETE CASCADE
        ) ENGINE=InnoDB
        DEFAULT CHARSET=utf8mb4
        COLLATE=utf8mb4_0900_ai_ci;
    """)
    cur.close()
    



def ensure_recommendation_options_schema(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS robotics_recommendation_options (
            id INT NOT NULL AUTO_INCREMENT,
            text VARCHAR(255) NOT NULL,
            is_active TINYINT(1) NOT NULL DEFAULT 1,
            PRIMARY KEY (id),
            UNIQUE KEY uq_text (text)
        ) ENGINE=InnoDB
          DEFAULT CHARSET=utf8mb4
          COLLATE=utf8mb4_0900_ai_ci;
    """)
    # seed with your defaults if empty
    cur.execute("SELECT COUNT(*) FROM robotics_recommendation_options")
    (cnt,) = cur.fetchone()
    if cnt == 0:
        for txt in RECOMMENDATION_OPTIONS:
            cur.execute(
                "INSERT IGNORE INTO robotics_recommendation_options(text) VALUES (%s)",
                (txt,),
            )
    cur.close()

def _fetch_checklist_meta(student_id: int, term: str, year: int) -> dict:
    """
    Returns robotics_checklist_meta for this learner/term/year if it exists,
    otherwise falls back to the most recent row for that learner.
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # 1) exact term/year (preferred)
    cur.execute("""
        SELECT overall_remark, special_communication,
               next_term_begin, next_term_end,
               school_fees, school_fees_daycare
        FROM robotics_checklist_meta
        WHERE student_id=%s AND term=%s AND year=%s
        LIMIT 1
    """, (student_id, term, year))
    meta = cur.fetchone()

    # 2) fallback: most recent meta row for this learner
    if not meta:
        cur.execute("""
            SELECT overall_remark, special_communication,
                   next_term_begin, next_term_end,
                   school_fees, school_fees_daycare
            FROM robotics_checklist_meta
            WHERE student_id=%s
            ORDER BY year DESC, term DESC, id DESC
            LIMIT 1
        """, (student_id,))
        meta = cur.fetchone()

    cur.close()
    conn.close()
    return meta or {}


def get_recommendation_options(conn):
    """Dynamic dropdown options for overall remark."""
    ensure_recommendation_options_schema(conn)
    cur = conn.cursor()
    cur.execute("""
        SELECT text
        FROM robotics_recommendation_options
        WHERE is_active=1
        ORDER BY text
    """)
    rows = [r[0] for r in cur.fetchall()]
    cur.close()
    # fallback to constants if somehow empty
    return rows or list(RECOMMENDATION_OPTIONS)

# ------------------ SCHEMA ENSURE KINDERGARTEN ------------------ #

# ------------------ KINDERGARTEN HELPERS ------------------ #

def _fetch_student_for_checklist(student_id: int):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Try with photo; if column missing, fall back gracefully
        try:
            cur.execute("""
                SELECT id, student_number,
                       first_name, COALESCE(Middle_name,'') AS middle_name,
                       last_name, class_name, stream, section,
                       date_joined, term_joined, year_of_joining,
                       house, photo
                FROM students
                WHERE id=%s
            """, (student_id,))
        except Exception:
            cur.execute("""
                SELECT id, student_number,
                       first_name, COALESCE(Middle_name,'') AS middle_name,
                       last_name, class_name, stream, section,
                       date_joined, term_joined, year_of_joining,
                       house
                FROM students
                WHERE id=%s
            """, (student_id,))

        return cur.fetchone()
    finally:
        cur.close()
        conn.close()



def _fetch_checklist_items():
    """
    Static checklist definition.
    Returns list of tuples:
      (area, area_code, section, label, competence)
    """
    items = []

    # GENERAL KNOWLEDGE
    items += [
        ("GENERAL KNOWLEDGE (G)", "G", "SOCIAL DEVELOPMENT", "",
         "I can identify and colour the Garden tools."),
        ("GENERAL KNOWLEDGE (G)", "G", "SOCIAL DEVELOPMENT", "",
         "I have the ability to Copy, Draw and match the garden tools."),
        ("GENERAL KNOWLEDGE (G)", "G", "SOCIAL DEVELOPMENT", "",
         "I can fill in the missing sounds and naming the given tool picture."),
        ("GENERAL KNOWLEDGE (G)", "G", "SOCIAL DEVELOPMENT", "",
         "I can identify, colour and paste domestic animals like Pig, Goat, Cow, Sheep, Dog, Cat"),
        ("GENERAL KNOWLEDGE (G)", "G", "SOCIAL DEVELOPMENT", "",
         "I can name, categorize animal according to their use. i.e. dog for protection,  cow for meat etc."),
        ("GENERAL KNOWLEDGE (G)", "G", "SOCIAL DEVELOPMENT", "",
         "I have the ability to copy, draw, and name the domestic animals."),
        ("GENERAL KNOWLEDGE (G)", "G", "SOCIAL DEVELOPMENT", "",
         "I can orally name common domestic birds like Cock, Duck Turkey etc."),
        ("GENERAL KNOWLEDGE (G)", "G", "SOCIAL DEVELOPMENT", "",
         "I have the ability to colour birds and differentiate them from animals."),
        ("GENERAL KNOWLEDGE (G)", "G", "SOCIAL DEVELOPMENT", "",
         "I can copy, draw and match animal names to their pictures."),
        ("GENERAL KNOWLEDGE (G)", "G", "SOCIAL DEVELOPMENT", "",
         "I can identify and colour important people in my community."),
        ("GENERAL KNOWLEDGE (G)", "G", "SOCIAL DEVELOPMENT", "",
         "I can match important people to their places of work."),
        ("GENERAL KNOWLEDGE (G)", "G", "SOCIAL DEVELOPMENT", "",
         "I can copy and draw important people and their tools."),
        ("GENERAL KNOWLEDGE (G)", "G", "HEALTH HABITS", "",
         "I can identify, colour and match the body organs."),
        ("GENERAL KNOWLEDGE (G)", "G", "HEALTH HABITS", "",
         "I can correctly match words to pictures, draw and mention body organs like the Tongue, nose, ears, eyes, and their uses."),
        ("GENERAL KNOWLEDGE (G)", "G", "HEALTH HABITS", "",
         "I can identify and colour the parts of the head i.e hair, face."),
        ("GENERAL KNOWLEDGE (G)", "G", "HEALTH HABITS", "",
         "I have the ability to draw, write and match the parts of the head"),
        ("GENERAL KNOWLEDGE (G)", "G", "HEALTH HABITS", "",
         "I can identify and colour things used to clean parts of the head like a combo, toothbrush etc."),
        ("GENERAL KNOWLEDGE (G)", "G", "HEALTH HABITS", "",
         "I can copy, draw and match things used to clean parts of the head"),
        ("GENERAL KNOWLEDGE (G)", "G", "HEALTH HABITS", "",
         "I can identify and colour food and fruits like fish, banana, mangoes"),
        ("GENERAL KNOWLEDGE (G)", "G", "HEALTH HABITS", "",
         "I have the ability to copy, draw and match food and fruits"),
        ("GENERAL KNOWLEDGE (G)", "G", "HEALTH HABITS", "",
         "I can tell the sources of food and fruits"),
    ]

    # LANGUAGE (examples – extend as needed)
    items += [
        ("LANGUAGE (L)", "L", "LISTENING", "",
         "I can listen and identify Bk3 sounds and pictures."),
        ("LANGUAGE (L)", "L", "LISTENING", "",
         "I have the ability to listen and demonstrate the sound’s action."),
        ("LANGUAGE (L)", "L", "LISTENING", "",
         "I can listen and follow simple instructions from the teacher"),
        ("LANGUAGE (L)", "L", "SPEAKING", "",
         "I can recite all vowel sounds a,e,i,o,u"),
        ("LANGUAGE (L)", "L", "SPEAKING", "",
         "I can recite different syllables ba, be, bi, bo, bu"),
        ("LANGUAGE (L)", "L", "SPEAKING", "",
         "I can sing all songs of the taught sounds"),
        ("LANGUAGE (L)", "L", "SPEAKING", "",
         "I can blend single sounds to form words"),
        ("LANGUAGE (L)", "L", "SPEAKING", "",
         "I can answer simple questions"),
        ("LANGUAGE (L)", "L", "SPEAKING", "",
         "I can use a lot of words to tell you about something I have done"),
        ("LANGUAGE (L)", "L", "SPEAKING", "",
         "I can retell short phrases and stories"),
        ("LANGUAGE (L)", "L", "READING", "",
         "I can sound most of Bk1 – Bk3 sounds (s,a,t,I,p,n), (c,k,e,h,r,m,d), (g,o,u,l,f,b)."),
        ("LANGUAGE (L)", "L", "READING", "",
         "I have the ability to read and form words from syllables. ba- bag."),
        ("LANGUAGE (L)", "L", "READING", "",
         "I can read simple sentences. (It is a cat. This is a mat..)"),
        ("LANGUAGE (L)", "L", "READING", "",
         "I can read and understand simple sentence"),
        ("LANGUAGE (L)", "L", "WRITING", "",
         "I can write sounds without the help of the teacher (s,a,t,i,p,n), (c,k,e,h,r,m), (d,g,o,u,l,f)"),
        ("LANGUAGE (L)", "L", "WRITING", "",
         "I have the ability to write and match sounds to their pictures"),
        ("LANGUAGE (L)", "L", "WRITING", "",
         "I have the ability to draw and match words to pictures."),
        ("LANGUAGE (L)", "L", "WRITING", "",
         "I have the ability to segment words using syllables"),
        ("LANGUAGE (L)", "L", "WRITING", "",
         "I can identify, name, match and draw thing in my classroom"),
        ("LANGUAGE (L)", "L", "WRITING", "",
         "I can state the uses of things in my classroom."),
        ("LANGUAGE (L)", "L", "WRITING", "",
         "I can identify and differentiate the colour, size and shape of objects round me."),
        ("LANGUAGE (L)", "L", "WRITING", "",
         "I can identify and tell the position of an object (on, under in, behind, over, near, above etc)"),
        ("LANGUAGE (L)", "L", "WRITING", "",
         "I can match drawn pictures to their prepositions "),
        ("LANGUAGE (L)", "L", "WRITING", "",
         "I can copy sentences from the white board"),
        ("LANGUAGE (L)", "L", "WRITING", "",
         "I can hold the writing tool properly and firmly."),
         
        
    ]

    # NUMBERS (examples)
    items += [
        ("NUMBERS (N)", "N", "ARITHMETICS", "",
         "I can count numbers orally 1-20"),
        ("NUMBERS (N)", "N", "ARITHMETICS", "",
         "I can count and recognize numbers 1-15"),
        ("NUMBERS (N)", "N", "ARITHMETICS", "",
         "I can correctly write numbers 1-15"),
        ("NUMBERS (N)", "N", "ARITHMETICS", "",
         "I can count and draw objects for the number"),
        ("NUMBERS (N)", "N", "ARITHMETICS", "",
         "I can write and tell the number before, after and in- between"),
        ("NUMBERS (N)", "N", "ARITHMETICS", "",
         "I can identify and match number figures to their number names"),
        ("NUMBERS (N)", "N", "ARITHMETICS", "",
         "I can add objects "),
        ("NUMBERS (N)", "N", "ARITHMETICS", "",
         "I can count and write the correct number on the abacus. "),
        ("NUMBERS (N)", "N", "ARITHMETICS", "",
         "I can form and classify sets according to items"),
        ("NUMBERS (N)", "N", "GEOMETRY", "",
         "I know my shapes Circle, Square, Triangle, Star, Oval etc"),
        ("NUMBERS (N)", "N", "GEOMETRY", "",
         "I can identify and match the same shapes"),
        ("NUMBERS (N)", "N", "GEOMETRY", "",
         "I can read the given shape name and draw"),

    ]

    # PHYSICAL DEVELOPMENT (examples)
    items += [
        ("PHYSICAL DEVELOPMENT (PD)", "PD", "FINE MOTOR SKILLS", "",
         "I have the ability to track the object with the eye and manipulate it with a hand."),
        ("PHYSICAL DEVELOPMENT (PD)", "PD", "FINE MOTOR SKILLS", "",
         "I have the ability to participate in basic activities like dancing, clapping, waving."),
        ("PHYSICAL DEVELOPMENT (PD)", "PD", "FINE MOTOR SKILLS", "",
         "I have the ability to use the fingers and hands to manipulate small objects i.e picking a bottle top/lid."),
        ("PHYSICAL DEVELOPMENT (PD)", "PD", "FINE MOTOR SKILLS", "",
         "I have the ability to balance the body in the activities i.e jumping, hopping."),
        ("PHYSICAL DEVELOPMENT (PD)", "PD", "LARGE MOTOR SKILLS", "",
         "I have the ability to climb up and down stairs or ladder and navigate obstacles."),
    ]

    return items


def _fetch_saved_checklist_map(student_id: int, term: str, year: int):
    """
    Returns:
      {(area, section, label, competence): {'tick': 0/1, 'remark': str}}
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    saved = {}
    try:
        cur.execute("""
            SELECT area, section, label, competence,
                   COALESCE(tick,0) AS tick,
                   COALESCE(remark,'') AS remark
            FROM robotics_checklist
            WHERE student_id=%s AND term=%s AND year=%s
        """, (student_id, term, year))
        for r in (cur.fetchall() or []):
            key = (r["area"], r["section"], r["label"], r["competence"])
            saved[key] = {
                "tick": int(r["tick"]),
                "remark": r["remark"],
            }
    finally:
        cur.close()
        conn.close()
    return saved


def _resolve_area_style(area: str):
    """
    Return (style_dict, canonical_key) by matching AREA_STYLES keys
    case-insensitively as substrings of the provided area label.
    Falls back to {} and None if no match.
    """
    a = (area or "").strip().lower()
    for k, v in AREA_STYLES.items():
        if k.lower() in a:
            return v, k # style, canonical key
    # try exact as last resort
    if area in AREA_STYLES:
        return AREA_STYLES[area], area
    return {}, None


def _compute_area_ranges_from_meta(row_meta_page, previous_area=None):
    """
    Build {area: [(start,end), ...]} for a single page.
    - row_meta_page[0] is the header.
    - If the page starts with section/skill (no 'area' yet), we use previous_area.
    """
    area_rows = {}
    eff_area = previous_area # carry-over when page starts mid-area

    for i, meta in enumerate(row_meta_page):
        if i == 0:
            continue # skip header row in this slice

        t = meta.get("type")

        if t == "area":
            eff_area = meta.get("area")
            area_rows.setdefault(eff_area, [])
            # note: the 'area' row itself has no visible cells, so we don't add i
            continue

        if t in ("section", "skill") and eff_area:
            area_rows.setdefault(eff_area, []).append(i)

    # collapse contiguous indices -> (start, end) ranges
    area_ranges = {}
    for area, rows in area_rows.items():
        if not rows:
            continue
        rows = sorted(set(rows))
        start = prev = rows[0]
        ranges = []
        for r in rows[1:]:
            if r == prev + 1:
                prev = r
            else:
                ranges.append((start, prev))
                start = prev = r
        ranges.append((start, prev))
        area_ranges[area] = ranges

    return area_ranges




def _draw_area_icons_vertical(
    c,
    left,
    table_y,
    row_heights,
    col_area_w,
    area_ranges,
    start_index=0, # lets us keep alternation across pages
    palette=None, # pass custom colors or use defaults below
):
    """
    Draw coloured vertical bands + icon + vertical label in the 'Area' column.

    - Primary colour comes from AREA_STYLES['bg'] (per area),
      falling back to an alternating palette (header blues) if missing.
    - Icon is rotated 90°, centered horizontally at ~30% of the band width.
    - Label is vertical near the right edge of the band (white, bold).
    - Returns updated band index so the caller can preserve alternation
      across pages.
    """
    # Fallback palette if not provided / COL_NAVY or COL_BLUE not defined
    try:
        default_palette = [COL_NAVY, COL_BLUE] # from your header config
    except NameError:
        default_palette = [colors.HexColor("#0b2a5b"), colors.HexColor("#1f66c5")]

    palette = palette or default_palette

    nrows = len(row_heights)
    table_height = sum(row_heights)

    # rowpositions[i] = top of row i (0 = header of this slice)
    rowpositions = [table_y + table_height]
    for rh in row_heights:
        rowpositions.append(rowpositions[-1] - rh)

    band_i = start_index

    for area, ranges in area_ranges.items():
        style, canon = _resolve_area_style(area)
        label_key = canon or area
        label_text = (style or {}).get("label", label_key)
        band_color = (style or {}).get("bg") or palette[band_i % len(palette)]

        for (start, end) in ranges:
            if not (0 <= start < nrows and 0 <= end < nrows and start <= end):
                continue

            y_top = rowpositions[start]
            y_bottom = rowpositions[end + 1]

            x = left
            w = col_area_w
            h = y_top - y_bottom

            # --- coloured band background ---
            c.saveState()
            c.setFillColor(band_color)
            c.setStrokeColor(band_color)
            c.rect(x, y_bottom, w, h, stroke=0, fill=1)
            c.restoreState()

            # --- centres ---
            y_center = y_bottom + h / 2.0
            icon_x = x + w * 0.30

            # dynamic icon size (kept square, rotated in helper)
            icon_size = min(14 * mm, h * 1.20)

            # Draw icon using vector graphics only
            _draw_area_icon(c, label_key, icon_x, y_center, icon_size)

            # --- vertical label near right edge ---
            label_x = x + w * 0.82
            c.saveState()
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 9)
            c.translate(label_x, y_center)
            c.rotate(90)
            txt = label_text if len(label_text) <= 25 else (label_text[:17] + "…")
            c.drawCentredString(0, 0, txt)
            c.restoreState()

            band_i += 1 # advance alternation per band segment

    return band_i


            



def _draw_area_icon(c, area, x, y, size):
    """
    Draw a rotated vector icon for the area (no image files).
    The canvas is rotated 90° so icons match the vertical card style.
    Icons assume a coloured background and are drawn in solid white.
    """
    key = (area or "").strip().lower()

    c.saveState()
    # centre and rotate 90° (to match vertical band)
    c.translate(x, y)
    c.rotate(90)

    c.setStrokeColor(colors.white)
    c.setFillColor(colors.white)
    c.setLineWidth(1.4)
    c.setLineJoin(1) # round joins
    c.setLineCap(1) # round caps

    r = size / 2.5 # base radius / scale

    # ===== KNOWLEDGE: solid open book with bookmark + small bulb =====
    if "knowledge" in key:
        cover_w = 2.1 * r
        cover_h = 1.3 * r

        # outer cover
        c.roundRect(-cover_w / 2, -cover_h / 2, cover_w, cover_h,
                    0.25 * r, stroke=1, fill=0)

        # inner pages
        c.roundRect(-cover_w / 2 + 0.2 * r,
                    -cover_h / 2 + 0.18 * r,
                    cover_w - 0.4 * r,
                    cover_h - 0.36 * r,
                    0.18 * r, stroke=1, fill=0)

        # spine
        c.line(0, -cover_h / 2 + 0.18 * r, 0, cover_h / 2 - 0.18 * r)

        # bookmark
        c.line(0.15 * r, cover_h / 2 - 0.18 * r,
               0.15 * r, 0)
        c.line(0.15 * r, 0, 0.05 * r, -0.2 * r)
        c.line(0.15 * r, 0, 0.25 * r, -0.2 * r)

        # bulb above book
        bulb_r = 0.35 * r
        bulb_y = cover_h / 2 + 0.4 * r
        c.circle(0, bulb_y, bulb_r, stroke=1, fill=0)
        # bulb base
        c.rect(-0.2 * r, bulb_y - 0.5 * bulb_r,
               0.4 * r, 0.25 * bulb_r, stroke=1, fill=1)
        # small rays
        for dx in (-0.9 * bulb_r, 0.9 * bulb_r):
            c.line(dx, bulb_y, dx * 1.25, bulb_y)
        c.line(0, bulb_y + bulb_r, 0, bulb_y + 1.35 * bulb_r)

    # ===== LANGUAGE: solid speech bubble with text lines =====
    elif "language" in key:
        bw = 2.2 * r
        bh = 1.4 * r

        # speech bubble body (filled)
        c.setFillColor(colors.white)
        c.setStrokeColor(colors.white)
        c.roundRect(-bw / 2, -bh / 2 + 0.1 * r,
                    bw, bh, 0.4 * r, stroke=1, fill=1)

        # tail as small triangle
        path = c.beginPath()
        path.moveTo(-0.3 * r, -bh / 2 + 0.1 * r)
        path.lineTo(-0.9 * r, -bh / 2 - 0.6 * r)
        path.lineTo(0.1 * r, -bh / 2 + 0.1 * r)
        path.close()
        c.drawPath(path, stroke=1, fill=1)

        # "text" lines inside bubble
        c.setLineWidth(1.0)
        text_left = -bw / 2 + 0.35 * r
        text_right = bw / 2 - 0.35 * r
        base_y = 0.2 * r
        gap = 0.35 * r
        for i in range(3):
            yy = base_y - i * gap
            # slightly shorter lines for nicer look
            inset = (0.0 if i == 0 else 0.15 * r)
            c.line(text_left + inset, yy,
                   text_right - inset, yy)

    # ===== NUMBERS: "123" inside a clean circular badge =====
    elif "number" in key or "numbers" in key:
        badge_r = 1.1 * r
        # circular badge (outline only, to keep "123" readable on bold bands)
        c.circle(0, 0, badge_r, stroke=1, fill=0)

        fs = max(6, int(size / 2.4))
        c.setFont("Helvetica-Bold", fs)
        # slightly raise text inside circle
        c.drawCentredString(0, -fs * 0.35, "123")

    # ===== PHYSICAL DEVELOPMENT: more solid running figure =====
    elif "physical" in key or "development" in key:
        # head
        head_r = 0.32 * r
        head_y = 0.8 * r
        c.circle(0, head_y, head_r, stroke=1, fill=0)

        # torso
        torso_top_y = head_y - 0.4 * r
        torso_bottom_y = -0.1 * r
        c.line(0, torso_top_y, 0, torso_bottom_y)

        # arms (one forward, one back)
        arm_y = torso_top_y - 0.1 * r
        c.line(0, arm_y, -0.9 * r, 0.3 * r) # back arm
        c.line(0, arm_y, 0.9 * r, 0.7 * r) # forward arm

        # legs (running / stride)
        hip_y = torso_bottom_y
        c.line(0, hip_y, -0.7 * r, -1.0 * r) # back leg
        c.line(0, hip_y, 0.8 * r, -0.8 * r) # front leg

        # small ground line for stability
        c.setLineWidth(1.0)
        c.line(-r, -1.05 * r, r, -1.05 * r)

    # ===== fallback: simple outlined circle =====
    else:
        c.circle(0, 0, r, stroke=1, fill=0)

    c.restoreState()

# ------------------ KINDERGARTEN HELPERS ------------------ #





def absolute_photo_path(photo_rel: str | None) -> str | None:
    """
    Turn whatever is stored in student['photo'] into a usable absolute path,
    working on both Windows and Linux.
    """
    if not photo_rel:
        return None
    # normalise Windows backslashes for Linux
    rel = photo_rel.replace("\\", "/").lstrip("/") # remove leading /
    return os.path.join(app.root_path, rel)




def ensure_record_score_table(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS record_score (
            id INT AUTO_INCREMENT PRIMARY KEY,
            student_id INT NOT NULL,
            subject_id INT NOT NULL,
            term VARCHAR(15) NOT NULL,
            initials VARCHAR(5),
            year INT NOT NULL,
            bot_mark DOUBLE,
            midterm_mark DOUBLE,
            eot_mark DOUBLE,
            holiday_mark DOUBLE,
            other_mark DOUBLE,
            ca_mark DOUBLE,
            average_mark DOUBLE,
            grade VARCHAR(10),
            comment VARCHAR(70),
            processed_on DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uq_record_score (student_id, subject_id, term, year),
            CONSTRAINT fk_rs_student FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE,
            CONSTRAINT fk_rs_subject FOREIGN KEY (subject_id) REFERENCES subjects(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def upgrade_record_score(conn):
    """
    Ensure record_score has columns: holiday_mark, other_mark, ca_mark.
    Uses MySQL SHOW COLUMNS; adds missing columns as DOUBLE (closest to REAL).
    """
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SHOW COLUMNS FROM record_score")
        # In MySQL, first column in the result is the column name
        existing = {row[0].lower() for row in cur.fetchall()}

        for col in ("holiday_mark", "other_mark", "ca_mark"):
            if col not in existing:
                cur.execute(
                    f"ALTER TABLE record_score ADD COLUMN {col} DOUBLE")
        conn.commit()
    finally:
        cur.close()


def ensure_results_table(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS results (
            id INT AUTO_INCREMENT PRIMARY KEY,
            student_id INT NOT NULL,
            subject_id INT NOT NULL,
            eot DOUBLE,
            total DOUBLE,
            grade VARCHAR(10),
            comment VARCHAR(50),
            initials VARCHAR(5),
            term VARCHAR(15) NOT NULL,
            year INT NOT NULL,
            UNIQUE KEY uq_results (student_id, subject_id, term, year),
            CONSTRAINT fk_res_student FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE,
            CONSTRAINT fk_res_subject FOREIGN KEY (subject_id) REFERENCES subjects(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def ensure_midterms_table(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS midterms (
            id INT AUTO_INCREMENT PRIMARY KEY,
            student_id INT NOT NULL,
            term VARCHAR(15) NOT NULL,
            year INT NOT NULL,
            assessment VARCHAR(64) NOT NULL,
            eng DOUBLE, mat DOUBLE, sci DOUBLE, sst DOUBLE,
            agg INT, total INT,
            UNIQUE KEY uq_mid (student_id, term, year, assessment),
            CONSTRAINT fk_mid_student FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def ensure_report_comments_table(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS report_comments (
            student_id INT NOT NULL,
            term VARCHAR(15) NOT NULL,
            year INT NOT NULL,
            teacher_comment VARCHAR(70),
            head_comment VARCHAR(70),
            PRIMARY KEY (student_id, term, `year`),
            CONSTRAINT fk_report_comments_student FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def ensure_reports_table(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS reports (
            id INT AUTO_INCREMENT PRIMARY KEY,
            student_id INT NOT NULL,
            class_name VARCHAR(15) NOT NULL,
            stream VARCHAR(15),
            subject_id INT NOT NULL,
            term VARCHAR(10) NOT NULL,
            year INT NOT NULL,
            average_mark DOUBLE,
            grade VARCHAR(10),
            comment VARCHAR(70),
            teacher_remark VARCHAR(70),
            headteacher_remark VARCHAR(70),
            teacher_id INT,
            bot_mark DOUBLE,
            midterm_mark DOUBLE,
            eot_mark DOUBLE,
            holiday_mark DOUBLE,
            other_mark DOUBLE,
            ca_mark DOUBLE,
            teacher_initial VARCHAR(16),
            processed_on DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uq_reports (student_id, subject_id, year, term),
            CONSTRAINT fk_rep_student FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE,
            CONSTRAINT fk_rep_subject FOREIGN KEY (subject_id) REFERENCES subjects(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def ensure_expense_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS expense_categories (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(30) NOT NULL UNIQUE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS expenses (
            id INT AUTO_INCREMENT PRIMARY KEY,
            description VARCHAR(100),
            amount DOUBLE NOT NULL,
            term VARCHAR(32),
            year INT,
            date_spent DATE DEFAULT (CURRENT_DATE),
            category_id INT,
            recorded_by VARCHAR(25),
            type VARCHAR(30),
            CONSTRAINT fk_exp_cat FOREIGN KEY (category_id) REFERENCES expense_categories(id) ON DELETE SET NULL,
            INDEX ix_expenses_date (date_spent)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def ensure_payroll_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS payroll (
            id INT AUTO_INCREMENT PRIMARY KEY,
            employee_id INT,
            teacher_id INT,
            term VARCHAR(10) NOT NULL,
            year INT NOT NULL,
            expected_salary DOUBLE NOT NULL,
            bonus DOUBLE DEFAULT 0,
            allowance DOUBLE DEFAULT 0,
            total DOUBLE NOT NULL,
            paid_amount DOUBLE DEFAULT 0,
            status ENUM('fully_paid','partially_paid','not_paid') NOT NULL DEFAULT 'not_paid',
            date_paid DATE DEFAULT (CURRENT_DATE),
            INDEX ix_payroll_term_year (term, year),
            CONSTRAINT fk_pay_teacher FOREIGN KEY (teacher_id) REFERENCES teachers(id),
            CONSTRAINT fk_pay_employee FOREIGN KEY (employee_id) REFERENCES employees(id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def ensure_class_fees_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS class_fees (
            id INT AUTO_INCREMENT PRIMARY KEY,
            class_name VARCHAR(15) NOT NULL,
            section ENUM('Day','Boarding') NOT NULL,
            level VARCHAR(25),
            amount DOUBLE NOT NULL,
            UNIQUE KEY uq_class_fees_class_section (class_name, section)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def ensure_classes_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS classes (
            id INT AUTO_INCREMENT PRIMARY KEY,
            class_name VARCHAR(15) NOT NULL,
            level VARCHAR(20),
            stream VARCHAR(15) NOT NULL,
            UNIQUE KEY uq_classes_class_stream (class_name, stream)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


# ---- seed: default classes ----

def seed_default_classes(conn):
    """Creates Baby–P7, stream A (primary) if missing, and ensures req index."""
    ensure_classes_schema(conn)

    cur = conn.cursor(dictionary=True)
    rows = [
        ("Baby", "Nursery", "A"),
        ("Middle", "Nursery", "A"),
        ("Top", "Nursery", "A"),
        ("P1", "Primary", "A"),
        ("P2", "Primary", "A"),
        ("P3", "Primary", "A"),
        ("P4", "Primary", "A"),
        ("P5", "Primary", "A"),
        ("P6", "Primary", "A"),
        ("P7", "Primary", "A"),
    ]

    # Insert classes if missing
    for class_name, level, stream in rows:
        cur.execute(
            "INSERT IGNORE INTO classes (class_name, level, stream) VALUES (%s, %s, %s)",
            (class_name, level, stream),
        )
    conn.commit()

    # Ensure unique index on requirements table (if missing)
    if not _index_exists(conn, "requirements", "uq_requirements_class_term_name"):
        try:
            cur.execute("""
                CREATE UNIQUE INDEX uq_requirements_class_term_name
                ON requirements(class_name, term, name)
            """)
            conn.commit()
        except Exception as e:
            print(f"Index creation skipped: {e}")

    cur.close()
    
# ---------- Utility ----------
def _years_between(d1: date, d2: date) -> float:
    return (d2 - d1).days / 365.25

def _parse_date(s: str | None) -> date | None:
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def _auto_category(purchase_date_str: str | None, useful_life_years: int | float | None) -> str:
    """
    Category = 'current' if remaining useful life <= 1 year, else 'non-current'.
    When info is missing, default to 'non-current'.
    """
    try:
        if not purchase_date_str or not useful_life_years or float(useful_life_years) <= 0:
            return 'non-current'
        pd = _parse_date(purchase_date_str)
        if not pd:
            return 'non-current'
        age = _years_between(pd, date.today())
        remaining = float(useful_life_years) - age
        return 'current' if remaining <= 1.0 else 'non-current'
    except Exception:
        return 'non-current'

def _fmt_money(v) -> str:
    try:
        return f"{float(v):,.0f}"
    except Exception:
        return ""


# ---------- Query Builders ----------
def _build_asset_query(from_date: str, to_date: str, asset_name: str, category: str):
    q = "SELECT * FROM assets WHERE 1=1"
    p = []
    if from_date and to_date:
        q += " AND (purchase_date BETWEEN %s AND %s)"
        p += [from_date, to_date]
    elif from_date:
        q += " AND purchase_date >= %s"
        p += [from_date]
    elif to_date:
        q += " AND purchase_date <= %s"
        p += [to_date]

    if asset_name:
        q += " AND LOWER(asset_name) LIKE %s"
        p.append(f"%{asset_name.lower()}%")

    if category in ("current", "non-current"):
        q += " AND category=%s"
        p.append(category)

    q += " ORDER BY purchase_date DESC, id DESC"
    return q, p

def _refresh_categories(conn, rows):
    """
    Ensure category column is consistent (auto-compute & update if wrong).
    Returns possibly updated list of rows.
    """
    if not rows:
        return rows
    cur = conn.cursor(dictionary=True)
    dirty = []
    for r in rows:
        purchase_date_str = r.get("purchase_date").strftime("%Y-%m-%d") if r.get("purchase_date") else (r.get("year_purchased") or "")
        new_cat = _auto_category(purchase_date_str, r.get("useful_life_years"))
        if new_cat and new_cat != (r.get("category") or ""):
            dirty.append((new_cat, r["id"]))
    if dirty:
        cur.executemany("UPDATE assets SET category=%s WHERE id=%s", dirty)
        conn.commit()
        # Re-fetch to reflect updates
        ids = [r[1] for r in dirty]
        fmt = ",".join(["%s"] * len(ids))
        cur.execute(f"SELECT * FROM assets WHERE id IN ({fmt})", ids)
        updated = {r["id"]: r for r in (cur.fetchall() or [])}
        for i, r in enumerate(rows):
            if r["id"] in updated:
                rows[i] = updated[r["id"]]
    cur.close()
    return rows


def ensure_assets_schema(conn):
    cur = conn.cursor(dictionary=True)
    # Base create
    cur.execute("""
        CREATE TABLE IF NOT EXISTS assets (
            id INT AUTO_INCREMENT PRIMARY KEY,
            asset_name VARCHAR(100),
            description VARCHAR(100),
            model VARCHAR(30),
            value DOUBLE DEFAULT 0,
            year_purchased VARCHAR(10), -- legacy
            purchase_date DATE, -- new canonical
            asset_condition VARCHAR(50),
            qty INT,
            location VARCHAR(100),
            asset_code VARCHAR(50),
            company_number VARCHAR(50),
            useful_life_years INT,
            category VARCHAR(20),
            archived_reason VARCHAR(70)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()
 


def ensure_teachers_employees_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            id INT AUTO_INCREMENT PRIMARY KEY,
            first_name VARCHAR(50) NOT NULL,
            Middle_name VARCHAR(50),
            last_name VARCHAR(50) NOT NULL,
            gender VARCHAR(10),
            contact VARCHAR(30),
            email VARCHAR(50),
            residence VARCHAR(70),
            department VARCHAR(50),
            designation VARCHAR(30),
            hire_date VARCHAR(15),
            status ENUM('active','archived') DEFAULT 'active',
            base_salary DOUBLE DEFAULT 0,
            allowance DOUBLE DEFAULT 0,
            bonus DOUBLE DEFAULT 0,
            pay_cycle VARCHAR(32) DEFAULT 'monthly',
            bank_name VARCHAR(25),
            bank_account VARCHAR(25),
            tin VARCHAR(20),
            notes VARCHAR(50),
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS teachers (
            id INT AUTO_INCREMENT PRIMARY KEY,
            employee_id INT UNIQUE,
            initials VARCHAR(5),
            subjects VARCHAR(20),
            class_name VARCHAR(15),
            can_reset_password TINYINT DEFAULT 0,
            status VARCHAR(20) DEFAULT 'active',
            CONSTRAINT fk_teacher_employee FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS teacher_subjects (
            id INT AUTO_INCREMENT PRIMARY KEY,
            teacher_id INT NOT NULL,
            subject_id INT NOT NULL,
            class_name VARCHAR(15) NOT NULL,
            UNIQUE KEY uq_teacher_subjects (teacher_id, subject_id, class_name),
            INDEX ix_teacher_subjects_teacher (teacher_id),
            INDEX ix_teacher_subjects_subject (subject_id),
            INDEX ix_teacher_subjects_class (class_name),
            CONSTRAINT fk_ts_teacher FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE CASCADE,
            CONSTRAINT fk_ts_subject FOREIGN KEY (subject_id) REFERENCES subjects(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def add_created_at_if_missing():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SHOW COLUMNS FROM users")
        cols = [c["user"] for r in cur.fetchall()]
        if "created_at" not in cols:
            cur.execute(
                "ALTER TABLE users ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP")
            conn.commit()
            current_app.logger.info("[migrate] users.created_at added")
        else:
            current_app.logger.info(
                "[migrate] users.created_at already present")
    finally:
        cur.close()
        conn.close()


def ensure_academic_years_schema(conn):
    this_year = int(datetime.now().strftime("%Y"))
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS academic_years (
            id INT AUTO_INCREMENT PRIMARY KEY,
            year INT UNIQUE NOT NULL,
            current_term VARCHAR(15) DEFAULT 'Term 1',
            is_active TINYINT DEFAULT 0
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS term_dates (
            year INT NOT NULL,
            term VARCHAR(15) NOT NULL,
            next_term VARCHAR(15),
            next_term_date VARCHAR(20),
            PRIMARY KEY (year, term)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    cur.execute(
        "INSERT IGNORE INTO academic_years (year, current_term, is_active) VALUES (%s, 'Term 1', 1)",
        (this_year,),
    )
    cur.execute(
        "UPDATE academic_years SET is_active = CASE WHEN year = %s THEN 1 ELSE 0 END", (this_year,))
    conn.commit()
    cur.close()


def ensure_audit_trail_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS audit_trail (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT,
            action VARCHAR(30) NOT NULL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            role VARCHAR(25), outcome VARCHAR(30), severity VARCHAR(30),
            route VARCHAR(50), method VARCHAR(25), ip_address VARCHAR(25),
            target_table VARCHAR(20), target_id INT,
            details_json JSON, http_status INT,
            KEY idx_audit_user (user_id),
            CONSTRAINT fk_audit_user FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


# ---------- schema guards helpers ----------

def _has_column(conn, table: str, column: str) -> bool:
    """Return True if `table`.`column` exists in the current database."""
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COUNT(*)
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
          AND COLUMN_NAME = %s
    """, (table, column))
    exists = (cur.fetchone()[0] or 0) > 0
    cur.close()
    return exists


def add_column_if_missing(conn, table: str, ddl: str) -> None:
    """
    Add a column to `table` if it's missing.
    `ddl` should be the column definition starting with the column name,
    e.g. "created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP".
    """
    # first token of ddl is the column name
    col = ddl.split()[0].strip('`')
    if not _has_column(conn, table, col):
        cur = conn.cursor(dictionary=True)
        cur.execute(f"ALTER TABLE `{table}` ADD COLUMN {ddl}")
        cur.close()


# def apply_schema_guards(app=None) -> None:
 #   """
  #  Make sure certain columns exist with safe defaults.
   # - users.status: VARCHAR with default 'active'
    # - users.role: VARCHAR with default 'teacher' (use ENUM if you prefer)
    # - users.created_at: DATETIME with DEFAULT CURRENT_TIMESTAMP
   # """
   # conn = get_db_connection()
    # try:
   #     add_column_if_missing(conn, "users",
   #                           "status VARCHAR(20) NOT NULL DEFAULT 'active'")
   #     add_column_if_missing(conn, "users",
   #                           "role VARCHAR(30) NOT NULL DEFAULT 'teacher'")
   #     add_column_if_missing(conn, "users",
   #                           "created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP")
   #     conn.commit()
   # finally:
   #     conn.close()


def ensure_requirements_table_alias(conn):
    # kept for compatibility if some places call a different name
    ensure_requirements_schema(conn)


def ensure_other_income_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS other_income (
            id INT AUTO_INCREMENT PRIMARY KEY,
            description VARCHAR(50) NOT NULL,
            `source` VARCHAR(50),
            amount DOUBLE NOT NULL,
            term VARCHAR(10),
            year INT,
            date_received DATE DEFAULT (CURRENT_DATE),
            recorded_by VARCHAR(30),
            INDEX ix_other_income (term, year)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def ensure_fees_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS fees (
          id INT AUTO_INCREMENT PRIMARY KEY,
          student_id INT NOT NULL,
          term VARCHAR(10) NOT NULL,
          year INT NOT NULL,
          amount_paid DOUBLE NOT NULL DEFAULT 0,
          requirement_name VARCHAR(50),
          req_term VARCHAR(15),
          payment_item VARCHAR(50),
          bursary_amount DOUBLE NOT NULL DEFAULT 0,
          carried_forward DOUBLE NOT NULL DEFAULT 0,
          expected_amount DOUBLE NOT NULL DEFAULT 0,
          date_paid DATE NULL, -- DATE so BETWEEN filters work correctly
          comment VARCHAR(50),
          receipt_no VARCHAR(20),
          processed_on DATETIME NULL,
          method VARCHAR(20) NOT NULL DEFAULT 'N/A',
          payment_type VARCHAR(50) NOT NULL DEFAULT 'school_fees',
          recorded_by VARCHAR(20),
          UNIQUE KEY uq_fees_receipt(receipt_no),
          INDEX ix_fees_student (student_id),
          INDEX ix_fees_term_year (term, year),
          INDEX ix_fees_payment_type (payment_type),
          INDEX ix_fees_student_period_type(student_id, term, year, payment_type),
          INDEX ix_fees_date(date_paid),
          CONSTRAINT fk_fees_student FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;

    """)
    conn.commit()
    cur.close()


def seed_expense_categories(conn):
    cur = conn.cursor(dictionary=True)
    cats = [
        ('Salaries',), ('Stationery',), ('Utilities',), ('Transport',),
        ('Maintenance',), ('Service Providers',), ('Uniforms',),
        ('Examinations',), ('Meals',), ('Office supplies',), ('Medical',),
        ('Bonus',), ('Allowance',), ('Miscellaneous',)
    ]
    for name in cats:
        try:
            cur.execute(
                "INSERT IGNORE INTO expense_categories(name) VALUES (%s)", (name,))
        except Exception:
            pass
    conn.commit()
    cur.close()


def seed_grading_scale(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT COUNT(*) FROM grading_scale")
    (count,) = cur.fetchone()
    if count and count > 0:
        cur.close()
        return

    # Example PLE-like bands (adjust to your scheme)
    bands = [
        ('D1', 90, 100, 'Excellent'),
        ('D2', 80, 89, 'Very good'),
        ('C3', 75, 79, 'Good'),
        ('C4', 70, 74, 'Good'),
        ('C5', 65, 69, 'Fair'),
        ('C6', 60, 64, 'Fair'),
        ('P7', 50, 59, 'Pass'),
        ('P8', 40, 49, 'Basic'),
        ('F9', 0, 39, 'Fail'),
    ]
    cur.executemany(
        "INSERT INTO grading_scale (grade, lower_limit, upper_limit, comment) VALUES (%s, %s, %s, %s)",
        bands
    )
    conn.commit()
    cur.close()


def run_migrations(conn):
    # place versioned migrations here if needed
    pass


def ensure_class_comments_table(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS class_comments (
            id INT AUTO_INCREMENT PRIMARY KEY,
            class_name VARCHAR(15),
            term VARCHAR(10),
            year INT,
            comment VARCHAR(70)
        )
    """)
    conn.commit()
    cur.close()


def ensure_archived_students_table(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS archived_students (
            id INT AUTO_INCREMENT PRIMARY KEY,
            student_id INT,
            student_number VARCHAR(25),
            full_name VARCHAR(50),
            class_name VARCHAR(15),
            year_completed INT NOT NULL,
            completed_stage VARCHAR(15),
            outstanding_balance DOUBLE DEFAULT 0,
            archived_on DATETIME DEFAULT CURRENT_TIMESTAMP,
            INDEX ix_archived_students_yr (year_completed),
            INDEX ix_archived_students_sn (student_number),
            INDEX ix_archived_students_name (full_name),
            CONSTRAINT fk_arch_student FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def ensure_promotions_log_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS promotions_log (
            id INT AUTO_INCREMENT PRIMARY KEY,
            student_id INT NOT NULL,
            from_class VARCHAR(15) NOT NULL,
            to_class VARCHAR(15) NOT NULL,
            actor VARCHAR(20),
            batch_id VARCHAR(30),
            reversed TINYINT DEFAULT 0,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            INDEX ix_promolog_student (student_id, created_at)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def ensure_promotion_lock(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS promotion_lock (
            year INT PRIMARY KEY,
            executed_by VARCHAR(30),
            executed_at DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def promotion_already_done(year: int) -> bool:
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT 1 FROM promotion_lock WHERE year=%s", (year,))
    r = cur.fetchone()
    cur.close
    conn.close()
    return bool(r)


def mark_promotion_done(year: int, actor: str):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "INSERT IGNORE INTO promotion_lock(year, executed_by) VALUES(%s,%s)", (year, actor))
    conn.commit()
    cur.close()
    conn.close()


ORDER = ["Baby", "Middle", "Top", "P1", "P2", "P3", "P4", "P5", "P6", "P7"]


def next_class_name(current: str) -> str | None:
    cur = (current or "").strip()
    if cur not in ORDER:
        return None
    idx = ORDER.index(cur)
    return ORDER[idx+1] if idx+1 < len(ORDER) else None


def prev_class_name(current: str) -> str | None:
    cur = (current or "").strip()
    if cur not in ORDER:
        return None
    idx = ORDER.index(cur)
    return ORDER[idx-1] if idx-1 >= 0 else None


def write_audit(conn, *, user_id=None, role=None, action="", outcome="success",
                severity="info", route=None, method=None, ip=None,
                target_table=None, target_id=None, details=None, http_status=None):
    try:
        try:
            ensure_audit_trail_schema(conn)
        except Exception:
            pass

        import json
        dj = None
        if details is not None:
            try:
                dj = json.dumps(details, ensure_ascii=False, default=str)
            except Exception:
                dj = str(details)

        cur = conn.cursor(dictionary=True)
        cur.execute("""
            INSERT INTO audit_trail
              (user_id, role, action, outcome, severity,
               route, method, ip_address, target_table, target_id,
               details_json, http_status, timestamp)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, NOW())
        """, (user_id, role, action, outcome, severity, route, method, ip,
              target_table, target_id, dj, http_status))
        conn.commit()
        cur.close()
    except Exception:
        pass


def audit_from_request(conn, *, action, outcome="success", severity="info",
                       target_table=None, target_id=None, details=None, http_status=None):
    """Convenience wrapper that pulls request/session info."""
    try:
        uid = session.get("user_id")
        role = session.get("role")
        route = request.path
        method = request.method
        ip = request.headers.get("X-Forwarded-For", request.remote_addr)
        write_audit(conn,
                    user_id=uid, role=role, action=action, outcome=outcome, severity=severity,
                    route=route, method=method, ip=ip,
                    target_table=target_table, target_id=target_id,
                    details=details, http_status=http_status
                    )
    except Exception:
        pass


def _subjects_lookup(conn) -> dict[int, dict]:
    """id -> {id, name, code}"""
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, name, COALESCE(code, name) AS code FROM subjects")
    rows = cur.fetchall()
    cur.close()
    return {r["id"]: r for r in rows}


def _class_students(conn, class_name: str, stream: str):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS mid, last_name
        FROM students
        WHERE archived=0 AND class_name=%s AND COALESCE(stream,'')=%s
        ORDER BY last_name, first_name
    """, (class_name, stream))
    rows = cur.fetchall()
    cur.close()
    return rows


def _scores_for_class(conn, class_name: str, stream: str, term: str, year: int):
    """
    Return {sid: {code: {"score": x, "grade": g}}}.
    Uses EOT if available; otherwise average_mark (blended).
    """
    subs = _subjects_lookup(conn)
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT rs.student_id AS sid,
               rs.subject_id AS sub_id,
               COALESCE(rs.eot_mark, rs.average_mark) AS score
        FROM record_score rs
        JOIN students s ON s.id=rs.student_id
        WHERE s.archived=0
          AND s.class_name=%s AND COALESCE(s.stream,'')=%s
          AND rs.term=%s AND rs.year=%s
    """, (class_name, stream, term, year))
    rows = cur.fetchall()
    cur.close()

    out = {}
    for r in rows:
        sid = r["sid"]
        sub = subs.get(r["sub_id"])
        if not sub:
            continue
        code = (sub["code"] or sub["name"]).strip().upper()
        score = None if r["score"] is None else float(r["score"])
        grade = grade_for_score(conn, score) if score is not None else ""
        out.setdefault(sid, {})[code] = {"score": score, "grade": grade}
    return out


def subjects_with_marks(conn, class_name: str, stream: str | None, term: str, year: int):
    """
    Return subjects that have at least one row in record_score for any student
    in the selected class (+ optional stream) for the given term/year.
    """
    sql = """
        SELECT DISTINCT sub.id, sub.name, sub.code
        FROM subjects sub
        JOIN record_score rs ON rs.subject_id = sub.id
        JOIN students st ON st.id = rs.student_id
        WHERE st.class_name = %s
          AND (%s IS NULL OR %s = '' OR st.stream = %s)
          AND rs.term = %s
          AND rs.year = %s
        ORDER BY sub.name
    """
    cur = conn.cursor(dictionary=True)
    cur.execute(sql, (class_name, stream, stream, stream, term, year))
    rows = cur.fetchall() or []
    cur.close()
    return rows


def compute_outstanding_balance(student_id: int) -> float:
    # If you already have compute_student_financials(...), call it.
    # Otherwise, quick placeholder that sums fees - payments. Replace as needed.
    try:
        fin = compute_student_financials(student_id, None, None, None)
        return float(fin.get("overall_balance", 0)) if isinstance(fin, dict) else float(fin.overall_balance or 0)
    except Exception:
        return 0.0


# safe helper (put once in your codebase)


def _archive_student(
    conn,
    student_id: int,
    new_status: str = "completed",
    stage: str = "Manual Archive",
    year_completed: int | None = None,
) -> int:
    """
    Soft-archive one student (students.archived=1 + allowed status) and
    ensure a row exists in archived_students for the same student/year.
    """
    ensure_archived_students_table(conn)  # no-op if already exists
    from datetime import datetime

    ay = (get_active_academic_year() or {})

    # --- compute academic year (yc) robustly ---
    try:
        yc = int(year_completed or ay.get("year") or ay.get(
            "active_year") or datetime.now().year)
    except Exception:
        yc = datetime.now().year

    cur = conn.cursor(dictionary=True)

    # --- fetch student (ALWAYS do this; was previously inside the except) ---
    cur.execute("""
        SELECT id, student_number, first_name, last_name, class_name,
               COALESCE(Middle_name, '') AS middle_name
        FROM students
        WHERE id = %s
        LIMIT 1
    """, (student_id,))
    s = cur.fetchone()
    if not s:
        cur.close()
        return 0

    # 1) flip archived & use an allowed status
    cur.execute(
        "UPDATE students SET archived=1, status=%s WHERE id=%s",
        (new_status, student_id)
    )
    changed = cur.rowcount or 0

    # 2) optional best-effort outstanding (guarded)
    outstanding = 0.0
    try:
        term = (ay.get("current_term") or ay.get("term") or "Term 1")
        fin = compute_student_financials(student_id, s["class_name"], term, yc)
        outstanding = float(
            (fin.get("overall_balance") if isinstance(fin, dict)
             else getattr(fin, "overall_balance", 0)) or 0
        )
    except Exception:
        pass

    # build full_name safely
    full_name = f"{s['first_name']} {s.get('middle_name', '')} {s['last_name']}".strip(
    )

    # 3) write to archive table (skip duplicate for same student/year)
    cur.execute("""
        INSERT INTO archived_students
            (student_id, student_number, full_name, class_name,
             year_completed, completed_stage, outstanding_balance)
        SELECT %s, %s, %s, %s, %s, %s, %s
        WHERE NOT EXISTS (
            SELECT 1
            FROM archived_students
            WHERE student_id=%s AND year_completed=%s
        )
    """, (
        s["id"], s["student_number"], full_name, s["class_name"],
        yc, stage, outstanding, s["id"], yc
    ))

    conn.commit()
    cur.close()
    return changed


def _unarchive_student(conn, student_id: int, *, remove_archive_rows: bool = True) -> int:
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "UPDATE students SET archived=0, status='active' WHERE id=%s",
        (student_id,)
    )
    changed = cur.rowcount or 0

    if remove_archive_rows:
        ensure_archived_students_table(conn)
        cur.execute(
            "DELETE FROM archived_students WHERE student_id=%s", (student_id,))

    conn.commit()
    cur.close()
    return changed


def ensure_subject_papers_schema(conn):
    cur = conn.cursor(dictionary=True)
    # Base table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS subject_papers(
            id INT AUTO_INCREMENT PRIMARY KEY,
            subject_id INT NOT NULL,
            paper_name VARCHAR(15) NOT NULL,
            paper_initial VARCHAR(7),
            FOREIGN KEY(subject_id) REFERENCES subjects(id) ON DELETE CASCADE
        )ENGINE = InnoDB DEFAULT CHARSET = utf8mb4;
    """)
    for stmt in ("CREATE INDEX ix_subject_papers_subject ON subject_papers(subject_id)",
                 "CREATE UNIQUE INDEX uq_subject_papers_unique ON subject_papers(subject_id, paper_name)",
                 ):
        try:
            cur.execute(stmt)
        except Exception:
            pass
    conn.commit()
    cur.close()


def ensure_streams_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS streams(
        id INT AUTO_INCREMENT PRIMARY KEY,
        name VARCHAR(15) NOT NULL,
        UNIQUE KEY uq_stream(name, stream)
    ) ENGINE = InnoDB DEFAULT CHARSET = utf8mb4;
    """)
    conn.commit()
    cur.close()




def fmt_report_date(val: str | None) -> str:
    """
    Format 'YYYY-MM-DD' to '01/Feb/2026'.
    If parsing fails, return the original string.
    """
    if not val:
        return ""
    try:
        d = datetime.strptime(val, "%Y-%m-%d")
        return d.strftime("%d/%b/%Y")
    except Exception:
        return val

def ensure_term_dates_schema(conn=None):
    """
    Create term_dates table if missing.
    Works with or without a supplied connection(backward compatible).
    """
    must_close = False
    if conn is None:
        conn = get_db_connection()
        must_close = True

    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS term_dates(
            year INT NOT NULL,
            term VARCHAR(15) NOT NULL,
            next_term VARCHAR(15),
            next_term_date VARCHAR(20), -- store as 'YYYY-MM-DD' or any display string
            next_term_end_date VARCHAR(20), -- new: also 'YYYY-MM-DD'
            PRIMARY KEY(year, term)
        )ENGINE = InnoDB DEFAULT CHARSET = utf8mb4;
    """)
    conn.commit()
    cur.close()
    if must_close:
        conn.close()
        
        



def ensure_grading_scale_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS grading_scale(
            id INT AUTO_INCREMENT PRIMARY KEY,
            grade VARCHAR(5) NOT NULL,
            lower_limit INT NOT NULL,
            upper_limit INT NOT NULL,
            comment VARCHAR(50),
            INDEX ix_grading_bounds(lower_limit, upper_limit)
        )ENGINE = InnoDB DEFAULT CHARSET = utf8mb4;
    """)
    conn.commit()
    cur.close()


def ensure_users_table(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users(
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(15) NOT NULL UNIQUE,
            password_hash VARCHAR(255) NOT NULL,
            role ENUM('admin', 'bursar', 'teacher', 'headteacher', 'director', 'clerk', 'deputyheadteacher', 'dos') NOT NULL,
            status ENUM('active', 'archived') NOT NULL DEFAULT 'active',
            employee_id INT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_users_employee_id(employee_id),
            CONSTRAINT fk_user_employee FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE SET NULL
        ) ENGINE = InnoDB DEFAULT CHARSET = utf8mb4
        """)
    conn.commit()
    cur.close()


def seed_default_admin(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users(
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(15) NOT NULL UNIQUE,
            password_hash VARCHAR(255) NOT NULL,
            role ENUM('admin', 'bursar', 'teacher', 'headteacher', 'director', 'clerk', 'deputyheadteacher', 'dos') NOT NULL,
            status ENUM('active', 'archived') NOT NULL DEFAULT 'active',
            employee_id INT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_users_employee_id(employee_id),
            CONSTRAINT fk_user_employee FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE SET NULL
            ) ENGINE = InnoDB DEFAULT CHARSET = utf8mb4;
    """)
    cur.execute("SELECT 1 FROM users WHERE username='admin'")
    row = cur.fetchone()
    if not row:
        cur.execute("""
            INSERT INTO users(username, password_hash, role, status)
            VALUES ( % s, % s, 'admin', 'active')
        """, ("admin", generate_password_hash("admin123")))
    conn.commit()
    cur.close()


def next_class_name(current: str) -> str | None:
    """Lightweight promotion map; extend as needed."""
    order = ["Baby", "Middle", "Top", "P1", "P2", "P3", "P4", "P5", "P6", "P7"]
    if current not in order:
        return None
    idx = order.index(current)
    return order[idx+1] if idx+1 < len(order) else None


def generate_term_fees(student_row, term, year, c):
    sid = student_row["id"]
    class_name = student_row["class_name"]
    level = student_row.get("level") if hasattr(
        student_row, "get") else student_row["level"]
    section = (student_row.get("section") if hasattr(student_row, "get") else student_row.get("section", None)) or \
              (student_row.get("section") if hasattr(student_row, "get") else None)
    section = (section or '').strip()
    conn.commit()
    cur.close()
    cur.execute("""
        SELECT amount
        FROM class_fees
        WHERE class_name= %s
          AND LOWER(section) = LOWER(%s)
          AND (level IS NULL OR level=%s)
        LIMIT 1
    """, (class_name, section, level))
    row = cur.fetchone()
    cur.close()
    expected = float(
        row["amount"]) if row and row["amount"] is not None else 0.0
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM bursaries
        WHERE student_id= %s AND term = %s AND year = %s
    """, (sid, term, year))
    b = cur.fetchone()
    cur.close()
    bursary_amount = float(b["amount"] or 0)

    cur = conn.cursor(dictionary=True)
    cur.execute("""
        INSERT INTO fees(student_id, term, year, payment_type,
                          expected_amount, bursary_amount, carried_forward, amount_paid)
        VALUES (% s, % s, % s, 'school_fees', % s, % s, 0, 0)
        ON DUPLICATE KEY UPDATE
            expected_amount= VALUES(expected_amount),
            bursary_amount= VALUES(bursary_amount)
    """, (sid, term, year, expected, bursary_amount))
    cur.close()

    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id FROM fees
        WHERE student_id= %s AND term = %s AND year = %s AND payment_type = 'school_fees'
        LIMIT 1
    """, (sid, term, year))
    rid = cur.fetchone()
    cur.close()
    return rid["id"] if rid else None

# ---------------- FIX FEES (RECALC EXPECTED / BURSARY / CARRY-FORWARD) ----------------


def _recalc_all_fees(conn) -> int:
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, student_id, term, year
        FROM fees
        WHERE LOWER(payment_type) IN('school_fees', 'fees')
    """)
    rows = cur.fetchall() or []
    updated = 0

    for fee in rows:
        sid = fee["student_id"]
        term = (fee["term"] or "")
        year = int(fee["year"])

        cur.execute("""
            SELECT class_name, NULLIF(TRIM(section), '') AS sec
            FROM students
            WHERE id= %s
        """, (sid,))
        st = cur.fetchone()
        if not st:
            continue

        class_name = (st["class_name"] or '').strip()
        section = st.get("sec") if hasattr(st, "get") else st["sec"]

        expected = 0.0
        if class_name and section:
            cur.execute("""
                SELECT amount
                FROM class_fees
                WHERE class_name= %s
                  AND LOWER(section) = LOWER(% s)
                LIMIT 1
            """, (class_name, section))
            cr = cur.fetchone()
            if cr and (cr.get("amount") if hasattr(cr, "get") else None) is not None:
                expected = float(cr.get("amount") if hasattr(
                    cr, "get") else cr["amount"])

        cur.execute("""
            SELECT COALESCE(SUM(amount), 0) AS total
            FROM bursaries
            WHERE student_id= %s AND term = %s AND year = %s
        """, (sid, term, year))
        bursary = float((cur.fetchone() or {}).get("total", 0.0) or 0.0)

        cur.execute("""
            SELECT expected_amount AS exp,
                   bursary_amount AS bur,
                   amount_paid AS paid,
                   term, year
            FROM fees
            WHERE student_id= %s
              AND LOWER(payment_type) IN('school_fees', 'fees', 'opening_balance')
              AND(year < %s OR(year = % s AND
                    CASE LOWER(term)
                      WHEN 'term 1' THEN 1
                      WHEN 'term 2' THEN 2
                      WHEN 'term 3' THEN 3
                      ELSE 0
                    END < %s))
            ORDER BY year DESC,
                     CASE LOWER(term)
                       WHEN 'term 3' THEN 3
                       WHEN 'term 2' THEN 2
                       WHEN 'term 1' THEN 1
                       ELSE 0
                     END DESC
            LIMIT 1
        """, (sid, year, year, _term_rank(term)))
        prev = cur.fetchone()

        carried = 0.0
        if prev:
            carried = max(
                (float(prev.get("exp") or 0) -
                 float(prev.get("bur") or 0) - float(prev.get("paid") or 0)),
                0.0,
            )

        cur.execute("""
            UPDATE fees
               SET expected_amount= %s,
                   bursary_amount= %s,
                   carried_forward= %s
             WHERE id= %s
        """, (expected, bursary, carried, fee["id"]))
        updated += 1

    conn.commit()
    cur.close()
    return updated
    
def _mean_nonnull(values):
    nums = [float(v) for v in values if v is not None]
    return (sum(nums) / len(nums)) if nums else None

def division_from_aggregate(agg: int | None) -> str:
    if agg is None:
        return "NG"
    a = int(agg)
    if 4 <= a <= 12: return "Div 1"
    if 13 <= a <= 23: return "Div 2"
    if 24 <= a <= 29: return "Div 3"
    if 30 <= a <= 34: return "Div 4"
    return "U"

def core_name_from(name: str, code: str) -> str:
    c = (code or "").strip().upper()
    if c in CORE_CODES:
        return c.lower() # eng, math, sci, sst
    n = (name or "").strip().lower()
    if n.startswith("eng"): return "eng"
    if n.startswith(("mat", "math")): return "math"
    if n.startswith("sci"): return "sci"
    if n in {"sst", "soc. studies", "social studies", "social std", "socialstudies"}:
        return "sst"
    return ""


def ensure_fee_rows_for_all(conn, term: str, year: int) -> int:
    """
    Insert a 'school_fees' row for every active student for (term, year)
    if it doesn't already exist. Returns number of rows inserted.
    """
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        INSERT INTO fees(
            student_id, payment_type, term, year,
            expected_amount, bursary_amount, amount_paid, carried_forward, date_paid
        )
        SELECT s.id, 'school_fees', % s, % s, 0, 0, 0, 0, CURDATE()
        FROM students s
        WHERE COALESCE(LOWER(s.status), 'active') = 'active'
          AND NOT EXISTS(
            SELECT 1 FROM fees f
            WHERE f.student_id=s.id
              AND f.year=% s
              AND LOWER(f.term) = LOWER(% s)
              AND LOWER(f.payment_type) IN('school_fees', 'fees')
          )
    """, (term, year, year, term))
    inserted = cur.rowcount or 0
    conn.commit()
    cur.close
    return inserted


# --- Helper: find where raw scores live (record_score or results) ---


def detect_scores_table(conn):
    """
    Detect whether 'record_score' or 'results' table exists.
    Returns the table name as a string, or None if neither exist.
    """
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SHOW TABLES LIKE 'record_score'")
        if cur.fetchone():
            return "record_score"

        cur.execute("SHOW TABLES LIKE 'results'")
        if cur.fetchone():
            return "results"

        return None
    finally:
        cur.close()


# --- Helper: process snapshot for a class/term/year (idempotent) ---
#=============MANUAL COMMENTS ON REPORTS====================


def fetch_overall_overrides(conn, student_id, term, year):
    """
    Return a dict with any saved manual comments/special comms for this learner.
    Keys:
      teacher_overall_custom, head_overall_custom, special_communication
    or {} if none.
    """
    ensure_report_overall_overrides_schema(conn)
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT teacher_overall_custom,
               head_overall_custom,
               special_communication
          FROM report_overall_overrides
         WHERE student_id=%s AND term=%s AND year=%s
         LIMIT 1
    """, (student_id, term, year))
    row = cur.fetchone()
    cur.close()
    return row or {}


def save_overall_overrides(conn, student_id, term, year,
                           teacher_text=None,
                           head_text=None,
                           special_text=None):
    """
    Upsert helper – called from a form route.
    Any None value keeps the existing value.
    """
    ensure_report_overall_overrides_schema(conn)
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO report_overall_overrides
            (student_id, term, year,
             teacher_overall_custom,
             head_overall_custom,
             special_communication)
        VALUES (%s,%s,%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
            teacher_overall_custom =
                COALESCE(VALUES(teacher_overall_custom), teacher_overall_custom),
            head_overall_custom =
                COALESCE(VALUES(head_overall_custom), head_overall_custom),
            special_communication =
                COALESCE(VALUES(special_communication), special_communication)
    """, (student_id, term, year,
          teacher_text, head_text, special_text))
    conn.commit()
    cur.close()
    

def ensure_midterm_overall_comments_schema(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS midterm_overall_comments (
            id INT AUTO_INCREMENT PRIMARY KEY,
            student_id INT NOT NULL,
            class_name VARCHAR(50),
            term VARCHAR(20) NOT NULL,
            year INT NOT NULL,
            teacher_comment TEXT,
            headteacher_comment TEXT,
            special_communication TEXT,
            UNIQUE KEY uniq_midterm_overall (student_id, term, year)
        )
    """)
    conn.commit()
    cur.close()


def load_midterm_manual_comments(conn, student_id, term, year):
    """
    Returns {
      'teacher_comment': str,
      'headteacher_comment': str,
      'special_communication': str
    } – all may be empty strings if nothing saved.
    """
    ensure_midterm_overall_comments_schema(conn)
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT teacher_comment, headteacher_comment, special_communication
        FROM midterm_overall_comments
        WHERE student_id=%s AND term=%s AND year=%s
        LIMIT 1
    """, (student_id, term, year))
    row = cur.fetchone()
    cur.close()
    if not row:
        return {
            "teacher_comment": "",
            "headteacher_comment": "",
            "special_communication": "",
        }
    return {
        "teacher_comment": row.get("teacher_comment") or "",
        "headteacher_comment": row.get("headteacher_comment") or "",
        "special_communication": row.get("special_communication") or "",
    }


def merge_midterm_comments(auto_teacher, auto_head, manual):
    """
    Start from automatic comments; override with manual where provided.
    special_communication is purely manual.
    Returns (teacher_comment, head_comment, special_communication)
    """
    t = auto_teacher or ""
    h = auto_head or ""
    s = manual.get("special_communication") or ""

    if manual.get("teacher_comment"):
        t = manual["teacher_comment"]
    if manual.get("headteacher_comment"):
        h = manual["headteacher_comment"]

    return t, h, s
    
    
#=============MANUAL COMMENTS ON REPORTS====================


def process_reports_snapshot(conn, class_name, term, year):
    """
    Build/refresh snapshot rows in `reports` for one class+term+year.

    Rules aligned with mark sheet & performance summary:
      - Per subject TOTAL(%) = average_mark if present; else mean of available
        (other, holiday, BOT, MID, EOT, CA).
      - Grade strictly from grading_scale.
      - Use the LATEST record_score row (by id) per (student, subject, term, year).
    """
    ensure_reports_table(conn) # must create UNIQUE(student_id,subject_id,term,year)
    ensure_record_score_table(conn)

    # --- helpers ---
    def _mean_nonnull(values):
        nums = [float(v) for v in values if v is not None]
        return (sum(nums) / len(nums)) if nums else None

    # --- pull students of this class (active only) ---
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, class_name, COALESCE(stream,'') AS stream
        FROM students
        WHERE archived=0 AND class_name=%s
        ORDER BY last_name, first_name
    """, (class_name,))
    students = cur.fetchall() or []
    cur.close()

    # --- wipe this class/term/year snapshot (safe because we rebuild below) ---
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM reports WHERE class_name=%s AND term=%s AND year=%s",
        (class_name, term, year)
    )
    cur.close()

    # --- upsert per student ---
    for s in students:
        sid = s["id"]

        # latest record_score per subject for this student/term/year
        # (If you have an updated_at column, add it to the MAX() ordering.)
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT rs.*
            FROM record_score rs
            JOIN (
                SELECT subject_id, MAX(id) AS max_id
                FROM record_score
                WHERE student_id=%s AND term=%s AND year=%s
                GROUP BY subject_id
            ) m ON m.max_id = rs.id
            ORDER BY rs.subject_id
        """, (sid, term, year))
        rows = cur.fetchall() or []
        cur.close()

        for r in rows:
            # authoritative numeric total (same rule used by other routes)
            if r.get("average_mark") is not None:
                total_num = float(r["average_mark"])
            else:
                total_num = _mean_nonnull([
                    r.get("other_mark"),
                    r.get("holiday_mark"),
                    r.get("bot_mark"),
                    r.get("midterm_mark"),
                    r.get("eot_mark"),
                    r.get("ca_mark"),
                ])

            grd = grade_for_score(conn, total_num) if total_num is not None else None
            cmt = comment_for_grade(conn, grd)

            # prefer stored initials; else guess from class/subject
            initials = (r.get("initials") or r.get("teacher_initial") or "") \
                       or guess_teacher_initials(conn, s["class_name"], r["subject_id"])

            # MySQL upsert
            sql = """
                INSERT INTO reports (
                    student_id, class_name, stream, subject_id, term, year,
                    average_mark, grade, comment,
                    teacher_remark, headteacher_remark,
                    teacher_id,
                    bot_mark, midterm_mark, eot_mark,
                    holiday_mark, other_mark, ca_mark,
                    teacher_initial, processed_on
                )
                VALUES (%s,%s,%s,%s,%s,%s,
                        %s,%s,%s,
                        %s,%s,
                        %s,
                        %s,%s,%s,
                        %s,%s,%s,
                        %s, CURRENT_TIMESTAMP)
                ON DUPLICATE KEY UPDATE
                    class_name = VALUES(class_name),
                    stream = VALUES(stream),
                    average_mark = VALUES(average_mark),
                    grade = VALUES(grade),
                    comment = VALUES(comment),
                    teacher_remark = VALUES(teacher_remark),
                    headteacher_remark = VALUES(headteacher_remark),
                    teacher_id = VALUES(teacher_id),
                    bot_mark = VALUES(bot_mark),
                    midterm_mark = VALUES(midterm_mark),
                    eot_mark = VALUES(eot_mark),
                    holiday_mark = VALUES(holiday_mark),
                    other_mark = VALUES(other_mark),
                    ca_mark = VALUES(ca_mark),
                    teacher_initial = VALUES(teacher_initial),
                    processed_on = CURRENT_TIMESTAMP
            """
            params = (
                sid, s["class_name"], s["stream"], r["subject_id"], term, year,
                total_num, grd, cmt,
                None, None, # teacher_remark, headteacher_remark (kept null here)
                None, # teacher_id
                r.get("bot_mark"), r.get("midterm_mark"), r.get("eot_mark"),
                r.get("holiday_mark"), r.get("other_mark"), r.get("ca_mark"),
                initials
            )
            cur = conn.cursor()
            cur.execute(sql, params)
            cur.close()

    conn.commit()



# ---------- Single report card ----------

# =========================
# CHARACTER ASSESSMENT HELPERS
# =========================

def _fetch_student_basic(student_id: int):
    """Small helper to get learner info for headers/nav."""
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT id, first_name, middle_name, last_name,
               student_number, class_name, stream, photo, photo_blob
        FROM students
        WHERE id = %s
        """,
        (student_id,),
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row

def _term_to_no(term: str) -> int:
    # supports "Term 1", "Term 2", "Term 3"
    m = re.search(r"(\d+)", term or "")
    n = int(m.group(1)) if m else 1
    return n if n in (1, 2, 3) else 1

def _fetch_character_items(term: str, class_name: str | None = None):
    """
    Term-based items.
    - item.term_no can be NULL => usable in any term
    - item.class_name can be NULL => usable in any class
    """
    term_no = _term_to_no(term)

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute(
        """
        SELECT id, area, section, term_no, class_name, skill_label, description, sort_order, is_active
        FROM character_items
        WHERE is_active = 1
          AND (term_no = %s OR term_no IS NULL)
          AND (%s IS NULL OR class_name IS NULL OR class_name = %s)
        ORDER BY area, section, sort_order, id
        """,
        (term_no, class_name, class_name),
    )

    rows = cur.fetchall() or []
    cur.close()
    conn.close()
    return rows




def _fetch_character_scores_map(student_id: int, term: str, year: int):
    """
    Returns {item_id: level_text} for this learner / term / year.
    Level text is the FULL phrase from REMARK_OPTIONS.
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT item_id, level
        FROM character_scores
        WHERE student_id = %s AND term = %s AND year = %s
        """,
        (student_id, term, year),
    )
    rows = cur.fetchall() or []
    cur.close()
    conn.close()
    return {row["item_id"]: (row["level"] or "") for row in rows}


def _fetch_character_meta(student_id: int, term: str, year: int):
    """Meta info: comments, dates, fees, special comm."""
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT class_teacher_comment,
               head_teacher_comment,
               next_term_begin,
               next_term_end,
               school_fees,
               school_fees_daycare,
               special_communication
        FROM character_meta
        WHERE student_id = %s AND term = %s AND year = %s
        """,
        (student_id, term, year),
    )
    meta = cur.fetchone() or {}
    cur.close()
    conn.close()
    return meta
    

def _draw_character_report_page(c, student, term, year, items, scores_map, meta):
    """
    Draw one complete Character Progressive Report on the given canvas `c`,
    allowing the table to flow onto extra pages. Area rows (CONFIDENCE,
    EMOTIONAL INTELLIGENCE, TIME MANAGEMENT, PREPARATION) are shaded grey.
    """
    import os
    from datetime import datetime
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.utils import ImageReader
    import io
    from reportlab.lib.units import mm

    width, height = A4
    styles = getSampleStyleSheet()
    left = 40
    right = 40
    bottom_margin = 60

    # =================== BRANDED HEADER (copied from checklist) ===================

    banner_h = 30 * mm
    banner_y = height - 35 * mm
    left_margin, right_margin = left, right
    strip_w = width - left_margin - right_margin
    navy_w = strip_w * 0.73
    blue_w = strip_w - navy_w

    c.saveState()
    # navy and blue strips
    c.setFillColor(COL_NAVY)
    c.rect(left_margin, banner_y, navy_w, banner_h, stroke=0, fill=1)
    c.setFillColor(COL_BLUE)
    c.rect(left_margin + navy_w, banner_y, blue_w, banner_h, stroke=0, fill=1)

    # folded-paper joint
    fold_depth = 11 * mm
    fold_lip = 6 * mm
    c.setFillColor(COL_BLUE2)
    ps = c.beginPath()
    ps.moveTo(left_margin + navy_w, banner_y)
    ps.lineTo(left_margin + navy_w + fold_depth, banner_y + banner_h)
    ps.lineTo(left_margin + navy_w + fold_depth + 2 * mm, banner_y + banner_h)
    ps.lineTo(left_margin + navy_w + 2 * mm, banner_y)
    ps.close()
    c.drawPath(ps, stroke=0, fill=1)
    flap_col = colors.HexColor("#3a86e0")
    c.setFillColor(flap_col)
    pf = c.beginPath()
    pf.moveTo(left_margin + navy_w - fold_lip, banner_y)
    pf.lineTo(left_margin + navy_w, banner_y)
    pf.lineTo(left_margin + navy_w + fold_depth, banner_y + banner_h)
    pf.lineTo(left_margin + navy_w - fold_lip, banner_y + banner_h)
    pf.close()
    c.drawPath(pf, stroke=0, fill=1)

    # --- logo (ensure path defined like EOT header) ---
    SCHOOL_LOGO_PATH = os.path.join(current_app.static_folder, "logo.jpg")
    logo_box = 24 * mm
    logo_x = left_margin + 6 * mm
    logo_y = banner_y + (banner_h - logo_box) / 2
    if os.path.exists(SCHOOL_LOGO_PATH):
        try:
            c.drawImage(
                SCHOOL_LOGO_PATH,
                logo_x,
                logo_y,
                width=logo_box,
                height=logo_box,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass

    # --- school name, tagline, P.O Box centred on navy ---
    name_left = logo_x + logo_box + 6 * mm
    name_right = left_margin + navy_w - 6 * mm
    name_box_w = max(10, name_right - name_left)

    # centre of the text block (between logo-right & navy-right)
    center_x = (name_left + name_right) / 2.0

    name_text = SCHOOL_NAME or ""
    name_fs = 18
    while (
        name_fs >= 10
        and c.stringWidth(name_text, "Helvetica-Bold", name_fs) > name_box_w
    ):
        name_fs -= 1

    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", name_fs)
    # a bit lower + more room for lines below
    name_y = banner_y + banner_h - 5 * mm
    c.drawCentredString(center_x, name_y, name_text)

    sub_text = SCHOOL_SUB or ""
    addr_text = SCHOOL_ADDRESS or "" # P.O Box line

    # tagline line
    if sub_text:
        sub_fs = 12
        while (
            sub_fs >= 8
            and c.stringWidth(sub_text, "Helvetica-Bold", sub_fs) > name_box_w
        ):
            sub_fs -= 1
        c.setFont("Helvetica-Bold", sub_fs)
        tagline_y = name_y - (name_fs * 1.15) # <- bigger = more spacing
        c.drawCentredString(center_x, tagline_y, sub_text)
    else:
        tagline_y = name_y

    # P.O Box line
    if addr_text:
        addr_fs = max(8, (sub_fs - 1) if sub_text else 10)
        c.setFont("Helvetica-Bold", addr_fs)
        addr_y = tagline_y - (addr_fs * 1.1) # extra spacing again
        c.drawCentredString(center_x, addr_y, addr_text)

    # --- contacts on BLUE: phones + email only, each on new line ---
   
    c.setFillColor(colors.white)
    c.setFont("Helvetica", 9)
    right_pad = 6 * mm
    text_right = left_margin + strip_w - right_pad
    line_gap = 5.5 * mm
    y_cursor = banner_y + banner_h - 8 * mm

    # -------- Contacts block (RIGHT SIDE, one per line) --------
    raw = SCHOOL_PHONE_LINES or ""

    # Normalize to one comma-separated string first
    if isinstance(raw, (list, tuple)):
        combined = ", ".join(str(p) for p in raw)
    else:
        combined = str(raw)

    # Now definitely split into separate phone numbers
    phone_lines = [p.strip() for p in combined.split(",") if p.strip()]

    # draw each phone on its own line
    for ph in phone_lines:
        c.drawRightString(text_right, y_cursor, ph)
        y_cursor -= line_gap

    # small extra gap, then email if present
    if SCHOOL_EMAIL:
        y_cursor -= 2.5 * mm
        c.drawRightString(text_right, y_cursor, SCHOOL_EMAIL)
        
    c.restoreState()


    # ========== LEARNER INFO TABLE ==========
    info_top = banner_y - 6 * mm
    info_left = left_margin
    info_width = width - left_margin - right_margin - (40 * mm)

    lab = ParagraphStyle(
        "lab",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=colors.black,
    )
    val = ParagraphStyle(
        "val",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=11,
    )

    full_name = f"{student.get('first_name','')} {student.get('middle_name','')} {student.get('last_name','')}".strip()

    info_rows = [
        [Paragraph("Learner's Name:", lab), Paragraph(full_name or "-", val)],
        [Paragraph("Student No.:", lab), Paragraph(student.get("student_number") or "-", val)],
        [
            Paragraph("Class / Stream:", lab),
            Paragraph(f"{student.get('class_name','-')} {student.get('stream','') or ''}", val),
        ],
        [Paragraph("Term / Year:", lab), Paragraph(f"{term} / {year}", val)],
    ]
    info_tbl = Table(
        info_rows,
        colWidths=[35 * mm, info_width - 35 * mm],
        hAlign="LEFT",
    )
    info_tbl.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ]
        )
    )
    w_info, h_info = info_tbl.wrapOn(c, info_width, 9999)
    info_tbl.drawOn(c, info_left, info_top - h_info)

    # ---- PHOTO BLOCK (DB blob or file path, same style as other reports) ----
    box_w = box_h = 32 * mm
    photo_x = width - right_margin - box_w
    photo_y = info_top - (h_info - box_h) / 2

    # Always draw frame + "Photo" text first
    c.setStrokeColor(colors.grey)
    c.rect(photo_x, photo_y - box_h, box_w, box_h, stroke=1, fill=0)
    c.setFont("Helvetica", 7)
    c.drawCentredString(photo_x + box_w / 2, photo_y - box_h / 2, "Photo")

    photo_blob = student.get("photo_blob")
    photo_path = student.get("photo")

    try:
        if photo_blob:
            img_reader = ImageReader(io.BytesIO(photo_blob))
            c.drawImage(
                img_reader,
                photo_x + 2,
                photo_y - box_h + 2,
                box_w - 4,
                box_h - 4,
                preserveAspectRatio=True,
                mask="auto",
            )
        elif photo_path:
            full_path = os.path.join(app.root_path, photo_path)
            if os.path.exists(full_path):
                img_reader = ImageReader(full_path)
                c.drawImage(
                    img_reader,
                    photo_x + 2,
                    photo_y - box_h + 2,
                    box_w - 4,
                    box_h - 4,
                    preserveAspectRatio=True,
                    mask="auto",
                )
    except Exception:
        # if image fails, we just keep the placeholder box
        pass

    # ====== Title ======
    title_y = (info_top - h_info) - 6 * mm
    c.setFont("Helvetica-Bold", 13)
    c.setFillColor(colors.black)
    c.drawString(left, title_y, "Learner's Character Progressive Report")

    table_top = title_y - 6 * mm
    avail_width = width - left - right

    # ---------- MAIN TABLE (with pagination & grey area rows) ----------
    p_area = ParagraphStyle(
        "p_area",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
    )
    p_section = ParagraphStyle(
        "p_section",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
    )
    p_skill = ParagraphStyle(
        "p_skill",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=9,
    )
    p_desc = ParagraphStyle(
        "p_desc",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=9,
    )
    p_level = ParagraphStyle(
        "p_level",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=9,
    )
    head_style = ParagraphStyle(
        "head",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.white,
        alignment=1,
    )

    data = [
        [
            Paragraph("Area", head_style),
            Paragraph("Section", head_style),
            Paragraph("Skill", head_style),
            Paragraph("Description", head_style),
            Paragraph("Level", head_style),
        ]
    ]
    row_meta = [{"type": "header"}]
    last_area = None
    last_section = None

    for it in items:
        lvl_text = scores_map.get(it["id"]) or ""

        # area row
        if it["area"] != last_area:
            data.append([Paragraph(it["area"], p_area), "", "", "", ""])
            row_meta.append({"type": "area", "area": it["area"]})
            last_area = it["area"]
            last_section = None

        # section row
        if it["section"] and it["section"] != last_section:
            data.append(["", Paragraph(it["section"], p_section), "", "", ""])
            row_meta.append(
                {"type": "section", "area": it["area"], "section": it["section"]}
            )
            last_section = it["section"]

        # skill row
        data.append(
            [
                "",
                "",
                Paragraph(it["skill_label"], p_skill),
                Paragraph(it["description"], p_desc),
                Paragraph(lvl_text, p_level),
            ]
        )
        row_meta.append(
            {
                "type": "skill",
                "area": it["area"],
                "section": it["section"],
                "skill": it["skill_label"],
            }
        )

    col_area_w = 23 * mm
    col_section_w = 28 * mm
    col_skill_w = 30 * mm
    col_level_w = 35 * mm
    col_desc_w = avail_width - (col_area_w + col_section_w + col_skill_w + col_level_w)

    base_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.0, 0.45, 0.80)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
        ]
    )

    base_table = Table(
        data,
        colWidths=[col_area_w, col_section_w, col_skill_w, col_desc_w, col_level_w],
        repeatRows=1,
    )
    base_table.setStyle(base_style)

    avail_height_first = table_top - bottom_margin
    base_table.wrapOn(c, avail_width, avail_height_first)
    full_rows = base_table._cellvalues
    full_heights = list(base_table._rowHeights)
    header_row = full_rows[0]
    header_meta = row_meta[0]

    # paginate
    pages = []
    max_height = avail_height_first
    cur_rows = [header_row]
    cur_meta_page = [header_meta]
    cur_height = full_heights[0]

    for i in range(1, len(full_rows)):
        rh = full_heights[i]
        meta_i = row_meta[i]
        if cur_height + rh > max_height and len(cur_rows) > 1:
            pages.append((cur_rows, cur_meta_page))
            cur_rows = [header_row, full_rows[i]]
            cur_meta_page = [header_meta, meta_i]
            cur_height = full_heights[0] + rh
            max_height = height - 80
        else:
            cur_rows.append(full_rows[i])
            cur_meta_page.append(meta_i)
            cur_height += rh

    pages.append((cur_rows, cur_meta_page))

    try:
        base_cmds = list(base_style.getCommands())
    except AttributeError:
        base_cmds = list(base_style._cmds)

    last_table_y = table_top
    for page_index, (page_rows, page_meta) in enumerate(pages):
        if page_index == 0:
            top_y = table_top
        else:
            c.showPage()
            cont_y = height - 60
            c.setFont("Helvetica-Bold", 11)
            c.setFillColor(colors.black)
            c.drawString(left, cont_y, "Learner's Character Progressive Report (continued)")
            top_y = cont_y - 18

        page_table = Table(
            page_rows,
            colWidths=[col_area_w, col_section_w, col_skill_w, col_desc_w, col_level_w],
            repeatRows=1,
        )

        extra_cmds = []
        for i, meta_row in enumerate(page_meta):
            if meta_row.get("type") == "area":
                extra_cmds.append(
                    ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#e0e0e0"))
                )
                extra_cmds.append(("TEXTCOLOR", (0, i), (-1, i), colors.black))
                extra_cmds.append(("FONT", (0, i), (-1, i), "Helvetica-Bold", 8))

        ts_page = TableStyle(base_cmds + extra_cmds)
        page_table.setStyle(ts_page)
        avail_h = top_y - bottom_margin
        w, h = page_table.wrapOn(c, avail_width, avail_h)
        table_y = top_y - h
        page_table.drawOn(c, left, table_y)
        last_table_y = table_y

    # ---------- FOOTER ----------
    footer_top = last_table_y - 15
    if footer_top < 80:
        c.showPage()
        footer_top = height - 80

    def _fmt_date(val):
        """Return dd/mm/yy string or '__________'."""
        if not val:
            return "__________"
        if hasattr(val, "strftime"):
            return val.strftime("%d/%m/%y")
        if isinstance(val, str):
            s = val.strip()
            if not s:
                return "__________"
            try:
                d = datetime.strptime(s[:10], "%Y-%m-%d").date()
                return d.strftime("%d/%m/%y")
            except Exception:
                return s
        return "__________"


    # ----- next-term dates: use meta, then fall back to term-dates helper -----
    next_begin = meta.get("next_term_begin")
    next_end = meta.get("next_term_end")

    if not next_begin or not next_end:
        try:
            nti = get_next_term_info(term, year) # same helper as report cards
        except Exception:
            nti = None
        if nti:
            if not next_begin:
                next_begin = nti.get("next_term_date") or next_begin
            if not next_end:
                next_end = nti.get("next_term_end_date") or next_end

    ntb_str = _fmt_date(next_begin)
    nte_str = _fmt_date(next_end)

    # ----- comments stay with their existing logic -----
    class_comment = (meta.get("class_teacher_comment") or "").strip()
    head_comment = (meta.get("head_teacher_comment") or "").strip()
    #fees = (meta.get("school_fees") or "").strip()
    #fees_dc = (meta.get("school_fees_daycare") or "").strip()

    # ----- SPECIAL COMMUNICATION: meta first, then overall overrides -----
    special = (meta.get("special_communication") or "").strip()
    if not special:
        sid = student.get("id") or student.get("student_id")
        if sid:
            try:
                conn = get_db_connection()
                overrides = fetch_overall_overrides(conn, sid, term, year)
                conn.close()
            except Exception:
                overrides = {}
            if overrides:
                special = (overrides.get("special_communication") or "").strip()


    class_comment_html = (class_comment or " ").replace("\n", "<br/>")
    head_comment_html = (head_comment or " ").replace("\n", "<br/>")

    lbl_style = ParagraphStyle(
        "lbl",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
    )
    txt_style = ParagraphStyle(
        "txt",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
    )

    footer_data = [
        [Paragraph("Class Teacher's comment:", lbl_style),
         Paragraph(class_comment_html, txt_style)],
        [Paragraph("Head Teacher's comment:", lbl_style),
         Paragraph(head_comment_html, txt_style)],
        [Paragraph("Next term begins on:", lbl_style),
         Paragraph(f"{ntb_str} will end on: {nte_str}", txt_style)],
        #[Paragraph("School fees:", lbl_style),
         #Paragraph(fees or " ", txt_style)],
        #[Paragraph("School fees + daycare:", lbl_style),
         #Paragraph(fees_dc or " ", txt_style)],
        [Paragraph("Special Communication:", lbl_style),
         Paragraph(special or " ", txt_style)],
    ]

    avail_footer_w = width - left - right
    footer_tbl = Table(
        footer_data,
        colWidths=[avail_footer_w * 0.25, avail_footer_w * 0.75],
    )
    footer_tbl.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    fw, fh = footer_tbl.wrapOn(c, avail_footer_w, 200)
    footer_y = footer_top - fh
    footer_tbl.drawOn(c, left, footer_y)
    c.showPage()




# =========================
# CHARACTER ASSESSMENT HELPERS END
# =========================


# ==== Payroll Hub & Actions ====


def _payroll_status(total, paid):
    """Return status string based on amounts."""
    paid = paid or 0
    total = total or 0
    if paid <= 0:
        return "not_paid"
    if paid < total:
        return "partially_paid"
    return "fully_paid"


def get_or_create_expense_category(conn, name="Salaries"):
    """Return category_id for `name`, creating it if needed(case-sensitive match)."""
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id FROM expense_categories WHERE name=%s", (name,))
    row = cur.fetchone()
    if row:
        cur.close()
        return row["id"]

    cur2 = conn.cursor()
    cur2.execute("INSERT INTO expense_categories(name) VALUES (%s)", (name,))
    conn.commit()
    new_id = cur2.lastrowid
    cur2.close()
    cur.close()
    return new_id


# -------- Financial Hub helpers --------

def _parse_finance_filters(req, ay):
    """Return a dict of filters with sensible defaults."""
    from datetime import date, timedelta

    f = {}
    f["term"] = (req.values.get("term") or ay["current_term"]).strip()
    f["year"] = int(req.values.get("year") or ay["year"])

    # Optional date overrides (YYYY-MM-DD); if provided, they apply to date filters
    f["from_date"] = (req.values.get("from_date") or "").strip() or None
    f["to_date"] = (req.values.get("to_date") or "").strip() or None

    # If no explicit dates, we filter by term/year; if dates given, we ignore term/year date constraints
    f["use_dates"] = bool(f["from_date"] and f["to_date"])
    return f


def _date_where(fragment_date_col, f):
    """Return(where_sql, params) for date range OR term/year filter."""
    if f["use_dates"]:
        return f"({fragment_date_col} BETWEEN %s AND %s)", [f["from_date"], f["to_date"]]
    else:
        # term/year filter
        return "(term = %s AND year = %s)", [f["term"], f["year"]]

def rebuild_summary_for_active_term():
    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)  # << dictionary cursor
    # active period
    cur.execute("SELECT year, current_term FROM academic_years WHERE is_active=1 LIMIT 1")
    act = cur.fetchone()
    if not act:
        raise RuntimeError("No active academic year/term.")
    year = int(act["year"])
    term = (act["current_term"] or "").strip().lower()
    term_no = 1 if term=='term 1' else 2 if term=='term 2' else 3 if term=='term 3' else None
    if term_no is None:
        raise RuntimeError(f"Unknown term label: {act['current_term']}")

    # get students
    cur.execute("SELECT id FROM students WHERE archived=0")
    ids = [row["id"] for row in (cur.fetchall() or [])]

    # recompute per student using callproc (auto-consumes results)
    for sid in ids:
        cur.callproc("recompute_fee_term_summary", (sid, year, term_no))

        # Some MySQL servers still emit resultsets; consume if present:
        for _ in cur.stored_results():
            _ = _.fetchall()

    conn.commit()
    cur.close(); conn.close()



def _class_options():
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)  # << important
        cur.execute("""
            SELECT DISTINCT class_name
            FROM students
            WHERE class_name IS NOT NULL
            ORDER BY class_name
        """)
        rows = cur.fetchall() or []
        return [r["class_name"] for r in rows if r.get("class_name")]
    finally:
        try: cur.close()
        except: pass
        conn.close()
        
def _term_no(t: str) -> int | None:
    t = (t or "").strip().lower()
    return 1 if t == "term 1" else 2 if t == "term 2" else 3 if t == "term 3" else None

   

def term_to_no(term_label: str | None) -> int | None:
    if not term_label:
        return None
    t = term_label.strip().lower()
    return {"term 1": 1, "term 2": 2, "term 3": 3}.get(t)

def term_no_to_label(n: int | None) -> str | None:
    if not n:
        return None
    return {1: "Term 1", 2: "Term 2", 3: "Term 3"}.get(int(n))
    
def get_active_year_term():
    ay = get_active_academic_year() or {}
    active_year = int(ay.get("year") or ay.get("active_year") or datetime.now().year)
    active_term = (ay.get("current_term") or ay.get("term") or "Term 1").strip()
    active_term_no = {"term 1": 1, "term 2": 2, "term 3": 3}.get(active_term.lower(), 1)
    return active_year, active_term, active_term_no
    
def is_academic_year_locked(year: int) -> bool:
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT is_locked FROM academic_years WHERE year=%s LIMIT 1", (int(year),))
        row = cur.fetchone() or {}
        cur.close()
        return bool(row.get("is_locked", 0))
    finally:
        conn.close()

def call_proc(sql: str, params: tuple | None = None):
    """Safe CALL wrapper for mysql-connector (buffered cursor, no next_result needed)."""
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True, buffered=True)
    try:
        cur.execute(sql, params or ())

        # Drain all returned result sets safely
        while True:
            try:
                _ = cur.fetchall()
            except Exception:
                pass

            # mysql.connector uses nextset(), not next_result()
            if not cur.nextset():
                break

        conn.commit()
    finally:
        try: cur.close()
        except: pass
        conn.close()


def _active_year() -> int:
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT year FROM academic_years WHERE is_active=1 LIMIT 1")
        return int((cur.fetchone() or {}).get("year") or 0)
    finally:
        try: cur.close()
        except: pass
        conn.close()

def _recompute_class_fallback(class_name: str, term_no: int | None):
    """
    Fallback when the class-level proc isn't installed.
    Recomputes active students in class for the active year.
    """
    year = _active_year()
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id FROM students WHERE archived=0 AND class_name=%s", (class_name,))
        sids = [r["id"] for r in (cur.fetchall() or [])]
    finally:
        try: cur.close()
        except: pass
        conn.close()

    for sid in sids:
        if term_no is None:
            for t in (1, 2, 3):
                call_proc("CALL recompute_fee_term_summary(%s,%s,%s)", (sid, year, t))
        else:
            call_proc("CALL recompute_fee_term_summary(%s,%s,%s)", (sid, year, term_no))

def _recompute_class(class_name: str, term_no: int | None):
    """
    Try fast class proc first; if missing, fall back to per-student recompute.
    """
    try:
        call_proc("CALL recompute_by_class_for_active_year(%s,%s)", (class_name, term_no))
    except Exception:
        _recompute_class_fallback(class_name, term_no)


def _fetch_finance_data(f):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    if f.get("use_dates"):
        date_from, date_to = f["from_date"], f["to_date"]

        # FEES (school fees)
        cur.execute("""
            SELECT p.id, p.student_id, s.student_number,
                   CONCAT(s.first_name,' ',COALESCE(s.Middle_name,''),' ',s.last_name) AS full_name,
                   p.method, p.term, p.year, p.date_paid, p.amount_paid
            FROM fees p
            JOIN students s ON s.id=p.student_id
            WHERE p.payment_type_norm IN ('school_fees','fees')
              AND (p.comment IS NULL OR LOWER(p.comment) NOT LIKE '%void%')
              AND p.date_paid BETWEEN %s AND %s
            ORDER BY p.date_paid, p.id
        """, (date_from, date_to))
        fees_rows = cur.fetchall() or []

        # REQUIREMENTS
        cur.execute("""
            SELECT p.id, p.student_id, s.student_number,
                   CONCAT(s.first_name,' ',COALESCE(s.Middle_name,''),' ',s.last_name) AS full_name,
                   COALESCE(p.requirement_name,'') AS requirement_name,
                   p.method, p.term, p.year, p.date_paid, p.amount_paid
            FROM fees p
            JOIN students s ON s.id=p.student_id
            WHERE p.payment_type_norm IN ('requirements','requirement')
              AND (p.comment IS NULL OR LOWER(p.comment) NOT LIKE '%void%')
              AND p.date_paid BETWEEN %s AND %s
            ORDER BY p.date_paid, p.id
        """, (date_from, date_to))
        req_rows = cur.fetchall() or []

        # OTHER INCOME
        cur.execute("""
            SELECT id, source, description, recorded_by, term, year, date_received, amount
            FROM other_income
            WHERE date_received BETWEEN %s AND %s
            ORDER BY date_received, id
        """, (date_from, date_to))
        other_rows = cur.fetchall() or []

        # EXPENSES
        cur.execute("""
            SELECT e.id, e.date_spent,
                   COALESCE(ec.name, '') AS category,
                   e.description, e.type, e.recorded_by, e.term, e.year, e.amount
            FROM expenses e
            LEFT JOIN expense_categories ec ON ec.id = e.category_id
            WHERE e.date_spent BETWEEN %s AND %s
            ORDER BY e.date_spent, e.id
        """, (date_from, date_to))
        exp_rows = cur.fetchall() or []

    else:
        year, term_no = f["year"], _term_no(f["term"])

        cur.execute("""
            SELECT p.id, p.student_id, s.student_number,
                   CONCAT(s.first_name,' ',COALESCE(s.Middle_name,''),' ',s.last_name) AS full_name,
                   p.method, p.term, p.year, p.date_paid, p.amount_paid
            FROM fees p
            JOIN students s ON s.id=p.student_id
            WHERE p.payment_type_norm IN ('school_fees','fees')
              AND (p.comment IS NULL OR LOWER(p.comment) NOT LIKE '%void%')
              AND p.year=%s AND p.term_no=%s
            ORDER BY p.date_paid, p.id
        """, (year, term_no))
        fees_rows = cur.fetchall() or []

        cur.execute("""
            SELECT p.id, p.student_id, s.student_number,
                   CONCAT(s.first_name,' ',COALESCE(s.Middle_name,''),' ',s.last_name) AS full_name,
                   COALESCE(p.requirement_name,'') AS requirement_name,
                   p.method, p.term, p.year, p.date_paid, p.amount_paid
            FROM fees p
            JOIN students s ON s.id=p.student_id
            WHERE p.payment_type_norm IN ('requirements','requirement')
              AND (p.comment IS NULL OR LOWER(p.comment) NOT LIKE '%void%')
              AND p.year=%s AND p.term_no=%s
            ORDER BY p.date_paid, p.id
        """, (year, term_no))
        req_rows = cur.fetchall() or []

        cur.execute("""
            SELECT id, source, description, recorded_by, term, year, date_received, amount
            FROM other_income
            WHERE year=%s AND term_no=%s
            ORDER BY date_received, id
        """, (year, term_no))
        other_rows = cur.fetchall() or []

        cur.execute("""
            SELECT e.id, e.date_spent,
                   COALESCE(ec.name, '') AS category,
                   e.description, e.type, e.recorded_by, e.term, e.year, e.amount
            FROM expenses e
            LEFT JOIN expense_categories ec ON ec.id = e.category_id
            WHERE e.year=%s AND e.term_no=%s
            ORDER BY e.date_spent, e.id
        """, (year, term_no))
        exp_rows = cur.fetchall() or []

    cur.close(); conn.close()

    fees_total         = sum(r.get("amount_paid") or 0 for r in fees_rows)
    requirements_total = sum(r.get("amount_paid") or 0 for r in req_rows)
    other_income_total = sum(r.get("amount") or 0 for r in other_rows)
    expenses_total     = sum(r.get("amount") or 0 for r in exp_rows)
    income_total       = fees_total + requirements_total + other_income_total
    net_total          = income_total - expenses_total

    totals = {
        "fees_total": round(fees_total, 2),
        "requirements_total": round(requirements_total, 2),
        "other_income_total": round(other_income_total, 2),
        "income_total": round(income_total, 2),
        "expenses_total": round(expenses_total, 2),
        "net_total": round(net_total, 2),
    }
    return fees_rows, req_rows, other_rows, exp_rows, totals







def _balance_sheet_snapshot(f):
    """
    Collation-safe & fast snapshot.
    - Uses normalized payment_type_norm with explicit COLLATE to avoid 'illegal mix of collations'
    - Sums period cash-in (fees+requirements+other_income), period expenses
    - Receivables from fee_term_summary (precomputed)
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # choose one collation you consistently want to use in comparisons
    # utf8mb4_unicode_ci is a safe general choice
    COLL = "utf8mb4_unicode_ci"

    if f.get("use_dates"):
        params = [f["from_date"], f["to_date"]]

        # FEES (school fees)
        cur.execute(f"""
            SELECT COALESCE(SUM(p.amount_paid),0) AS t
            FROM fees p
            WHERE (p.comment IS NULL OR LOWER(p.comment) NOT LIKE '%void%')
              AND (p.date_paid BETWEEN %s AND %s)
              AND (p.payment_type_norm COLLATE {COLL}) IN ('fees','school_fees')
        """, params)
        fees_in_all = float((cur.fetchone() or {}).get("t") or 0.0)

        # REQUIREMENTS
        cur.execute(f"""
            SELECT COALESCE(SUM(p.amount_paid),0) AS t
            FROM fees p
            WHERE (p.comment IS NULL OR LOWER(p.comment) NOT LIKE '%void%')
              AND (p.date_paid BETWEEN %s AND %s)
              AND (p.payment_type_norm COLLATE {COLL}) IN ('requirements','requirement')
        """, params)
        req_in_all = float((cur.fetchone() or {}).get("t") or 0.0)

        # OTHER INCOME
        cur.execute("""
            SELECT COALESCE(SUM(oi.amount),0) AS t
            FROM other_income oi
            WHERE oi.date_received BETWEEN %s AND %s
        """, params)
        other_in = float((cur.fetchone() or {}).get("t") or 0.0)

        # EXPENSES (period)
        cur.execute("""
            SELECT COALESCE(SUM(e.amount),0) AS t
            FROM expenses e
            WHERE e.date_spent BETWEEN %s AND %s
        """, params)
        expenses = float((cur.fetchone() or {}).get("t") or 0.0)

        # ACTIVE-only cash-in for fees+requirements
        cur.execute(f"""
            SELECT COALESCE(SUM(p.amount_paid),0) AS t
            FROM fees p
            JOIN students s ON s.id = p.student_id AND s.archived=0
            WHERE (p.comment IS NULL OR LOWER(p.comment) NOT LIKE '%void%')
              AND (p.date_paid BETWEEN %s AND %s)
              AND (p.payment_type_norm COLLATE {COLL}) IN ('fees','school_fees','requirements','requirement')
        """, params)
        fr_active = float((cur.fetchone() or {}).get("t") or 0.0)

        cash_in_all = fees_in_all + req_in_all + other_in
        cash_in_active = fr_active + other_in

        # Receivables (use active term/year derived from dates? keep from current filters)
        # If your UI always passes f["term"]/f["year"] with use_dates, prefer those.
        year = f.get("year")
        tno  = _term_no(f.get("term")) if f.get("term") else None

    else:
        year = int(f["year"])
        tno  = _term_no(f["term"])

        # FEES (school fees)
        cur.execute(f"""
            SELECT COALESCE(SUM(p.amount_paid),0) AS t
            FROM fees p
            WHERE (p.comment IS NULL OR LOWER(p.comment) NOT LIKE '%void%')
              AND p.year=%s AND p.term_no=%s
              AND (p.payment_type_norm COLLATE {COLL}) IN ('fees','school_fees')
        """, (year, tno))
        fees_in_all = float((cur.fetchone() or {}).get("t") or 0.0)

        # REQUIREMENTS
        cur.execute(f"""
            SELECT COALESCE(SUM(p.amount_paid),0) AS t
            FROM fees p
            WHERE (p.comment IS NULL OR LOWER(p.comment) NOT LIKE '%void%')
              AND p.year=%s AND p.term_no=%s
              AND (p.payment_type_norm COLLATE {COLL}) IN ('requirements','requirement')
        """, (year, tno))
        req_in_all = float((cur.fetchone() or {}).get("t") or 0.0)

        # OTHER INCOME
        cur.execute("""
            SELECT COALESCE(SUM(oi.amount),0) AS t
            FROM other_income oi
            WHERE oi.year=%s AND oi.term_no=%s
        """, (year, tno))
        other_in = float((cur.fetchone() or {}).get("t") or 0.0)

        # EXPENSES
        cur.execute("""
            SELECT COALESCE(SUM(e.amount),0) AS t
            FROM expenses e
            WHERE e.year=%s AND e.term_no=%s
        """, (year, tno))
        expenses = float((cur.fetchone() or {}).get("t") or 0.0)

        # ACTIVE only
        cur.execute(f"""
            SELECT COALESCE(SUM(p.amount_paid),0) AS t
            FROM fees p
            JOIN students s ON s.id = p.student_id AND s.archived=0
            WHERE (p.comment IS NULL OR LOWER(p.comment) NOT LIKE '%void%')
              AND p.year=%s AND p.term_no=%s
              AND (p.payment_type_norm COLLATE {COLL}) IN ('fees','school_fees','requirements','requirement')
        """, (year, tno))
        fr_active = float((cur.fetchone() or {}).get("t") or 0.0)

        cash_in_all = fees_in_all + req_in_all + other_in
        cash_in_active = fr_active + other_in

    # Receivables (precomputed)
    cur.execute("""
        SELECT COALESCE(SUM(overall_outstanding),0) AS t
        FROM fee_term_summary
        WHERE year=%s AND term_no=%s
    """, (year, tno))
    receivables_all = float((cur.fetchone() or {}).get("t") or 0.0)

    cur.execute("""
        SELECT COALESCE(SUM(fs.overall_outstanding),0) AS t
        FROM fee_term_summary fs
        JOIN students s ON s.id = fs.student_id AND s.archived=0
        WHERE fs.year=%s AND fs.term_no=%s
    """, (year, tno))
    receivables_active = float((cur.fetchone() or {}).get("t") or 0.0)

    cur.close(); conn.close()

    net_all    = (cash_in_all    - expenses) + receivables_all
    net_active = (cash_in_active - expenses) + receivables_active

    return {
        "cash_in_all": round(cash_in_all, 2),
        "cash_in_active": round(cash_in_active, 2),
        "expenses": round(expenses, 2),
        "receivables_all": round(receivables_all, 2),
        "receivables_active": round(receivables_active, 2),
        "net_all": round(net_all, 2),
        "net_active": round(net_active, 2),
        "cash_in": round(cash_in_all, 2),          # back-compat
        "receivables": round(receivables_all, 2),  # back-compat
        "net_position": round(net_all, 2),         # back-compat
    }


# --- Helper: compute ranking list for class/term/year (for single view) ---
def class_ranking(conn, class_name, term, year):
    """
    Rank students in a class by overall average of displayed totals:
    EOT if present else blended average.
    """
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT r.student_id AS sid,
               AVG(COALESCE(r.eot_mark, r.average_mark)) AS overall
        FROM reports r
        JOIN students s ON s.id= r.student_id
        WHERE s.class_name= %s AND r.term = %s AND r.year = %s
        GROUP BY r.student_id
        ORDER BY overall DESC
    """, (class_name, term, year))
    rows = cur.fetchall()
    cur.close()
    return rows  # list of rows with sid, overall

# --- Helper: grading legend rows ---


def fetch_grading_scale(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT grade, lower_limit, upper_limit, COALESCE(comment, '') AS comment
        FROM grading_scale
        ORDER BY lower_limit
    """)
    rows = cur.fetchall()
    cur.close()
    return rows


def ensure_join_columns():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    for name, typ in [
        ("date_joined", "VARCHAR(15)"),
        ("term_joined", "VARCHAR(10)"),
        ("year_of_joining", "INT"),
    ]:
        try:
            cur.execute(f"ALTER TABLE students ADD COLUMN {name} {typ}")
        except Exception:
            pass
    conn.commit()
    cur.close()
    conn.close()


def ensure_fees_method_column():
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("ALTER TABLE fees ADD COLUMN method TEXT DEFAULT 'N/A'")
        conn.commit()
    except mysql.connector.Error:
        pass
    finally:
        cur.close()
        conn.close()


def _is_safe_url(target):
    # Only allow redirects to our own host
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target or ""))
    return (test_url.scheme in ("http", "https")) and (ref_url.netloc == test_url.netloc)


def get_student(student_number=None, last_name=None):
    """Load ONE student by student_number OR(first matching) last_name."""
    conn = get_db_connection()
    if student_number:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, student_number, first_name, Middle_name, last_name,
                   class_name, stream, section
            FROM students
            WHERE student_number= %s AND archived = 0
        """, (student_number,))
        row = cur.fetchone()
        cur.close()
    elif last_name:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, student_number, first_name, Middle_name, last_name,
                   class_name, stream, section
            FROM students
            WHERE last_name LIKE % s AND archived = 0
            ORDER BY last_name, first_name
            LIMIT 1
        """, (f"%{last_name}%",))
        row = cur.fetchone()
        cur.close()
    else:
        row = None
    conn.close()
    return row


def ensure_comment_rules_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS comment_rules (
            id INT NOT NULL AUTO_INCREMENT,
            role ENUM('teacher','headteacher') NOT NULL,
            scope ENUM('subject','overall') NOT NULL,
            match_type ENUM('grade','division','range') NOT NULL,
            grade VARCHAR(10),
            division VARCHAR(15),
            lower_limit DECIMAL(10,2),
            upper_limit DECIMAL(10,2),
            class_name VARCHAR(15),
            level VARCHAR(20),
            term VARCHAR(20),
            template_text VARCHAR(100) NOT NULL,
            priority INT DEFAULT 100,
            active TINYINT NOT NULL DEFAULT 1,
            PRIMARY KEY (id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    if not _index_exists(conn, "comment_rules", "ix_comment_rules_lookup"):
        cur.execute("""
            CREATE INDEX ix_comment_rules_lookup
            ON comment_rules(role, scope, active, match_type)
        """)
    conn.commit()
    cur.close()



def ensure_comment_library_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS comment_library (
            id INT NOT NULL AUTO_INCREMENT,
            category ENUM('good','moderate','poor') NOT NULL,
            text VARCHAR(255) NOT NULL,
            role ENUM('teacher','headteacher') DEFAULT NULL,
            scope ENUM('subject','overall') DEFAULT NULL,
            uses INT NOT NULL DEFAULT 0,
            PRIMARY KEY (id),
            UNIQUE KEY uq_comment_library_text_role_scope (text, role, scope)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    conn.commit()
    cur.close()
    
def ensure_report_overall_overrides_schema(conn):
    """
    Stores per-learner manual overall comments and special communication
    for a given term/year.
    """
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS report_overall_overrides (
            id INT NOT NULL AUTO_INCREMENT,
            student_id INT NOT NULL,
            term VARCHAR(20) NOT NULL,
            year INT NOT NULL,
            teacher_overall_custom TEXT,
            head_overall_custom TEXT,
            special_communication TEXT,
            PRIMARY KEY (id),
            UNIQUE KEY uq_rep_overall_learner (student_id, term, year),
            CONSTRAINT fk_rep_overall_student
            FOREIGN KEY (student_id) REFERENCES students(id)
            ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    conn.commit()
    cur.close()


# optional: simple hierarchy (admin can do everything)
ROLE_IMPLIES = {
    "admin": {"admin", "bursar", "director", "headteacher", "dos", "clerk", "teacher"},
    # add more if you have seniors: e.g. "headteacher": {"teacher"}
}


def _normalize_roles(value) -> set[str]:
    """
    Accepts a string like 'bursar, clerk' or a list; returns a lowercased set.
    """
    if not value:
        return set()
    if isinstance(value, (list, tuple, set)):
        items = value
    else:
        items = str(value).split(",")
    return {r.strip().lower() for r in items if str(r).strip()}


def require_role(*allowed_roles):
    allowed = _normalize_roles(allowed_roles)

    @wraps(require_role)
    def decorator(fn):
        @wraps(fn)
        def wrapped(*args, **kwargs):
            # must be logged in
            user_id = session.get("user_id")
            raw_role = session.get("role")  # may be 'bursar, clerk'
            if not user_id or raw_role is None:
                flash("Please sign in.", "warning")
                return redirect(url_for("login", next=request.path))

            user_roles = _normalize_roles(raw_role)

            # expand with hierarchy: admin -> all, etc.
            expanded = set(user_roles)
            for r in list(user_roles):
                expanded |= ROLE_IMPLIES.get(r, set())

            if not (expanded & allowed):
                flash("You don't have permission to access this page.", "danger")
                return redirect(url_for("dashboard"))

            return fn(*args, **kwargs)
        return wrapped
    return decorator


def _asdict(row):
    """Return a plain dict for mapping-like rows(dictionary cursor rows)."""
    if row is None:
        return {}
    return dict(row) if isinstance(row, dict) else {}


def _mean(seq):
    vals = [float(x) for x in seq if x is not None]
    return round(sum(vals) / len(vals), 2) if vals else None


def ordinal(n):
    try:
        n = int(n)
    except Exception:
        return ""
    if 10 <= (n % 100) <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def _code_map_from_record_score(conn, student_id: int, term: str, year: int, column: str) -> dict:
    """
    Returns {'ENG': 80, 'MAT': 75, 'SCI': 90, 'SST': 88} for given column in record_score.
    Missing subject codes are omitted.
    """
    cur = conn.cursor(dictionary=True)
    cur.execute(
        f"""
        SELECT UPPER(COALESCE(sub.code, SUBSTR(sub.name, 1, 3))) AS code,
               r.{column} AS val
        FROM record_score r
        JOIN subjects sub ON sub.id= r.subject_id
        WHERE r.student_id= %s AND r.term = %s AND r.year = %s
        """,
        (student_id, term, year)
    )
    rows = cur.fetchall()
    cur.close()
    out = {}
    for r in rows:
        d = _asdict(r)
        code = (d.get("code") or "").upper().strip()
        if code in BIG4_CODES:
            out[code] = d.get("val")
    return out


def _code_map_from_midterms(conn, student_id: int, term: str, year: int, assessment_name: str) -> dict:
    """
    Reads a midterms row for the given assessment and returns a code -> mark map.
    """
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT eng, mat, sci, sst
        FROM midterms
        WHERE student_id = %s AND term = %s AND year = %s AND LOWER(TRIM(assessment)) = LOWER(TRIM(% s))
        LIMIT 1
    """, (student_id, term, year, assessment_name))
    r = cur.fetchone()
    cur.close()
    if not r:
        return {}
    return {
        "ENG": r.get("eng"),
        "MAT": r.get("mat"),
        "SCI": r.get("sci"),
        "SST": r.get("sst"),
    }


def _build_midterm_panel_dynamic(conn, student_id: int, term: str, year: int, include_eot_row: bool = True):
    """
    Builds a dynamic midterm panel for *all * subjects present in subjects table,
    but computes AGG/TOTAL only from the core four: ENG, MATH, SCI, SST.

    Returns a list of rows like:
      [
        {
          "assessment": "OTH" | "HP" | "BOT" | "MID" | "EOT",
          "per_subj": { <subject_id>: {"mark": x, "grade": "C3"} , ... },
          "agg": <sum of points for core subjects or None>,
          "total": <sum of marks for core subjects or None>,
        },
        ...
      ]
    """

    cur = conn.cursor(dictionary=True)

    # Get all subjects once (dynamic)
    cur.execute("SELECT id AS sid, name, code FROM subjects ORDER BY name")
    subs = cur.fetchall()
    cur.close()
    if not subs:
        return []

    # map subject_id -> True if core
    core_codes = {"ENG", "MATH", "SCI", "SST"}
    is_core_by_sid = {s["sid"]: (
        (s["code"] or "").upper() in core_codes) for s in subs}

    # helper: grading
    def grade_of(v):
        return grade_for_score(conn, v) if v is not None else None

    # helper: fetch values for a given column name into {sid: mark}
    def colvals_for(colname: str):
        # Pull only subjects that exist for this student/term/year
        cur.execute(
            f"""
            SELECT subject_id AS sid, MAX({colname}) AS val
            FROM record_score
            WHERE student_id=%s AND term=%s AND year=%s
            GROUP BY subject_id
            """,
            (student_id, term, year)
        )
        rows = cue.fetchall()
        return {r["sid"]: r["val"] for r in rows}

    # Build each assessment in the desired order
    spec = [
        ("OTH", "other_mark"),
        ("HP", "holiday_mark"),
        ("BOT", "bot_mark"),
        ("MID", "midterm_mark"),
        ("EOT", "eot_mark"),
    ]

    out = []
    for label, col in spec:
        colvals = colvals_for(col)
        if not colvals:
            # nothing recorded for this assessment at all
            continue

        # Build per-subject cells and accumulate only cores
        per_subj = {}
        core_marks = []
        core_points = []

        for s in subs:
            sid = s["sid"]
            v = colvals.get(sid)
            g = grade_of(v)
            per_subj[sid] = {"mark": v, "grade": g}

            if is_core_by_sid.get(sid) and v is not None:
                core_marks.append(v)
                if g in {'D1', 'D2', 'C3', 'C4', 'C5', 'C6', 'P7', 'P8', 'F9'}:
                    _pts = {'D1': 1, 'D2': 2, 'C3': 3, 'C4': 4,
                            'C5': 5, 'C6': 6, 'P7': 7, 'P8': 8, 'F9': 9}[g]
                    core_points.append(_pts)

        if not any(v is not None for v in colvals.values()):
            # no actual numbers anywhere — skip row
            continue

        out.append(dict(
            assessment=label,
            per_subj=per_subj,
            agg=(sum(core_points) if core_points else None),
            total=(sum(core_marks) if core_marks else None),
        ))

    if not include_eot_row:
        out = [r for r in out if r["assessment"] != "EOT"]

    return out


def _norm_section(sec: str) -> str:
    s = (sec or '').strip().lower()
    if s in ('day', 'd'):
        return 'Day'
    if s in ('boarding', 'board', 'b'):
        return 'Boarding'
    return (sec or '').strip()


def _expected_fees_for_student(conn, student_id: int, term: str, year: int) -> float:
    """
    Priority:
      1) fee_codes.amount when student has a fees_code (prefers year match if column exists)
      2) class_fees.amount for class_name + section (prefers year match if column exists)
      3) 0.0
    """
    cur = conn.cursor(dictionary=True)
    # --- student basics
    cur.execute(
        "SELECT class_name, section, fees_code FROM students WHERE id=%s",
        (student_id,)
    )
    st = cur.fetchone()
    cur.close()
    if not st:
        return 0.0

    class_name = (st["class_name"] or "").strip()
    section_raw = (st["section"] or "").strip().lower()
    if section_raw in ("day", "d"):
        section = "Day"
    elif section_raw in ("boarding", "board", "b"):
        section = "Boarding"
    else:
        section = ""  # unknown → let queries ignore/handle

    # --- 1) fee_codes by student's fees_code
    fees_code = ((st["fees_code"] or "").strip()
                 if "fees_code" in st.keys() else "")
    if fees_code and _table_exists(conn, "fee_codes"):
        # prefer year-specific if the column exists
        if column_exists(conn, "fee_codes", "year"):
            cur = conn.cursor(dictionary=True)
            cur.execute(
                "SELECT amount FROM fee_codes WHERE code=%s AND year=%s LIMIT 1",
                (fees_code, year)
            )
            row = cur.fetchone()
            cur.close()
            if row and row["amount"] is not None:
                return float(row["amount"])
        # fallback without year
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT amount FROM fee_codes WHERE code=%s LIMIT 1",
            (fees_code,)
        )
        row = cur.fetchone()
        cur.close()
        if row and row["amount"] is not None:
            return float(row["amount"])

    # --- 2) class_fees by class + section (prefer year if present)
    if _table_exists(conn, "class_fees"):
        if column_exists(conn, "class_fees", "year"):
            cur = conn.cursor(dictionary=True)
            cur.execute(
                """
                SELECT amount
                FROM class_fees
                WHERE class_name = %s
                  AND LOWER(section) = LOWER(%s)
                  AND year = %s
                LIMIT 1
                """,
                (class_name, section, year)
            )
            row = cur.fetchone()
            cur.close()
            if row and row["amount"] is not None:
                return float(row["amount"])
        # fallback without year
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT amount
            FROM class_fees
            WHERE class_name = %s
              AND LOWER(section) = LOWER(%s)
            LIMIT 1
            """,
            (class_name, section)
        )
        row = cur.fetchone()
        cur.close()
        if row and row["amount"] is not None:
            return float(row["amount"])

    # --- 3) nothing matched
    return 0.0


def _build_mid_row(conn, title: str, by_code: dict) -> dict:
    """
    by_code: {'ENG': mark, 'MAT': mark, 'SCI': mark, 'SST': mark}
    Returns a dict with marks + *_grade + *_comment + agg + total for the row.
    """
    row = {"assessment": title}
    points_map = {'D1': 1, 'D2': 2, 'C3': 3, 'C4': 4,
                  'C5': 5, 'C6': 6, 'P7': 7, 'P8': 8, 'F9': 9}
    agg_points = []

    for code in BIG4_CODES:
        val = by_code.get(code)
        grd = grade_for_score(conn, val) if val is not None else None
        cmt = comment_for_grade(conn, grd) if grd else None

        # pack with lowercase keys the template expects
        key = code.lower()  # eng/mat/sci/sst
        row[key] = val
        row[f"{key}_grade"] = grd or ""
        row[f"{key}_comment"] = cmt or ""

        if grd in points_map:
            agg_points.append(points_map[grd])

    # aggregate (only if all 4 grades exist)
    row["agg"] = sum(agg_points) if len(agg_points) == 4 else None

    # total (sum of numeric big-4 marks present)
    total_vals = [by_code.get(c)
                  for c in BIG4_CODES if by_code.get(c) is not None]
    row["total"] = round(sum(total_vals), 0) if total_vals else None
    return row

# ---------- main builder for the Mid-Term panel ----------


def midterm_rows_for_student(conn, student_id: int, term: str, year: int) -> list[dict]:
    """
    Returns a list of dict rows for the Mid-Term table:
      - Beginning of Term (BOT)
      - Holiday Package (if present)
      - Other Assessments (if present)
      - Mid of Term (stored midterm_mark or mean of BOT/HP/Other)
    Each row includes eng/mat/sci/sst, *_grade, *_comment, agg, total.
    """
    # source maps
    bot_map = _code_map_from_record_score(
        conn, student_id, term, year, "bot_mark")
    mid_map = _code_map_from_record_score(
        conn, student_id, term, year, "midterm_mark")
    holiday_map = _code_map_from_midterms(
        conn, student_id, term, year, "Holiday Package")
    other_map = _code_map_from_midterms(
        conn, student_id, term, year, "Other Assessments")

    rows = []

    if bot_map:
        rows.append(_build_mid_row(conn, "Beginning of Term (BOT)", bot_map))

    if holiday_map:
        rows.append(_build_mid_row(conn, "Holiday Package", holiday_map))

    if other_map:
        rows.append(_build_mid_row(conn, "Other Assessments", other_map))

    # Mid of Term (prefer stored midterm_mark; else mean of BOT/HP/Other per subject)
    final_mid = {}
    for code in BIG4_CODES:
        stored = mid_map.get(code)
        if stored is not None:
            final_mid[code] = stored
        else:
            final_mid[code] = _mean(
                [bot_map.get(code), holiday_map.get(code), other_map.get(code)])

    # Only add the Mid row if at least one subject has a value
    if any(final_mid.get(c) is not None for c in BIG4_CODES):
        rows.append(_build_mid_row(conn, "Mid of Term Exams", final_mid))

    return rows


def _ordinal(n: int | None) -> str | None:
    if n is None:
        return None
    s = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    if 10 <= (n % 100) <= 20:
        s = "th"
    return f"{n}{s}"


def _next_term_name_and_year(cur_term: str, cur_year: int):
    order = ["Term 1", "Term 2", "Term 3"]
    try:
        i = order.index(cur_term)
    except ValueError:
        return "Term 2", cur_year
    return (order[i+1], cur_year) if i < 2 else (order[0], cur_year + 1)


def grade_for_score(conn, score: float) -> str | None:
    """Look up D1..F9 from grading_scale where min_score <= score <= max_score."""
    if score is None:
        return None
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT grade
        FROM grading_scale
        WHERE %s BETWEEN lower_limit AND upper_limit
        ORDER BY (upper_limit - lower_limit) ASC
        LIMIT 1
    """, (score,))
    row = cur.fetchone()
    cur.close()
    return (row["grade"] if row else None)

# Used to calculate average for both single and batch


def fetch_report_rows(conn, student_id, term, year):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT sub.name AS subject, sub.code AS subject_code,
               r.eot_mark AS eot,
               COALESCE(r.average_mark, r.eot_mark) AS total_100,
               r.grade, r.comment, r.teacher_initial AS initials,
               r.teacher_remark, r.headteacher_remark
        FROM reports r
        JOIN subjects sub ON sub.id = r.subject_id
        WHERE r.student_id=%s AND r.term=%s AND r.year=%s
        ORDER BY sub.name
    """, (student_id, term, year))
    rows = cur.fetchall()
    cur.close()
    return rows


def _bot_mid_by_sid(conn, class_name, term, year):
    """returns dict[sid][code] = {'bot':x, 'mid':y} (codes are ENG/MATH/SCI/SST)"""
    src = detect_scores_table(conn)
    if not src:
        return {}
    # join subjects to derive codes
    if not col_exists(conn, "subjects", "code"):
        code_sql = "UPPER(SUBSTR(sub.name,1,3))"
    else:
        code_sql = "UPPER(sub.code)"
    cur = conn.cursor(dictionary=True)
    cur.execute(f"""
        SELECT r.student_id, {code_sql} AS code,
               MAX(r.bot_mark) AS bot, MAX(r.midterm_mark) AS mid
        FROM {src} r
        JOIN students s ON s.id = r.student_id
        JOIN subjects sub ON sub.id = r.subject_id
        WHERE s.archived=0 AND s.class_name=%s AND r.term=%s AND r.year=%s
        GROUP BY r.student_id, code
    """, (class_name, term, year))
    rows = cur.fetchall()
    cur.close()
    out = {}
    for r in rows:
        sid = r["student_id"]
        code = (r["code"] or "").upper()
        if code not in BIG4_CODES:  # only big-4 here for mid-term panel
            continue
        out.setdefault(sid, {})[code] = {"bot": _to_num(
            r["bot"]), "mid": _to_num(r["mid"])}
    return out


def _midterm_rows(conn, student_id, term, year):
    """
    Assemble the optional midterm table from record_score columns.
    We create up to 3 logical rows if any data exists:
      - 'Beginning of Term'
      - 'Mid of Term'
      - 'Holiday Package'
    Each row shows ENG/MAT/SCI/SST marks + derived grade.
    """
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT sub.code AS code,
               rs.bot_mark, rs.midterm_mark, rs.eot_mark,
               rs.holiday_mark, rs.other_mark, rs.ca_mark
        FROM record_score rs
        JOIN subjects sub ON sub.id = rs.subject_id
        WHERE rs.student_id=%s AND rs.term=%s AND rs.year=%s
    """, (student_id, term, year))
    r = cur.fetchall()
    cur.close()

    # Build code->marks map
    by_code = {}
    for row in r:
        code = (row["code"] or "").upper().strip()
        if code in BIG4_CODES:
            by_code[code] = dict(bot=row["bot_mark"],
                                 mid=row["midterm_mark"],
                                 eot=row["eot_mark"],
                                 holiday=row["holiday_mark"],
                                 other=row["other_mark"],
                                 ca=row["ca_mark"])

    def _row(label, key):
        """key in {'bot','mid','holiday'} -> row dict or None if empty"""
        vals = {c: (by_code.get(c, {}).get(key)) for c in BIG4_CODES}
        if all(v is None for v in vals.values()):
            return None
        # derive grade per subject
        grades = {c: grade_for_score(conn, vals[c]) for c in BIG4_CODES}
        # aggregate (Big4 points)
        points_map = {'D1': 1, 'D2': 2, 'C3': 3, 'C4': 4,
                      'C5': 5, 'C6': 6, 'P7': 7, 'P8': 8, 'F9': 9}
        pts = [points_map[g] for g in grades.values() if g in points_map]
        agg = sum(pts) if len(pts) == 4 else None
        total = sum([v for v in vals.values() if v is not None]
                    ) if any(vals.values()) else None
        return dict(
            assessment=label,
            eng=vals["ENG"], eng_grade=grades["ENG"],
            mat=vals["MATH"], mat_grade=grades["MATH"],
            sci=vals["SCI"], sci_grade=grades["SCI"],
            sst=vals["SST"], sst_grade=grades["SST"],
            agg=agg, total=total
        )

    rows = []
    for label, key in (("Beginning of Term", "bot"),
                       ("Mid of Term Exams", "mid"),
                       ("Holiday Package", "holiday")):
        rr = _row(label, key)
        if rr:
            rows.append(rr)
    return rows


# --- helpers used below (you already have these) ---
# grade_for_score(conn, score) -> 'D1'/'C3'/...
# comment_for_grade(conn, grade) -> 'Excellent'/'Good'/...
# BIG4_CODES = ['ENG','MATH','SCI','SST']

def _mid_panel_for_student(conn, student_id: int, term: str, year: int):
    """
    Build the midterm panel rows (BOT, Mid of Term, Holiday Package, Other Assessments)
    from *record_score* only. We never select *_grade columns from SQL; we compute them here.
    Returns a list of dicts with keys:
      assessment, eng, eng_grade, mat, mat_grade, sci, sci_grade, sst, sst_grade, agg, total
    Only includes assessments that actually have at least one score.
    """
    # Map subject_id -> code for this student (BIG4 only)
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, UPPER(TRIM(code)) AS code
        FROM subjects
        WHERE UPPER(TRIM(code)) IN ('ENG','MATH','SCI','SST')
    """)
    code_map = {r["id"]: r["code"] for r in cur.fetchall()}
    cur.close()

    # Pull all record_score rows for this student/term/year (raw marks only)
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT subject_id, bot_mark, midterm_mark, holiday_mark, other_mark, eot_mark
        FROM record_score
        WHERE student_id=%s AND term=%s AND year=%s
    """, (student_id, term, year))
    rows = cur.fetchall()
    cur.close()

    # Bucket marks by subject code
    marks = {code: {"BOT": None, "MID": None, "HP": None, "OTHER": None, "EOT": None}
             for code in ("ENG", "MATH", "SCI", "SST")}
    for r in rows:
        code = code_map.get(r["subject_id"])
        if not code:  # ignore non-BIG4
            continue
        if r["bot_mark"] is not None:
            marks[code]["BOT"] = r["bot_mark"]
        if r["midterm_mark"] is not None:
            marks[code]["MID"] = r["midterm_mark"]
        if r["holiday_mark"] is not None:
            marks[code]["HP"] = r["holiday_mark"]
        if r["other_mark"] is not None:
            marks[code]["OTHER"] = r["other_mark"]
        if r["eot_mark"] is not None:
            marks[code]["EOT"] = r["eot_mark"]

    def _row_from_bucket(name, key):
        """Build one output row for an assessment key (BOT/MID/HP/OTHER) if any value exists."""
        eng = marks["ENG"][key]
        mat = marks["MATH"][key]
        sci = marks["SCI"][key]
        sst = marks["SST"][key]
        # skip completely empty rows
        if all(v is None for v in (eng, mat, sci, sst)):
            return None

        def g(v):  # grade string for a mark (or empty)
            return grade_for_score(conn, v) if v is not None else ""

        # Optional agg/total on BIG4 for this assessment
        grade_points = {'D1': 1, 'D2': 2, 'C3': 3, 'C4': 4,
                        'C5': 5, 'C6': 6, 'P7': 7, 'P8': 8, 'F9': 9}
        gp = [grade_points.get(g(eng)), grade_points.get(g(mat)),
              grade_points.get(g(sci)), grade_points.get(g(sst))]
        agg = (sum(x for x in gp if isinstance(x, int))
               if all(isinstance(x, int) for x in gp) else "")
        total = (sum(v for v in (eng, mat, sci, sst) if v is not None)
                 if any(v is not None for v in (eng, mat, sci, sst)) else "")
        return {
            "assessment": name,
            "eng": eng, "eng_grade": g(eng),
            "mat": mat, "mat_grade": g(mat),
            "sci": sci, "sci_grade": g(sci),
            "sst": sst, "sst_grade": g(sst),
            "agg": agg, "total": total,
        }

    out = []
    # Only append rows that actually exist
    for label, key in (("Beginning of Term", "BOT"),
                       ("Mid of Term Exams", "MID"),
                       ("Holiday Package", "HP"),
                       ("Other Assessments", "OTHER")):
        row = _row_from_bucket(label, key)
        if row:
            out.append(row)
    return out


def _mid_panel_from_record_score(conn, student_id, term, year):
    """
    Builds a compact mid-term panel:
      - Only shows columns among BOT / HOLIDAY / MID that actually have data.
      - No initials.
      - One concise comment per subject: prefer MID, else HOLIDAY, else BOT.
      - Grades come from grading_scale per mark.
    """
    q = """
      SELECT sub.name AS subject, sub.code AS code, r.subject_id,
             MAX(r.bot_mark) AS bot,
             MAX(r.midterm_mark) AS mid,
             MAX(r.holiday_mark) AS holiday
      FROM record_score r
      JOIN subjects sub ON sub.id = r.subject_id
      WHERE r.student_id=%s AND r.term=%s AND r.year=%s
      GROUP BY r.subject_id
      ORDER BY sub.name
    """
    cur = conn.cursor(dictionary=True)
    cur.execute(q, (student_id, term, year))
    rows = cur.fetchall()
    cur.close()

    def pack(mark):
        if mark is None:
            return {"mark": None, "grade": "", "comment": ""}
        g = grade_for_score(conn, mark)
        c = comment_for_grade(conn, g) or ""
        return {"mark": mark, "grade": g, "comment": c}

    panel = []
    any_bot = any_mid = any_hol = False
    for r in rows:
        bot = pack(r["bot"])
        mid = pack(r["mid"])
        hol = pack(r["holiday"])
        if r["bot"] is not None:
            any_bot = True
        if r["mid"] is not None:
            any_mid = True
        if r["holiday"] is not None:
            any_hol = True
        comment_choice = mid["comment"] or hol["comment"] or bot["comment"] or ""
        panel.append({
            "subject": r["subject"],
            "code": r["code"],
            "BOT": bot,
            "HOLIDAY": hol,
            "MID": mid,
            "comment_choice": comment_choice
        })

    mid_cols = []
    if any_bot:
        mid_cols.append("BOT")
    if any_hol:
        mid_cols.append("HOLIDAY")
    if any_mid:
        mid_cols.append("MID")
    return panel, mid_cols


def comment_for_grade(conn, grade):
    """Use grading_scale.comment if present; else empty."""
    if not grade:
        return ""
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT comment FROM grading_scale WHERE grade=%s LIMIT 1", (grade,))
    row = cur.fetchone()
    cur.close()
    return (row["comment"] if row and row["comment"] else "") or ""


def guess_teacher_initials(conn, class_name, subject_id):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT initials FROM teachers
        WHERE class_name=%s AND (subjects LIKE '%'||%s||'%' OR subjects IS NULL)
        AND TRIM(COALESCE(initials,'')) <> ''
        LIMIT 1
    """, (class_name, str(subject_id)))
    row = cur.fetchone()
    cur.close()
    return row["initials"] if row else ""


def _is_big4(code_or_name: str) -> str | None:
    s = (code_or_name or "").upper().strip()
    # try code first
    if s in {"ENG", "ENGLISH"}:
        return "ENG"
    if s in {"MATH", "MATHEMATICS", "MAT"}:
        return "MATH"
    if s in {"SCI", "SCIENCE"}:
        return "SCI"
    if s in {"SST", "SOCIAL STUDIES", "SOCIAL-STUDIES"}:
        return "SST"
    # try name patterns
    if "ENGLISH" in s:
        return "ENG"
    if "MATH" in s or "MATHEMAT" in s:
        return "MATH"
    if "SCIENCE" in s:
        return "SCI"
    if "SOCIAL" in s or "SST" in s:
        return "SST"
    return None


# ---------- Overall average + Division (Big-4 only) ----------
def compute_overall_for_student(conn, student_id, term, year):
    """
    Overall average: mean of COALESCE(EOT, blended average) across all subjects.
    Aggregate/Division: ONLY ENG, MATH, SCI, SST; map grades -> points.
    Division bands: 4–12 => 1, 13–24 => 2, 25–29 => 3, 30–34 => 4, else U.
    """
    BIG4 = {"ENG", "MATH", "SCI", "SST"}
    points_map = {'D1': 1, 'D2': 2, 'C3': 3, 'C4': 4,
                  'C5': 5, 'C6': 6, 'P7': 7, 'P8': 8, 'F9': 9}

    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(sub.code, TRIM(UPPER(sub.name))) AS code,
               r.grade,
               COALESCE(r.eot_mark, r.average_mark) AS total_100
        FROM reports r
        JOIN subjects sub ON sub.id = r.subject_id
        WHERE r.student_id=%s AND r.term=%s AND r.year=%s
    """, (student_id, term, year))
    rows = cur.fetchall()
    cur.close()

    disp = [r["total_100"] for r in rows if r["total_100"] is not None]
    avg_overall = round(sum(disp)/len(disp), 2) if disp else None

    agg_points = []
    for r in rows:
        code = (r["code"] or "").upper().strip()
        grd = (r["grade"] or "").upper().strip()
        if code in BIG4 and grd in points_map:
            agg_points.append(points_map[grd])

    aggregate = sum(agg_points) if len(agg_points) == 4 else None

    division = None
    if aggregate is not None:
        if 4 <= aggregate <= 12:
            division = "1"
        elif 13 <= aggregate <= 24:
            division = "2"
        elif 25 <= aggregate <= 29:
            division = "3"
        elif 30 <= aggregate <= 34:
            division = "4"
        else:
            division = "U"

    return avg_overall, division, aggregate
    


def load_comment_library_groups(conn):
    """
    Load teacher + headteacher overall comments from comment_library
    and group them by performance band.

    Returns: (teacher_lib, head_lib)
      teacher_lib = {"excellent": [...], "moderate": [...], "poor": [...]}
      head_lib = same structure
    """
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, category, text, role, scope, uses
        FROM comment_library
        WHERE scope = 'overall' AND role IN ('teacher','headteacher')
        ORDER BY category, id
    """)
    rows = cur.fetchall() or []
    cur.close()

    def blank_groups():
        return {"excellent": [], "moderate": [], "poor": []}

    teacher_lib = blank_groups()
    head_lib = blank_groups()

    for r in rows:
        # map DB category -> band key
        cat = (r.get("category") or "").lower()
        if cat == "good":
            band = "excellent"
        elif cat == "moderate":
            band = "moderate"
        elif cat == "poor":
            band = "poor"
        else:
            continue

        text = r["text"] or ""
        if not text:
            continue

        target = teacher_lib if r["role"] == "teacher" else head_lib
        target[band].append(text)

    return teacher_lib, head_lib


def pick_comment_template(
    conn,
    *,
    role: str,
    scope: str,
    division=None,
    average=None,
    class_name=None,
    term=None,
    student_id=None # NEW
):
    """
    Returns the best-matching template_text (string) or None.
    Caller owns `conn` (we do not open/close it here).
    """
    ensure_comment_rules_schema(conn)

    q = """
        SELECT *
          FROM comment_rules
         WHERE role=%s AND scope=%s AND active=1
    """
    params = [role, scope]
    cur = conn.cursor(dictionary=True)
    cur.execute(q, params)
    rows = cur.fetchall()
    cur.close()

    ranked = []

    for r in rows:
        mtype = r["match_type"]

        # --- student filter: specific learner rules only apply to that learner
        if r.get("student_id") is not None:
            if student_id is None or int(r["student_id"]) != int(student_id):
                continue

        # validate essentials
        if mtype == "division" and r["division"] is None:
            continue
        if mtype == "range" and (r["lower_limit"] is None or r["upper_limit"] is None):
            continue

        # logical match
        fits = False
        if mtype == "division" and division is not None:
            try:
                fits = int(division) == int(r["division"])
            except Exception:
                fits = False
        elif mtype == "range" and average is not None:
            try:
                a = float(average)
                fits = float(r["lower_limit"]) <= a <= float(r["upper_limit"])
            except Exception:
                fits = False
        elif mtype == "grade":
            if isinstance(average, str) and r.get("grade"):
                fits = (
                    average.strip().upper()
                    == r["grade"].strip().upper()
                )

        if not fits:
            continue

        # specificity: student > class > any, term > any
        student_spec = (
            0
            if (r.get("student_id") and student_id and int(r["student_id"]) == int(student_id))
            else (1 if r.get("student_id") else 2)
        )
        class_spec = (
            0
            if (r.get("class_name") and class_name and r["class_name"] == class_name)
            else (1 if r.get("class_name") else 2)
        )
        term_spec = (
            0
            if (r.get("term") and term and r["term"] == term)
            else (1 if r.get("term") else 2)
        )
        priority = int(r["priority"] or 100)

        rank = (student_spec, priority, class_spec, term_spec, int(r["id"]))
        ranked.append((rank, r))

    if not ranked:
        return None

    ranked.sort(key=lambda x: x[0])
    base_text = (ranked[0][1].get("template_text") or "").strip()
    if not base_text:
        return None

    # ---------------- PERSONALISATION BLOCK ----------------
    # If we know the learner, prefix their first name.
    if student_id is not None:
        try:
            c2 = conn.cursor(dictionary=True)
            c2.execute(
                "SELECT first_name FROM students WHERE id=%s LIMIT 1",
                (student_id,),
            )
            row = c2.fetchone()
            c2.close()

            first = (row.get("first_name") or "").strip() if row else ""
            if first:
                # Avoid double-name if template already starts with the name
                if not base_text.lower().startswith(first.lower()):
                    # Make sure the bit after the name starts nicely
                    # e.g. "is an outstanding learner."
                    # -> "Moses is an outstanding learner."
                    # If template already starts with capital, leave as is.
                    return f"{first} {base_text[0].lower() + base_text[1:] if base_text and base_text[0].isupper() else base_text}"
        except Exception:
            # Any error -> just fall back to base_text
            pass
    # -------------------------------------------------------

    return base_text


def autofill_head_comment(student_id, class_name, term, year):
    conn = get_db_connection()
    avgm, division = compute_overall_for_student(conn, student_id, term, year)

    text = pick_comment_template(
        conn,
        role="headteacher",
        scope="overall",
        division=division,
        average=avgm,
        class_name=class_name,
        term=term
    )
    if text:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
          UPDATE reports
          SET headteacher_remark = %s
          WHERE student_id=%s AND term=%s AND year=%s
        """, (text, student_id, term, year))
        conn.commit()
        cur.close()
    conn.close()


def bursary_total(student_id, term, year):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(SUM(amount),0) AS total
        FROM bursaries
        WHERE student_id=%s AND term=%s AND year=%s
    """, (student_id, term, year))
    r = cur.fetchone()
    cur.close()
    conn.close()
    return float(r["total"] or 0)


def paid_sum(student_id, term, year, payment_type="school_fees") -> float:
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT COALESCE(SUM(amount_paid), 0) AS t
            FROM fees
            WHERE student_id = %s AND term = %s AND year = %s AND payment_type = %s
        """, (student_id, term, year, payment_type))
        r = cur.fetchone()  # tuple cursor -> (t,)
        cur.close()
        val = r[0] if r else 0
        return float(val or 0.0)
    finally:
        conn.close()


def carried_forward(student_id, term, year):
    """Outstanding before the active term/year."""
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(expected_amount,0) AS expected_amount,
               COALESCE(bursary_amount,0) AS bursary_amount,
               COALESCE(amount_paid,0) AS amount_paid,
               term, year
        FROM fees
        WHERE student_id=%s AND (
              year < %s
           OR (year = %s AND
               (CASE term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 99 END)
             < (CASE %s WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 99 END))
        )
    """, (student_id, year, year, term))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    outstanding = 0.0
    for r in rows:
        outstanding += (float(r["expected_amount"]) -
                        float(r["bursary_amount"]) - float(r["amount_paid"]))
    return max(outstanding, 0.0)


def requirements_due(student):
    """Requirements configured for the student's class (term-aware if column exists)."""
    ay = get_active_ay()
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SHOW COLUMNS FROM requirements")
    cols = [c["name"] for c in cur.fetchall()]
    cur.close()
    if "term" in cols:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, name, qty, amount, COALESCE(term,'') AS term
            FROM requirements
            WHERE class_name = %s
              AND (term IS NULL OR term = %s)
            ORDER BY name
        """, (student["class_name"], ay["term"]))
        rows = cur.fetchall()
        cur.close()
    else:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, name, qty, amount, '' AS term
            FROM requirements
            WHERE class_name = %s
            ORDER BY name
        """, (student["class_name"],))
        rows = cur.fetchall()
        cur.close()
    conn.close()
    return rows


def requirements_paid_sum(student_id, term, year):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(SUM(amount_paid),0) AS total
        FROM fees
        WHERE student_id=%s AND term=%s AND year=%s AND payment_type='requirements'
    """, (student_id, term, year))
    r = cur.fetchone()
    cur.close()
    conn.close()
    return float(r["total"] or 0)


def log_action(user_id, action):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "INSERT INTO audit_trail (user_id, action) VALUES (%s, %s)", (user_id, action))
    conn.commit()
    cur.close()
    conn.close()


def seed_classes():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute('''CREATE TABLE IF NOT EXISTS classes (
        id INT AUTO_INCREMENT PRIMARY KEY,
        class_name VARCHAR(15) NOT NULL,
        level VARCHAR(20),
        stream VARCHAR(15),
        UNIQUE KEY uq_classes_class_stream (class_name, COALESCE(stream, ''))
    )''')

    # Clear old records
    cur.execute("DELETE FROM classes")

    # Insert fresh P1–P7 Stream A
    cur.executemany(
        "INSERT INTO classes (class_name, level, stream) VALUES (%s, %s, %s)",
        [
            ('Baby', 'Nursery', 'A'),
            ('Middle', 'Nursery', 'A'),
            ('Top', 'Nursery', 'A'),
            ('P1', 'Primary', 'A'),
            ('P2', 'Primary', 'A'),
            ('P3', 'Primary', 'A'),
            ('P4', 'Primary', 'A'),
            ('P5', 'Primary', 'A'),
            ('P6', 'Primary', 'A'),
            ('P7', 'Primary', 'A')
        ]
    )

    # Unique index
    cur.execute('''CREATE UNIQUE INDEX IF NOT EXISTS uq_classes_class_stream
                 ON classes(class_name, stream)''')

    conn.commit()
    cur.close()
    conn.close()
    print("✅ Classes seeded successfully")

    # helpful uniqueness to avoid duplicates per class/term/item
    cur = conn.cursor(dictionary=True)
    cur.execute("""
      CREATE UNIQUE INDEX IF NOT EXISTS uq_requirements_class_term_name
      ON requirements(class_name, COALESCE(term,''), name)
    """)
    cur.close()
    # fees table: add requirement_name column (optional, nice to keep what was paid)
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("ALTER TABLE fees ADD COLUMN requirement_name TEXT")
        cur.close()
    except Exception:
        pass  # already exists or not needed

    conn.commit()
    conn.close()


def resolve_subject_id(conn, subject_code=None, subject_name=None):
    if subject_code:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id FROM subjects WHERE code = %s", (subject_code,))
        row = cur.fetchone()
        cur.close()
        if row:
            return row["id"]
    if subject_name:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id FROM subjects WHERE name = %s", (subject_name,))
        row = cur.fetchone()
        cur.close()
        if row:
            return row["id"]
    return None


def resolve_student_id(conn, student_number=None):
    if not student_number:
        return None
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id FROM students WHERE student_number = %s",
                (student_number,))
    row = cur.fetchone()
    cur.close()
    return row["id"] if row else None


def get_class_requirements(class_name: str, term: str | None):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, name, qty, amount, term
        FROM requirements
        WHERE class_name = %s
          AND (term = %s OR term IS NULL OR term = '')
        ORDER BY name
    """, (class_name, term))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def get_student_by_id(sid: int):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, student_number, first_name, Middle_name, last_name,
               class_name, stream, section
        FROM students WHERE id = %s
    """, (sid,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row


def populate_default_expense_categories():
    default_categories = [
        'Salaries',
        'Stationery',
        'Utilities',
        'Transport',
        'Maintenance',
        'Service Providers',
        'Uniforms',
        'Examinations',
        'Meals',
        'Office supplies',
        'Medical',
        'Bonus',
        'Allowance',
        'Electricity',
        'Teachers Rent',
        'Water',
        'Religion',
        'Staf Welfare',
        'P7 Budget',
        'Outing',
        'Directors Budget',
        'Loans',
        'School Functions',
        'Donations',
        'Construction 1',
        'Construction 2',
        'Sports',
        'Computers and Printers',
        'Medical',
        'Sanitation',
        'Vans Repair',
        'Fuel',
        'Kitchen',
        'Miscellaneous'
    ]

    conn = get_db_connection()

    for category in default_categories:
        try:
            cur = conn.cursor(dictionary=True)
            cur.execute(
                "INSERT IGNORE INTO expense_categories (name) VALUES (%s)", (category,))
        except:
            continue  # Skip any insert error
    conn.commit()
    cur.close()
    conn.close()
    print("Default expense categories inserted.")


def get_class_fee(conn, class_name: str, section: str | None, level: str | None = None) -> float:
    """Return amount from class_fees for (class_name, section[, level]). 0.0 if not found."""
    sec = norm_section(section) or ""
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT amount
          FROM class_fees
         WHERE class_name = %s
           AND LOWER(section) = LOWER(%s)
           AND (level IS NULL OR level = %s)
         LIMIT 1
        """,
        (class_name or "", sec, level),
    )
    row = cur.fetchone()
    cur.close()
    return float(row["amount"]) if row and row["amount"] is not None else 0.0


def ensure_bursaries_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS bursaries (
            id INT AUTO_INCREMENT PRIMARY KEY,
            student_id INT NOT NULL,
            term VARCHAR(10) NOT NULL,
            year INT NOT NULL,
            amount DOUBLE NOT NULL,
            sponsor_name VARCHAR(30) NOT NULL,
            granted_on DATE DEFAULT (CURRENT_DATE),
            UNIQUE KEY uq_bursary(student_id, year, term),
            INDEX ix_bursaries_student (student_id),
            INDEX ix_bursaries_period (term, year),
            CONSTRAINT fk_bur_student FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()




# ======================= TRANSPORT (as Other Income) =========================
# Uses your existing `other_income` table exactly.


TRANSPORT_DESC = "Transport"  # stored in other_income.description

# ---------- 1) Schema (unchanged from your version) ----------


def ensure_transport_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
      CREATE TABLE IF NOT EXISTS transport_routes (
        id INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
        name VARCHAR(30) NOT NULL,
        tr DOUBLE NOT NULL DEFAULT 0,
        UNIQUE KEY uq_transport_routes_name (name)
      ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
    """)

    cur.execute("""
      CREATE TABLE IF NOT EXISTS transport_subscriptions (
        id INT AUTO_INCREMENT PRIMARY KEY,
        student_id INT NOT NULL,
        route_id INT NOT NULL,
        start_term VARCHAR(15) NOT NULL,
        start_year INT NOT NULL,
        active TINYINT NOT NULL DEFAULT 1,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(student_id, route_id, start_term, start_year),
        FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE,
        FOREIGN KEY(route_id) REFERENCES transport_routes(id) ON DELETE CASCADE
      ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
    """)
    conn.commit()
    cur.close()


def transport_get_routes(conn=None):
    """
    Return all transport routes ordered by name.
    Ensures the transport tables exist before querying.
    """
    close_after = False
    if conn is None:
        conn = get_db_connection()
        close_after = True

    # make sure tables exist
    ensure_transport_schema(conn)

    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT id, name, fare_per_term FROM transport_routes ORDER BY name"
    )
    rows = cur.fetchall()
    cur.close()
    if close_after:
        conn.close()
    return rows


def get_class_requirements_without_transport(class_name: str, term: str):
    """
    Your original requirements (no transport). We simply filter names that look like 'Transport (...)'.
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, class_name, term, name, qty, amount
        FROM requirements
        WHERE class_name=%s AND term=%s
          AND (name NOT LIKE 'Transport (%' AND name NOT LIKE 'Transport% - %')
        ORDER BY name
    """, (class_name, term))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def _term_no_from_term(term: str) -> int:
    """Fallback if term_to_no() isn't available everywhere."""
    try:
        tn = term_to_no(term)
        if tn:
            return int(tn)
    except Exception:
        pass

    t = (term or "").strip().lower()
    if t == "term 1":
        return 1
    if t == "term 2":
        return 2
    if t == "term 3":
        return 3
    return 1


def transport_subscribe(student_id: int, route_id: int, term: str, year: int):
    """
    Subscribe (or re-activate) a student's transport.
    Requires a UNIQUE KEY to support ON DUPLICATE KEY UPDATE.
    Recommended: UNIQUE KEY uq_transport_student_active (student_id)
    or (student_id, route_id) depending on your design.
    """
    term_no = _term_no_from_term(term)

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            INSERT INTO transport_subscriptions
              (student_id, route_id, start_term, start_year, active, term_no)
            VALUES (%s, %s, %s, %s, 1, %s)
            ON DUPLICATE KEY UPDATE
              route_id = VALUES(route_id),
              start_term = VALUES(start_term),
              start_year = VALUES(start_year),
              term_no = VALUES(term_no),
              active = 1
        """, (student_id, route_id, term, year, term_no))
        conn.commit()
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def transport_unsubscribe(student_id: int, route_id: int | None = None):
    """
    Unsubscribe transport (disable active subscription).
    If route_id is provided, it disables that specific route.
    If route_id is None, it disables ANY active subscription for the student.
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        if route_id:
            cur.execute("""
              UPDATE transport_subscriptions
                 SET active = 0
               WHERE student_id = %s
                 AND route_id = %s
                 AND active = 1
            """, (student_id, route_id))
        else:
            cur.execute("""
              UPDATE transport_subscriptions
                 SET active = 0
               WHERE student_id = %s
                 AND active = 1
            """, (student_id,))
        conn.commit()
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def transport_subscription_info(student_id: int, term: str, year: int) -> dict:
    """
    Returns:
      {
        'has_sub': bool,
        'route_id': int,
        'route_name': str,
        'fare_per_term': float
      }

    Logic: active subscription that started on/before the selected term/year.
    Uses term_no for correct ordering/comparison.
    """
    sel_term_no = _term_no_from_term(term)

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT tr.id AS route_id, tr.name AS route_name, tr.fare_per_term
            FROM transport_subscriptions ts
            JOIN transport_routes tr ON tr.id = ts.route_id
            WHERE ts.student_id = %s
              AND ts.active = 1
              AND (
                    ts.start_year < %s
                 OR (ts.start_year = %s AND COALESCE(ts.term_no, 0) <= %s)
              )
            ORDER BY ts.start_year DESC,
                     COALESCE(ts.term_no, 0) DESC,
                     ts.created_at DESC
            LIMIT 1
        """, (student_id, year, year, sel_term_no))
        row = cur.fetchone()
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()

    if not row:
        return {"has_sub": False, "route_id": 0, "route_name": "", "fare_per_term": 0.0}

    return {
        "has_sub": True,
        "route_id": int(row.get("route_id") or 0),
        "route_name": (row.get("route_name") or ""),
        "fare_per_term": float(row.get("fare_per_term") or 0.0),
    }


def build_virtual_transport_requirement(student_id: int, term: str, year: int):
    """
    If subscribed, returns a synthetic requirement row for UI:
      {id: 't-<route_id>', name: 'Transport - <route>', qty:1, amount: fare}
    Else returns None.
    """
    info = transport_subscription_info(student_id, term, year)
    if not info["has_sub"]:
        return None

    return {
        "id": f"t-{info['route_id']}",
        "class_name": "",
        "term": term,
        "name": f"Transport - {info['route_name']}",
        "qty": 1,
        "amount": info["fare_per_term"] or 0.0,
    }


def transport_active_for_student(conn, student_id: int, term: str, year: int):
    """
    Returns None if not subscribed for this term.
    Else returns row with route_name, fare_per_term, route_id.
    Uses term_no comparison for correctness.
    """
    sel_term_no = _term_no_from_term(term)

    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
          SELECT tr.id as route_id, tr.name AS route_name, tr.fare_per_term
          FROM transport_subscriptions ts
          JOIN transport_routes tr ON tr.id = ts.route_id
          WHERE ts.active = 1
            AND ts.student_id = %s
            AND (
                  ts.start_year < %s
               OR (ts.start_year = %s AND COALESCE(ts.term_no,0) <= %s)
            )
          ORDER BY ts.start_year DESC, COALESCE(ts.term_no,0) DESC, ts.created_at DESC
          LIMIT 1
        """, (student_id, year, year, sel_term_no))
        return cur.fetchone()
    finally:
        try:
            cur.close()
        except Exception:
            pass




def _find_student_by_sn_or_ln(student_number, last_name):
    conn = get_db_connection()
    r = None
    if student_number:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, student_number, first_name, Middle_name, last_name,
                   class_name, stream, section
            FROM students
            WHERE student_number = %s AND archived = 0
        """, (student_number,))
        r = cur.fetchone()
        cur.close()
    elif last_name:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, student_number, first_name, Middle_name, last_name,
                   class_name, stream, section
            FROM students
            WHERE last_name LIKE %s AND archived = 0
            ORDER BY last_name, first_name
            LIMIT 1
        """, (f"%{last_name}%",))
        r = cur.fetchone()
        cur.close()
    conn.close()
    return r


def _term_order_val(t: str) -> int:
    order = {"Term 1": 1, "Term 2": 2, "Term 3": 3}
    return order.get(t, 99)

def student_fee_group(stu: dict, current_year: int) -> str:
    """New if joined in current academic year, else old."""
    try:
        yoj = int((stu.get("year_of_joining") or "").strip())
    except Exception:
        yoj = None
    return "new" if (yoj == int(current_year)) else "old"



def compute_student_financials(student_id: int, class_name: str, term: str, year: int) -> dict:
    """
    SAFE update (non-breaking keys):
      - If fee_term_summary exists for (student, year, term_no): use it (authoritative),
        including credits (negative outstanding/carry_forward).
      - Otherwise fall back to live computation.
      - IMPORTANT (as requested):
          * Transport is INCLUDED inside expected_requirements
          * overall_balance must reflect transport after subscribe/unsubscribe
          * Transport is still returned separately for UI display
      - Keys returned remain unchanged.
    """

    def _term_order_val(t: str) -> int:
        t = (t or "").strip().lower()
        return 1 if t == "term 1" else 2 if t == "term 2" else 3 if t == "term 3" else 1

    term_no = term_to_no(term) or _term_order_val(term)

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # ---------- 1) Try authoritative summary first ----------
        cur.execute("""
            SELECT
              COALESCE(expected_fees,0) AS expected_fees,
              COALESCE(expected_reqs_base,0) AS expected_reqs_base,
              COALESCE(transport_due_term,0) AS transport_due_term,
              COALESCE(transport_paid_term,0) AS transport_paid_term,
              COALESCE(bursary_current,0) AS bursary_current,
              COALESCE(paid_fees_nonvoid,0) AS paid_fees_nonvoid,
              COALESCE(paid_reqs_nonvoid,0) AS paid_reqs_nonvoid,
              COALESCE(fees_expected_net,0) AS fees_expected_net,
              COALESCE(req_expected_final,0) AS req_expected_final,
              COALESCE(carry_forward,0) AS carry_forward,
              COALESCE(overall_expected,0) AS overall_expected,
              COALESCE(overall_paid,0) AS overall_paid,
              COALESCE(overall_outstanding,0) AS overall_outstanding,
              COALESCE(opening_balance_nv,0) AS opening_balance_nv,
              COALESCE(prior_arrears,0) AS prior_arrears
            FROM fee_term_summary
            WHERE student_id=%s AND year=%s AND term_no=%s
            LIMIT 1
        """, (student_id, year, term_no))
        fs = cur.fetchone()

        if fs:
            expected_fees = float(fs["expected_fees"] or 0.0)
            expected_requirements_base = float(fs["expected_reqs_base"] or 0.0)

            bursary_current = float(fs["bursary_current"] or 0.0)
            paid_fees = float(fs["paid_fees_nonvoid"] or 0.0)
            paid_requirements = float(fs["paid_reqs_nonvoid"] or 0.0)

            carry_forward = float(fs["carry_forward"] or 0.0) # signed ok
            fees_expected_net = float(fs["fees_expected_net"] or 0.0)

            opening_balance = float(fs["opening_balance_nv"] or 0.0) # informational
            prior_arrears = float(fs["prior_arrears"] or 0.0) # informational

            # ---- LIVE transport so subscribe/unsubscribe reflects instantly ----
            transport_fare = 0.0
            try:
                tinfo = transport_subscription_info(student_id, term, year)
            except Exception:
                tinfo = None

            if tinfo and tinfo.get("has_sub"):
                transport_fare = float(tinfo.get("fare_per_term") or 0.0)

            # Keep transport paid term as informational (it may come from summary or your own calc)
            transport_paid_term = float(fs["transport_paid_term"] or 0.0)

            # ✅ expected_requirements must include transport
            expected_requirements = expected_requirements_base + transport_fare

            # ✅ total due includes expected_requirements (already contains transport)
            total_due_this_term = fees_expected_net + expected_requirements

            # ✅ do NOT subtract transport_paid_term separately (your design pays transport via requirements channel)
            balance_this_term = total_due_this_term - (paid_fees + paid_requirements)

            # ✅ overall balance must reflect live transport
            overall_balance = carry_forward + balance_this_term

            # safe display helpers
            overall_outstanding_safe = max(overall_balance, 0.0)
            overall_credit = max(-overall_balance, 0.0)
            carry_forward_safe = max(carry_forward, 0.0)
            carry_credit = max(-carry_forward, 0.0)

            return {
                "expected_fees": expected_fees,
                "expected_requirements": expected_requirements, # includes transport
                "bursary_current": bursary_current,
                "paid_fees": paid_fees,
                "paid_requirements": paid_requirements,
                "carry_forward": carry_forward,
                "total_due_this_term": total_due_this_term,
                "balance_this_term": balance_this_term,
                "overall_balance": overall_balance,

                "opening_balance_raw": opening_balance,
                "prior_arrears_raw": prior_arrears,
                "transport_due_term": transport_fare,
                "transport_paid_term": transport_paid_term,
                "transport_balance_term": transport_fare - transport_paid_term,
                "fee_group_used": None,
                "term_no": term_no,

                "carry_forward_safe": carry_forward_safe,
                "carry_credit": carry_credit,
                "overall_outstanding_safe": overall_outstanding_safe,
                "overall_credit": overall_credit,
            }

        # ---------- 2) FALLBACK: live computation ----------
        cur.execute("""
            SELECT class_name, section, year_of_joining
            FROM students
            WHERE id=%s
            LIMIT 1
        """, (student_id,))
        stu = cur.fetchone() or {}

        eff_class = (stu.get("class_name") or class_name or "").strip()

        # normalize section
        try:
            sec_norm = norm_section((stu.get("section") or "").strip())
        except Exception:
            s = (stu.get("section") or "").strip().lower()
            sec_norm = "Day" if s in ("day", "d") else ("Boarding" if s in ("boarding", "board", "b") else "Day")

        fee_group = student_fee_group(stu, year).title() # 'New' / 'Old'

        # A) Expected fees
        cur.execute("""
            SELECT amount
            FROM class_fees
            WHERE class_name=%s
              AND LOWER(section)=LOWER(%s)
              AND year=%s
              AND term_no=%s
              AND LOWER(fee_group)=LOWER(%s)
            LIMIT 1
        """, (eff_class, sec_norm, year, term_no, fee_group))
        row = cur.fetchone()
        expected_fees = float(row["amount"]) if row and row["amount"] is not None else 0.0

        if expected_fees == 0.0 and fee_group.lower() == "new":
            cur.execute("""
                SELECT amount
                FROM class_fees
                WHERE class_name=%s
                  AND LOWER(section)=LOWER(%s)
                  AND year=%s
                  AND term_no=%s
                  AND LOWER(fee_group)='old'
                LIMIT 1
            """, (eff_class, sec_norm, year, term_no))
            row2 = cur.fetchone()
            expected_fees = float(row2["amount"]) if row2 and row2["amount"] is not None else expected_fees

        # B) Expected requirements (base)
        cur.execute("""
            SELECT COALESCE(SUM(r.amount * COALESCE(r.qty,1)),0) AS total, COUNT(*) AS c
            FROM requirements r
            WHERE r.class_name=%s
              AND r.year=%s
              AND (r.term_no=%s OR r.term_no IS NULL)
              AND LOWER(r.fee_group)=LOWER(%s)
              AND (r.section IS NULL OR LOWER(r.section)=LOWER(%s))
        """, (eff_class, year, term_no, fee_group, sec_norm))
        rr = cur.fetchone() or {}
        expected_requirements_base = float(rr.get("total") or 0.0)
        req_rows = int(rr.get("c") or 0)

        if req_rows == 0 and fee_group.lower() == "new":
            cur.execute("""
                SELECT COALESCE(SUM(r.amount * COALESCE(r.qty,1)),0) AS total
                FROM requirements r
                WHERE r.class_name=%s
                  AND r.year=%s
                  AND (r.term_no=%s OR r.term_no IS NULL)
                  AND LOWER(r.fee_group)='old'
                  AND (r.section IS NULL OR LOWER(r.section)=LOWER(%s))
            """, (eff_class, year, term_no, sec_norm))
            rr2 = cur.fetchone() or {}
            expected_requirements_base = float(rr2.get("total") or 0.0)

        # C) Transport overlay (included)
        transport_fare = 0.0
        transport_paid_term = 0.0
        try:
            tinfo = transport_subscription_info(student_id, term, year)
        except Exception:
            tinfo = None

        if tinfo and tinfo.get("has_sub"):
            transport_fare = float(tinfo.get("fare_per_term") or 0.0)
            if transport_fare > 0:
                try:
                    transport_paid_term = float(transport_paid_via_requirements(conn, student_id, term, year) or 0.0)
                except Exception:
                    transport_paid_term = 0.0

        expected_requirements = expected_requirements_base + transport_fare

        # D) Bursary
        bursary_current = 0.0
        try:
            cur.execute("""
                SELECT COALESCE(SUM(amount),0) AS total
                FROM bursaries
                WHERE student_id=%s AND year=%s AND term_no=%s
            """, (student_id, year, term_no))
            bursary_current = float((cur.fetchone() or {}).get("total") or 0.0)
        except Exception:
            cur.execute("""
                SELECT COALESCE(SUM(amount),0) AS total
                FROM bursaries
                WHERE student_id=%s AND year=%s AND term=%s
            """, (student_id, year, term))
            bursary_current = float((cur.fetchone() or {}).get("total") or 0.0)

        # E) Payments this term (non-void)
        cur.execute("""
            SELECT COALESCE(SUM(amount_paid),0) AS total
            FROM fees
            WHERE student_id=%s
              AND year=%s AND term_no=%s
              AND payment_type_norm IN ('school_fees','fees')
              AND (comment IS NULL OR LOWER(comment) NOT LIKE '%void%')
        """, (student_id, year, term_no))
        paid_fees = float((cur.fetchone() or {}).get("total") or 0.0)

        cur.execute("""
            SELECT COALESCE(SUM(amount_paid),0) AS total
            FROM fees
            WHERE student_id=%s
              AND year=%s AND term_no=%s
              AND payment_type_norm IN ('requirements','requirement')
              AND (comment IS NULL OR LOWER(comment) NOT LIKE '%void%')
        """, (student_id, year, term_no))
        paid_requirements = float((cur.fetchone() or {}).get("total") or 0.0)

        # F) Opening balance (informational)
        cur.execute("""
            SELECT COALESCE(SUM(COALESCE(expected_amount,0) - COALESCE(amount_paid,0)), 0) AS total
            FROM fees
            WHERE student_id=%s
              AND payment_type_norm IN ('opening_balance','ob')
              AND (comment IS NULL OR LOWER(comment) NOT LIKE '%void%')
        """, (student_id,))
        opening_balance = float((cur.fetchone() or {}).get("total") or 0.0)

        # G) Prior arrears (informational)
        cur.execute("""
            SELECT COALESCE(SUM(exp - bur - paid),0) AS total
            FROM (
                SELECT
                  COALESCE(expected_amount,0) AS exp,
                  COALESCE(bursary_amount,0) AS bur,
                  COALESCE(amount_paid,0) AS paid
                FROM fees
                WHERE student_id=%s
                  AND payment_type_norm IN ('school_fees','fees')
                  AND (comment IS NULL OR LOWER(comment) NOT LIKE '%void%')
                  AND (year < %s OR (year=%s AND term_no < %s))
            ) t
        """, (student_id, year, year, term_no))
        prior_arrears = float((cur.fetchone() or {}).get("total") or 0.0)
        if prior_arrears < 0:
            prior_arrears = 0.0

        # H) Carry forward from previous summary
        prev_year, prev_term = year, term_no - 1
        if term_no == 1:
            prev_year, prev_term = year - 1, 3

        cur.execute("""
            SELECT COALESCE(overall_outstanding,0) AS cf
            FROM fee_term_summary
            WHERE student_id=%s AND year=%s AND term_no=%s
            LIMIT 1
        """, (student_id, prev_year, prev_term))
        carry_forward = float((cur.fetchone() or {}).get("cf") or 0.0)

        fees_expected_net = max(expected_fees - bursary_current, 0.0)
        total_due_this_term = fees_expected_net + expected_requirements
        balance_this_term = total_due_this_term - (paid_fees + paid_requirements)
        overall_balance = carry_forward + balance_this_term

        return {
            "expected_fees": expected_fees,
            "expected_requirements": expected_requirements,
            "bursary_current": bursary_current,
            "paid_fees": paid_fees,
            "paid_requirements": paid_requirements,
            "carry_forward": carry_forward,
            "total_due_this_term": total_due_this_term,
            "balance_this_term": balance_this_term,
            "overall_balance": overall_balance,

            "opening_balance_raw": opening_balance,
            "prior_arrears_raw": prior_arrears,
            "transport_due_term": transport_fare,
            "transport_paid_term": transport_paid_term,
            "transport_balance_term": transport_fare - transport_paid_term,

            "fee_group_used": fee_group,
            "term_no": term_no,
        }

    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()



def _expand_terms_from_row(row: dict):
    """
    Accepts:
      - term = 'Term 1'
      - terms = 'Term 1, Term 2' (comma/semicolon separated)
      - apply_to = 'year' -> Term 1..3
      - apply_to   = 'two' with term_one & term_two
    Returns list[str] terms.
    """
    def norm(t):
        t = (t or "").strip()
        return t if t in TERMS else None

    # multiple terms column
    terms_mult = row.get("terms") or row.get("Terms") or row.get("TERMS") or ""
    if terms_mult:
        parts = [p.strip() for p in terms_mult.replace(";", ",").split(",")]
        expanded = [p for p in (norm(p) for p in parts) if p]
        if expanded:
            return expanded

    # single term
    single = norm(row.get("term") or row.get("Term") or row.get("TERM"))
    if single:
        return [single]

    # apply_to
    apply_to = (row.get("apply_to") or row.get(
        "Apply_To") or "").lower().strip()
    if apply_to == "year":
        return TERMS[:]
    if apply_to == "two":
        t1 = norm(row.get("term_one") or row.get("Term_One"))
        t2 = norm(row.get("term_two") or row.get("Term_Two"))
        return [t for t in (t1, t2) if t]
    return []


def _term_order_val(t: str) -> int:
    t = (t or "").strip()
    return 1 if t == "Term 1" else 2 if t == "Term 2" else 3 if t == "Term 3" else 99


def _is_subscribed_this_term(conn, student_id: int, term: str, year: int):
    """
    Returns (is_subscribed: bool, route_name: str or '', fare: float).
    Safe if transport tables don't exist (returns False, '', 0.0).
    """
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT tr.name AS route_name, tr.fare_per_term AS fare, ts.start_term, ts.start_year
            FROM transport_subscriptions ts
            JOIN transport_routes tr ON tr.id = ts.route_id
            WHERE ts.student_id=%s AND ts.active=1
        """, (student_id,))
        row = cur.fetchone()
        cur.close()
        if not row:
            return False, "", 0.0
        # active if start <= current (term/year)
        start_rank = (_term_order_val(
            row["start_term"]), int(row["start_year"]))
        now_rank = (_term_order_val(term), int(year))
        if start_rank <= now_rank:
            return True, row["route_name"] or "", float(row["fare"] or 0.0)
        return False, "", 0.0
    except Exception:
        # transport tables may not exist yet
        return False, "", 0.0


def _transport_paid_total(conn, student_number: str, term: str, year: int) -> float:
    like_token = f"%SN: {student_number}%"
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(SUM(amount),0) AS total
        FROM other_income
        WHERE description='Transport' AND term=%s AND year=%s AND source LIKE %s
    """, (term, int(year), like_token))
    row = cur.fetchone()
    cur.close()
    return float(row["total"] if row else 0.0)


def _expected_fees_for_class(conn, class_name: str) -> float:
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT amount
        FROM class_fees
        WHERE class_name = %s
        ORDER BY id DESC LIMIT 1
    """, (class_name,))
    row = cur.fetchone()
    cur.close()
    return float(row["amount"]) if row and row["amount"] is not None else 0.0


def _expected_requirements_for_class(conn, class_name: str, term: str) -> float:
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM requirements
        WHERE class_name = %s
          AND (term = %s OR term IS NULL OR term = '')
    """, (class_name, term))
    row = cur.fetchone()
    cur.close()
    return float(row["total"] if row else 0.0)


def _bursary_for_term(conn, student_id: int, term: str, year: int) -> float:
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM bursaries
        WHERE student_id=%s AND term=%s AND year=%s
    """, (student_id, term, int(year)))
    row = cur.fetchone()
    cur.close()
    return float(row["total"] if row else 0.0)


def _fees_paid_for_term(conn, student_id: int, term: str, year: int) -> float:
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(SUM(amount_paid), 0) AS total
        FROM fees
        WHERE student_id=%s AND term=%s AND year=%s AND payment_type='fees'
    """, (student_id, term, int(year)))
    row = cur.fetchone()
    cur.close()
    return float(row["total"] if row else 0.0)


def _requirements_paid_for_term(conn, student_id: int, term: str, year: int) -> float:
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(SUM(amount_paid), 0) AS total
        FROM fees
        WHERE student_id=%s AND term=%s AND year=%s AND payment_type='requirements'
    """, (student_id, term, int(year)))
    row = cur.fetchone()
    cur.close()
    return float(row["total"] if row else 0.0)


def _filters_from_request():
    ay = get_active_academic_year()
    default_term = ay.get("current_term") or ay.get("term") or "Term 1"
    default_year = int(ay.get("year") or datetime.now().year)

    return {
        "student_number": (request.args.get("student_number") or "").strip(),
        "last_name": (request.args.get("last_name") or "").strip(),
        "class_name": (request.args.get("class_name") or "").strip(),
        "term": (request.args.get("term") or default_term).strip(),
        "year": int(request.args.get("year") or default_year),
        "export": (request.args.get("export") == "1"),
    }


def _fetch_employees_for_dropdown(conn):
    """
    Returns employees for the 'Link Employee' select.
    Order: active first, then last_name, first_name.
    Fields: id, first_name, middle_name, last_name, designation, status
    """
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT
            id,
            first_name,
            COALESCE(Middle_name, '') AS middle_name,
            last_name,
            COALESCE(designation,'') AS designation,
            COALESCE(status,'active') AS status
        FROM employees
        ORDER BY
            CASE WHEN status='active' THEN 0 ELSE 1 END,
            last_name, first_name
    """)
    rows = cur.fetchall()
    cur.close()
    return rows


def _suggest_username(conn, first_name: str, last_name: str) -> str:
    """
    Base username: first letter of last_name + first_name (lowercase).
    Ensure uniqueness by appending a number if needed.
    """
    fn = re.sub(r'[^a-z0-9]', '', (first_name or '').strip().lower())
    ln = re.sub(r'[^a-z0-9]', '', (last_name or '').strip().lower())
    base = (ln[:1] + fn) or "user"

    cur = conn.cursor(dictionary=True)
    # if base free, use it
    cur.execute("SELECT 1 FROM users WHERE username=%s LIMIT 1", (base,))
    if not cur.fetchone():
        cur.close()
        return base

    # otherwise find next available suffix
    suffix = 2
    while True:
        cand = f"{base}{suffix}"
        cur.execute("SELECT 1 FROM users WHERE username=%s LIMIT 1", (cand,))
        if not cur.fetchone():
            cur.close()
            return cand
        suffix += 1


ALLOWED_EXTS = {"csv", "xlsx"}


def _safe_int(v):
    try:
        if v is None or v == "":
            return None
        # excel numeric loads may be floats; allow float->int too
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _safe_float(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _avg(*parts):
    nums = [p for p in parts if isinstance(p, (int, float))]
    return float(sum(nums) / len(nums)) if nums else None


def get_user_initials(conn, user_id):
    """
    Returns initials for the logged-in user via users -> employees -> teachers.
    Falls back to '' if not found.
    """
    if not user_id:
        return ""
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(t.initials,'') AS initials
          FROM users u
          LEFT JOIN employees e ON e.id = u.employee_id
          LEFT JOIN teachers t ON t.employee_id = e.id AND t.status='active'
         WHERE u.id = %s
         LIMIT 1
    """, (user_id,))
    row = cur.fetchone()
    cur.close()
    return (row["initials"] if row else "") or ""


def resolve_student_id(conn, student_number):
    if not student_number:
        return None
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT id FROM students WHERE student_number=%s AND archived=0",
        (student_number,)
    )
    r = cur.fetchone()
    cur.close()
    return r["id"] if r else None


def resolve_subject_id(conn, subject_id=None, subject_code=None, subject_name=None):
    # prefer explicit id
    sid = _safe_int(subject_id)
    if sid:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT 1 FROM subjects WHERE id=%s", (sid,))
        exists = cur.fetchone()
        cur.close()
        return sid if exists else None

    # fallbacks
    if subject_code:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id FROM subjects WHERE code=%s", (subject_code,))
        r = cur.fetchone()
        cur.close()
        if r:
            return r["id"]

    if subject_name:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id FROM subjects WHERE name=%s", (subject_name,))
        r = cur.fetchone()
        cur.close()
        if r:
            return r[0]

    return None


def new_namedtuple_like(d):
    """Allow dot-access in Jinja for next_term_info even when None/dict."""
    if not d:
        return None
    return type("NTI", (), d)


def compute_agg_div(core_grades: dict[str, str]):
    """
    core_grades: mapping like {'ENG':'D2','MATH':'C3','SCI':'D1','SST':'C4'} (any order)
    Returns (aggregate:int|None, division:str|None)
    """
    if len(core_grades) < 4:
        return (None, None)
    agg = sum(AGG_MAP[g] for g in core_grades.values() if g in AGG_MAP)
    if 4 <= agg <= 12:
        div = "1"
    elif 13 <= agg <= 23:
        div = "2"
    elif 24 <= agg <= 29:
        div = "3"
    elif 30 <= agg <= 34:
        div = "4"
    else:
        div = "U"
    return (agg, div)




def build_report_payload(conn, student_id, term, year, include_mid=False):
    """
    Returns a dict with the SAME keys your single report passes to
    report_card_citizen.html, so report_batch_citizen.html can loop it.
    Uses live record_score for TOTAL(%) and recomputes grades.
    """
    ensure_term_dates_schema(conn)

    # ---- Student ----
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM students WHERE id=%s", (student_id,))
    student = cur.fetchone()
    cur.close()
    if not student:
        return None

    # ---- Ensure snapshot exists ----
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT 1 FROM reports WHERE student_id=%s AND term=%s AND year=%s "
        "LIMIT 1",
        (student_id, term, year),
    )
    if not cur.fetchone():
        process_reports_snapshot(conn, student["class_name"], term, year)
    cur.close()

    # =========================
    # 1) Snapshot rows
    # =========================
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT sub.name AS subject, sub.code AS subject_code,
               r.subject_id,
               r.eot_mark AS eot,
               COALESCE(r.average_mark, r.eot_mark) AS total_100_snapshot,
               r.grade AS grade_snapshot,
               r.comment, r.teacher_initial AS initials,
               r.teacher_remark, r.headteacher_remark
        FROM reports r
        JOIN subjects sub ON sub.id = r.subject_id
        WHERE r.student_id=%s AND r.term=%s AND r.year=%s
        ORDER BY sub.name
        """,
        (student_id, term, year),
    )
    rows = cur.fetchall() or []
    cur.close()

    # =========================
    # 2) Live record_score
    # =========================
    COMPONENT_FIELDS = (
        "other_mark", "holiday_mark", "bot_mark",
        "midterm_mark", "eot_mark", "ca_mark"
    )
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT subject_id, average_mark,
               other_mark, holiday_mark, bot_mark,
               midterm_mark, eot_mark, ca_mark
        FROM record_score
        WHERE student_id=%s AND term=%s AND year=%s
        """,
        (student_id, term, year),
    )
    live_score_rows = cur.fetchall() or []
    cur.close()

    def _mean_nonnull(values):
        nums = [float(v) for v in values if v is not None]
        return (sum(nums) / len(nums)) if nums else None

    def _pick_average_row_score(r: dict):
        if r.get("average_mark") is not None:
            return float(r["average_mark"])
        return _mean_nonnull([r.get(k) for k in COMPONENT_FIELDS])

    live_numeric_by_subj = {}
    for r in live_score_rows:
        sc = _pick_average_row_score(r)
        if sc is not None:
            live_numeric_by_subj[int(r["subject_id"])] = sc

    def _grade_num(x):
        return grade_for_score(conn, x) if x is not None else None

    # =========================
    # 3) Final subject rows
    # =========================
    final_rows = []
    for r in rows:
        sid = int(r["subject_id"])
        total_num = live_numeric_by_subj.get(sid, None)
        if total_num is None:
            total_num = r.get("total_100_snapshot")
        new_grade = _grade_num(total_num)
        final_rows.append({
            "subject": r["subject"],
            "subject_code": r["subject_code"],
            "eot": r.get("eot"),
            "total_100": total_num,
            "grade": new_grade,
            "comment": r.get("comment"),
            "initials": r.get("initials"),
            "teacher_remark": r.get("teacher_remark"),
            "headteacher_remark": r.get("headteacher_remark"),
        })

    # =========================
    # 4) Totals / averages
    # =========================
    nums = [x["total_100"] for x in final_rows if x["total_100"] is not None]
    total_sum = round(sum(nums), 2) if nums else None
    avg_overall = round(sum(nums) / len(nums), 2) if nums else None

    # =========================
    # 5) Aggregate & Division
    # =========================
    CORE_CODES = {"ENG", "MATH", "SCI", "SST"}
    core_grades = {
        (r["subject_code"] or "").upper(): r["grade"]
        for r in final_rows
        if (r["subject_code"] or "").upper() in CORE_CODES and r["grade"]
    }
    aggregate, division = compute_agg_div(core_grades)

    # =========================
    # 6) Class position
    # =========================
    rk = class_ranking(conn, student["class_name"], term, year)
    class_size = len(rk)
    position = next(
        (i for i, rr in enumerate(rk, start=1) if rr["sid"] == student_id),
        None
    )

    # =========================
    # 7) MIDTERM PANEL (same as before if needed)
    # =========================
    midterms = []
    midterm_subjects = []
    CORE_CODES = {"ENG", "MATH", "SCI", "SST"}

    if include_mid:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT UPPER(code) AS code, id
            FROM subjects
            WHERE code IS NOT NULL AND TRIM(code) <> ''
        """)
        code_rows = cur.fetchall() or []
        cur.close()
        code_to_id = {r["code"]: r["id"] for r in code_rows}
        all_codes = list(code_to_id.keys())
        others = sorted([c for c in all_codes if c not in CORE_CODES])
        midterm_subjects = [c for c in ("ENG", "MATH", "SCI", "SST")
                            if c in all_codes] + others

        def fetch_marks(colname: str, *, round0: bool = False) -> dict:
            out = {}
            for sc in midterm_subjects:
                sid2 = code_to_id.get(sc)
                if not sid2:
                    out[sc] = None
                    continue
                cur2 = conn.cursor(dictionary=True)
                cur2.execute(
                    f"""SELECT MAX({colname}) AS v
                        FROM record_score
                        WHERE student_id=%s AND subject_id=%s
                          AND term=%s AND year=%s""",
                    (student_id, sid2, term, year)
                )
                row2 = cur2.fetchone()
                cur2.close()
                v = row2["v"] if row2 and row2["v"] is not None else None
                if v is not None:
                    try:
                        v = int(round(float(v))) if round0 else int(v)
                    except Exception:
                        try:
                            v = int(float(v))
                        except Exception:
                            v = None
                out[sc] = v
            return out

        def grade_of(score):
            return grade_for_score(conn, score) if score is not None else None

        panels = [
            ("OTH", fetch_marks("other_mark", round0=True)),
            ("HP", fetch_marks("holiday_mark", round0=True)),
            ("BOT", fetch_marks("bot_mark")),
            ("MID", fetch_marks("midterm_mark")),
        ]
        for label, score_map in panels:
            if any(v is not None for v in score_map.values()):
                total_all = sum(int(v) for v in score_map.values() if v is not None)
                midterms.append({
                    "assessment": label,
                    "scores": score_map,
                    "grades": {sc: grade_of(v) for sc, v in score_map.items()},
                    "total": total_all
                })

    # =========================
    # 8) Payment number
    # =========================
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT id AS payment_number
        FROM fees
        WHERE student_id=%s AND term=%s AND year=%s
        ORDER BY date_paid DESC, id DESC
        LIMIT 1
        """,
        (student_id, term, year),
    )
    pay = cur.fetchone()
    cur.close()
    payment_number = pay["payment_number"] if pay else None

    # =========================
    # 9) Overall comments (with overrides)
    # =========================
    overrides = fetch_overall_overrides(conn, student["id"], term, year)

    from collections import Counter

    auto_head = (
        pick_comment_template(
            conn,
            role="headteacher", scope="overall",
            division=(int(division) if division and str(division).isdigit() else None),
            average=avg_overall, class_name=student["class_name"], term=term, student_id=student_id
        )
        or (comment_for_grade(conn, grade_for_score(conn, avg_overall)) or "")
    )
    head_comment = (overrides.get("head_overall_custom") or "").strip() or auto_head

    auto_teacher = pick_comment_template(
        conn,
        role="teacher", scope="overall",
        division=(int(division) if division and str(division).isdigit() else None),
        average=avg_overall, class_name=student["class_name"], term=term, student_id=student_id
    )
    if not auto_teacher:
        per_subj = [r["teacher_remark"] for r in final_rows
                    if r.get("teacher_remark")]
        auto_teacher = (Counter(per_subj).most_common(1)[0][0]
                        if per_subj else None)
    if not auto_teacher:
        auto_teacher = (comment_for_grade(conn, grade_for_score(conn, avg_overall))
                        or "")

    teacher_comment = (overrides.get("teacher_overall_custom") or "").strip() or auto_teacher
    special_communication = (overrides.get("special_communication") or "").strip()

    # =========================
    # 10) Next-term info
    # =========================

    def _next_term_name_and_year(cur_term: str, cur_year: int):
        order = ["Term 1", "Term 2", "Term 3"]
        try:
            i = order.index(cur_term)
        except ValueError:
            return "Term 2", cur_year
        return (order[i+1], cur_year) if i < 2 else (order[0], cur_year + 1)

    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT next_term, next_term_date, next_term_end_date
        FROM term_dates
        WHERE year=%s AND term=%s
        LIMIT 1
        """,
        (year, term),
    )
    row = cur.fetchone()
    cur.close()

    if row and (row["next_term"] or row["next_term_date"] or row["next_term_end_date"]):
        next_term_info = dict(
            next_term=row["next_term"],
            next_term_date=row["next_term_date"],
            next_term_end_date=row["next_term_end_date"],
        )
    else:
        nt_name, nt_year = _next_term_name_and_year(term, year)
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT next_term, next_term_date, next_term_end_date
            FROM term_dates
            WHERE year=%s AND term=%s
            LIMIT 1
            """,
            (nt_year, nt_name),
        )
        fb = cur.fetchone()
        cur.close()
        next_term_info = (
            dict(
                next_term=nt_name,
                next_term_date=fb["next_term_date"],
                next_term_end_date=fb["next_term_end_date"],
            )
            if fb and (fb["next_term_date"] or fb["next_term_end_date"])
            else None
        )

    grading = fetch_grading_scale(conn)

    return dict(
        school=dict(
            name="DEMO DAY & BOARDING",
            tagline="PRIMARY SCHOOL – KAMPALA",
            motto="Strive for the best",
            phones="+256778878411, +256759685640, +256773589232, +256750347624",
            pobox="P.O Box 1X1X1 Kampala",       
        ),
        student=student,
        term=term, year=year,
        rows=final_rows,
        total_sum=total_sum,
        avg_overall=avg_overall,
        aggregate=aggregate,
        division=division,
        position=position,
        class_size=class_size,
        midterms=midterms,
        midterm_subjects=[r["subject_code"] for r in final_rows if r.get("subject_code")],
        payment_number=payment_number,
        comments={"teacher_comment": teacher_comment, "head_comment": head_comment},
        special_communication=special_communication,
        next_term_info=new_namedtuple_like(next_term_info),
        grading=grading,
    )
    
    


     

def build_midterm_payload(conn, student_id, term, year):
    """
    Build one mid-term report dict for a single learner.
    Includes:
      - subject rows with OTH/BOT/MID/AVG/GRADE/comment/initials
      - total_mid_sum, avg_mid_overall
      - aggregate + division (cores only)
      - comments (auto from rules, overridden by manual if present)
      - grading scale, column visibility flags
    **NOTE**: Will still return a payload even if there are NO marks,
              so the view never 404s just because marks are empty.
    """
    ensure_midterm_overall_comments_schema(conn)
    grading = fetch_grading_scale(conn) or []

    # ---- student ----
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, student_number,
               first_name, COALESCE(Middle_name,'') AS middle_name,
               last_name, class_name, COALESCE(stream,'') AS stream,
               COALESCE(photo,'') AS photo
        FROM students
        WHERE id=%s AND archived=0
        LIMIT 1
    """, (student_id,))
    stu = cur.fetchone()
    cur.close()
    if not stu:
        return None

    # ---- subjects meta ----
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id,
               UPPER(COALESCE(code,'')) AS code,
               name
        FROM subjects
        ORDER BY name
    """)
    subject_meta = {r["id"]: r for r in (cur.fetchall() or [])}
    cur.close()

    # ---- scores for this learner (latest via MAX) ----
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT student_id, subject_id,
               MAX(other_mark)   AS oth,
               MAX(bot_mark)     AS bot,
               MAX(midterm_mark) AS mid,
               MAX(initials)     AS initials,
               MAX(comment)      AS subj_comment
        FROM record_score
        WHERE term=%s AND year=%s AND student_id=%s
        GROUP BY student_id, subject_id
    """, (term, year, student_id))
    raw = cur.fetchall() or []
    cur.close()

    from collections import defaultdict, Counter
    per = defaultdict(lambda: defaultdict(dict))
    for r in raw:
        per[r["student_id"]][r["subject_id"]] = {
            "OTH": r["oth"],
            "BOT": r["bot"],
            "MID": r["mid"],
            "initials": (r.get("initials") or None),
            "subj_comment": (r.get("subj_comment") or None),
        }


    CORE_CODES = {"ENG", "MATH", "SCI", "SST"}
    AGG_MAP = {"D1": 1, "D2": 2, "C3": 3, "C4": 4,
               "C5": 5, "C6": 6, "P7": 7, "P8": 8, "F9": 9}

    def mean_present(vals):
        nums = [float(v) for v in vals if v is not None]
        return round(sum(nums) / len(nums), 1) if nums else None

    subj_scores = per[student_id] # may be {} if no scores at all

    # subjects that have any OTH/BOT/MID
    present_ids = [
        sj for sj, m in subj_scores.items()
        if any([m.get("OTH"), m.get("BOT"), m.get("MID")])
    ]

    # --- CORE detection (same style as EOT) -----------------------------
    def _is_core_meta(meta: dict) -> bool:
        code_up = (meta.get("code") or "").upper()
        name = (meta.get("name") or "").strip().lower()
        if code_up in CORE_CODES:
            return True
        if name.startswith("eng"):
            return True
        if name.startswith(("mat", "math")):
            return True
        if name.startswith("sci"):
            return True
        if name in {
            "sst", "soc. studies", "social studies",
            "social std", "socialstudies"
        }:
            return True
        return False

    def is_core(sj_id: int) -> bool:
        return _is_core_meta(subject_meta.get(sj_id, {}))

    # even if present_ids is empty, we proceed (empty table)
    present_ids.sort(
        key=lambda sj: (
            0 if is_core(sj) else 1,
            subject_meta.get(sj, {}).get("name") or ""
        )
    )

    rows = []
    core_grades = []
    show_oth = show_bot = show_mid = False
    for sj in present_ids:
        meta = subject_meta.get(sj, {})
        code_up = (meta.get("code") or "").upper()
        name = meta.get("name") or f"Subject {sj}"
        marks = subj_scores.get(sj, {})
        oth = marks.get("OTH")
        bot = marks.get("BOT")
        midm = marks.get("MID")
        show_oth |= (oth is not None)
        show_bot |= (bot is not None)
        show_mid |= (midm is not None)
        avg_mid = mean_present([oth, bot, midm])
        grade = grade_for_score(conn, avg_mid) if avg_mid is not None else ""

        # per-subject comment – stored or from grade rule
        subj_comment = marks.get("subj_comment") or (
            comment_for_grade(conn, grade_for_score(conn, avg_mid))
            if avg_mid is not None else ""
        )

        initials = (marks.get("initials") or "") or \
                   guess_teacher_initials(conn, stu["class_name"], sj) or ""

        rows.append(dict(
            name=name,
            code=code_up,
            OTH=oth,
            BOT=bot,
            MID=midm,
            AVG=avg_mid,
            grade=grade or "",
            comment=subj_comment or "",
            initials=initials or "",
        ))

        # Only these 4 cores are used for aggregate/division
        if is_core(sj) and grade:
            core_grades.append(grade)


    # totals / overall average
    avgs = [r["AVG"] for r in rows if r["AVG"] is not None]
    total_mid_sum = round(sum(avgs), 1) if avgs else None
    avg_mid_overall = round(sum(avgs) / len(avgs), 1) if avgs else None

    # aggregate & division (cores only)
    def compute_agg_div(core_grades_list):
        if len(core_grades_list) != 4:
            return None, "NG"
        pts = [AGG_MAP[g] for g in core_grades_list if g in AGG_MAP]
        if len(pts) != 4:
            return None, "NG"
        agg = sum(pts)
        if 4 <= agg <= 12:
            div = "1"
        elif 13 <= agg <= 24:
            div = "2"
        elif 25 <= agg <= 29:
            div = "3"
        elif 30 <= agg <= 34:
            div = "4"
        else:
            div = "U"
        return agg, f"Div {div}"

    aggregate, division = compute_agg_div(core_grades) if core_grades else (None, "NG")

    # automatic comments from rules
    head_comment = (
        pick_comment_template(
            conn,
            role="headteacher", scope="overall",
            division=(int(division[-1]) if division and division.startswith("Div ")
                      and division[-1].isdigit() else None),
            average=avg_mid_overall,
            class_name=stu["class_name"], term=term, student_id=student_id
        )
        or (comment_for_grade(conn, grade_for_score(conn, avg_mid_overall))
            if avg_mid_overall is not None else "")
        or ""
    )

    from collections import Counter
    teacher_comment = pick_comment_template(
        conn,
        role="teacher", scope="overall",
        division=(int(division[-1]) if division and division.startswith("Div ")
                  and division[-1].isdigit() else None),
        average=avg_mid_overall,
        class_name=stu["class_name"], term=term, student_id=student_id
    )
    if not teacher_comment:
        per_subj_cmts = [r["comment"] for r in rows if r.get("comment")]
        teacher_comment = (
            Counter(per_subj_cmts).most_common(1)[0][0]
            if per_subj_cmts else ""
        ) or (
            comment_for_grade(conn, grade_for_score(conn, avg_mid_overall))
            if avg_mid_overall is not None else ""
        )

    # manual overrides (and special communication)
    manual = load_midterm_manual_comments(conn, student_id, term, year)
    teacher_comment, head_comment, special_comm = merge_midterm_comments(
        teacher_comment, head_comment, manual
    )

    school = dict(
        name=current_app.config.get("SCHOOL_NAME", "DEMO DAY & BOARDING"),
        tagline=current_app.config.get("SCHOOL_TAGLINE", "PRIMARY SCHOOL – KAMPALA"),
        motto=current_app.config.get("SCHOOL_MOTTO", "Code the future"),
        phones=current_app.config.get(
            "SCHOOL_PHONES",
            "+256778878411, +256759685640, +256773589232, +256750347624"
        ),
        pobox=current_app.config.get("P.O Box 1X1X1 Kampala"),
    )
    return dict(
        school=school,
        student=stu,
        term=term,
        year=year,
        rows=rows,
        aggregate=aggregate,
        division=division or "NG",
        total_mid_sum=total_mid_sum,
        avg_mid_overall=avg_mid_overall,
        comments={
            "teacher_comment": teacher_comment,
            "head_comment": head_comment,
            "special_communication": special_comm,
        },
        grading=grading,
        show_oth=show_oth,
        show_bot=show_bot,
        show_mid=show_mid,
    )


def _teacher_classes_list():
    # Baby, Middle, Top, P1..P7
    return ["Baby", "Middle", "Top", "P1", "P2", "P3", "P4", "P5", "P6", "P7"]


def _streams_list(conn):
    cur.execute("SELECT name FROM streams ORDER BY name")
    row = cur.fetchall()
    return [r["name"] for r in rows] or ["A"]


def _subjects_list(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT id, name, COALESCE(code,'') AS code FROM subjects ORDER BY name")
    rows = cur.fetchall()
    cur.close()
    return rows


def _employees_list(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, first_name, COALESCE(Middle_name,'') AS Middle_name, last_name, designation, status
        FROM employees
        ORDER BY (status='active') DESC, last_name, first_name
    """)
    rows = cur.fetchall()
    cur.close()
    return rows


# =========================
# TEACHERS: routes (CRUD + subject/class assignments)
# =========================

# ---------- TEACHER MANAGEMENT (no login/user creation here) ----------

def _get_employees_without_teacher(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT e.id, e.first_name, e.middle_name, e.last_name, e.designation
        FROM employees e
        LEFT JOIN teachers t ON t.employee_id = e.id
        WHERE t.id IS NULL AND e.status='active'
        ORDER BY e.last_name, e.first_name
    """)
    rows = cur.fetchall()
    cur.close()
    return rows
    
def _username():
    return (session.get("username")
            or (getattr(g, "current_user", None) and getattr(g.current_user, "username", None))
            or "system")

def _default_term_year():
    ay = get_active_academic_year() or {}
    term = (request.args.get("term") or ay.get("current_term") or ay.get("term") or "").strip()
    year = int(request.args.get("year") or ay.get("year") or ay.get("active_year") or datetime.now().year)
    return term, year

def _query_expenses(conn, *, term=None, year=None, category_id=None, q=None, date_from=None, date_to=None, limit=500):
    where, params = ["1=1"], []
    if term:
        where.append("e.term = %s"); params.append(term)
    if year:
        where.append("e.year = %s"); params.append(int(year))
    if category_id:
        where.append("e.category_id = %s"); params.append(int(category_id))
    if date_from:
        where.append("e.date_spent >= %s"); params.append(date_from)
    if date_to:
        where.append("e.date_spent <= %s"); params.append(date_to)
    if q:
        where.append("(e.description LIKE %s OR COALESCE(ec.name,'') LIKE %s)")
        like = f"%{q}%"; params += [like, like]

    cur = conn.cursor(dictionary=True)
    cur.execute(f"""
        SELECT
            e.id,
            DATE_FORMAT(e.date_spent,'%%Y-%%m-%%d') AS date_spent,
            e.description, e.amount, e.term, e.year, e.type,
            e.category_id, COALESCE(ec.name,'—') AS category_name,
            e.recorded_by
        FROM expenses e
        LEFT JOIN expense_categories ec ON ec.id = e.category_id
        WHERE {" AND ".join(where)}
        ORDER BY e.date_spent DESC, e.id DESC
        LIMIT {int(limit)}
    """, tuple(params) if params else ())
    rows = cur.fetchall() or []
    cur.close()
    return rows

def _get_or_create_expense_category(conn, name: str):
    name = (name or "").strip()
    if not name:
        return None
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id FROM expense_categories WHERE name=%s", (name,))
    row = cur.fetchone()
    if row:
        cur.close()
        return row["id"]
    cur.execute("INSERT INTO expense_categories (name) VALUES (%s)", (name,))
    cid = cur.lastrowid
    cur.close()
    conn.commit()
    return cid


def _get_all_subjects(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT id, name, COALESCE(code,'') AS code FROM subjects ORDER BY name")
    rows = cur.fetchall()
    cur.close()
    return rows


def _get_all_classes(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name")
    rows = cur.fetchall()
    return rows

# --- Active Academic Year guard ---------------------------------------------


def _redirect_to_existing(*endpoints):
    """Redirect to the first endpoint that exists; fall back to dashboard."""
    for ep in endpoints:
        if ep in app.view_functions:
            return redirect(url_for(ep))
    return redirect(url_for("dashboard"))


def require_active_academic_year():
    """
    Use like:
        guard = require_active_academic_year()
        if guard: return guard
    Returns None if an active AY exists; otherwise returns a redirect Response.
    """
    try:
        ay = get_active_academic_year()  # your existing helper
        year = ay.get("year")
        term = ay.get("current_term") or ay.get("term")
        if year and term:
            return None
    except Exception:
        pass

    flash("Please create/activate an academic year first.", "warning")
    # Try a few likely endpoints you might have; falls back to dashboard.
    return _redirect_to_existing("manage_academic_years", "academic_years", "settings_academic_year", "dashboard")


def _password_col_name(conn) -> str:
    """
    Detect which column stores the hash: 'password_hash' preferred, fall back to 'password'.
    """
    cur.execute("SHOW COLUMNS FROM users")
    cols = [c["Field"] for c in cur.fetchall()]
    if "password_hash" in cols:
        return "password_hash"
    if "password" in cols:
        return "password"  # assume it already stores a hash
    # If neither exists, fail clearly:
    raise RuntimeError(
        "No password column found in 'users' table. Expected 'password_hash' or 'password'.")


def _is_safe_url(target: str) -> bool:
    """Prevent open redirects."""
    if not target:
        return False
    host_url = urlparse(request.host_url)
    redirect_url = urlparse(urljoin(request.host_url, target))
    return (redirect_url.scheme in ("http", "https")
            and host_url.netloc == redirect_url.netloc)


def get_user_profile(user_id: int) -> dict:
    """
    Returns {"initials": str, "full_name": str} for the given users.id.
    - Looks up users.employee_id -> employees (name) and teachers (initials).
    - Opens & closes its own connection (so it never uses a closed DB).
    """
    initials = ""
    full_name = ""

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("SELECT employee_id FROM users WHERE id=%s", (user_id,))
    u = cur.fetchone()
    emp_id = u["employee_id"] if u and "employee_id" in u.keys() else None

    if emp_id:
        cur.execute(
            "SELECT first_name, COALESCE(Middle_name,'') AS Middle_name, last_name "
            "FROM employees WHERE id=%s", (emp_id,)
        )
        emp = cur.fetchone()
        if emp:
            full_name = " ".join([
                (emp["first_name"] or "").strip(),
                (emp["Middle_name"] or "").strip(),
                (emp["last_name"] or "").strip(),
            ]).strip()

        cur.execute(
            "SELECT initials FROM teachers WHERE employee_id=%s", (emp_id,)
        )
        tch = cur.fetchone()
        if tch and (tch["initials"] or "").strip():
            initials = tch["initials"].strip()

    cur.close()
    conn.close()
    return {"initials": initials, "full_name": full_name}


# --- Role helpers ---

def _norm_role(val):
    """Return a canonical, lowercase role string."""
    return (str(val or "").strip().lower())


def require_login(f):
    @wraps(f)
    def _inner(*args, **kwargs):
        if "user_id" not in session:
            flash("Please login.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return _inner


def require_role(*roles):
    # normalize decorator inputs once
    wanted = tuple(_norm_role(r) for r in roles if r)

    def wrapper(f):
        @wraps(f)
        def inner(*args, **kwargs):
            if "user_id" not in session or "role" not in session:
                flash("Please login.", "warning")
                return redirect(url_for("login"))

            srole = _norm_role(session.get("role"))
            if wanted and srole not in wanted:
                flash("Access denied.", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return inner
    return wrapper


# allowed roles
ALLOWED_ROLES = (
    "admin", "bursar", "teacher", "headteacher",
    "director", "clerk", "dos", "deputyheadteacher", "classmanager"
)


def login_post():
    username = request.form["username"].strip()
    password = request.form["password"].strip()
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute(
        "SELECT * FROM users WHERE username=%s AND status='active' LIMIT 1",
        (username,)
    )
    user = cur.fetchone()
    cur.close()
    conn.close()

    if not user or not check_password_hash(user["password_hash"], password):
        flash("Invalid credentials.", "danger")
        return redirect(url_for("login"))

    session.clear()
    session["user_id"] = user["id"]
    session["username"] = user["username"]
    session["role"] = _norm_role(user["role"])  # normalized
    return redirect(url_for("dashboard"))


TERMS = ["Term 1", "Term 2", "Term 3"]  # or import your existing TERMS


def ensure_requirements_schema(conn):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS requirements (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(50) NOT NULL,
            class_name VARCHAR(15),
            section ENUM('Day','Boarding'),
            term VARCHAR(20), -- required by get_class_requirements()
            year INT NOT NULL,
            qty INT NOT NULL,
            amount DOUBLE NOT NULL DEFAULT 0,
            UNIQUE KEY uq_req (name, class_name, term, year)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)
    conn.commit()
    cur.close()


def _classes_for_dropdown(conn):
    # prefer classes table; fall back to students if needed
    cur.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name")
    rows = cur.fetchall()
    classes = [r[0] for r in rows]
    if not classes:
        cur.execute(
            "SELECT DISTINCT class_name FROM students ORDER BY class_name")
        rows = cur.fetchall()
        classes = [r[0] for r in rows]
    return classes


CLASS_ORDER = ["Baby", "Middle", "Top",
               "P1", "P2", "P3", "P4", "P5", "P6", "P7"]


def prev_class_name(current: str) -> str | None:
    if current not in CLASS_ORDER:
        return None
    i = CLASS_ORDER.index(current)
    return CLASS_ORDER[i-1] if i-1 >= 0 else None


def _distinct_classes(conn) -> list[str]:
    """Collect classes from classes table + students, dedup + canonical order."""
    seen, out = set(), []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT DISTINCT class_name FROM classes WHERE class_name IS NOT NULL")
        for r in cur.fetchall():
            cn = (r["class_name"] or "").strip()
            if cn and cn not in seen:
                out.append(cn)
                seen.add(cn)
        cur.close()
    except mysql.connector.Error:
        pass
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT DISTINCT class_name FROM students WHERE class_name IS NOT NULL")
        for r in cur.fetchall():
            cn = (r["class_name"] or "").strip()
            if cn and cn not in seen:
                out.append(cn)
                seen.add(cn)
        cur.close()
    except mysql.connector.Error:
        pass
    # order by canonical list first, then name
    return sorted(out, key=lambda x: (CLASS_ORDER.index(x) if x in CLASS_ORDER else 999, x))


def _classes_for_dropdown(conn) -> list[str]:
    """Collect distinct classes for the class dropdown."""
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT DISTINCT class_name FROM students WHERE class_name IS NOT NULL ORDER BY class_name"
    )
    rows = cur.fetchall()
    cur.close()
    return [r["class_name"] for r in rows]


def years_for_dropdown(conn) -> list[int]:
    """
    Collect distinct years from the fees table for a dropdown (DESC order).
    Falls back to active academic year or current year if no data found.
    """
    cur = conn.cursor(dictionary=True)
    try:
        # Collect distinct non-null years
        cur.execute(
            "SELECT DISTINCT year FROM fees WHERE year IS NOT NULL ORDER BY year DESC")
        rows = cur.fetchall()
        yrs = [r["year"] for r in rows]
    except Exception as e:
        print("Error fetching years:", e)
        yrs = []
    finally:
        cur.close()

    # Fallback logic
    if not yrs:
        try:
            from mysql_db import get_active_academic_year  # if already defined
            ay = get_active_academic_year()
            if ay and "year" in ay:
                yrs = [int(ay["year"])]
            else:
                yrs = [datetime.now().year]
        except Exception:
            yrs = [datetime.now().year]

    return yrs


def _safe_set_status(conn, student_id: int, new_status: str) -> None:
    cur = conn.cursor(dictionary=True)
    s = (new_status or "").strip().lower()
    if s in ALLOWED_STATUS:
        cur.execute("UPDATE students SET status=%s WHERE id=%s",
                    (s, student_id))
        conn.commit()
        cur.close()


def _term_order(t: str) -> int:
    t = (t or "").strip().lower()
    return 1 if t == "term 1" else 2 if t == "term 2" else 3 if t == "term 3" else 0


def _employees_query_and_params(q=None, status=None, department=None, designation=None):
    sql = ["SELECT * FROM employees"]
    params = []
    where = []

    if q:
        where.append(
            "(first_name LIKE %s OR middle_name LIKE %s OR last_name LIKE %s OR department LIKE %s OR designation LIKE %s)")
        like = f"%{q}%"
        params += [like, like, like, like, like]
    if status:
        where.append("status = %s")
        params.append(status)
    if department:
        where.append("department LIKE %s")
        params.append(f"%{department}%")
    if designation:
        where.append("designation LIKE %s")
        params.append(f"%{designation}%")

    if where:
        sql.append("WHERE " + " AND ".join(where))

    sql.append("ORDER BY (status='active') DESC, last_name, first_name")
    return " ".join(sql), params


def _img_to_escpos_raster(img: "Image.Image", max_width_dots: int = 576) -> bytes:
    """
    Convert a PIL image to ESC/POS Raster format (GS v 0).
    max_width_dots: 58mm ≈ 384; 80mm ≈ 576 (some printers use 640).
    """
    # Resize keeping aspect ratio
    w, h = img.size
    if w > max_width_dots:
        new_h = int(h * (max_width_dots / float(w)))
        img = img.resize((max_width_dots, max(1, new_h)), Image.LANCZOS)

    # 1-bit dithered image works best for thermal
    img = img.convert("L")
    img = img.point(lambda x: 0 if x < 160 else 255, "1")  # threshold
    w, h = img.size

    row_bytes = (w + 7) // 8
    pixels = img.load()
    data = bytearray(row_bytes * h)
    idx = 0

    for y in range(h):
        byte = 0
        bit_count = 0
        for x in range(w):
            # In mode "1": black=0, white=255. For ESC/POS, 1 = black dot.
            bit = 1 if pixels[x, y] == 0 else 0
            byte = (byte << 1) | bit
            bit_count += 1
            if bit_count == 8:
                data[idx] = byte
                idx += 1
                byte = 0
                bit_count = 0
        if bit_count:  # pad final byte
            byte <<= (8 - bit_count)
            data[idx] = byte
            idx += 1

    # GS v 0 m=0 xL xH yL yH data...
    xL = row_bytes & 0xFF
    xH = (row_bytes >> 8) & 0xFF
    yL = h & 0xFF
    yH = (h >> 8) & 0xFF
    header = b"\x1d\x76\x30\x00" + bytes([xL, xH, yL, yH])
    return header + bytes(data)


def _logo_payload(logo_path: str, max_width_dots: int) -> bytes:
    """Safely open JPG/PNG and return ESC/POS raster bytes. Empty bytes if not available."""
    if not (logo_path and Image):
        return b""
    try:
        path = logo_path if os.path.isabs(
            logo_path) else os.path.join(os.getcwd(), logo_path)
        if not os.path.exists(path):
            return b""
        with Image.open(path) as im:
            return _img_to_escpos_raster(im, max_width_dots)
    except Exception:
        return b""


def _send_raw_to_printer(payload: bytes, printer_name: str) -> bool:
    """Send RAW bytes to a Windows printer."""
    try:
        import win32print
        h = win32print.OpenPrinter(printer_name)
        try:
            win32print.StartDocPrinter(h, 1, ("ESC/POS Receipt", None, "RAW"))
            win32print.StartPagePrinter(h)
            win32print.WritePrinter(h, payload)
            win32print.EndPagePrinter(h)
            win32print.EndDocPrinter(h)
            return True
        finally:
            try:
                win32print.ClosePrinter(h)
            except Exception:
                pass
    except Exception as e:
        try:
            from flask import current_app
            current_app.logger.exception(f"[PRINT] RAW send failed: {e}")
        except Exception:
            print("[PRINT ERROR]", e)
        return False

# --------- Public helpers your code calls ----------


def print_receipt_windows_raw(text: str, printer_name: str) -> bool:
    """
    Print plain text as ESC/POS RAW (no logo).
    """
    body = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    payload = bytearray()
    payload += ESC_INIT
    payload += ALIGN_LEFT
    payload += body.encode("utf-8") + b"\n"
    payload += FEED_6 + CUT_FULL
    return _send_raw_to_printer(bytes(payload), printer_name)


def print_receipt_with_logo_windows_raw(
    fee, stu,
    printer_name: str,
    logo_path: str = "",
    school_name: str = "",
    paper_width_dots: int = 576
) -> bool:
    """
    Same logic as before; ensures the logo prints once at the top.
    Adds bold for: school name, receipt title, receipt no, student name, amounts.
    """
    # Build the body
    try:
        body_text = build_receipt_text_clean(
            fee, stu,
            school_name=current_app.config.get("SCHOOL_NAME", school_name),
            school_address=current_app.config.get("SCHOOL_ADDRESS_LINE1", ""),
            school_tagline=current_app.config.get("SCHOOL_TAGLINE", ""),
            width=int(current_app.config.get("RECEIPT_CHARS", 42)),
        )
    except TypeError:
        body_text = build_receipt_text_clean(fee, stu)

    body_text = (body_text or "").replace(
        "\r\n", "\n").replace("\r", "\n").strip("\n")

    # Resolve logo path
    logo_cfg = (logo_path or current_app.config.get(
        "RECEIPT_LOGO_PATH", "")).strip()
    if logo_cfg and not os.path.isabs(logo_cfg):
        logo_cfg = os.path.join(current_app.root_path, logo_cfg)

    # Cap the logo width
    max_logo = int(current_app.config.get("RECEIPT_LOGO_MAX_DOTS", 200))
    max_dots = min(max_logo, int(paper_width_dots or 576))

    # Build payload
    payload = bytearray()
    payload += ESC_INIT

    # --- Logo ---
    try:
        logo_bytes = _logo_payload(logo_cfg, max_dots) if logo_cfg else b""
    except Exception:
        logo_bytes = b""

    if logo_bytes:
        payload += ALIGN_CTR
        payload += logo_bytes + b"\n"

    # --- Body with selective bolding ---
    payload += ALIGN_LEFT
    for line in body_text.split("\n"):
        if school_name and school_name.upper() in line.upper():
            payload += (ALIGN_CTR + TXT_BOLD_ON +
                        line.encode("utf-8", "ignore") + TXT_BOLD_OFF + b"\n")
        elif "PAYMENT RECEIPT" in line.upper():
            payload += (ALIGN_CTR + TXT_BOLD_ON +
                        line.encode("utf-8", "ignore") + TXT_BOLD_OFF + b"\n")
        elif line.strip().startswith("Receipt No:"):
            payload += (TXT_BOLD_ON +
                        line.encode("utf-8", "ignore") + TXT_BOLD_OFF + b"\n")
        elif line.strip().startswith("Name :"):
            payload += (TXT_BOLD_ON +
                        line.encode("utf-8", "ignore") + TXT_BOLD_OFF + b"\n")
        elif "This Payment" in line or "Amount Due" in line:
            payload += (TXT_BOLD_ON +
                        line.encode("utf-8", "ignore") + TXT_BOLD_OFF + b"\n")
        else:
            payload += line.encode("utf-8", "ignore") + b"\n"

    # Feed and cut
    payload += FEED_6 + CUT_FULL

    return _send_raw_to_printer(bytes(payload), printer_name)


def handle_payment_and_print(fee_id: int) -> bool:
    """
    Print a receipt for fee_id.
    - Uses the receipt number stored in fees.receipt_no.
    - If fees.receipt_no is NULL/blank, generate one, SAVE it, then print.
    - Never overwrites an existing receipt_no.
    """
    from datetime import datetime
    import os
    from flask import current_app

    try:
        # Load fee + student
        fee, stu = load_payment_with_student(fee_id)
        if not (fee and stu):
            current_app.logger.warning(
                f"[PRINT] No fee/student for id={fee_id}")
            return False

        # --- ALWAYS come from DB ---
        rec_no = None
        try:
            rec_no = fee["receipt_no"]
        except Exception:
            rec_no = getattr(fee, "receipt_no", None)

        # If missing, generate once and persist to DB, then use that value
        if not rec_no or str(rec_no).strip() == "":
            try:
                # Use your helper if present
                rec_no = generate_receipt_no(None, fee_id)
            except Exception:
                today = datetime.now().strftime("%Y%m%d")
                rec_no = f"RCPT-{today}-{int(fee_id):06d}"

            conn = get_db_connection()
            try:
                cur = conn.cursor(dictionary=True)
                cur.execute(
                    "UPDATE fees SET receipt_no=%s WHERE id=%s", (rec_no, fee_id))
                conn.commit()
                cur.close()
            finally:
                conn.close()

            # reflect into row we pass to the builder
            fee_for_print = dict(fee)
            fee_for_print["receipt_no"] = rec_no
        else:
            fee_for_print = fee  # already has table value

        # --- Print config ---
        cfg = current_app.config
        printer_name = cfg.get("RECEIPT_PRINTER_NAME", r"GP-80220(Cut) Series")
        logo_rel = cfg.get("RECEIPT_LOGO_PATH", "") or ""
        school_name = cfg.get("SCHOOL_NAME", "") or ""
        paper_dots = int(cfg.get("RECEIPT_PAPER_DOTS", 576))

        logo_abs = ""
        if logo_rel:
            logo_abs = logo_rel if os.path.isabs(
                logo_rel) else os.path.join(current_app.root_path, logo_rel)
            if not os.path.exists(logo_abs):
                current_app.logger.warning(
                    f"[PRINT] Logo file not found: {logo_abs}")
                logo_abs = ""  # pass empty to skip

        current_app.logger.info(
            f"[PRINT] send payment_id={fee_id}; printer='{printer_name}', rec_no='{rec_no}', logo='{logo_abs or logo_rel}'"
        )

        # --- Send to printer (header/title handled in helper) ---
        ok = print_receipt_with_logo_windows_raw(
            fee_for_print,
            stu,
            printer_name=printer_name,
            logo_path=logo_abs,
            school_name=school_name,
            paper_width_dots=paper_dots,
        )

        current_app.logger.info(
            f"[PRINT] done payment_id={fee_id}, success={ok}")
        return bool(ok)

    except Exception as e:
        try:
            current_app.logger.exception(
                f"[PRINT] send failed for payment_id={fee_id}: {e}")
        except Exception:
            print(f"[PRINT ERROR] payment_id={fee_id}: {e}")
        return False


# ===================== RECEIPT NUMBER UTILITIES (drop-in) =====================
# Requirements:
# - You already have: get_db_connection()
# - Table: fees (with columns id, date_paid, receipt_no TEXT, ...)
# If receipt_no column / unique index might be missing, call
# ensure_fees_has_receipt_no_column_and_index() once at startup.


def ensure_fees_has_receipt_no_column_and_index():
    """
    Make sure fees.receipt_no exists and is (uniquely) indexed.
    Safe to call on every startup.
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cols = [c[0] for c in cur.execute("SHOW COLUMNS FROM fees").fetchall()]
    if "receipt_no" not in cols:
        cur.execute("ALTER TABLE fees ADD COLUMN receipt_no VARCHAR(20)")
        conn.commit()
    # Unique for non-NULL values; NULLs allowed to repeat
    cur.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_fees_receipt_no ON fees(receipt_no)")
    conn.commit()
    conn.close()


def generate_receipt_no(conn, fee_id: int) -> str:
    """
    Generate or fetch a receipt number from the fees table.
    Always ensure it's consistent with what is stored.
    """
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT receipt_no FROM fees WHERE id=%s", (fee_id,))
    row = cur.fetchone()
    if row and row["receipt_no"]:
        return row["receipt_no"]

    # If no receipt_no yet, create one based on date + fee_id
    from datetime import datetime
    today = datetime.now().strftime("%Y%m%d")
    new_no = f"RCPT-{today}-{fee_id:06d}"

    # Save back to DB
    cur.execute("UPDATE fees SET receipt_no=%s WHERE id=%s", (new_no, fee_id))
    conn.commit()
    cur.close()
    return new_no


def ensure_fee_has_receipt_no(conn, fee_id: int) -> None:
    """
    If fees.receipt_no is NULL/blank for the given id, set it.
    Leaves existing receipt numbers untouched.
    """
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT id, receipt_no, date_paid FROM fees WHERE id = %s", (fee_id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        return

    current = (row["receipt_no"] or "").strip()
    if current:
        cur.close()
        return  # already set

    new_no = generate_receipt_no_for_row(row["id"], row["date_paid"])
    cur2 = conn.cursor()
    cur2.execute("UPDATE fees SET receipt_no = %s WHERE id = %s",
                 (new_no, row["id"]))
    conn.commit()
    cur2.close()
    cur.close()


def backfill_missing_receipt_numbers() -> int:
    """
    Assign receipt numbers to all existing fees rows that are missing one.
    Returns number of rows updated.
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, date_paid
        FROM fees
        WHERE receipt_no IS NULL OR TRIM(receipt_no) = ''
        ORDER BY id
    """)
    rows = cur.fetchall()

    updated = 0
    for r in rows:
        rec = generate_receipt_no_for_row(r["id"], r["date_paid"])
        cur.execute(
            "UPDATE fees SET receipt_no = %s WHERE id = %s", (rec, r["id"]))
        updated += 1

    conn.commit()
    cur.close()
    conn.close()
    return updated

# ---------------------- OPTIONAL HOOKS / USAGE EXAMPLES ----------------------
# 1) Call once at startup (after your schema migrations):
# ensure_fees_has_receipt_no_column_and_index()
# backfill_missing_receipt_numbers()

# 2) Right after inserting a new payment (so printing sees the number immediately):
# conn = get_db_connection()
# cur = conn.cursor(dictionary=True)
# cur.execute("INSERT INTO fees (...) VALUES (...)", (...,))
# fee_id = cur.lastrowid
# ensure_fee_has_receipt_no(conn, fee_id)
# conn.commit()
# conn.close()

# 3) Extra safety inside your print flow (before loading fee+student):
# def handle_payment_and_print(fee_id: int) -> bool:
# try:
# conn = get_db_connection()
# ensure_fee_has_receipt_no(conn, fee_id) # guarantees receipt_no exists
# conn.close()
# fee, stu = load_payment_with_student(fee_id)
# # ... continue with your existing printing logic ...
# except Exception:
# ...
# ===========================================================================


def _logo_bytes(path, max_width):
    if not Image:
        return b""
    try:
        if not os.path.isabs(path):
            path = os.path.join(os.getcwd(), path)
        if os.path.exists(path):
            with Image.open(path) as im:
                return _img_to_escpos_raster_bytes(im, max_width=max_width)
    except Exception:
        pass
    return b""


# --- JPG/PNG logo → ESC/POS + print helper -----------------------------------


def finalize_new_payment(fee_id: int):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Set receipt_no if missing
    cur.execute("SELECT id, receipt_no FROM fees WHERE id=%s", (fee_id,))
    fee = cur.fetchone()
    if fee and not (fee["receipt_no"] or "").strip():
        rcpt = generate_receipt_no(conn, fee_id)
        cur.execute("UPDATE fees SET receipt_no=%s WHERE id=%s",
                    (rcpt, fee_id))

    # Set recorded_by from the logged-in operator (fallback System)
    op_name = get_current_operator_name()  # see helper below
    cur.execute(
        "UPDATE fees SET recorded_by=COALESCE(%s, recorded_by) WHERE id=%s", (op_name, fee_id))

    conn.commit()
    cur.close()
    conn.close()


def get_current_operator_name() -> str:
    # adapt keys to your login/session shape
    u = session.get("user") or {}
    name = (u.get("full_name") or u.get("username")
            or u.get("email") or "").strip()
    return name or "System"


# ================= Pretty Receipt Builder (text for ESC/POS) =================

# School constants for the header lines (already printed by the print helper too)
SCHOOL_TITLE = "DEMO DAY & BOARDING"
tagline="PRIMARY SCHOOL – KAMPALA"
SCHOOL_POBOX = "P.O Box 1X1X1 Kampala"

# --- tiny helpers ---


def _nz(v, d=0):
    try:
        return d if v is None else v
    except Exception:
        return d


def _get(row, key, default=None):
    try:
        return row[key] if row is not None else default
    except Exception:
        return default


def _fmt_money(x):
    try:
        return f"UGX {float(x):,.0f}"
    except Exception:
        return f"UGX {x}"


def _term_short(t):
    # e.g. "Term 3" -> "III", else keep as-is
    mapping = {"Term 1": "I", "Term 2": "II", "Term 3": "III"}
    return mapping.get((t or "").strip(), t or "")


def build_receipt_text_clean(
    fee,
    stu,
    *,
    school_name: str = "",
    school_address: str = "",
    school_tagline: str = "",
    width: int = 42,
) -> str:
    """Mono receipt body; uses fees.receipt_no and recorded_by. (no duplicate school name)"""
    W = max(30, int(width))

    # helpers
    def line(ch: str = "-") -> str: return ch * W
    def center(s: str) -> str: return (s or "").center(W)

    def money(x) -> str:
        try:
            return f"UGX {float(x or 0):,.0f}"
        except:
            return f"UGX {x}"

    def cols(left: str, right: str, mid: int = 24) -> str:
        mid = min(max(10, mid), W - 6)
        L = (left or "")[:mid]
        Rw = max(0, W - mid - 1)
        R = (right or "")[:Rw].rjust(Rw)
        return f"{L} {R}"

    def g(row, key, default=None):
        try:
            if isinstance(row, dict):
                return row.get(key, default)
            return row[key]
        except Exception:
            return default

    from datetime import datetime

    # ---- fee + student fields ----
    rec_no = g(fee, "receipt_no")
    rec_id = rec_no if rec_no else f"ID:{g(fee, 'id', '')}"
    try:
        amount = float(g(fee, "amount_paid", 0.0) or 0.0)
    except Exception:
        amount = 0.0
    term = g(fee, "term", "") or ""  # keep “Term 1/2/3” as-is
    year = g(fee, "year", "") or ""
    paid_dt = g(fee, "date_paid") or datetime.now().strftime("%Y-%m-%d")
    method = g(fee, "method") or "N/A"
    cashier = g(fee, "recorded_by") or "System"

    stu_no = g(stu, "student_number", "") or ""
    first = g(stu, "first_name", "") or ""
    middle = g(stu, "Middle_name", "") or ""
    last = g(stu, "last_name", "") or ""
    klass = g(stu, "class_name", "") or ""
    stream = g(stu, "stream", "") or ""
    sid = g(stu, "id", 0)

    # Amount Due (overall balance)
    try:
        fin = compute_student_financials(
            sid, klass, term, int(year or datetime.now().year)) or {}
    except Exception:
        fin = {}
    try:
        overall = float(fin.get("overall_balance", 0) or 0)
    except Exception:
        overall = 0.0

    full_name = " ".join(p for p in [first, middle, last] if p).strip()
    cls_line = f"{klass}{(' ' + stream) if stream else ''}".strip()

    # ---- assemble ----
    L = []
    # Header block: (logo is added by the print function), then school name/address/tagline
    # Then the single title line and a divider. No duplicate school name afterwards.
    if school_name:
        L += [center(school_name)]
        if school_address:
            L.append(center(school_address))
        if school_tagline:
            L.append(center(school_tagline))
    # Title (the print function will render this line bold/bigger)
    L += [center("*** PAYMENT RECEIPT ***"), line("-")]

    # Body
    L += [
        cols("Receipt No:", str(rec_id)),
        "",
        cols("Name :", full_name),
        cols("Number :", stu_no),
        cols("Class :", cls_line),
        cols("Term :", f"{term}, {year}"),
        cols("Date :", paid_dt),
        cols("Method :", method),
        cols("Cashier :", cashier),
        line("-"),
        " # Payment Details".ljust(W),
        line("-"),
        cols("This Payment", money(amount)),  # ONLY what is paid now
        line("-"),
        cols("Amount Due :", money(overall)),
        line("-"),
        "Thank you.".ljust(W),
    ]
    return "\n".join(L) + "\n"


# EXAMPLE usage inside your payment route (you already have fee_id, load helpers):
#
# fee, stu = load_payment_with_student(fee_id)
# if fee and stu:
# ok = print_receipt_with_logo_windows_raw(
# fee, stu,
# printer_name=r"GP-80220(Cut) Series", # EXACT Windows name
# logo_path="static/logo.jpg", # jpg or png
# school_name=current_app.config.get("SCHOOL_NAME", "My School"),
# paper_width_dots=576 # 80mm = 576; 58mm = 384
# )
# if ok:
# flash("Payment saved and sent to printer.", "success")
# else:
# flash("Payment saved. Printer not confirmed — open receipt and Reprint.", "warning")


def load_payment_with_student(payment_id: int):
    """Read one fees row + its student record."""
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, student_id, term, year, amount_paid, method, payment_type,
               expected_amount, carried_forward, date_paid
        FROM fees
        WHERE id = %s
    """, (payment_id,))
    fee = cur.fetchone()
    stu = None
    if fee:
        cur.execute("""
            SELECT id, student_number, first_name, Middle_name, last_name,
                   class_name, stream
            FROM students
            WHERE id = %s
        """, (fee["student_id"],))
        stu = cur.fetchone()
    cur.close()
    conn.close()
    return fee, stu


def handle_payment_and_print(fee_id: int) -> bool:
    """
    Print a receipt for fee_id.
    - Uses the receipt number stored in fees.receipt_no.
    - If fees.receipt_no is NULL/blank, generate one, SAVE it, then print.
    - Never overwrites an existing receipt_no.
    """
    from datetime import datetime
    import os
    from flask import current_app

    try:
        # Load fee + student
        fee, stu = load_payment_with_student(fee_id)
        if not (fee and stu):
            current_app.logger.warning(
                f"[PRINT] No fee/student for id={fee_id}")
            return False

        # --- ALWAYS come from DB ---
        rec_no = None
        try:
            rec_no = fee["receipt_no"]
        except Exception:
            rec_no = getattr(fee, "receipt_no", None)

        # If missing, generate once and persist to DB, then use that value
        if not rec_no or str(rec_no).strip() == "":
            try:
                # Use your helper if present
                rec_no = generate_receipt_no(None, fee_id)
            except Exception:
                today = datetime.now().strftime("%Y%m%d")
                rec_no = f"RCPT-{today}-{int(fee_id):06d}"

            conn = get_db_connection()
            try:
                cur = conn.cursor(dictionary=True)
                cur.execute(
                    "UPDATE fees SET receipt_no=%s WHERE id=%s", (rec_no, fee_id))
                conn.commit()
                cur.close()
            finally:
                conn.close()

            # reflect into row we pass to the builder
            fee_for_print = dict(fee)
            fee_for_print["receipt_no"] = rec_no
        else:
            fee_for_print = fee  # already has table value

        # --- Print config ---
        cfg = current_app.config
        printer_name = cfg.get("RECEIPT_PRINTER_NAME", r"GP-80220(Cut) Series")
        logo_rel = cfg.get("RECEIPT_LOGO_PATH", "") or ""
        school_name = cfg.get("SCHOOL_NAME", "") or ""
        paper_dots = int(cfg.get("RECEIPT_PAPER_DOTS", 576))

        logo_abs = ""
        if logo_rel:
            logo_abs = logo_rel if os.path.isabs(
                logo_rel) else os.path.join(current_app.root_path, logo_rel)
            if not os.path.exists(logo_abs):
                current_app.logger.warning(
                    f"[PRINT] Logo file not found: {logo_abs}")
                logo_abs = ""  # pass empty to skip

        current_app.logger.info(
            f"[PRINT] send payment_id={fee_id}; printer='{printer_name}', rec_no='{rec_no}', logo='{logo_abs or logo_rel}'"
        )

        # --- Send to printer (header/title handled in helper) ---
        ok = print_receipt_with_logo_windows_raw(
            fee_for_print,
            stu,
            printer_name=printer_name,
            logo_path=logo_abs,
            school_name=school_name,
            paper_width_dots=paper_dots,
        )

        current_app.logger.info(
            f"[PRINT] done payment_id={fee_id}, success={ok}")
        return bool(ok)

    except Exception as e:
        try:
            current_app.logger.exception(
                f"[PRINT] send failed for payment_id={fee_id}: {e}")
        except Exception:
            print(f"[PRINT ERROR] payment_id={fee_id}: {e}")
        return False


# =============================== ROUTE UPDATES ==========================================


# =================== TRANSPORT as REQUIREMENT (Single Start Payment) ===================
# Minimal additions: new tables + helpers + 2 routes + a small change in start_payment GET

# ---- 1) Schemas (idempotent) ---------------------------------------------------------


def get_student_requirements(class_name: str, term: str, student_id: int, year: int):
    """
    Combine class requirements (your existing 'requirements' table) with any
    active student extras for the term/year. Return as a list of rows shaped
    like your template expects: id, name, amount, qty(optional).
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Existing class requirements (include generic/no-term items, preserving your logic)
    cur.execute("""
        SELECT id, name, amount, qty
        FROM requirements
        WHERE class_name = %s
          AND (term = %s OR term IS NULL OR term = '')
    """, (class_name, term))
    class_rows = cur.fetchall()

    # Active extras for this student/term/year (Transport lives here)
    cur.execute("""
        SELECT id, item_name AS name, amount, NULL AS qty
        FROM student_extra_requirements
        WHERE student_id=%s AND term=%s AND year=%s AND active=1
    """, (student_id, term, year))
    extra_rows = cur.fetchall()

    cur.close()
    conn.close()

    # Return a single list (template treats them equally)
    # To let you identify extras in the template if desired, add a flag in memory:
    result = [dict(r) | {"_extra": False} for r in class_rows] + \
        [dict(r) | {"_extra": True} for r in extra_rows]
    return result


# use the same text you have in requirements names
TRANSPORT_PREFIX = "Transport — "


def _get_term_index(term: str) -> int:
    return {"Term 1": 1, "Term 2": 2, "Term 3": 3}.get(term, 1)


def compute_transport_term_status(student_id: int, class_name: str, term: str, year: int):
    """
    Compute Transport due/paid/balance for balances card.
    Paid is pulled from your existing FEES table where payment_type='requirements'
    and requirement_name matches Transport (...).
    """
    conn = get_db_connection()

    sub = transport_active_for_student(conn, student_id, term, year)
    if not sub:
        conn.close()
        return {
            "has_sub": False,
            "route_name": None,
            "fare_per_term": 0.0,
            "paid": 0.0,
            "balance": 0.0,
            "route_id": None,
        }

    req_name = f"Transport ({sub['route_name']})"

    # Due (from requirements table)
    cur = conn.cursor(dictionary=True)
    cur.execute("""
      SELECT amount FROM requirements
      WHERE class_name=%s AND term=%s AND year=%s AND name=%s
      LIMIT 1
    """, (class_name, term, year, req_name))
    due_row = cur.fetchone()
    due = float(due_row["amount"]) if due_row else float(
        sub["fare_per_term"] or 0)
    cur.close()

    # Paid (from fees entries recorded as requirements)
    cur = conn.cursor(dictionary=True)
    cur.execute("""
      SELECT COALESCE(SUM(amount_paid), 0) AS tot
      FROM fees
      WHERE student_id=%s AND term=%s AND year=%s
        AND payment_type='requirements' AND requirement_name=%s
    """, (student_id, term, year, req_name))
    paid_row = cur.fetchone()
    paid = float(paid_row["tot"] if paid_row else 0.0)

    cur.close()
    conn.close()
    return {
        "has_sub": True,
        "route_name": sub["route_name"],
        "fare_per_term": due,
        "paid": paid,
        "balance": max(due - paid, 0.0),
        "route_id": sub["route_id"],
    }










def transport_paid_via_requirements(conn, student_id: int, term: str, year: int) -> float:
    """
    Sums payments captured in 'fees' table as payment_type='requirements'
    whose requirement_name starts with 'Transport (' for the given term/year.
    """
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(SUM(amount_paid),0) AS total
          FROM fees
         WHERE student_id=%s AND term=%s AND year=%s AND LOWER(payment_type)='requirements'
           AND requirement_name LIKE 'Transport (%'
    """, (student_id, term, year))
    row = cur.fetchone()
    return float(row["total"]) if row else 0.0


def transport_get_active_subscribers(conn, term: str, year: int, route_id: int | None = None,):
    base = """
      SELECT ts.*, s.student_number, s.first_name, s.Middle_name, s.last_name, s.class_name, s.stream,
             tr.name AS route_name, tr.fare_per_term
      FROM transport_subscriptions ts
      JOIN students s ON s.id = ts.student_id
      JOIN transport_routes tr ON tr.id = ts.route_id
      WHERE ts.active=1
        AND (ts.start_year < %s OR (ts.start_year = %s AND ts.start_term <= %s))
    """
    params = [year, year, term]
    if route_id:
        base += " AND ts.route_id=%s"
        params.append(route_id)
    base += " ORDER BY tr.name, s.class_name, s.last_name"
    cur = conn.cursor(dictionary=True)
    cur.execute(base, params)
    rows = cur.fetchall()
    cur.close()
    return rows


def transport_is_already_subscribed(conn, student_id: int, route_id: int) -> bool:
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT 1 FROM transport_subscriptions WHERE active=1 AND student_id=%s AND route_id=%s",
        (student_id, route_id)
    )
    r = cur.fetchone()
    cur.close()
    return bool(r)


def transport_has_active_subscription(conn, student_id: int, term: str, year: int, route_id: int | None = None) -> bool:
    """
    True iff student has an active subscription that applies to (term, year).
    If route_id is provided, it must match; otherwise any active route qualifies.
    """
    base = """
      SELECT 1
      FROM transport_subscriptions
      WHERE active=1
        AND student_id=%s
        AND (start_year < %s OR (start_year = %s AND start_term <= %s))
    """
    params = [student_id, year, year, term]
    if route_id:
        base += " AND route_id=%s"
        params.append(route_id)

    base += "LIMIT 1"
    cur = conn.cursor(dictionary=True)
    cur.execute(base, params)
    row = cur.fetchone()
    cur.close()
    return bool(row)


def _transport_source_string(sn: str, route_name: str) -> str:
    return f"Transport (SN: {sn}) - {route_name}"


def transport_record_payment_for_student_number(student_number: str, route_name: str, term: str, year: int,
                                                amount: float, method: str, recorded_by: str):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
      INSERT INTO other_income (source, amount, term, year, description, recorded_by, date_received)
      VALUES (%s, %s, %s, %s, %s, %s, NOW())
    """, (_transport_source_string(student_number, route_name), amount, term, year, TRANSPORT_DESC, recorded_by))
    conn.commit()
    cur.close()
    conn.close()


def transport_paid_total_for_sn(conn, student_number: str, term: str, year: int) -> float:
    like_token = f"%SN: {student_number}%"
    cur = conn.cursor(dictionary=True)
    cur.execute("""
      SELECT COALESCE(SUM(amount),0) AS total
      FROM other_income
      WHERE description=%s AND term=%s AND year=%s AND source LIKE %s
    """, (TRANSPORT_DESC, term, year, like_token))
    row = cur.fetchone()
    return float(row["total"]) if row else 0.0


# must match the names you inject in requirements
TRANSPORT_PREFIX = "Transport — "


def _term_index(t: str) -> int:
    return 1 if t == "Term 1" else 2 if t == "Term 2" else 3


def transport_subscription_info(student_id: int, term: str, year: int):
    """
    Return the active subscription for this student at/before (term,year).
    {has_sub: bool, route_id: int|None, route_name: str, fare_per_term: float}
    """
    conn = get_db_connection()
    t_idx = _term_index(term)
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT ts.route_id, tr.name AS route_name, tr.fare_per_term
        FROM transport_subscriptions ts
        JOIN transport_routes tr ON tr.id = ts.route_id
        WHERE ts.student_id = %s AND ts.active=1
          AND (ts.start_year < %s
               OR (ts.start_year = %s AND
                   CASE ts.start_term
                     WHEN 'Term 1' THEN 1
                     WHEN 'Term 2' THEN 2
                     WHEN 'Term 3' THEN 3
                     ELSE 1
                   END <= %s)
              )
        ORDER BY ts.created_at DESC
        LIMIT 1
        """,
        (student_id, year, year, t_idx)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        return {"has_sub": False, "route_id": None, "route_name": "", "fare_per_term": 0.0}
    return {
        "has_sub": True,
        "route_id": int(row["route_id"]),
        "route_name": row["route_name"],
        "fare_per_term": float(row["fare_per_term"] or 0.0),
    }


def transport_subscribe(student_id: int, route_id: int, term: str, year: int):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
      INSERT IGNORE INTO transport_subscriptions (student_id, route_id, start_term, start_year, active)
      VALUES (%s, %s, %s, %s, 1)
    """, (student_id, route_id, term, year))
    # If there was an inactive sub for the same route earlier, make it active:
    cur.execute("""
      UPDATE transport_subscriptions
         SET active=1
       WHERE student_id=%s AND route_id=%s AND start_term=%s AND start_year=%s AND active=0
    """, (student_id, route_id, term, year))
    conn.commit()
    cur.close()
    conn.close()


def _normalize_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_")


def _read_opening_rows(file_storage, filename: str):
    """
    Returns a list of dict rows with keys:
      student_number (str), amount (float), asof_year (int|None)
    Accepts .csv or .xlsx
    """
    name = (filename or "").lower()
    raw = file_storage.read()

    rows = []
    if name.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        headers = [_normalize_header(h) for h in reader.fieldnames or []]
        for r in reader:
            row = {_normalize_header(k): (v or "").strip()
                   for k, v in r.items()}
            rows.append(row)

    elif name.endswith(".xlsx"):
        bio = io.BytesIO(raw)
        df = pd.read_excel(bio, dtype=str)  # read everything as text first
        df.columns = [_normalize_header(c) for c in df.columns]
        for _, r in df.fillna("").iterrows():
            rows.append({k: str(v).strip() for k, v in r.to_dict().items()})

    else:
        raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")

    # Map & coerce
    normalized = []
    for i, r in enumerate(rows, start=2):  # start=2: header is row 1
        sn = (r.get("student_number") or r.get("studentno") or "").strip()
        amt_raw = (r.get("amount") or r.get("balance") or "").strip()
        asof_raw = (r.get("asof_year") or r.get(
            "as_of_year") or r.get("year") or "").strip()

        if not sn or not amt_raw:
            normalized.append({
                "_row": i, "student_number": sn, "amount": None, "asof_year": None,
                "_error": "Missing student_number or amount"
            })
            continue

        try:
            amt = float(amt_raw)
        except Exception:
            normalized.append({
                "_row": i, "student_number": sn, "amount": None, "asof_year": None,
                "_error": f"Amount '{amt_raw}' is not a number"
            })
            continue

        asof_year = None
        if asof_raw:
            try:
                asof_year = int(asof_raw)
            except Exception:
                normalized.append({
                    "_row": i, "student_number": sn, "amount": amt, "asof_year": None,
                    "_error": f"asof_year '{asof_raw}' is not a valid year"
                })
                continue

        normalized.append({
            "_row": i, "student_number": sn, "amount": amt, "asof_year": asof_year
        })
    return normalized


def set_opening_balance(conn, student_id: int, amount: float, year: int, note: str = ""):
    """
    Insert an opening balance record into the fees table.
    Uses payment_type='opening_balance' so it's distinguishable.
    """
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        INSERT INTO fees (
            student_id, term, year,
            expected_amount, bursary_amount, amount_paid,
            date_paid, method, payment_type, recorded_by, payment_item
        )
        VALUES (%s, 'Term 3', %s, %s, 0, 0, NOW(), 'N/A', 'opening_balance', %s, %s)
    """, (
        student_id, int(year), float(amount),
        session.get("username") or "system",
        note
    ))
    cur.close()


def carried_forward(student_id, term, year):
    """Outstanding before the active term/year + any opening_balance rows (always included)."""
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT expected_amount, bursary_amount, amount_paid, term, year
        FROM fees
        WHERE student_id=%s AND (
              year < %s
           OR (year = %s AND
               (CASE term
                 WHEN 'Term 1' THEN 1
                 WHEN 'Term 2' THEN 2
                 WHEN 'Term 3' THEN 3
                 ELSE 99
               END)
               < (CASE %s WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 99 END))
        ) AND (payment_type IS NULL OR payment_type!='requirements')
    """, (student_id, year, year, term))
    prev_rows = cur.fetchall()
    cur.close()

    out_prev = 0.0
    for r in prev_rows:
        out_prev += (float(r["expected_amount"] or 0) -
                     float(r["bursary_amount"] or 0) -
                     float(r["amount_paid"] or 0))

    # Always include all opening_balance rows (no date/term filter)
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT COALESCE(SUM(expected_amount - bursary_amount - amount_paid), 0) AS t
        FROM fees
        WHERE student_id=%s AND payment_type='opening_balance'
    """, (student_id,))
    ob_row = cur.fetchone()
    cur.close()
    conn.close()

    total = max(out_prev, 0.0) + float(ob_row["t"] or 0.0)
    return max(total, 0.0)


def _hp_ensure():
    try:
        ensure_holiday_package_schema()
    except Exception:
        pass

# 2) Subject helper


def get_or_create_subject_by_name(name: str) -> int:
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT id FROM subjects WHERE LOWER(name)=LOWER(%s)", (name.strip(),))
    row = cur.fetchone()
    if row:
        sid = row["id"]
    else:
        cur.execute("INSERT INTO subjects(name) VALUES (%s)", (name.strip(),))
        sid = c.lastrowid
        conn.commit()
    cur.close()
    conn.close()
    return sid


# 4) Aggregation
def hp_aggregate_student_subject(conn, student_id: int, subject_id: int, term: str, year: int) -> float | None:
    cur = conn.cursor(dictionary=True)
    cur.execute("""
      SELECT score, max_score, weight
      FROM holiday_package_scores
      WHERE student_id=%s AND subject_id=%s AND term=%s AND year=%s
    """, (student_id, subject_id, term, year))
    rows = cur.fetchall()
    cur.close()
    if not rows:
        return None
    any_weight = any(r["weight"] is not None for r in rows)
    if any_weight:
        total, total_w = 0.0, 0.0
        for r in rows:
            w = float(r["weight"] or 0.0)
            if w <= 0:
                continue
            pct = (float(r["score"] or 0.0) /
                   float(r["max_score"] or 100.0)) * 100.0
            total += pct * w
            total_w += w
        return (total / total_w) if total_w > 0 else None
    else:
        vals = [(float(r["score"]) / float(r["max_score"] or 100.0))
                * 100.0 for r in rows if r["max_score"]]
        return (sum(vals)/len(vals)) if vals else None

# 5) Sync into record_score


def hp_sync_into_record_score(class_name: str, term: str, year: int, initials: str = "HP"):
    ensure_record_score_table(get_db_connection())
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT id FROM students WHERE archived=0 AND class_name=%s",
        (class_name,)
    )
    students = cur.fetchall()
    cur.execute("""
      SELECT DISTINCT h.subject_id
      FROM holiday_package_scores h
      JOIN students s ON s.id=h.student_id
      WHERE s.class_name=%s AND h.term=%s AND h.year=%s
    """, (class_name, term, year))
    subject_rows = cur.fetchall()
    subject_ids = [r["subject_id"] for r in subject_rows]
    if not students or not subject_ids:
        conn.close()
        return 0

    cur.execute("SELECT grade, lower_limit, upper_limit FROM grading_scale")
    grading = cur.fetchall()

    def _grade_from_mark(mark):
        if mark is None:
            return None
        for g in grading:
            if float(g["lower_limit"]) <= mark <= float(g["upper_limit"]):
                return g["grade"]
        return None

    n = 0
    for s in students:
        for subj_id in subject_ids:
            avg = hp_aggregate_student_subject(
                conn, s["id"], subj_id, term, year)
            if avg is None:
                continue
            grd = _grade_from_mark(avg)
            cur.execute("""
              INSERT INTO record_score
              (student_id, subject_id, term, year, average_mark,
               grade, comment, initials, processed_on)
              VALUES (%s, %s, %s, %s, %s, %s, 'Holiday Package', %s, CURRENT_TIMESTAMP)
              ON DUPLICATE KEY UPDATE(student_id, subject_id, term, year) DO UPDATE SET
                average_mark=excluded.average_mark,
                grade=excluded.grade,
                comment=excluded.comment,
                initials=excluded.initials,
                processed_on=CURRENT_TIMESTAMP
            """, (s["id"], subj_id, term, year, avg, grd, initials))
            n += 1
    conn.commit()
    cur.close()
    conn.close()
    return n


# 2) Create app and load config


def create_app():
    app = Flask(
        __name__,
        template_folder="templates",
        static_folder="static",
        static_url_path="/static",  # serve /static/... URLs
    )
    from reports.reports import fee_reports
    app.register_blueprint(fee_reports)

    # ---- load config by env
    env = os.getenv("FLASK_ENV", "production").lower()
    if env == "development":
        from config import DevConfig
        app.config.from_object(DevConfig)
    elif env == "testing":
        from config import TestConfig
        app.config.from_object(TestConfig)
    else:
        from config import ProdConfig
        app.config.from_object(ProdConfig)

    configure_logging(app)

    # ---- allow static files to bypass any auth checks
    @app.before_request
    def _skip_static():
        # IMPORTANT: return NOTHING for static; do not redirect/abort
        if request.path.startswith("/static/"):
            return None
            
    

    return app


app = create_app()


@app.before_request
def _coerce_session_role():
    if "role" in session:
        session["role"] = _norm_role(session["role"])

  # ✅ allow static files to bypass login checks


@app.before_request
def _sync_active_term_year():
    if "user_id" in session:
        try:
            conn = get_db_connection()
            t, y = get_active_academic_year().get(
                "current_term"), get_active_academic_year().get("year")
            conn.close()
            if t and y:
                if session.get("current_term") != t or session.get("current_year") != y:
                    session["current_term"] = t
                    session["current_year"] = y
        except Exception:
            pass


@app.context_processor
def inject_current_role():
    from flask import session
    return {"current_role": (session.get("role") or "").lower()}


@app.route('/api')
def api_home():
    return "School Management System API"


@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # --- get active academic session (safe fallback) ---
    ay = get_active_academic_year() or {}
    term = (ay.get("current_term") or ay.get("term") or "Term 1")
    try:
        year = int(ay.get("year") or ay.get(
            "active_year") or datetime.now().year)
    except Exception:
        year = datetime.now().year

    role = session.get("role", "teacher")
    is_admin = role in ("admin", "director", "headteacher", "bursar")
    stats = {}

    if is_admin:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        try:
            # totals (use aliases so dict access works)
            cur.execute(
                "SELECT COUNT(*) AS total_students FROM students WHERE archived=0")
            total_students = (cur.fetchone() or {}).get("total_students", 0)

            cur.execute(
                "SELECT COALESCE(SUM(amount_paid),0) AS fees_in_term "
                "FROM fees WHERE term=%s AND year=%s",
                (term, year),
            )
            fees_in_term = (cur.fetchone() or {}).get("fees_in_term", 0)

            cur.execute(
                "SELECT COALESCE(SUM(amount),0) AS other_in_term "
                "FROM other_income WHERE term=%s AND year=%s",
                (term, year),
            )
            other_in_term = (cur.fetchone() or {}).get("other_in_term", 0)

            cur.execute(
                "SELECT COALESCE(SUM(amount),0) AS exp_in_term "
                "FROM expenses WHERE term=%s AND year=%s",
                (term, year),
            )
            exp_in_term = (cur.fetchone() or {}).get("exp_in_term", 0)

            term_net_income = (fees_in_term or 0) + \
                (other_in_term or 0) - (exp_in_term or 0)

            cur.execute(
                "SELECT COALESCE(SUM(amount_paid),0) AS cumulative_fees FROM fees")
            cumulative_fees = (cur.fetchone() or {}).get("cumulative_fees", 0)

            cur.execute(
                "SELECT COALESCE(SUM(amount),0) AS cumulative_other FROM other_income")
            cumulative_other = (cur.fetchone() or {}).get(
                "cumulative_other", 0)

            cumulative_income = (cumulative_fees or 0) + \
                (cumulative_other or 0)

            cur.execute(
                "SELECT COALESCE(SUM(amount),0) AS total_expenses FROM expenses")
            total_expenses = (cur.fetchone() or {}).get("total_expenses", 0)

            stats = {
                "total_students": total_students or 0,
                "term_net_income": term_net_income or 0,
                "cumulative_net_income": cumulative_income or 0,
                "total_expenses": total_expenses or 0,
            }
        finally:
            try:
                cur.close()
                conn.close()
            except Exception:
                pass

    return render_template(
        "dashboard.html",
        username=session.get("full_name") or session.get("username") or "User",
        user_id=session.get("user_id"),
        role=role,
        show_admin=is_admin,
        active_term=term,
        active_year=year,
        stats=stats,
    )


@app.route('/add_subject', methods=['GET', 'POST'])
@require_role("admin", "headteacher", "deputyheadteacher", "dos")
def add_subject():
    conn = get_db_connection()

    if request.method == 'POST':
        subject_name = request.form['name']
        part_names = request.form.getlist('parts')

        # Insert subject
        cur = conn.cursor(dictionary=True)
        cur.execute("INSERT INTO subjects (name) VALUES (%s)", (subject_name,))
        subject_id = c.lastrowid

        # Insert parts (if any)
        for part in part_names:
            if part.strip():
                cur = conn.cursor(dictionary=True)
                cur.execute(
                    "INSERT INTO subject_papers (subject_id, paper_name) VALUES (%s, %s)", (subject_id, part.strip()))

        conn.commit()
        cur.close()
        conn.close()
        flash("Subject and parts added successfully", "success")
        return redirect(url_for('add_subject'))

    conn.close()
    return render_template("add_subject.html")


# --------- SUBJECTS: list/search/add/edit/delete, import/export ----------
@app.route("/subjects", methods=["GET", "POST"])
@require_role("admin", "headteacher", "dos", "classmanager", "deputyheadteacher")  # adjust roles as you prefer
def manage_subjects():
    conn = get_db_connection()

    # Add new subject
    if request.method == "POST" and request.form.get("action") == "create":
        name = (request.form.get("name") or "").strip()
        code = (request.form.get("code") or "").strip().upper() or None
        if not name:
            flash("Subject name is required.", "warning")
        else:
            try:
                cur = conn.cursor(dictionary=True)
                cur.execute(
                    "INSERT INTO subjects (name, code) VALUES (%s, %s)", (name, code))
                conn.commit()
                cur.close()
                flash("Subject added.", "success")
            except Exception as e:
                flash(f"Could not add subject: {e}", "danger")
        conn.close()
        return redirect(url_for("manage_subjects"))

    # Search/filter
    q = (request.args.get("q") or "").strip()
    params = []
    sql = "SELECT id, name, code FROM subjects"
    if q:
        sql += " WHERE name LIKE %s OR code LIKE %s"
        like = f"%{q}%"
        params.extend([like, like])
    sql += " ORDER BY name"

    cur = conn.cursor(dictionary=True)
    cur.execute(sql, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("manage_subjects.html", rows=rows, q=q)


@app.route("/subjects/<int:sid>/edit", methods=["GET", "POST"])
@require_role("admin", "headteacher", "deputyheadteacher")
def edit_subject(sid):
    conn = get_db_connection()

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        code = (request.form.get("code") or "").strip().upper() or None
        if not name:
            flash("Subject name is required.", "warning")
        else:
            try:
                cur = conn.cursor(dictionary=True)
                cur.execute(
                    "UPDATE subjects SET name=%s, code=%s WHERE id=%s", (name, code, sid))
                conn.commit()
                cur.close()
                flash("Subject updated.", "success")
                conn.close()
                return redirect(url_for("manage_subjects"))
            except Exception as e:
                flash(f"Could not update subject: {e}", "danger")

    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, name, code FROM subjects WHERE id=%s", (sid,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        flash("Subject not found.", "warning")
        return redirect(url_for("manage_subjects"))
    return render_template("edit_subject.html", row=row)


@app.route("/subjects/<int:sid>/delete", methods=["POST"])
@require_role("admin", "headteacher", "dos", "deputyheadteacher")
def delete_subject(sid):
    conn = get_db_connection()
    try:
        # if FK constraints exist (e.g., record_score.subject_id), this will error when referenced—catch it
        cur = conn.cursor(dictionary=True)
        cur.execute("DELETE FROM subjects WHERE id=%s", (sid,))
        conn.commit()
        cur.close()
        flash("Subject deleted.", "success")
    except Exception as e:
        flash(
            f"Cannot delete: subject is referenced by marks/records. ({e})", "danger")
    finally:
        conn.close()
    return redirect(url_for("manage_subjects"))

# --------- Import/Export ----------


@app.route("/subjects/export")
@require_role("admin", "headteacher", "dos", "deputyheadteacher")
def subjects_export():
    conn = get_db_connection()
    df = pd.read_sql_query(
        "SELECT name, code FROM subjects ORDER BY name", conn)
    conn.close()

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Subjects")
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="subjects.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/subjects/import", methods=["POST"])
@require_role("admin", "headteacher")
def subjects_import():
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("Choose a file to import.", "warning")
        return redirect(url_for("manage_subjects"))

    ext = "." + file.filename.rsplit(".", 1)[-1].lower()
    try:
        if ext in (".xlsx", ".xls"):
            df = pd.read_excel(file)
        elif ext == ".csv":
            df = pd.read_csv(file)
        else:
            flash("Unsupported file type. Use .xlsx / .xls / .csv", "danger")
            return redirect(url_for("manage_subjects"))

        cols = {c.lower(): c for c in df.columns}
        need = ["name", "code"]
        missing = [k for k in need if k not in cols]
        if missing:
            flash(f"Missing columns: {', '.join(missing)}", "danger")
            return redirect(url_for("manage_subjects"))

        conn = get_db_connection()
        added = updated = 0
        for _, r in df.iterrows():
            name = str(r[cols["name"]]).strip() if pd.notna(
                r[cols["name"]]) else None
            code = str(r[cols["code"]]).strip().upper(
            ) if pd.notna(r[cols["code"]]) else None
            if not name:
                continue
            # Upsert by name; if you prefer by code, switch WHERE clause
            cur = conn.cursor(dictionary=True)
            cur.execute("SELECT id FROM subjects WHERE name=%s", (name,))
            existing = cur.fetchone()
            cur.close()
            if existing:
                cur = conn.cursor(dictionary=True)
                cur.execute("UPDATE subjects SET code=%s WHERE id=%s",
                            (code or None, existing["id"]))
                updated += 1
                cur.close()
            else:
                cur = conn.cursor(dictionary=True)
                cur.execute(
                    "INSERT INTO subjects (name, code) VALUES (%s, %s)", (name, code or None))
                added += 1
                cur.close()
        conn.commit()
        conn.close()
        flash(f"Import done. Added: {added}, Updated: {updated}.", "success")
    except Exception as e:
        flash(f"Import failed: {e}", "danger")

    return redirect(url_for("manage_subjects"))


# ---------- Safe DB helper (commit/rollback/close handled) ----------

@contextmanager
def mysql_conn(*, dictionary=True):
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        try:
            # heal idle/stale connections if possible
            conn.ping(reconnect=True, attempts=3, delay=1)
        except Exception:
            pass
        cur = conn.cursor(dictionary=dictionary)
        yield conn, cur
        conn.commit()
    except Exception:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        raise
    finally:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        if conn:
            try:
                conn.close()
            except Exception:
                pass


# ======================= USERS: CREATE + LIST =======================
@app.route("/users", methods=["GET", "POST"])
@require_role("admin")
def manage_users():
    conn = get_db_connection()

    # employees for dropdown + preselect via ?employee_id=...
    employees = _fetch_employees_for_dropdown(conn)
    pre_emp_id = (request.args.get("employee_id") or "").strip()

    if request.method == "POST":
        f = request.form
        username = (f.get("username") or "").strip()
        password = (f.get("password") or "").strip()
        role = (f.get("role") or "").strip()
        status = (f.get("status") or "active").strip()
        employee_id = f.get("employee_id")
        employee_id = int(employee_id) if (
            employee_id and employee_id.isdigit()) else None

        # auto-suggest username if blank and employee chosen
        if not username and employee_id:
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                SELECT first_name, COALESCE(Middle_name,'') AS Middle_name, last_name
                FROM employees WHERE id=%s
            """, (employee_id,))
            emp = cur.fetchone()
            cur.close()
            if emp:
                username = _suggest_username(
                    conn, emp["first_name"], emp["last_name"])

        # basic validations
        if not username or not password:
            conn.close()
            flash("Username and password are required.", "danger")
            return redirect(url_for("manage_users"))
        if role not in ALLOWED_ROLES:
            conn.close()
            flash("Invalid role.", "danger")
            return redirect(url_for("manage_users"))
        if status not in ("active", "archived"):
            conn.close()
            flash("Invalid status.", "danger")
            return redirect(url_for("manage_users"))

        # unique username check
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT 1 FROM users WHERE username=%s LIMIT 1", (username,))
        if cur.fetchone():
            cur.close()
            conn.close()
            flash("Username already exists.", "warning")
            return redirect(url_for("manage_users"))
        cur.close()

        # create user
        try:
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                INSERT INTO users (username, password_hash, role, status, employee_id)
                VALUES (%s, %s, %s, %s, %s)
            """, (username, generate_password_hash(password), role, status, employee_id))
            conn.commit()
            cur.close()
            flash("User created.", "success")
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            flash(f"Could not create user: {e}", "danger")
        finally:
            conn.close()
        return redirect(url_for("manage_users"))

    # GET list (+ optional search)
    q = (request.args.get("q") or "").strip()
    cur = conn.cursor(dictionary=True)
    if q:
        cur.execute("""
            SELECT u.*, e.first_name, e.Middle_name AS middle_name, e.last_name, e.designation
              FROM users u
              LEFT JOIN employees e ON e.id = u.employee_id
             WHERE u.username LIKE %s OR u.role LIKE %s
             ORDER BY u.id DESC
        """, (f"%{q}%", f"%{q}%"))
    else:
        cur.execute("""
            SELECT u.*, e.first_name, e.Middle_name AS middle_name, e.last_name, e.designation
              FROM users u
              LEFT JOIN employees e ON e.id = u.employee_id
             ORDER BY u.id DESC
        """)
    users = cur.fetchall()
    cur.close()
    conn.close()

    return render_template(
        "users.html",
        users=users,
        employees=employees,
        q=q,
        pre_emp_id=pre_emp_id,
        ALLOWED_ROLES=ALLOWED_ROLES
    )

# ============================ USERS: EDIT ===========================


@app.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@require_role("admin")
def edit_user(user_id):
    conn = get_db_connection()

    # Load current user row
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM users WHERE id=%s", (user_id,))
    user = cur.fetchone()
    cur.close()
    if not user:
        conn.close()
        flash("User not found.", "warning")
        return redirect(url_for("manage_users"))

    # Dropdown data
    employees = _fetch_employees_for_dropdown(conn)

    if request.method == "POST":
        f = request.form
        username = (f.get("username") or "").strip()
        role = (f.get("role") or "").strip()
        status = (f.get("status") or "active").strip()
        employee_id = f.get("employee_id")
        employee_id = int(employee_id) if (
            employee_id and employee_id.isdigit()) else None
        new_pass = (f.get("new_password") or "").strip()

        # Auto-suggest if username left blank but an employee is chosen
        if not username and employee_id:
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                SELECT first_name, COALESCE(Middle_name,'') AS Middle_name, last_name
                FROM employees WHERE id=%s
            """, (employee_id,))
            emp = cur.fetchone()
            cur.close()
            if emp:
                username = _suggest_username(
                    conn, emp["first_name"], emp["last_name"])

        # Validations
        if not username:
            conn.close()
            flash("Username is required.", "danger")
            return redirect(url_for("edit_user", user_id=user_id))
        if role not in ALLOWED_ROLES:
            conn.close()
            flash("Invalid role.", "danger")
            return redirect(url_for("edit_user", user_id=user_id))
        if status not in ("active", "archived"):
            conn.close()
            flash("Invalid status.", "danger")
            return redirect(url_for("edit_user", user_id=user_id))

        # Unique username check (exclude this user)
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT 1 FROM users WHERE username=%s AND id<>%s", (username, user_id))
        taken = cur.fetchone()
        cur.close()
        if taken:
            conn.close()
            flash("Username already taken by another account.", "warning")
            return redirect(url_for("edit_user", user_id=user_id))

        # Update
        try:
            cur = conn.cursor(dictionary=True)
            if new_pass:
                cur.execute("""
                    UPDATE users
                       SET username=%s,
                           password_hash=%s,
                           role=%s,
                           status=%s,
                           employee_id=%s
                     WHERE id=%s
                """, (username, generate_password_hash(new_pass), role, status, employee_id, user_id))
            else:
                cur.execute("""
                    UPDATE users
                       SET username=%s,
                           role=%s,
                           status=%s,
                           employee_id=%s
                     WHERE id=%s
                """, (username, role, status, employee_id, user_id))
            conn.commit()
            cur.close()
            flash("User updated.", "success")
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            flash(f"Could not update user: {e}", "danger")
        finally:
            conn.close()

        return redirect(url_for("manage_users"))

    # GET
    conn.close()
    return render_template(
        "user_edit.html",
        user=user,
        employees=employees,
        ALLOWED_ROLES=ALLOWED_ROLES
    )


# ============================ USERS: DELETE =========================
@app.route("/users/<int:user_id>/delete", methods=["POST"])
@require_role("admin")
def delete_user(user_id):
    if session.get("user_id") == user_id:
        flash("You cannot delete your own account.", "warning")
        return redirect(url_for("manage_users"))

    try:
        with mysql_conn() as (conn, cur):
            cur.execute("DELETE FROM users WHERE id=%s", (user_id,))
        flash("User deleted.", "info")
    except Exception as e:
        flash(f"Could not delete user: {e}", "danger")

    return redirect(url_for("manage_users"))


# ====================== USERS: TOGGLE ACTIVE/ARCHIVED ===============
@app.route("/users/<int:user_id>/toggle", methods=["POST"])
@require_role("admin")
def toggle_user(user_id):
    if session.get("user_id") == user_id:
        flash("You cannot archive/activate your own account.", "warning")
        return redirect(url_for("manage_users"))

    with mysql_conn() as (conn, cur):
        cur.execute(
            "SELECT id, username, status FROM users WHERE id=%s", (user_id,))
        row = cur.fetchone()

    if not row:
        flash("User not found.", "warning")
        return redirect(url_for("manage_users"))

    new_status = "archived" if row["status"] == "active" else "active"
    try:
        with mysql_conn() as (conn, cur):
            cur.execute("UPDATE users SET status=%s WHERE id=%s",
                        (new_status, user_id))
        flash(f"User {row['username']} is now {new_status}.", "info")
    except Exception as e:
        flash(f"Could not change status: {e}", "danger")

    return redirect(url_for("manage_users"))


@app.route('/subjects', methods=['GET', 'POST'])
@require_role('admin')
def subjects():
    conn = get_db_connection()
    message = ""

    # Handle adding subject
    if request.method == 'POST':
        if 'add_subject' in request.form:
            name = request.form['name'].strip()
            initial = request.form['initial'].strip().upper()
            if name and initial:
                try:
                    cur = conn.cursor(dictionary=True)
                    cur.execute(
                        "INSERT INTO subjects (name, initial) VALUES (%s, %s)", (name, initial))
                    conn.commit()
                    cur.close()
                    message = "Subject added successfully!"
                except mysql.connector.Error:
                    message = "Subject or initial already exists!"
            else:
                message = "Please fill in all fields."

        # Handle adding paper
        elif 'add_paper' in request.form:
            subject_id = request.form['subject_id']
            paper_name = request.form['paper_name'].strip()
            paper_initial = request.form['paper_initial'].strip().upper()
            if paper_name and paper_initial:
                try:
                    cur = conn.cursor(dictionary=True)
                    cur.execute('''
                        INSERT INTO subject_papers (subject_id, paper_name, paper_initial)
                        VALUES (%s, %s, %s)
                    ''', (subject_id, paper_name, paper_initial))
                    conn.commit()
                    cur.close()
                    message = "Paper added successfully!"
                except mysql.connector.Error:
                    message = "Paper already exists for this subject!"
            else:
                message = "Please fill in all paper fields."

    # Fetch subjects with papers
    cur = conn.cursor(dictionary=True)
    cur.execute('SELECT * FROM subjects ORDER BY name ASC')
    subjects = c.fetchall()
    cur.close()

    # Fetch all papers grouped by subject
    papers_dict = {}
    for subj in subjects:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            'SELECT * FROM subject_papers WHERE subject_id = %s', (subj['id'],))
        papers_dict[subj['id']] = cur.fetchall()
        cur.close()

    conn.close()
    return render_template('subjects.html', subjects=subjects, papers_dict=papers_dict, message=message)


@app.route('/record_score', methods=['GET', 'POST'])
@require_role('admin', 'teacher','classmanager','deputyheadteacher', 'headteacher')
def record_score():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("SELECT id, CONCAT_WS(' ', first_name, COALESCE(Middle_name, ''), last_name) AS student_name,, class FROM students WHERE status='active'")
    students = cur.fetchall()
    cur.execute("SELECT id, name FROM subjects")
    subjects = cur.fetchall()
    cur.close()
    conn.close()

    if request.method == 'POST':
        student_id = request.form.get('student_id')
        subject_id = request.form.get('subject_id')
        subject_part_id = request.form.get('subject_part_id')
        term = request.form.get('term')
        year = request.form.get('year')
        score = request.form.get('score')

        if not all([student_id, subject_id, subject_part_id, term, year, score]):
            flash('Please fill all fields.')
            return redirect(url_for('record_score'))

        try:
            year = int(year)
            score = float(score)
        except ValueError:
            flash('Year must be integer and score must be a number.')
            return redirect(url_for('record_score'))

        conn = get_db_connection()
        try:
            cur = conn.cursor(dictionary=True)
            cur.execute('''
                INSERT INTO record_score (student_id, subject_part_id, term, year, score)
                VALUES (%s, %s, %s, %s, %s)
            ''', (student_id, subject_part_id, term, year, score))
            conn.commit()
            cur.close()
            flash('Score recorded successfully.')
        except Exception as e:
            flash(f'Error: {str(e)}')
        finally:
            conn.close()

        return redirect(url_for('record_score'))

    return render_template('record_score.html', students=students, subjects=subjects)


@app.route('/record_batch_score', methods=['GET', 'POST'])
@require_role('admin', 'headteacher','classmanager', 'deputyheadteacher')
def record_batch_score():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Get active term and year
    cur.execute(
        "SELECT year, current_term FROM academic_years WHERE is_active = 1")
    academic = cur.fetchone()
    cur.close()
    if not academic:
        flash("No active academic year found.", "warning")
        return redirect(url_for('dashboard'))

    year, term = academic['year'], academic['current_term']

    # Load filters from GET or POST
    class_name = request.args.get(
        'class_name') or request.form.get('class_name')
    stream = request.args.get('stream') or request.form.get('stream')
    subject_id = request.args.get(
        'subject_id') or request.form.get('subject_id')
    part_id = request.args.get(
        'subject_part_id') or request.form.get('subject_part_id')

    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM subjects")
    subjects = cur.fetchall()
    cur.close()
    parts = []
    if subject_id:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT * FROM subject_papers WHERE subject_id = V", (subject_id,))
        parts = cur.fetchall()
        cur.close()

    students = []
    scores = {}

    if all([class_name, stream, subject_id]):
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT * FROM students
            WHERE class_name = %s AND stream = %s AND status = 'active' AND archived = 0
        """, (class_name, stream))
        students = cur.fetchall()

        cur.execute("""
            SELECT * FROM record_score
            WHERE year = %s AND term = %s AND subject_id = %s
                  AND (subject_part_id = %s OR (%s IS NULL AND subject_part_id IS NULL))
        """, (year, term, subject_id, part_id, part_id))
        rows = cur.fetchall()

        for row in rows:
            scores[str(row['student_id'])] = row
        cur.close()

    if request.method == 'POST' and students:
        for student in students:
            sid = str(student['id'])
            bot = request.form.get(f'bot_{sid}')
            mid = request.form.get(f'mid_{sid}')
            eot = request.form.get(f'eot_{sid}')

            marks = [int(m) for m in [bot, mid, eot] if m and m.isdigit()]
            avg = sum(marks) / len(marks) if marks else None

            if avg is not None:
                cur = conn.cursor(dictionary=True)
                cur.execute("""
                    SELECT id FROM record_score
                    WHERE student_id = %s AND subject_id = %s AND year = %s AND term = %s
                          AND (subject_part_id = %s OR (%s IS NULL AND subject_part_id IS NULL))
                """, (student['id'], subject_id, year, term, part_id, part_id))
                exists = cur.fetchone()
                cur.close()

                if exists:
                    cur = conn.cursor(dictionary=True)
                    cur.execute("""
                        UPDATE record_score
                        SET bot_mark = %s, midterm_mark = %s, eot_mark = %s, average_mark = %s, processed_on = CURRENT_TIMESTAMP
                        WHERE id = %s
                    """, (bot or None, mid or None, eot or None, avg, exists['id']))
                    cur.close()
                else:
                    cur = conn.cursor(dictionary=True)
                    cur.execute("""
                        INSERT INTO record_score (
                            student_id, subject_id, subject_part_id,
                            year, term, bot_mark, midterm_mark, eot_mark, average_mark
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        student['id'], subject_id, part_id or None, year, term,
                        bot or None, mid or None, eot or None, avg
                    ))
                    cur.close()

        conn.commit()
        flash("Scores recorded successfully.", "success")
        return redirect(url_for('record_batch_score'))

    conn.close()
    return render_template(
        'record_batch_score.html',
        students=students,
        class_name=class_name,
        stream=stream,
        term=term,
        year=year,
        subject_id=subject_id,
        subject_part_id=part_id,
        subjects=subjects,
        parts=parts,
        scores=scores
    )


@app.route('/view_scores', methods=['GET'])
@require_role('admin', 'headteacher', 'bursar', 'dos', 'classmanager', 'deputyheadteacher')
def view_scores():
    conn = get_db_connection()

    class_name = request.args.get('class_name', '')
    stream = request.args.get('stream', '')
    year = request.args.get('year', '')
    term = request.args.get('term', '')
    subject_id = request.args.get('subject_id', '')

    query = """
        SELECT rs.*, s.student_number, s.first_name, s.Middle_name, s.last_name,
               s.class_name, s.stream, subj.name AS subject, sp.paper_name
        FROM record_score rs
        JOIN students s ON rs.student_id = s.id
        JOIN subjects subj ON rs.subject_id = subj.id
        LEFT JOIN subject_papers sp ON rs.subject_part_id = sp.id
        WHERE s.status = 'active' AND s.archived = 0
    """
    params = []

    if class_name:
        query += " AND s.class_name = %s"
        params.append(class_name)
    if stream:
        query += " AND s.stream = %s"
        params.append(stream)
    if year:
        query += " AND rs.year = %s"
        params.append(year)
    if term:
        query += " AND rs.term = %s"
        params.append(term)
    if subject_id:
        query += " AND rs.subject_id = %s"
        params.append(subject_id)

    query += " ORDER BY s.last_name, sp.paper_name"
    cur = conn.cursor(dictionary=True)
    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close()

    # Group scores by student and then by paper
    results_by_student = {}
    for row in rows:
        sid = row['student_id']
        if sid not in results_by_student:
            results_by_student[sid] = {
                'student_number': row['student_number'],
                'full_name': f"{row['first_name']} {row['Middle_name'] or ''} {row['last_name']}",
                'class_name': row['class_name'],
                'stream': row['stream'],
                'subject': row['subject'],
                'parts': [],
                'total_avg': 0,
                'total_parts': 0
            }

        avg = row['average_mark'] or 0
        results_by_student[sid]['parts'].append({
            'record_id': row['id'],
            'paper_name': row['paper_name'] or '-',
            'bot_mark': row['bot_mark'],
            'midterm_mark': row['midterm_mark'],
            'eot_mark': row['eot_mark'],
            'average_mark': avg,
            'grade': row['grade'],
            'comment': row['comment']
        })

        if avg > 0:
            results_by_student[sid]['total_avg'] += avg
            results_by_student[sid]['total_parts'] += 1

    for s in results_by_student.values():
        if s['total_parts']:
            s['total_avg'] = round(s['total_avg'] / s['total_parts'], 1)
        else:
            s['total_avg'] = 0

    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, name FROM subjects")
    subjects = cur.fetchall()
    cur.execute(
        "SELECT DISTINCT year FROM academic_years ORDER BY year DESC"
    )
    available_years = [row['year'] for row in cur.fetchall()]
    cur.close()
    conn.close()

    return render_template(
        "view_scores_grouped.html",
        results=results_by_student,
        class_name=class_name,
        stream=stream,
        year=year,
        term=term,
        subject_id=subject_id,
        subjects=subjects,
        available_years=available_years
    )


@app.route('/view_scores_grouped', methods=['GET', 'POST'])
@require_role('admin', 'headteacher')
def view_scores_grouped():
    conn = get_db_connection()

    # Get active academic year
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT year, current_term FROM academic_years WHERE is_active = 1")
    academic = cur.fetchone()
    year = academic['year'] if academic else ''
    term = academic['current_term'] if academic else ''

    class_name = request.args.get('class_name', '')
    stream = request.args.get('stream', '')
    subject_id = request.args.get('subject_id', '')

    cur.execute("SELECT id, name FROM subjects")
    subjects = cur.fetchall()
    grouped_data = []
    cur.close()

    if class_name and stream:
        query = '''
            SELECT rs.*, s.student_number,
                   CONCAT_WS(' ',s.first_name, COALESCE(s.Middle_name, ''), s.last_name) AS full_name,
                   s.class_name, s.stream, subj.name AS subject, sp.paper_name
            FROM record_score rs
            JOIN students s ON s.id = rs.student_id
            JOIN subjects subj ON subj.id = rs.subject_id
            LEFT JOIN subject_papers sp ON rs.subject_part_id = sp.id
            WHERE s.class_name = %s AND s.stream = %s AND s.status = 'active' AND s.archived = 0
              AND rs.year = %s AND rs.term = %s
        '''
        params = [class_name, stream, year, term]

        if subject_id:
            query += " AND rs.subject_id = %s"
            params.append(subject_id)

        cur = conn.cursor(dictionary=True)
        cur.execute(query, params)
        rows = cur.fetchall()
        grouped = {}
        cur.close()

        for r in rows:
            key = (r['student_number'], r['subject'])

            if key not in grouped:
                grouped[key] = {
                    'student_number': r['student_number'],
                    'full_name': r['full_name'],
                    'class_name': r['class_name'],
                    'stream': r['stream'],
                    'subject': r['subject'],
                    'parts': [],
                    'overall_avg': 0,
                    'grade': '',
                    'comment': ''
                }

            marks = [r['bot_mark'], r['midterm_mark'], r['eot_mark']]
            marks = [m for m in marks if m is not None]
            avg = sum(marks) / len(marks) if marks else None

            grouped[key]['parts'].append({
                'record_id': r['id'],
                'paper': r['paper_name'] or 'N/A',
                'bot': r['bot_mark'],
                'mid': r['midterm_mark'],
                'eot': r['eot_mark'],
                'average': round(avg, 2) if avg else None
            })

        for g in grouped.values():
            part_avgs = [p['average']
                         for p in g['parts'] if p['average'] is not None]
            if part_avgs:
                overall = sum(part_avgs) / len(part_avgs)
                g['overall_avg'] = round(overall, 2)
                cur = conn.cursor(dictionary=True)
                cur.execute(
                    "SELECT grade, comment FROM grading_scale WHERE %s BETWEEN lower_limit AND upper_limit LIMIT 1",
                    (overall,)
                )
                row = cur.fetchone()
                g['grade'] = row['grade'] if row else ''
                g['comment'] = row['comment'] if row else ''
                cur.close()

        grouped_data = list(grouped.values())

    conn.close()

    return render_template(
        'view_scores_grouped.html',
        class_name=class_name,
        stream=stream,
        term=term,
        year=year,
        subject_id=subject_id,
        subjects=subjects,
        grouped_scores=grouped_data
    )


@app.route('/save_scores_grouped', methods=['POST'])
@require_role('admin', 'headteacher')
def save_scores_grouped():
    conn = get_db_connection()

    for key in request.form:
        if key.startswith(('bot_', 'mid_', 'eot_')):
            prefix, record_id = key.split('_')
            mark = request.form.get(key)
            if mark:
                try:
                    cur = conn.cursor(dictionary=True)
                    cur.execute(
                        f"UPDATE record_score SET {prefix}_mark = %s WHERE id = %s",
                        (int(mark), int(record_id))
                    )
                    cur.close()
                except Exception as e:
                    flash(f"Error updating record {record_id}: {e}", "danger")

    conn.commit()
    conn.close()
    flash("Scores updated successfully.", "success")
    return redirect(url_for('view_scores_grouped', **request.args))


@app.route('/export_grouped_scores')
@require_role('admin', 'headteacher')
def export_grouped_scores():
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Grouped Scores")

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    class_name = request.args.get('class_name')
    stream = request.args.get('stream')
    subject_id = request.args.get('subject_id')

    cur.execute(
        "SELECT year, current_term FROM academic_years WHERE is_active = 1")
    academic = cur.fetchone()
    year = academic['year'] if academic else ''
    term = academic['current_term'] if academic else ''

    query = '''
        SELECT rs.*, s.student_number,
               CONCAT_WS(' ',s.first_name, COALESCE(s.Middle_name, ''), s.last_name) AS full_name,
               s.class_name, s.stream, subj.name AS subject, sp.paper_name
        FROM record_score rs
        JOIN students s ON s.id = rs.student_id
        JOIN subjects subj ON subj.id = rs.subject_id
        LEFT JOIN subject_papers sp ON rs.subject_part_id = sp.id
        WHERE s.class_name = %s AND s.stream = %s AND rs.year = %s AND rs.term = %s
    '''
    params = [class_name, stream, year, term]

    if subject_id:
        query += " AND rs.subject_id = %s"
        params.append(subject_id)

    cur.execute(query, params)
    rows = cur.fetchall()

    worksheet.write_row(0, 0, [
        'Student No', 'Name', 'Class', 'Stream',
        'Subject', 'Paper', 'BOT', 'MID', 'EOT', 'Average'
    ])

    for row_num, r in enumerate(rows, start=1):
        marks = [r['bot_mark'], r['midterm_mark'], r['eot_mark']]
        valid_marks = [m for m in marks if m is not None]
        avg = sum(valid_marks) / len(valid_marks) if valid_marks else 0

        worksheet.write_row(row_num, 0, [
            r['student_number'], r['full_name'], r['class_name'], r['stream'],
            r['subject'], r['paper_name'] or 'N/A',
            r['bot_mark'], r['midterm_mark'], r['eot_mark'],
            round(avg, 2)
        ])

    workbook.close()
    output.seek(0)
    cur.close()
    conn.close()

    return send_file(
        output,
        download_name="grouped_scores.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/update_grouped_scores', methods=['POST'])
@require_role('admin', 'headteacher')
def update_grouped_scores():
    conn = get_db_connection()

    for key, value in request.form.items():
        if key.startswith('bot_') or key.startswith('midterm_') or key.startswith('eot_'):
            field, record_id = key.split('_')

            cur = conn.cursor(dictionary=True)
            cur.execute(
                f"UPDATE record_score SET {field}_mark = %s WHERE id = %s", (value or None, record_id))
            cur.close()

    conn.commit()
    conn.close()
    flash("Scores updated successfully.", "success")
    return redirect(url_for('view_scores_grouped', **request.args.to_dict()))


@app.route('/results')
def results():
    results = Result.query.options(
        joinedload(Result.student),
        joinedload(Result.subject)
    ).all()

    grouped_results = defaultdict(
        lambda: {"student_info": None, "subject": None, "papers": [], "total_avg": 0})

    for result in results:
        key = (result.student_id, result.subject_id)
        avg = sum(filter(None, [result.bot_mark, result.midterm_mark, result.eot_mark])) / \
            3 if all([result.bot_mark, result.midterm_mark,
                     result.eot_mark]) else None

        grading = get_grade_comment(avg) if avg is not None else None
        grade = grading.grade if grading else ''
        comment = grading.comment if grading else ''

        if not grouped_results[key]["student_info"]:
            grouped_results[key]["student_info"] = result.student
            grouped_results[key]["subject"] = result.subject.name

        grouped_results[key]["papers"].append({
            "id": result.id,
            "subject_part": result.subject_part,
            "bot_mark": result.bot_mark,
            "midterm_mark": result.midterm_mark,
            "eot_mark": result.eot_mark,
            "average_mark": round(avg, 2) if avg is not None else '',
            "grade": grade,
            "comment": comment
        })

    # Add total average per group
    for group in grouped_results.values():
        valid_avgs = [float(p['average_mark'])
                      for p in group['papers'] if p['average_mark'] != '']
        group['total_avg'] = round(
            sum(valid_avgs) / len(valid_avgs), 2) if valid_avgs else ''

    return render_template('your_template.html', grouped_results=grouped_results)


@app.route('/import_scores', methods=['POST'])
@require_role('admin', 'headteacher')
def import_scores():
    class_name = request.form.get('class_name')
    stream = request.form.get('stream')
    subject_id = request.form.get('subject_id')
    subject_part_id = request.form.get('subject_part_id') or None

    # Get active academic year and term
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT year, current_term FROM academic_years WHERE is_active = 1")
    academic = cur.fetchone()
    year, term = academic['year'], academic['current_term']
    file = request.files.get('score_file')
    cur.close()
    if not file:
        flash("No file uploaded", "danger")
        return redirect(url_for('record_batch_score'))

    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()

    try:
        if ext == '.csv':
            df = pd.read_csv(file)
        elif ext in ['.xls', '.xlsx']:
            df = pd.read_excel(
                file, engine='openpyxl' if ext == '.xlsx' else 'xlrd')
        else:
            flash("Unsupported file format. Use .xlsx, .xls or .csv", "danger")
            return redirect(url_for('record_batch_score'))
    except Exception as e:
        flash(f"Failed to read file: {e}", "danger")
        return redirect(url_for('record_batch_score'))

    required_cols = ['Student Number', 'BOT', 'MIDTERM', 'EOT']
    if not all(col in df.columns for col in required_cols):
        flash("Missing required columns: Student Number, BOT, MIDTERM, EOT", "danger")
        return redirect(url_for('record_batch_score'))

    imported = 0
    for _, row in df.iterrows():
        student_number = row['Student Number']
        bot = row.get('BOT')
        midterm = row.get('MIDTERM')
        eot = row.get('EOT')

        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id FROM students
            WHERE student_number = %s AND class_name = %s AND stream = %s AND archived = 0
        """, (student_number, class_name, stream))
        student = cur.fetchone()
        cur.close()
        if not student:
            continue

        student_id = student['id']
        marks = [m for m in [bot, midterm, eot] if pd.notnull(m)]
        avg = sum(marks) / len(marks) if marks else None

        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id FROM record_score
            WHERE student_id = %s AND subject_id = %s AND term = %s AND year = %s
            AND (subject_part_id = %s OR (%s IS NULL AND subject_part_id IS NULL))
        """, (student_id, subject_id, term, year, subject_part_id, subject_part_id))
        exists = cur.fetchone()
        cur.close()

        if exists:
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                UPDATE record_score SET bot_mark = %s, midterm_mark = %s, eot_mark = %s, average_mark = %s
                WHERE id = %s
            """, (bot, midterm, eot, avg, exists['id']))
            cur.close()
        else:
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                INSERT INTO record_score (
                    student_id, subject_id, subject_part_id, year, term,
                    bot_mark, midterm_mark, eot_mark, average_mark
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (student_id, subject_id, subject_part_id, year, term, bot, midterm, eot, avg))
            cur.close()
        imported += 1

    conn.commit()
    conn.close()

    flash(f"{imported} scores imported successfully.", "success")
    return redirect(url_for('record_batch_score',
                            class_name=class_name,
                            stream=stream,
                            subject_id=subject_id,
                            subject_part_id=subject_part_id))


@app.route('/download_score_template')
@require_role('admin', 'headteacher','dos','teacher','classmanager')
def download_score_template():
    class_name = request.args.get('class_name')
    stream = request.args.get('stream')

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT student_number, first_name, Middle_name, last_name
        FROM students
        WHERE class_name = %s AND stream = %s AND status = 'active' AND archived = 0
        ORDER BY student_number
    """, (class_name, stream))
    students = cur.fetchall()
    cur.close()
    conn.close()

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"
    ws.append(["Student Number", "Student Name", "BOT", "MIDTERM", "EOT"])

    for s in students:
        names = [s['first_name'], s['Middle_name'], s['last_name']]
        full_name = " ".join(filter(None, names)).strip()
        ws.append([s['student_number'], full_name, "", "", ""])

    # Save to in-memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"scores_template_{class_name}_{stream}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/get_subject_parts/<int:subject_id>')
def get_subject_parts(subject_id):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT id, paper_name AS name FROM subject_papers WHERE subject_id = %s", (subject_id,))
    parts = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify(parts=[dict(p) for p in parts])


@app.route('/end_of_year_process')
def end_of_year_process():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    current_year = 2025  # You can make this dynamic

    # Get active students
    cur.execute(
        "SELECT id, name, class_name FROM students WHERE status='active'")
    students = cur.fetchall()
    cur.close()

    for student in students:
        student_id = student['id']
        class_name = student['class_name']

        # Get distinct subject_ids the student has results for
        cur = conn.cursor(dictionary=True)
        cur.execute('''
            SELECT sp.subject_id
            FROM student_results sr
            JOIN subject_parts sp ON sr.subject_part_id = sp.id
            WHERE sr.student_id = %s AND sr.year = %s
            GROUP BY sp.subject_id
        ''', (student_id, current_year))
        subjects = cur.fetchall()
        cur.close()

        total_avg = 0
        subject_count = 0

        for subject in subjects:
            subject_id = subject['subject_id']

            # Average all part scores for this subject
            cur = conn.cursor(dictionary=True)
            cur.execute('''
                SELECT AVG(score) as avg_score
                FROM student_results sr
                JOIN subject_parts sp ON sr.subject_part_id = sp.id
                WHERE sr.student_id = %s AND sp.subject_id = %s AND sr.year = %s
            ''', (student_id, subject_id, current_year))
            avg_score = cur.fetchone()['avg_score']
            cur.close()

            if avg_score is not None:
                total_avg += avg_score
                subject_count += 1

                # Archive result
                cur = conn.cursor(dictionary=True)
                cur.execute('''
                    INSERT INTO archived_results (student_id, subject_id, average_score, class_name, year)
                    VALUES (%s, %s, %s, %s, %s)
                ''', (student_id, subject_id, avg_score, class_name, current_year))
                cur.close()

        # Determine promotion
        if subject_count == 0:
            continue  # No subjects found

        final_avg = total_avg / subject_count
        class_name_number = int(''.join(filter(str.isdigit, class_name)))
        class_name_prefix = ''.join(filter(str.isalpha, class_name)).upper()

        # Determine next class
        if final_avg >= 50:
            new_class_name = f"{class_name_prefix}{class_number_number + 1}"
        else:
            new_class_name = class_name  # Repeat

        # If already in P7, mark graduated
        if class_name.upper() == 'P7':
            new_status = 'graduated'
        else:
            new_status = 'active'

        # Update student
        cur = conn.cursor(dictionary=True)
        cur.execute('''
            UPDATE students SET class = %s, status = %s
            WHERE id = %s
        ''', (new_class, new_status, student_id))
        cur.close()

    conn.commit()
    conn.close()

    return 'End-of-year processing complete: results archived, students promoted or repeated.'


# ===== MARKS HUB (main page) ==============================================


@app.route("/marks/hub", methods=["GET", "POST"])
@require_role("admin", "headteacher", "teacher", "dos", "classmanager", "deputyheadteacher")
def marks_hub():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # ---------- reference lists ----------
        cur.execute("""
            SELECT DISTINCT class_name
            FROM classes
            WHERE class_name IS NOT NULL
            ORDER BY class_name
        """)
        classes = [r["class_name"] for r in cur.fetchall()]

        cur.execute("""
            SELECT DISTINCT stream
            FROM classes
            WHERE stream IS NOT NULL
            ORDER BY stream
        """)
        streams = [r["stream"] for r in cur.fetchall()] or ["A"]

        cur.execute("SELECT id, name, code FROM subjects ORDER BY name")
        subjects = cur.fetchall()

        # ---------- active session ----------
        ay = get_active_academic_year()
        active_year = int(ay.get("year"))
        active_term = ay.get("current_term") or ay.get("term") or "Term 1"

        # ensure table exists (with extra columns)
        ensure_record_score_table(conn)

        # helper: average of present marks (OTH, HP, BOT, MID, EOT)
        def _avg5(oth, hp, bot, mid, eot):
            vals = [v for v in (oth, hp, bot, mid, eot) if v is not None]
            return round(sum(vals) / len(vals)) if vals else None

        # ---------- POST: save a single row ----------
        if request.method == "POST" and request.form.get("save_row") == "1":
            student_no = (request.form.get("student_number") or "").strip()
            subject_id = _safe_int(request.form.get("subject_id"))
            term = (request.form.get("term") or active_term).strip()
            year = _safe_int(request.form.get("year")) or active_year

            oth = _safe_int(request.form.get("other_mark"))
            hp = _safe_int(request.form.get("holiday_mark"))
            bot = _safe_int(request.form.get("bot_mark"))
            mid = _safe_int(request.form.get("midterm_mark"))
            eot = _safe_int(request.form.get("eot_mark"))
            avg = _avg5(oth, hp, bot, mid, eot)

            if not student_no or not subject_id:
                flash("Student number and subject are required.", "warning")
                return redirect(url_for("marks_hub"))

            student_id = resolve_student_id(conn, student_no)
            if not student_id:
                flash("Student not found or archived.", "warning")
                return redirect(url_for("marks_hub"))

            initials = get_user_initials(conn, session.get("user_id"))

            # Use a fresh cursor for writes (if you prefer)
            wcur = conn.cursor(dictionary=True)
            wcur.execute("""
                INSERT INTO record_score (
                    student_id, subject_id, term, year,
                    other_mark, holiday_mark, bot_mark, midterm_mark, eot_mark,
                    average_mark, initials
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE
                    other_mark = VALUES(other_mark),
                    holiday_mark = VALUES(holiday_mark),
                    bot_mark = VALUES(bot_mark),
                    midterm_mark = VALUES(midterm_mark),
                    eot_mark = VALUES(eot_mark),
                    average_mark = VALUES(average_mark),
                    initials = COALESCE(NULLIF(VALUES(initials),''), initials),
                    processed_on = CURRENT_TIMESTAMP
            """, (student_id, subject_id, term, year,
                  oth, hp, bot, mid, eot, avg, initials))

            conn.commit()
            wcur.close()

            flash("Saved.", "success")
            return redirect(url_for("marks_hub", **{k: v for k, v in request.args.items()}))

        # ---------- GET: filters ----------
        filter_class = (request.args.get("class") or "").strip()
        filter_stream = (request.args.get("stream") or "").strip()
        filter_subject = _safe_int(request.args.get("subject_id"))
        filter_term = (request.args.get("term") or active_term).strip()
        filter_year = _safe_int(request.args.get("year")) or active_year

        # ---------- student list ----------
        students = []
        if filter_class:
            q = (
                "SELECT id, student_number, first_name, "
                "COALESCE(Middle_name,'') AS m, last_name, "
                "class_name, stream "
                "FROM students WHERE class_name=%s AND archived=0"
            )
            args = [filter_class]
            if filter_stream:
                q += " AND stream=%s"
                args.append(filter_stream)
            q += " ORDER BY last_name, first_name"

            # execute then fetch
            cur.execute(q, args)
            students = cur.fetchall()

        # ---------- existing scores for this subject/term/year ----------
        scores_by_student = {}
        if students and filter_subject:
            sids = tuple(s["id"] for s in students)
            sid_clause = f"({sids[0]})" if len(sids) == 1 else str(sids)
            sql = f"""
                SELECT student_id,
                       other_mark, holiday_mark, bot_mark, midterm_mark, eot_mark,
                       average_mark, initials
                  FROM record_score
                 WHERE subject_id=%s AND term=%s AND year=%s
                   AND student_id IN {sid_clause}
            """
            cur.execute(sql, (filter_subject, filter_term, filter_year))
            for r in cur.fetchall():
                scores_by_student[r["student_id"]] = r

        my_initials = get_user_initials(conn, session.get("user_id"))

        return render_template(
            "marks_hub.html",
            classes=classes, streams=streams, subjects=subjects,
            active_term=active_term, active_year=active_year,
            filter_class=filter_class, filter_stream=filter_stream,
            filter_subject=filter_subject, filter_term=filter_term,
            filter_year=filter_year,
            students=students,
            scores_by_student=scores_by_student,
            my_initials=my_initials,
        )
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

# ===== Template download (pre-fills subject/term/year) =====================


@app.route("/marks/template")
@require_role("admin", "headteacher", "teacher", "dos","classmanager", "deputyheadteacher")
def marks_template():
    # ---- filters (pick the students you want in the file) -------------------
    class_name = (request.args.get("class_name") or "").strip()
    stream = (request.args.get("stream") or "").strip()
    subject_id = _safe_int(request.args.get("subject_id"))

    # ---- academic year context (with optional override) ---------------------
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute(
        "SELECT year, current_term FROM academic_years WHERE is_active = 1 LIMIT 1")
    ay = cur.fetchone()
    if not ay:
        cur.close()
        conn.close()
        flash("Please activate an academic year.", "warning")
        return redirect(url_for("dashboard"))

    default_year = int(ay["year"])
    default_term = ay["current_term"]

    term = (request.args.get("term") or default_term).strip()
    try:
        year = int(request.args.get("year") or default_year)
    except ValueError:
        year = default_year

    # ---- subject info (optional, just to prefill columns) -------------------
    subj_name = subj_code = ""
    if subject_id:
        cur.execute("SELECT name, code FROM subjects WHERE id=%s",
                    (subject_id,))
        srow = cur.fetchone()
        if srow:
            subj_name, subj_code = srow["name"], (srow["code"] or "")

    # ---- pull students (filtered) -------------------------------------------
    qs = """
        SELECT id,
               student_number,
               TRIM(CONCAT_WS(' ', s.first_name, COALESCE(s.Middle_name, ''), s.last_name)) AS full_name,
               class_name,
               stream
          FROM students s
         WHERE archived = 0
    """
    params = []
    if class_name:
        qs += " AND class_name = %s"
        params.append(class_name)
    if stream:
        qs += " AND stream = %s"
        params.append(stream)
    qs += " ORDER BY class_name, stream, last_name, first_name"
    cur.execute(qs, params)
    students = cur.fetchall()

    cur.close()
    conn.close()

    # ---- build workbook (headers include FULL NAME) -------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Marks"

    headers = [
        "student_number", "full_name", "class", "stream",
        "subject_id", "subject_name", "subject_code",
        "term", "year",
        # order: OTH, HP, BOT, MID, EOT
        "other_mark", "holiday_mark", "bot_mark", "midterm_mark", "eot_mark",
        "initials"
    ]
    ws.append(headers)

    if students:
        for s in students:
            ws.append([
                s["student_number"], s["full_name"], s["class_name"], s["stream"],
                subject_id or "", subj_name, subj_code,
                term, year,
                "", "", "", "", "",  # marks left blank
                ""
            ])
    else:
        # blank example row if no students matched the filter
        ws.append([
            "STD-YYYY-001", "Jane Doe", class_name or "", stream or "",
            subject_id or "", subj_name, subj_code,
            term, year,
            "", "", "", "", "",
            ""
        ])

    # nice column widths
    from openpyxl.utils import get_column_letter
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(
            i)].width = min(max(len(h) + 6, 14), 40)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"marks_template_{class_name or 'all'}_{term}_{year}.xlsx"
    return send_file(
        out,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ===== Upload (CSV/XLSX) – auto-fills initials if missing ==================


@app.route("/marks/upload", methods=["POST"])
@require_role("admin", "headteacher", "teacher", "dos", "classmanager", "deputyheadteacher")
def marks_upload():
    inserted = skipped = 0
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    def _avg5(oth, hp, bot, mid, eot):
        vals = [v for v in (oth, hp, bot, mid, eot) if v is not None]
        return round(sum(vals) / len(vals)) if vals else None

    try:
        ensure_record_score_table(conn)

        # active academic year for defaults
        cur.execute(
            "SELECT year, current_term FROM academic_years WHERE is_active=1 LIMIT 1")
        ay = cur.fetchone()
        if not ay:
            flash("Please activate an academic year.", "warning")
            return redirect(url_for("marks_hub", **request.args))
        default_year = int(ay["year"])
        default_term = ay["current_term"]

        auto_initials = get_user_initials(conn, session.get("user_id"))

        file = request.files.get("file")
        if not file:
            flash("No file uploaded.", "warning")
            return redirect(url_for("marks_hub", **request.args))

        name = (file.filename or "").lower()
        if name.endswith(".csv"):
            df = pd.read_csv(file)
        elif name.endswith((".xls", ".xlsx")):
            df = pd.read_excel(file)
        else:
            flash("Unsupported file format. Upload CSV or Excel.", "danger")
            return redirect(url_for("marks_hub", **request.args))

        # normalize headers to lower-case for safety
        df.columns = [str(c).strip().lower() for c in df.columns]

        for _, R in df.iterrows():
            # helpers that handle NaN/None
            def sval(key):
                v = R.get(key)
                return "" if pd.isna(v) else str(v).strip()

            def ival(key):
                v = R.get(key)
                try:
                    return int(v) if not pd.isna(v) and str(v).strip() != "" else None
                except Exception:
                    return None

            sn = sval("student_number")
            if not sn:
                skipped += 1
                continue

            # resolve subject (by id, code or name)
            subj_id = resolve_subject_id(
                conn,
                subject_id=ival("subject_id"),
                subject_code=(sval("subject_code") or None),
                subject_name=(sval("subject_name") or None),
            )
            if not subj_id:
                skipped += 1
                continue

            student_id = resolve_student_id(conn, sn)
            if not student_id:
                skipped += 1
                continue

            term = sval("term") or default_term
            year = ival("year") or default_year

            oth = ival("other_mark")
            hp = ival("holiday_mark")
            bot = ival("bot_mark")
            mid = ival("midterm_mark")
            eot = ival("eot_mark")
            avg = _avg5(oth, hp, bot, mid, eot)

            file_initials = sval("initials")
            initials = file_initials if file_initials else auto_initials

            # MySQL upsert syntax; relies on UNIQUE(student_id, subject_id, term, year)
            cur.execute("""
                INSERT INTO record_score (
                    student_id, subject_id, term, year,
                    other_mark, holiday_mark, bot_mark, midterm_mark, eot_mark,
                    average_mark, initials
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE
                    other_mark = COALESCE(VALUES(other_mark), other_mark),
                    holiday_mark = COALESCE(
                        VALUES(holiday_mark), holiday_mark),
                    bot_mark = COALESCE(VALUES(bot_mark), bot_mark),
                    midterm_mark = COALESCE(
                        VALUES(midterm_mark), midterm_mark),
                    eot_mark = COALESCE(VALUES(eot_mark), eot_mark),
                    average_mark = COALESCE(
                        VALUES(average_mark), average_mark),
                    initials = IF(VALUES(initials) IS NOT NULL AND VALUES(initials) <> '',
                                        VALUES(initials), initials),
                    processed_on = CURRENT_TIMESTAMP
            """, (student_id, subj_id, term, year,
                  oth, hp, bot, mid, eot,
                  avg, initials))
            inserted += 1

        conn.commit()
        flash(
            f"Upload complete: {inserted} rows processed; skipped {skipped}.", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Upload failed: {e}", "danger")
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    return redirect(url_for("marks_hub", **request.args))


# ===== Delete & Export =====================================================

@app.route("/marks/delete/<int:score_id>", methods=["POST"])
@require_role("admin", "headteacher", "dos", "classmanager", "deputyheadteacher")
def marks_delete(score_id):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        ensure_record_score_table(conn)
        cur.execute("DELETE FROM record_score WHERE id=%s", (score_id,))
        conn.commit()
        flash("Deleted.", "info")

        # ---- AUDIT ----
        audit_from_request(
            conn,
            action="marks_delete",
            target_table="record_score",
            target_id=score_id
        )
    except Exception as e:
        conn.rollback()
        flash(f"Delete failed: {e}", "danger")
        audit_from_request(
            conn,
            action="marks_delete",
            outcome="failure",
            severity="warning",
            target_table="record_score",
            target_id=score_id,
            details={"error": str(e)}
        )
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("marks_hub", **request.args))


@app.route("/marks/export")
@require_role("admin", "headteacher", "teacher", "dos","classmanger", "deputyheadteacher")
def marks_export():
    # ---- filters -------------------------------------------------------------
    class_name = (request.args.get("class_name") or "").strip()
    stream = (request.args.get("stream") or "").strip()
    subject_id = _safe_int(request.args.get("subject_id"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        ensure_record_score_table(conn)

        # ---- academic year (with override support) ---------------------------
        cur.execute(
            "SELECT year, current_term FROM academic_years WHERE is_active=1")
        ay = cur.fetchone()
        if not ay:
            flash("Please activate an academic year.", "warning")
            return redirect(url_for("dashboard"))
        default_year = int(ay["year"])
        default_term = ay["current_term"]

        term = (request.args.get("term") or default_term).strip()
        try:
            year = int(request.args.get("year") or default_year)
        except ValueError:
            year = default_year

        # ---- query (order kept: OTH, HP, BOT, MID, EOT in export columns) ----
        q = """
        SELECT st.student_number,
               st.first_name||' '||COALESCE(st.Middle_name,'')||' '||st.last_name AS full_name,
               st.class_name, st.stream,
               sub.name AS subject_name, sub.code AS subject_code,
               rs.term, rs.year,
               rs.other_mark, rs.holiday_mark, rs.bot_mark, rs.midterm_mark, rs.eot_mark,
               rs.average_mark, rs.initials
        FROM record_score rs
        JOIN students st ON st.id = rs.student_id
        JOIN subjects sub ON sub.id = rs.subject_id
        WHERE rs.term = %s AND rs.year = %s
        """
        params = [term, year]
        if class_name:
            q += " AND st.class_name = %s"
            params.append(class_name)
        if stream:
            q += " AND st.stream = %s"
            params.append(stream)
        if subject_id:
            q += " AND rs.subject_id = %s"
            params.append(subject_id)
        q += " ORDER BY st.class_name, st.stream, st.last_name, st.first_name, sub.name"

        cur.execute(q, params)
        rows = cur.fetchall()

        # ---- workbook --------------------------------------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = "Marks"
        headers = [
            "student_number", "full_name", "class", "stream", "subject_name", "subject_code",
            "term", "year",
            # OTH, HP, BOT, MID, EOT
            "other_mark", "holiday_mark", "bot_mark", "midterm_mark", "eot_mark",
            "average_mark", "initials"
        ]
        ws.append(headers)

        for r in rows:
            ws.append([
                r["student_number"], r["full_name"], r["class_name"], r["stream"],
                r["subject_name"], r["subject_code"], r["term"], r["year"],
                r["other_mark"] if r["other_mark"] is not None else "",
                r["holiday_mark"] if r["holiday_mark"] is not None else "",
                r["bot_mark"] if r["bot_mark"] is not None else "",
                r["midterm_mark"] if r["midterm_mark"] is not None else "",
                r["eot_mark"] if r["eot_mark"] is not None else "",
                r["average_mark"] if r["average_mark"] is not None else "",
                r["initials"] or ""
            ])

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        fname = f"marks_export_{class_name or 'all'}_{term}_{year}.xlsx"
        return send_file(
            out, as_attachment=True, download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    finally:
        cur.close()
        conn.close()


# (Optional) simple summary
@app.route("/marks/summary")
@require_role("admin", "headteacher", "teacher", "dos","classmanager", "deputyheadteacher")
def marks_summary():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT st.class_name, st.stream, sub.name AS subject, rs.term, rs.year,
                   COUNT(*) AS entries
              FROM record_score rs
              JOIN students st ON st.id = rs.student_id
              JOIN subjects sub ON sub.id = rs.subject_id
             GROUP BY st.class_name, st.stream, sub.name, rs.term, rs.year
             ORDER BY rs.year DESC,
                      CASE rs.term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 9 END,
                      st.class_name, st.stream, sub.name
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()
    return render_template("marks_summary.html", rows=rows)


@app.route("/performance/summary")
@require_role("admin", "headteacher", "dos", "bursar", "teacher", "classmanager", "deputyheadteacher")
def performance_summary():
    """
    Class-by-class performance summary for the ACTIVE term/year.
    - Average uses any available of OTH/HP/BOT/MID/EOT/CA (prefers average_mark).
    - Grades come from grading_scale (lower_limit..upper_limit).
    - Aggregate/Division use ONLY core subjects: ENG, MATH, SCI, SST.
      If any core is missing => NG.
    """
    ay = get_active_academic_year() or {}
    term = (ay.get("current_term") or ay.get("term") or "Term 1").strip()
    year = int(ay.get("year") or ay.get("active_year") or datetime.now().year)

    class_filter = (request.args.get("class_name") or "").strip()
    stream_filter = (request.args.get("stream") or "").strip()

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Dropdown options
    cur.execute(
        "SELECT DISTINCT class_name "
        "FROM students "
        "WHERE class_name IS NOT NULL AND TRIM(class_name)<>'' "
        "ORDER BY class_name"
    )
    class_options = [r["class_name"] for r in cur.fetchall()]

    cur.execute(
        "SELECT DISTINCT COALESCE(stream,'') AS stream "
        "FROM students "
        "WHERE stream IS NOT NULL AND TRIM(stream)<>'' "
        "ORDER BY stream"
    )
    stream_options = [r["stream"] for r in cur.fetchall()]

    # grading scale
    cur.execute("""
        SELECT grade, lower_limit, upper_limit
          FROM grading_scale
         ORDER BY lower_limit DESC
    """)
    SCALE = cur.fetchall() or []

    def grade_for_score(score: float) -> str:
        try:
            s = float(score)
        except Exception:
            return "NG"
        for r in SCALE:
            lo, hi = float(r["lower_limit"]), float(r["upper_limit"])
            if lo <= s <= hi:
                return (r["grade"] or "").strip()
        return "NG"

    # WHERE
    where = ["rs.term = %s", "rs.year = %s", "st.archived = 0"]
    params = [term, year]
    if class_filter:
        where.append("st.class_name = %s"); params.append(class_filter)
    if stream_filter:
        where.append("COALESCE(st.stream,'') = %s"); params.append(stream_filter)
    where_sql = " AND ".join(where)

    # Pull marks (components + subject code)
    cur.execute(f"""
        SELECT
          st.id AS student_id,
          st.first_name,
          COALESCE(st.Middle_name,'') AS middle_name,
          st.last_name,
          st.class_name,
          COALESCE(st.stream,'') AS stream,
          sj.name AS subject_name,
          UPPER(COALESCE(sj.code,'')) AS subject_code,
          rs.other_mark, rs.holiday_mark, rs.bot_mark, rs.midterm_mark, rs.eot_mark, rs.ca_mark,
          rs.average_mark
        FROM record_score rs
        JOIN students st ON st.id = rs.student_id
        LEFT JOIN subjects sj ON sj.id = rs.subject_id
        WHERE {where_sql}
        ORDER BY st.class_name, st.stream, st.last_name, st.first_name
    """, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    def pick_average_row_score(r: dict):
        if r.get("average_mark") is not None:
            return float(r["average_mark"])
        return _mean_nonnull([r.get(k) for k in COMPONENT_FIELDS])

    # Crunch per student
    per_student = {}
    for r in rows or []:
        sid = r["student_id"]
        stu = per_student.setdefault(sid, {
            "student_id": sid,
            "name": f"{r['first_name']} {r['middle_name']} {r['last_name']}".replace(" ", " ").strip(),
            "class_name": r["class_name"],
            "stream": r["stream"],
            "scores": [],
            "by_subject": {}
        })
        sc = pick_average_row_score(r)
        if sc is None:
            continue
        stu["scores"].append(sc)

        ck = core_name_from(r["subject_name"], r["subject_code"])
        g = grade_for_score(sc)
        p = AGG_MAP.get((g or "").upper().replace(" ", ""), 9)
        if ck:
            stu["by_subject"][ck] = {"score": sc, "grade": g, "points": p}

    classes = defaultdict(list)
    CORE = ("eng", "math", "sci", "sst")
    for s in per_student.values():
        n = len(s["scores"])
        total = sum(s["scores"]) if n else 0.0
        avg = (total / n) if n else 0.0

        has_all_core = all(k in s["by_subject"] for k in CORE)
        if has_all_core:
            aggregate = sum(s["by_subject"][k]["points"] for k in CORE)
            division = division_from_aggregate(aggregate)
        else:
            aggregate = None
            division = "NG"

        s["total"] = total
        s["average"] = avg
        s["aggregate"] = None if (aggregate is None) else aggregate
        s["division"] = division
        classes[(s["class_name"], s["stream"])].append(s)

    # Sort within each class: graded first, then by average desc
    def _sort_key(stu):
        ng_flag = (stu["division"] == "NG")
        return (ng_flag, -stu["average"])

    for key in classes:
        classes[key].sort(key=_sort_key)

    # Class stats + division counts
    DIV_LABELS = ["Div 1", "Div 2", "Div 3", "Div 4", "U", "NG"]
    class_stats = {}
    for key, lst in classes.items():
        graded = [x for x in lst if x["division"] != "NG"]
        avgs = [x["average"] for x in graded]
        counts = {d: 0 for d in DIV_LABELS}
        for s in lst:
            counts[s["division"]] = counts.get(s["division"], 0) + 1
        class_stats[key] = {
            "mean": (sum(avgs) / len(avgs)) if avgs else 0.0,
            "size": len(lst),
            "top_avg": max(avgs) if avgs else 0.0,
            "div_counts": counts,
        }

    return render_template(
        "performance_summary.html",
        term=term, year=year,
        classes=classes,
        class_stats=class_stats,
        class_filter=class_filter,
        stream_filter=stream_filter,
        class_options=class_options,
        stream_options=stream_options,
        BEST_N=None
    )

@app.route("/midterm/overview", methods=["GET"])
@require_role("admin", "headteacher", "teacher", "bursar", "dos", "classmanager", "deputyheadteacher")
def midterm_overview():
    class_name = (request.args.get("class_name") or "").strip()
    term = (request.args.get("term") or "Term 1").strip()
    year = int(request.args.get("year") or datetime.now().year)

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # roster
    cur.execute("""
        SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS Middle_name, last_name
        FROM students
        WHERE archived=0 AND class_name=%s
        ORDER BY last_name, first_name
    """, (class_name,))
    students = cur.fetchall()

    bm = _bot_mid_by_sid(conn, class_name, term, year)
    hol = _midterms_pick(conn, class_name, term, year, HOLIDAY_NAME)
    oth = _midterms_pick(conn, class_name, term, year, OTHER_NAME)

    table = []
    for s in students:
        sid = s["id"]
        name = f"{s['first_name']} {s['Middle_name']} {s['last_name']}".replace(
            " ", " ").strip()
        row = {"student_number": s["student_number"],
               "name": name, "subjects": []}
        for code in BIG4_CODES:
            bot = bm.get(sid, {}).get(code, {}).get("bot")
            mid_stored = bm.get(sid, {}).get(code, {}).get("mid")
            hp = hol.get(sid, {}).get(code)
            other = oth.get(sid, {}).get(code)
            mid_final = mid_stored if mid_stored is not None else _mean([
                                                                        bot, hp, other])

            def pack(val):
                g = grade_for_score(conn, val)
                return dict(mark=val, grade=g, comment=comment_for_grade(conn, g))

            row["subjects"].append(dict(
                code=code,
                bot=pack(bot),
                holiday=pack(hp),
                other=pack(other),
                mid=pack(mid_final)
            ))
        table.append(row)

    cur.close()
    conn.close()
    return render_template("midterm_overview.html",
                           class_name=class_name, term=term, year=year,
                           rows=table, subj_order=BIG4_CODES)



@app.route("/next_term", methods=["GET", "POST"])
@require_role('admin', 'headteacher', 'dos', 'classmanager', "deputyheadteacher")
def next_term_hub():
    # Defaults
    terms = ["Term 1", "Term 2", "Term 3"]

    try:
        # Try your active academic year helper if you have it
        ay = get_active_academic_year() or {}
        default_year = int(ay.get("year") or ay.get("active_year") or datetime.now().year)
    except Exception:
        default_year = datetime.now().year

    sel_year = request.values.get("year", type=int) or default_year
    sel_term = request.values.get("term") or terms[0]

    conn = get_db_connection()
    ensure_term_dates_schema(conn)

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        if action == "save":
            year = request.form.get("year", type=int)
            term = (request.form.get("term") or "").strip()
            next_term = (request.form.get("next_term") or "").strip() or None
            next_term_date = (request.form.get("next_term_date") or "").strip() or None
            next_term_end_date = (request.form.get("next_term_end_date") or "").strip() or None

            if not (year and term):
                flash("Please select a valid Year and Term.", "warning")
            else:
                try:
                    cur = conn.cursor(dictionary=True)
                    cur.execute(
                        """
                        INSERT INTO term_dates (
                            year, term, next_term, next_term_date, next_term_end_date
                        )
                        VALUES (%s, %s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                            next_term = VALUES(next_term),
                            next_term_date = VALUES(next_term_date),
                            next_term_end_date = VALUES(next_term_end_date)
                        """,
                        (year, term, next_term, next_term_date, next_term_end_date),
                    )
                    conn.commit()
                    cur.close()
                    flash("Next-term info saved.", "success")
                    # Persist selection in querystring
                    return redirect(url_for("next_term_hub", year=year, term=term))
                except Exception as e:
                    conn.rollback()
                    flash(f"Could not save: {e}", "danger")

    # Load current row (if any)
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT next_term, next_term_date, next_term_end_date
        FROM term_dates
        WHERE year=%s AND term=%s
        LIMIT 1
        """,
        (sel_year, sel_term),
    )
    row = cur.fetchone()
    cur.close()
    conn.close()

    return render_template(
        "next_term_hub.html",
        terms=terms,
        sel_year=sel_year,
        sel_term=sel_term,
        existing=row,
        fmt_date=fmt_report_date, # for 01/Feb/2026 display
    )

@app.route("/process_reports", methods=["GET", "POST"])
@require_role("admin", "headteacher", "bursar", "dos", "classmanager", "deputyheadteacher")
def process_reports():
    term = (request.values.get("term") or "Term 1").strip()
    try:
        year = int(request.values.get("year") or datetime.now().year)
    except Exception:
        year = datetime.now().year
    class_name = (request.values.get("class_name")
                  or request.values.get("class_id") or "").strip()

    # If you have heavy processing, call it here:
    # with get_db_connection() as conn:
    # process_reports_snapshot(conn, class_name, term, year)

    flash("Reports processed successfully.", "success")
    return redirect(url_for("reports_hub", class_name=class_name, term=term, year=year))


# ---------------------------
# Reports Hub (Process, Print, Next-term setup)

@app.route("/broadcast_special_comm", methods=["GET", "POST"])
@require_role("admin", "headteacher", "dos","classmanager","bursar","deputyheadteacher")
def broadcast_special_comm():
    conn = get_db_connection()

    # For dropdown
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT class_name
        FROM students
        WHERE class_name IS NOT NULL AND class_name <> ''
        ORDER BY class_name
    """)
    classes = [r[0] for r in cur.fetchall()]
    cur.close()

    terms = ["Term 1", "Term 2", "Term 3"]

    applied_mid = 0
    applied_eot = 0

    if request.method == "POST":
        term = request.form.get("term")
        year = request.form.get("year", type=int) or datetime.now().year
        class_name = request.form.get("class_name") or None # None = all classes
        message = (request.form.get("message") or "").strip()
        report_type = (request.form.get("report_type") or "both").lower()
        override_existing = bool(request.form.get("override_existing"))

        if not term or not message:
            flash("Term and message are required.", "warning")
        else:
            # --- get target learners ---
            cur = conn.cursor(dictionary=True)
            if class_name:
                cur.execute("""
                    SELECT id, class_name
                    FROM students
                    WHERE archived = 0 AND class_name = %s
                """, (class_name,))
            else:
                cur.execute("""
                    SELECT id, class_name
                    FROM students
                    WHERE archived = 0 AND class_name IS NOT NULL
                """)
            learners = cur.fetchall() or []
            cur.close()

            # ---------- MIDTERM (uses midterm_overall_comments) ----------
            if report_type in ("midterm", "both"):
                ensure_midterm_overall_comments_schema(conn)
                cur = conn.cursor(dictionary=True)
                for s in learners:
                    sid = s["id"]
                    cur.execute("""
                        SELECT id, special_communication
                        FROM midterm_overall_comments
                        WHERE student_id=%s AND term=%s AND year=%s
                        LIMIT 1
                    """, (sid, term, year))
                    row = cur.fetchone()
                    if row:
                        # skip if there is already a message and we don't want to override
                        if (row.get("special_communication") or "").strip() and not override_existing:
                            continue
                        cur.execute("""
                            UPDATE midterm_overall_comments
                            SET special_communication=%s
                            WHERE id=%s
                        """, (message, row["id"]))
                    else:
                        cur.execute("""
                            INSERT INTO midterm_overall_comments
                                (student_id, class_name, term, year, special_communication)
                            VALUES (%s,%s,%s,%s,%s)
                        """, (sid, s["class_name"], term, year, message))
                    applied_mid += 1
                conn.commit()
                cur.close()

            # ---------- END OF TERM (uses report_overall_overrides) ----------
            if report_type in ("eot", "both"):
                ensure_report_overall_overrides_schema(conn)
                cur = conn.cursor(dictionary=True)
                for s in learners:
                    sid = s["id"]
                    cur.execute("""
                        SELECT id, special_communication
                        FROM report_overall_overrides
                        WHERE student_id=%s AND term=%s AND year=%s
                        LIMIT 1
                    """, (sid, term, year))
                    row = cur.fetchone()
                    # if already has text and we don't override, skip
                    if row and (row.get("special_communication") or "").strip() and not override_existing:
                        continue

                    # use your existing helper (it already upserts + commits)
                    save_overall_overrides(
                        conn,
                        sid,
                        term,
                        year,
                        special_text=message,
                    )
                    applied_eot += 1
                # save_overall_overrides commits internally, so no extra commit needed
                cur.close()

            msg_bits = []
            if applied_mid:
                msg_bits.append(f"Midterm: {applied_mid} learners updated")
            if applied_eot:
                msg_bits.append(f"EOT: {applied_eot} learners updated")
            if not msg_bits:
                msg_bits.append("No records changed (possibly all had manual messages already).")
            flash(" | ".join(msg_bits), "success")

    conn.close()
    return render_template(
        "broadcast_special_comm.html",
        classes=classes,
        terms=terms,
    )

    

@app.route("/reports", methods=["GET", "POST"])
@require_role("admin", "headteacher", "dos", "bursar", "deputyheadteacher", "classmanager", "teacher")
def reports_hub():
    conn = get_db_connection()

    # --- dropdowns: classes, years, terms ---
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT DISTINCT class_name
        FROM students
        WHERE class_name IS NOT NULL AND TRIM(class_name) <> ''
        ORDER BY class_name
    """)
    classes = [r["class_name"] for r in cur.fetchall()]
    cur.close()

    # Which marks source do you use? returns "record_score" or "results"
    src_table = detect_scores_table(conn)

    cur = conn.cursor(dictionary=True)
    if src_table == "record_score":
        cur.execute("SELECT DISTINCT year FROM record_score ORDER BY 1 DESC")
    else:
        cur.execute("SELECT DISTINCT year FROM results ORDER BY 1 DESC")
    years = [r["year"] for r in cur.fetchall()]
    cur.close()

    terms = ["Term 1", "Term 2", "Term 3"]

    # --- selected filters ---
    sel_class = (request.values.get("class_name") or (
        classes[0] if classes else "")).strip()
    sel_term = (request.values.get("term") or (
        terms[0] if terms else "Term 1")).strip()
    try:
        sel_year = int(request.values.get("year") or (
            years[0] if years else datetime.now().year))
    except Exception:
        sel_year = datetime.now().year

    action = (request.form.get("action") or "").strip()

    # 1) Rebuild snapshot for this class/term/year
    if request.method == "POST" and action == "process":
        try:
            process_reports_snapshot(conn, sel_class, sel_term, sel_year)
            flash(
                f"Reports processed for {sel_class} — {sel_term} {sel_year}.",
                "success")
        except Exception as e:
            conn.rollback()
            flash(f"Processing failed: {e}", "danger")
        finally:
            conn.close()
        return redirect(url_for("reports_hub",
                                class_name=sel_class,
                                term=sel_term,
                                year=sel_year))

    # 2) Save next-term details
    
    if request.method == "POST" and action == "save_term_dates":
        ensure_term_dates_schema(conn)
        next_term = (request.form.get("next_term") or "").strip() or None
        next_term_date = (request.form.get("next_term_date") or "").strip() or None
        next_term_end_date = (request.form.get("next_term_end_date") or "").strip() or None

        try:
            cur = conn.cursor(dictionary=True)
            cur.execute(
                """
                INSERT INTO term_dates(year, term, next_term, next_term_date, next_term_end_date)
                VALUES (%s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    next_term = VALUES(next_term),
                    next_term_date = VALUES(next_term_date),
                    next_term_end_date = VALUES(next_term_end_date)
                """,
                (sel_year, sel_term, next_term, next_term_date, next_term_end_date),
            )
            conn.commit()
            cur.close()
            flash("Next-term details saved.", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Could not save next-term details: {e}", "danger")
        finally:
            conn.close()

        return redirect(
            url_for(
                "reports_hub",
                class_name=sel_class,
                term=sel_term,
                year=sel_year,
            )
        )


        # --- page data ---
    ensure_term_dates_schema(conn)

    KG_CLASSES = ("Baby", "Middle", "Top")
    cur = conn.cursor(dictionary=True)

    if sel_class in KG_CLASSES:
        # Baby–Top: show ALL active learners, even if they have no marks / reports yet
        cur.execute("""
            SELECT
                s.id AS student_id,
                s.student_number,
                s.first_name,
                s.last_name,
                s.class_name,
                %s AS term,
                %s AS year
            FROM students s
            WHERE s.archived = 0
              AND s.class_name = %s
            ORDER BY s.last_name, s.first_name
        """, (sel_term, sel_year, sel_class))
    else:
        # Other classes: keep existing logic (only those with processed reports)
        cur.execute("""
            SELECT DISTINCT
                s.id AS student_id,
                s.student_number,
                s.first_name,
                s.last_name,
                s.class_name,
                r.term,
                r.year
            FROM reports r
            JOIN students s ON s.id = r.student_id
            WHERE r.term = %s
              AND r.year = %s
              AND s.class_name = %s
            ORDER BY s.last_name, s.first_name
        """, (sel_term, sel_year, sel_class))

    students_ready = cur.fetchall() or []
    cur.close()


    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT next_term, next_term_date, next_term_end_date
        FROM term_dates
        WHERE year=%s AND term=%s
        LIMIT 1
    """, (sel_year, sel_term))
    existing_term = cur.fetchone()
    cur.close()
    conn.close()

    return render_template(
        "reports_hub.html",
        classes=classes, terms=terms, years=years,
        sel_class=sel_class, sel_term=sel_term, sel_year=sel_year,
        students_ready=students_ready,
        existing_term=existing_term
    )



    
@app.route("/report_card/<int:student_id>/<term>/<int:year>")
@require_role("admin", "teacher", "headteacher", "bursar", "dos", "deputyheadteacher", "classmanager")
def report_card(student_id, term, year):
    include_mid = (request.args.get("include_mid") == "1")

    conn = get_db_connection()
    ensure_term_dates_schema(conn)
    ensure_comment_rules_schema(conn)

    # --- student ---
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM students WHERE id=%s", (student_id,))
    student = cur.fetchone()
    cur.close()
    if not student:
        conn.close()
        abort(404)

    # --- ensure snapshot exists ---
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT 1 FROM reports WHERE student_id=%s AND term=%s AND year=%s "
        "LIMIT 1",
        (student_id, term, year)
    )
    if not cur.fetchone():
        process_reports_snapshot(conn, student["class_name"], term, year)
    cur.close()

    # --- subject results (EOT table) ---
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT sub.name AS subject, sub.code AS subject_code,
               r.eot_mark AS eot,
               COALESCE(r.average_mark, r.eot_mark) AS total_100,
               r.grade, r.comment, r.teacher_initial AS initials,
               r.teacher_remark, r.headteacher_remark
        FROM reports r
        JOIN subjects sub ON sub.id = r.subject_id
        WHERE r.student_id=%s AND r.term=%s AND r.year=%s
        ORDER BY sub.name
    """, (student_id, term, year))
    rows = cur.fetchall()
    cur.close()

    # --- total and average ---
    marks = [r["total_100"] for r in rows if r["total_100"] is not None]
    total_sum = round(sum(marks), 2) if marks else None
    avg_overall = round(sum(marks) / len(marks), 2) if marks else None

    # --- aggregate/division (STRICT cores) ---
    CORE_CODES = {"ENG", "MATH", "SCI", "SST"}
    AGG_MAP = {"D1": 1, "D2": 2, "C3": 3, "C4": 4,
               "C5": 5, "C6": 6, "P7": 7, "P8": 8, "F9": 9}

    def _is_core_row(r) -> bool:
        code = (r.get("subject_code") or "").strip().upper()
        name = (r.get("subject") or "").strip().lower()
        if code in CORE_CODES:
            return True
        if name.startswith("eng"):
            return True
        if name.startswith(("mat", "math")):
            return True
        if name.startswith("sci"):
            return True
        if name in {
            "sst", "soc. studies", "social studies",
            "social std", "socialstudies"
        }:
            return True
        return False

    core_grades = [r["grade"] for r in rows if _is_core_row(r) and r.get("grade")]
    if len(core_grades) == 4:
        aggregate = sum(AGG_MAP[g] for g in core_grades if g in AGG_MAP)
        if 4 <= aggregate <= 12:
            division = "1"
        elif 13 <= aggregate <= 23:
            division = "2"
        elif 24 <= aggregate <= 29:
            division = "3"
        elif 30 <= aggregate <= 34:
            division = "4"
        else:
            division = "U"
    else:
        aggregate, division = (None, None)

    # --- reorder rows: cores first, then others alphabetically ---
    core_rows = [r for r in rows if _is_core_row(r)]
    other_rows = [r for r in rows if not _is_core_row(r)]
    other_rows.sort(key=lambda r: (r.get("subject") or ""))
    rows = core_rows + other_rows

    # ------- MIDTERM PANEL (optional) -------------
    midterms = []
    midterm_subjects = []
    if include_mid:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT UPPER(code) AS code, id
            FROM subjects
            WHERE code IS NOT NULL AND TRIM(code) <> ''
        """)
        code_rows = cur.fetchall() or []
        cur.close()
        code_to_id = {r["code"]: r["id"] for r in code_rows}
        all_codes = list(code_to_id.keys())
        others = sorted([c for c in all_codes if c not in CORE_CODES])
        midterm_subjects = [c for c in ("ENG", "MATH", "SCI", "SST")
                            if c in all_codes] + others

        def fetch_marks(colname: str, *, round0: bool = False) -> dict:
            out = {}
            for sc in midterm_subjects:
                sid = code_to_id.get(sc)
                if not sid:
                    out[sc] = None
                    continue
                cur2 = conn.cursor(dictionary=True)
                cur2.execute(
                    f"""SELECT MAX({colname}) AS v
                        FROM record_score
                        WHERE student_id=%s AND subject_id=%s
                          AND term=%s AND year=%s""",
                    (student_id, sid, term, year)
                )
                row2 = cur2.fetchone()
                cur2.close()
                v = row2["v"] if row2 and row2["v"] is not None else None
                if v is not None:
                    try:
                        v = int(round(float(v))) if round0 else int(v)
                    except Exception:
                        try:
                            v = int(float(v))
                        except Exception:
                            v = None
                out[sc] = v
            return out

        def _grade_of(v):
            return grade_for_score(conn, v) if v is not None else None

        panels = [
            ("OTH", fetch_marks("other_mark", round0=True)),
            ("HP", fetch_marks("holiday_mark", round0=True)),
            ("BOT", fetch_marks("bot_mark")),
            ("MID", fetch_marks("midterm_mark")),
            ("EOT", fetch_marks("eot_mark")),
        ]
        for label, score_map in panels:
            if any(v is not None for v in score_map.values()):
                total_all = sum(int(v) for v in score_map.values()
                                if v is not None)
                midterms.append({
                    "assessment": label,
                    "scores": score_map,
                    "grades": {sc: _grade_of(v) for sc, v in score_map.items()},
                    "total": total_all
                })

    # --- latest payment no. ---
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id AS payment_number
        FROM fees
        WHERE student_id=%s AND term=%s AND year=%s
        ORDER BY date_paid DESC, id DESC
        LIMIT 1
    """, (student_id, term, year))
    pay = cur.fetchone()
    cur.close()
    payment_number = pay["payment_number"] if pay else None

    # --- overall comments (with overrides + special comm) ---
    overrides = fetch_overall_overrides(conn, student["id"], term, year)
    from collections import Counter

    auto_head = (
        pick_comment_template(
            conn,
            role="headteacher", scope="overall",
            division=(int(division) if division and division.isdigit() else None),
            average=avg_overall, class_name=student["class_name"], term=term, student_id=student_id
        )
        or (comment_for_grade(conn, grade_for_score(conn, avg_overall)) or "")
    )
    head_comment = (overrides.get("head_overall_custom") or "").strip() or auto_head

    auto_teacher = pick_comment_template(
        conn,
        role="teacher", scope="overall",
        division=(int(division) if division and division.isdigit() else None),
        average=avg_overall, class_name=student["class_name"], term=term, student_id=student_id
    )
    if not auto_teacher:
        per_subj = [r["teacher_remark"] for r in rows if r.get("teacher_remark")]
        auto_teacher = (Counter(per_subj).most_common(1)[0][0]
                        if per_subj else None)
    if not auto_teacher:
        auto_teacher = (comment_for_grade(conn, grade_for_score(conn, avg_overall))
                        or "")

    teacher_comment = (overrides.get("teacher_overall_custom") or "").strip() or auto_teacher
    special_communication = (overrides.get("special_communication") or "").strip()

    # --- next-term info (same logic as before) ---
    def _next_term_name_and_year(cur_term: str, cur_year: int):
        order = ["Term 1", "Term 2", "Term 3"]
        try:
            i = order.index(cur_term)
        except ValueError:
            return "Term 2", cur_year
        return (order[i + 1], cur_year) if i < 2 else (order[0], cur_year + 1)

    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT next_term, next_term_date
        FROM term_dates
        WHERE year=%s AND term=%s
        LIMIT 1
    """, (year, term))
    row = cur.fetchone()
    cur.close()

    if row and (row["next_term"] or row["next_term_date"]):
        next_term_info = dict(next_term=row["next_term"],
                              next_term_date=row["next_term_date"])
    else:
        nt_name, nt_year = _next_term_name_and_year(term, year)
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT next_term, next_term_date
            FROM term_dates
            WHERE year=%s AND term=%s
            LIMIT 1
        """, (nt_year, nt_name))
        fb = cur.fetchone()
        cur.close()
        next_term_info = (dict(next_term=nt_name, next_term_date=fb["next_term_date"])
                          if fb and fb["next_term_date"] else None)

    grading = fetch_grading_scale(conn)
    
    # Comment group
    teacher_lib, head_lib = load_comment_library_groups(conn)
    # --- navigation within same class (Prev/Next) ---
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id
        FROM students
        WHERE class_name=%s AND archived=0
        ORDER BY last_name, first_name
    """, (student["class_name"],))
    id_rows = cur.fetchall() or []
    cur.close()
    conn.close()

    all_ids = [r["id"] for r in id_rows]
    prev_id = next_id = None
    if student_id in all_ids:
        idx = all_ids.index(student_id)
        if idx > 0:
            prev_id = all_ids[idx - 1]
        if idx < len(all_ids) - 1:
            next_id = all_ids[idx + 1]

    return render_template(
        "report_card_citizen.html",
        school=dict(
            name="DEMO DAY & BOARDING",
            tagline="PRIMARY SCHOOL – KAMPALA",
            motto="Code the future",
            phones="+256778878411, +256759685640, +256773589232, +256750347624",
            pobox="P.O Box 1X1X1 Kampala",        
        ),
        student=student,
        term=term, year=year,
        rows=rows,
        total_sum=total_sum,
        avg_overall=avg_overall,
        aggregate=aggregate,
        division=division,
        midterms=midterms,
        midterm_subjects=midterm_subjects,
        payment_number=payment_number,
        comments={"teacher_comment": teacher_comment, "head_comment": head_comment},
        special_communication=special_communication,
        next_term_info=type("NTI", (), next_term_info) if next_term_info else None,
        grading=grading,
        include_mid=include_mid,
        prev_id=prev_id,
        next_id=next_id,
        teacher_library=teacher_lib,
        head_library=head_lib,
    )

@app.route("/reports/overall_comment/<int:student_id>/<term>/<int:year>",
           methods=["POST"])
@require_role("admin", "headteacher", "teacher", "dos")
def save_overall_comment(student_id, term, year):
    conn = get_db_connection()
    try:
        teacher_txt = (request.form.get("teacher_overall_custom") or "").strip() or None
        head_txt = (request.form.get("head_overall_custom") or "").strip() or None
        special_txt = (request.form.get("special_communication") or "").strip() or None

        save_overall_overrides(
            conn,
            student_id=student_id,
            term=term,
            year=year,
            teacher_text=teacher_txt,
            head_text=head_txt,
            special_text=special_txt
        )
        flash("Overall comments updated.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Could not save comments: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("report_card",
                            student_id=student_id,
                            term=term,
                            year=year))

# ---------------------------
# Batch printing
# ---------------------------

# ---------- batch print ----------


@app.route("/report_card/print_batch", methods=["POST"])
@require_role('admin', 'headteacher', 'bursar', 'dos', 'deputyheadteacher', 'classmanager', 'teacher')
def report_card_print_batch():
    class_name = (request.form.get("class_name") or "").strip()
    term = (request.form.get("term") or "").strip()
    year = request.form.get("year", type=int)
    include_mid = (request.form.get("include_midterms") == "1")

    try:
        ids = [int(x) for x in request.form.getlist("selected_ids")
               if str(x).strip().isdigit()]
    except Exception:
        ids = []

    if not ids:
        flash("Select at least one student to print.", "warning")
        return redirect(url_for("reports_hub",
                                class_name=class_name,
                                term=term, year=year))

    conn = get_db_connection()
    reports = []
    for sid in ids:
        payload = build_report_payload(conn, sid, term, year,
                                       include_mid=include_mid)
        if payload:
            reports.append(payload)
    conn.close()

    if not reports:
        flash("No printable reports found for the selected students.", "warning")
        return redirect(url_for("reports_hub",
                                class_name=class_name,
                                term=term, year=year))

    return render_template("report_batch_citizen.html", reports=reports)
    


@app.route("/report_card/print_batch_pdf", methods=["POST"])
@require_role('admin', 'headteacher', 'bursar', 'dos', 'deputyheadteacher', 'classmanager', 'teacher')
def report_card_print_batch_pdf():
    class_name = (request.form.get("class_name") or "").strip()
    term = (request.form.get("term") or "").strip()
    year = request.form.get("year", type=int) or datetime.now().year
    include_mid = (request.form.get("include_midterms") == "1")

    try:
        ids = [int(x) for x in request.form.getlist("selected_ids")
               if str(x).strip().isdigit()]
    except Exception:
        ids = []

    if not ids:
        flash("Select at least one student to print (End-of-Term PDF).", "warning")
        return redirect(url_for("reports_hub",
                                class_name=class_name,
                                term=term, year=year))

    conn = get_db_connection()
    reports = []
    for sid in ids:
        p = build_report_payload(conn, sid, term, year, include_mid=include_mid)
        if p:
            # ensure school details
            p.setdefault("school", dict(
                name=current_app.config.get("SCHOOL_NAME", "DEMO DAY & BOARDING"),
                tagline=current_app.config.get("SCHOOL_TAGLINE", "PRIMARY SCHOOL – KAMPALA"),
                motto=current_app.config.get("SCHOOL_MOTTO", "Code the future"),
                phones=current_app.config.get("SCHOOL_PHONES", "+256778878411, +256759685640, +256773589232, +256750347624"),
                pobox=current_app.config.get("SCHOOL_POBOX", "P.O Box 1X1X1 Kampala"),              
            ))
            p["next_term_date"] = p.get("next_term_date") or ""
            reports.append(p)
    conn.close()

    if not reports:
        flash("No printable reports found for the selected students.", "warning")
        return redirect(url_for("reports_hub",
                                class_name=class_name,
                                term=term, year=year))

    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)

    PAGE_W, PAGE_H = A4
    M = 10 * mm
    CONTENT_W = PAGE_W - 2 * M

    PAR_FONT_SIZE = 10.5
    LINE_SPACING_MULT = 1.5
    PAR_LEADING = PAR_FONT_SIZE * LINE_SPACING_MULT
    PAR_SPACING_AFTER = 6
    BOTTOM_SPACER = 10

    GAP_BEFORE_SUMMARY = PAR_LEADING
    GAP_AFTER_MIDTERM = PAR_LEADING

    MIN_BOTTOM_KEPT = 55 * mm
    SIGN_BLOCK_CLEAR = max(22 * mm, PAR_LEADING * 1.7)

    cell_style = ParagraphStyle(
        name="cell", fontName="Helvetica", fontSize=9,
        leading=12, alignment=TA_LEFT
    )
    hdr_style = ParagraphStyle(
        name="hdr", fontName="Helvetica-Bold", fontSize=10,
        leading=13, alignment=TA_LEFT
    )

    # ------------------------------------------------------------------
    # HEADER (banner + learner info + photo + title)
    # ------------------------------------------------------------------

    def draw_header_and_info(rep):
        student = rep["student"]
        term = rep["term"]
        year = rep["year"]
        school = rep["school"]

        width, height = PAGE_W, PAGE_H
        left = M
        right_margin = M

        # We keep these for positioning only
        banner_h = 30 * mm
        banner_y = height - 35 * mm
        left_margin, right_margin_local = left, left
        strip_w = width - left_margin - right_margin_local
        navy_w = strip_w * 0.62

        # --------- NO COLOURED SHAPES ----------
        # (Nothing drawn here – plain white background)

        # --------- Logo ----------
        SCHOOL_LOGO_PATH = os.path.join(current_app.static_folder, "logo.jpg")
        logo_box = 26 * mm
        logo_x = left_margin + 8 * mm
        logo_y = banner_y + (banner_h - logo_box) / 2

        if os.path.exists(SCHOOL_LOGO_PATH):
            try:
                c.drawImage(
                    SCHOOL_LOGO_PATH,
                    logo_x, logo_y,
                    width=logo_box, height=logo_box,
                    preserveAspectRatio=True, mask="auto",
                )
            except Exception:
                pass

        # --------- Centered school text from `school` dict ----------
        SCHOOL_NAME = school.get("name") or ""
        SCHOOL_TAGLINE = school.get("tagline") or ""
        SCHOOL_POBOX = school.get("pobox") or ""

        # We center across the whole page
        centre_x = width / 2.0

        # Fit school name
        max_text_width = width - 2 * left_margin
        name_fs = 18
        while name_fs >= 10 and c.stringWidth(
            SCHOOL_NAME, "Helvetica-Bold", name_fs
        ) > max_text_width:
            name_fs -= 1

        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", name_fs)
        name_y = banner_y + banner_h - 4 * mm
        if SCHOOL_NAME:
            c.drawCentredString(centre_x, name_y, SCHOOL_NAME)

        # Tagline just below name (bold, centered)
        if SCHOOL_TAGLINE:
            sub_fs = 13
            while sub_fs >= 8 and c.stringWidth(
                SCHOOL_TAGLINE, "Helvetica-Bold", sub_fs
            ) > max_text_width:
                sub_fs -= 1
            c.setFont("Helvetica-Bold", sub_fs)
            tagline_y = name_y - 6 * mm
            c.drawCentredString(centre_x, tagline_y, SCHOOL_TAGLINE)
        else:
            tagline_y = name_y

        # P.O Box centered just below tagline (bold)
        if SCHOOL_POBOX:
            pobox_fs = 11
            while pobox_fs >= 8 and c.stringWidth(
                SCHOOL_POBOX, "Helvetica-Bold", pobox_fs
            ) > max_text_width:
                pobox_fs -= 1
            c.setFont("Helvetica-Bold", pobox_fs)
            pobox_y = tagline_y - 6 * mm
            c.drawCentredString(centre_x, pobox_y, SCHOOL_POBOX)

        # --------- Contacts block (RIGHT SIDE) ----------
        phones_raw = (school.get("phones") or "").strip()
        SCHOOL_PHONE_LINES = [p.strip() for p in phones_raw.split(",") if p.strip()]
        SCHOOL_EMAIL = current_app.config.get("SCHOOL_EMAIL", "")

        c.setFillColor(colors.black)
        c.setFont("Helvetica", 9)

        right_pad = 6 * mm
        text_right = left_margin + strip_w - right_pad
        line_gap = 5.5 * mm
        y_cursor = banner_y + banner_h - 8 * mm

        for ph in SCHOOL_PHONE_LINES:
            c.drawRightString(text_right, y_cursor, ph)
            y_cursor -= line_gap

        if SCHOOL_EMAIL:
            y_cursor -= 2.5 * mm
            c.drawRightString(text_right, y_cursor, SCHOOL_EMAIL)

        # --- learner info block ---
        info_top = banner_y - 6 * mm
        info_left = left
        info_width = width - left - right_margin - (40 * mm)

        styles = getSampleStyleSheet()
        lab = ParagraphStyle(
            "lab", parent=styles["Normal"],
            fontName="Helvetica-Bold", fontSize=9, leading=11,
            textColor=colors.black,
        )
        val = ParagraphStyle(
            "val", parent=styles["Normal"],
            fontName="Helvetica", fontSize=9, leading=11,
        )

        full_name = f"{student.get('first_name','')} " \
                    f"{student.get('Middle_name') or student.get('middle_name','')} " \
                    f"{student.get('last_name','')}".strip()

        info_rows = [
            [Paragraph("Learner's Name:", lab),
             Paragraph(full_name or "-", val)],
            [Paragraph("Student No.:", lab),
             Paragraph(student.get("student_number") or "-", val)],
            [Paragraph("Class / Stream:", lab),
             Paragraph(f"{student.get('class_name','-')} "
                       f"{student.get('stream') or ''}", val)],
            [Paragraph("Term / Year:", lab),
             Paragraph(f"{term} / {year}", val)],
        ]
        info_tbl = Table(
            info_rows,
            colWidths=[35 * mm, info_width - 35 * mm],
            hAlign="LEFT",
        )
        info_tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ]))
        w_info, h_info = info_tbl.wrapOn(c, info_width, 9999)
        info_tbl.drawOn(c, info_left, info_top - h_info)


        # ---- PHOTO BLOCK (works with DB blob or file path) ----
        photo_blob = student.get("photo_blob")
        photo_path = student.get("photo")

        # Your existing box position
        box_w = box_h = 32 * mm
        photo_x = width - right_margin - box_w
        photo_y = info_top - (h_info - box_h) / 2

        if photo_blob:
            try:
                img_reader = ImageReader(io.BytesIO(photo_blob))
                c.drawImage(
                    img_reader,
                    photo_x + 2, photo_y - box_h + 2,
                    box_w - 4, box_h - 4,
                    preserveAspectRatio=True,
                    mask="auto"
                )
            except:
                pass

        elif photo_path:
            full_path = os.path.join(app.root_path, photo_path)
            if os.path.exists(full_path):
                try:
                    img_reader = ImageReader(full_path)
                    c.drawImage(
                        img_reader,
                        photo_x + 2, photo_y - box_h + 2,
                        box_w - 4, box_h - 4,
                        preserveAspectRatio=True,
                        mask="auto"
                    )
                except:
                    pass

        else:
            # placeholder
            c.setStrokeColor(colors.grey)
            c.rect(photo_x, photo_y - box_h, box_w, box_h)
            c.setFont("Helvetica", 7)
            c.drawCentredString(photo_x + box_w/2, photo_y - box_h/2, "Photo")


        # --- title ---
        title_y = (info_top - h_info) - 4 * mm
        c.setFont("Helvetica-Bold", 13)
        c.setFillColor(colors.black)
        c.drawCentredString(width / 2.0, title_y, "End of Term Report")

        table_top = title_y - 6 * mm
        return table_top

    # ------------------------------------------------------------------
    # TABLE DRAWING HELPER
    # ------------------------------------------------------------------
    def draw_table_autofit(tbl, x, y_top, content_w):
        BOTTOM_MARGIN = 36
        if y_top is None:
            try:
                _w, page_h = c._pagesize
                y_top = page_h - 150
            except Exception:
                y_top = 750

        w, h = tbl.wrapOn(c, content_w, PAGE_H)
        if not h:
            h = 0

        available = y_top - BOTTOM_MARGIN
        if h == 0 or h <= available:
            tbl.drawOn(c, x, y_top - h)
            return y_top - h

        scale = max(0.6, min(1.0, available / float(h)))
        c.saveState()
        c.translate(x, y_top - h * scale)
        c.scale(scale, scale)
        tbl.drawOn(c, 0, 0)
        c.restoreState()
        return y_top - h * scale

    # ------------------------------------------------------------------
    # EOT TABLE BUILDER (CORE SUBJECTS FIRST)
    # ------------------------------------------------------------------

    EOT_W, TOT_W, GRD_W, INIT_W = 16 * mm, 26 * mm, 18 * mm, 18 * mm
    CORE_CODES = {"ENG", "MATH", "SCI", "SST"}


    def _is_core_row_rep(r) -> bool:
        code = (r.get("subject_code") or r.get("code") or "").strip().upper()
        name = (r.get("subject") or r.get("name") or "").strip().lower()
        if code in CORE_CODES:
            return True
        if name.startswith("eng"):
            return True
        if name.startswith(("mat", "math")):
            return True
        if name.startswith("sci"):
            return True
        if name in {
            "sst", "soc. studies", "social studies",
            "social std", "socialstudies"
        }:
            return True
        return False


    def make_eot_table(rep):
        def fmt0(v):
            if v is None:
                return ""
            s = str(v).strip()
            if s == "":
                return ""
            try:
                return f"{int(round(float(s)))}"
            except Exception:
                return ""

        comment_w = 60 * mm
        fixed_total = EOT_W + TOT_W + GRD_W + INIT_W + comment_w
        subject_w = CONTENT_W - fixed_total
        if subject_w < 55 * mm:
            deficit = (55 * mm) - subject_w
            comment_w = max(40 * mm, comment_w - deficit)
            subject_w = CONTENT_W - (EOT_W + TOT_W + GRD_W + INIT_W + comment_w)

        data = [["SUBJECT", "EOT", "AVERAGE(%)", "GRADE", "COMMENT", "INITIALS"]]

        # ---- reorder: cores first, then others by name ----
        all_rows = list(rep["rows"] or [])
        cores = [r for r in all_rows if _is_core_row_rep(r)]
        others = [r for r in all_rows if not _is_core_row_rep(r)]
        others.sort(key=lambda r: (r.get("subject") or r.get("name") or ""))

        for r in cores + others:
            subj = (r.get("subject") or r.get("name") or "")
            code = (r.get("subject_code") or r.get("code") or "").upper()

            # format values
            eot = fmt0(r.get("eot"))
            tot = fmt0(r.get("total_100"))
            grade = (r.get("grade") or "").strip()
            comment_txt = (r.get("comment") or "").strip()
            initials_txt = (r.get("initials") or "").strip()

            # ---- SKIP completely empty subjects ----
            if not any([eot, tot, grade, comment_txt, initials_txt]):
                continue

            subj_cell = Paragraph(
                f"{subj}" + (
                    f" <font size='8' color='#666666'>({code})</font>"
                    if code else ""
                ),
                cell_style
            )

            initials_cell = Paragraph(initials_txt, cell_style) if initials_txt else ""

            data.append([
                subj_cell,
                eot,
                tot,
                grade,
                Paragraph(comment_txt, cell_style),
                initials_cell,
            ])

        total_str = fmt0(rep.get("total_sum"))
        data.append([
            Paragraph("<b>TOTAL</b>", hdr_style),
            "",
            total_str,
            "" if rep.get("aggregate") is None else str(rep["aggregate"]),
            "",
            ""
        ])

        widths = [subject_w, EOT_W, TOT_W, GRD_W, comment_w, INIT_W]
        t = Table(data, colWidths=widths, repeatRows=1)
        style = TableStyle([
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 10),
            ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
            ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
            ("ALIGN", (1, 0), (3, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWHEIGHT", (0, 0), (-1, 0), 18),
        ])
        for r_idx in range(1, len(data)):
            style.add("ROWHEIGHT", (0, r_idx), (-1, r_idx), 16)

        last_col = len(widths) - 1
        style.add("ALIGN", (last_col, 1), (last_col, len(data) - 1), "CENTER")
        style.add("VALIGN", (last_col, 1), (last_col, len(data) - 1), "MIDDLE")
        style.add("TOPPADDING", (last_col, 1), (last_col, len(data) - 1), 2)
        style.add("BOTTOMPADDING", (last_col, 1), (last_col, len(data) - 1), 2)

        t.setStyle(style)
        return t


    # ------------------------------------------------------------------
    # SUMMARY LINE (NO POSITION)
    # ------------------------------------------------------------------
    def draw_summary(rep, y):
        font = "Helvetica-Bold"
        fs = PAR_FONT_SIZE
        c.setFont(font, fs)

        avg = rep.get("avg_overall")
        left = f"Average: {'' if avg is None else f'{avg:.2f}'}"
        c.drawString(M, y, left)

        agg = rep.get("aggregate")
        div = rep.get("division") or ""
        mid = f"Aggregate: {'' if agg is None else agg} Division: {div}"
        mid_w = c.stringWidth(mid, font, fs)
        c.drawString((PAGE_W - mid_w) / 2.0, y, mid)

        # No class position text here
        return y - PAR_LEADING * 1.1

    # ------------------------------------------------------------------
    # MIDTERM PANEL (unchanged – uses same midterm table builder as before)
    # ------------------------------------------------------------------
    def filtered_midterm_codes(rep):
        mts = rep.get("midterms") or []
        codes = list(rep.get("midterm_subjects") or [])
        if not mts or not codes:
            return []
        keep = []
        for code in codes:
            if any((m.get("scores") or {}).get(code) is not None for m in mts):
                keep.append(code)
        return keep

    def make_midterm_table(rep):
        mts = rep.get("midterms") or []
        codes = filtered_midterm_codes(rep)
        if not mts or not codes:
            return None
        any_total = any(bool(m.get("total")) for m in mts)

        hdr = ["Assessment"]
        for code in codes:
            hdr.append(code)
            hdr.append("GR")
        if any_total:
            hdr.append("TOTAL")

        def _fmt(n):
            if n is None or n == "":
                return ""
            try:
                return str(int(round(float(n))))
            except Exception:
                return str(n)

        data = [hdr]
        for m in mts:
            has_any = any((m.get("scores") or {}).get(code) is not None
                          for code in codes) or bool(m.get("total"))
            if not has_any:
                continue
            row = [m.get("assessment")]
            for code in codes:
                sc = (m.get("scores") or {}).get(code)
                gr = (m.get("grades") or {}).get(code) or ""
                row.append(_fmt(sc))
                row.append(gr)
            if any_total:
                row.append(_fmt(m.get("total")))
            data.append(row)

        if len(data) == 1:
            return None


        n_pairs = len(codes)

        # reduce widths to allow more subjects to fit
        ASSESS_W = 20 * mm # was 34mm
        GR_W = 9 * mm # was 11mm
        MIN_SCORE_W = 12 * mm # was 16mm
        TOT_W = 16 * mm if any_total else 0

        # compute remaining width for score columns
        fixed_total = ASSESS_W + (n_pairs * GR_W) + TOT_W
        remaining_for_scores = CONTENT_W - fixed_total

        # distribute remaining space
        SCORE_W = max(MIN_SCORE_W, remaining_for_scores / max(1, n_pairs))


        col_w = [ASSESS_W]
        for _ in codes:
            col_w += [SCORE_W, GR_W]
        if any_total:
            col_w += [TOT_W]

        t = Table(data, colWidths=col_w, repeatRows=1)
        many = n_pairs >= 6
        hdr_font = 9 if not many else 8.5
        body_font = 8.5 if not many else 8.2
        hdr_h = 16 if not many else 15
        row_h = 14 if not many else 13

        style = TableStyle([
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", hdr_font),
            ("FONT", (0, 1), (-1, -1), "Helvetica", body_font),
            ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWHEIGHT", (0, 0), (-1, 0), hdr_h),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ])
        for i in range(1, len(data)):
            style.add("ROWHEIGHT", (0, i), (-1, i), row_h)
        if any_total:
            last_col = len(col_w) - 1
            style.add("BACKGROUND", (last_col, 0), (last_col, 0), colors.whitesmoke)
            style.add("FONT", (last_col, 0), (last_col, 0), "Helvetica-Bold", hdr_font)
        t.setStyle(style)
        return t

    def draw_midterm_panel(rep, y):
        t = make_midterm_table(rep)
        if not t:
            return y
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(PAGE_W / 2, y, "MID TERM EXAMS")
        y -= 6
        y = draw_table_autofit(t, M, y, CONTENT_W)
        return y - GAP_AFTER_MIDTERM

    # ------------------------------------------------------------------
    # COMMENTS + SIGNATURES + LEGEND + FOOTER (unchanged)
    # ------------------------------------------------------------------

    def draw_comments_signatures_legend_footer(rep, y):
        ps = ParagraphStyle(
            name="p", fontName="Helvetica", fontSize=PAR_FONT_SIZE,
            leading=PAR_LEADING, spaceAfter=PAR_SPACING_AFTER
        )

        label_ps = ParagraphStyle(
            name="label", parent=ps, fontName="Helvetica-Bold",
            fontSize=PAR_FONT_SIZE, leading=PAR_LEADING
        )
        text_ps = ParagraphStyle(
            name="text", parent=ps, fontName="Helvetica",
            fontSize=PAR_FONT_SIZE, leading=PAR_LEADING
        )
        text_italic_ps = ParagraphStyle(
            name="text_i", parent=text_ps, fontName="Helvetica-Oblique"
        )
        head_ps = ParagraphStyle(
            name="head_row", parent=ps, fontName="Helvetica-Bold",
            fontSize=PAR_FONT_SIZE, alignment=TA_LEFT
        )

        def ensure_space(min_h):
            nonlocal y
            if y - min_h < (M + BOTTOM_SPACER):
                c.showPage()
                y = draw_header_and_info(rep)

        # ---------- build comments + next-term table ----------
        comments = rep.get("comments") or {}
        teacher_comment = (comments.get("teacher_comment") or "").strip()
        head_comment = (comments.get("head_comment") or "").strip()
        special = (rep.get("special_communication") or "").strip()

        # next term info (same logic as before)
        nti = rep.get("next_term_info")
        def _field(obj, name):
            if obj is None:
                return None
            v = getattr(obj, name, None) if not isinstance(obj, dict) else obj.get(name)
            if v:
                return str(v).strip()
            return None

        nt_name = _field(nti, "next_term")
        nt_date = _field(nti, "next_term_date")
        nt_end = _field(nti, "next_term_end_date")

        next_term_line = ""
        if nt_name or nt_date or nt_end:
            line = f"Term {nt_name or ''}".strip()
            if nt_date:
                line += f" begins on: {nt_date}"
            if nt_end:
                if nt_date:
                    line += f" and will end on: {nt_end}"
                else:
                    line += f" will end on: {nt_end}"
            next_term_line = line.strip()

        rows = []
        if teacher_comment or head_comment or special or next_term_line:
            # header row
            rows.append([
                Paragraph("Comments, Communication & Next Term Information", head_ps),
                ""
            ])

            if teacher_comment:
                rows.append([
                    Paragraph("Class Teacher’s Remarks:", label_ps),
                    Paragraph(teacher_comment, text_italic_ps),
                ])
            if head_comment:
                rows.append([
                    Paragraph("Head Teacher’s Remarks:", label_ps),
                    Paragraph(head_comment, text_italic_ps),
                ])
            if special:
                rows.append([
                    Paragraph("Special Communication:", label_ps),
                    Paragraph(special, text_ps),
                ])
            if next_term_line:
                rows.append([
                    Paragraph("Next Term:", label_ps),
                    Paragraph(next_term_line, text_ps),
                ])

        if rows:
            col_w = [CONTENT_W * 0.28, CONTENT_W * 0.72]
            tbl = Table(rows, colWidths=col_w)
            ts = TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.75, colors.whitesmoke),
                ("GRID", (0, 1), (-1, -1), 0.25, colors.whitesmoke),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ])
            # header styling
            ts.add("SPAN", (0, 0), (-1, 0))
            ts.add("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke)
            ts.add("ALIGN", (0, 0), (-1, 0), "LEFT")
            tbl.setStyle(ts)

            w_tbl, h_tbl = tbl.wrapOn(c, CONTENT_W, PAGE_H)
            ensure_space(h_tbl)
            tbl.drawOn(c, M, y - h_tbl)
            y -= (h_tbl + PAR_SPACING_AFTER)

        # ---------- signatures (unchanged) ----------
        ensure_space(SIGN_BLOCK_CLEAR + 14)
        y -= (SIGN_BLOCK_CLEAR * 0.3)
        sig_y = y
        seg = (CONTENT_W - 2 * mm) / 2.0
        x1 = M
        c.setLineWidth(0.8)
        c.line(x1, sig_y, x1 + seg - 16 * mm, sig_y)
        c.line(x1 + seg + 2 * mm, sig_y, x1 + 2 * seg - 10 * mm, sig_y)
        c.setFont("Helvetica", PAR_FONT_SIZE - 1)
        drop = (PAR_FONT_SIZE * 1.1)
        c.drawString(x1, sig_y - drop, "Class Teacher’s Signature")
        c.drawString(x1 + seg + 2 * mm, sig_y - drop,
                     "Headteacher’s Signature & Stamp")
        y = sig_y - (PAR_LEADING * 1.5)

    # ---------- grading legend (same logic, but in a box) ----------
        grading = rep.get("grading") or []
        if grading:
            parts = []
            for g in grading:
                lo, hi = g.get("lower_limit"), g.get("upper_limit")
                gr = (g.get("grade") or "").strip()
                if lo is not None and hi is not None and gr:
                    parts.append(f"{int(lo)}–{int(hi)}: {gr}")

            if parts:
                leg_ps = ParagraphStyle(
                    name="legend", fontName="Helvetica",
                    fontSize=PAR_FONT_SIZE - 1,
                    leading=(PAR_FONT_SIZE - 1) * LINE_SPACING_MULT,
                    spaceAfter=PAR_SPACING_AFTER,
                )

                # same text as before
                leg_text = "Grading → " + ", ".join(parts)
                leg_para = Paragraph(leg_text, leg_ps)

                # put legend in a one-cell table (box)
                box_data = [[leg_para]]
                box_tbl = Table(box_data, colWidths=[CONTENT_W])
                box_style = TableStyle([
                    ("BOX", (0, 0), (-1, -1), 0.75, colors.lightgrey),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ])
                box_tbl.setStyle(box_style)

                w, h = box_tbl.wrapOn(c, CONTENT_W, PAGE_H)
                ensure_space(h)
                box_tbl.drawOn(c, M, y - h)
                y -= (h + PAR_SPACING_AFTER)
        

        # ---------- footer note + timestamp ----------
        c.setFont("Helvetica-Oblique", PAR_FONT_SIZE - 1)
        ensure_space(12)
        c.drawString(M, y, "Invalid without school Stamp")
        y -= (PAR_FONT_SIZE * LINE_SPACING_MULT)

        c.setFont("Helvetica", 8)
        c.setFillColor(colors.grey)
        c.drawRightString(
            PAGE_W - M,
            M + 6,
            f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )
        c.setFillColor(colors.black)
        return y


    # ------------------------------------------------------------------
    # RENDER ALL PAGES
    # ------------------------------------------------------------------
    for rep in reports:
        y = draw_header_and_info(rep)

        eot_tbl = make_eot_table(rep)
        y = draw_table_autofit(eot_tbl, M, y, CONTENT_W)

        y -= GAP_BEFORE_SUMMARY
        y = draw_summary(rep, y)

        if include_mid and (rep.get("midterms") or []):
            y = draw_midterm_panel(rep, y)

        y = draw_comments_signatures_legend_footer(rep, y)
        c.showPage()

    c.save()
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=f"EOT_{class_name}_{term}_{year}.pdf",
        mimetype="application/pdf",
    )
    
    
@app.route("/save_midterm_comments/<int:student_id>", methods=["POST"])
@require_role("admin", "headteacher", "dos", "teacher")
def save_midterm_comments(student_id):
    term = request.form.get("term")
    year = request.form.get("year", type=int) or datetime.now().year

    teacher_comment = (request.form.get("teacher_comment") or "").strip()
    head_comment = (request.form.get("head_comment") or "").strip()
    special_comm = (request.form.get("special_communication") or "").strip()

    action = (request.form.get("action") or "save_only").strip()
    next_id = request.form.get("next_id", type=int)

    conn = get_db_connection()
    ensure_midterm_overall_comments_schema(conn)
    cur = conn.cursor(dictionary=True)

    # upsert into your midterm overall comments table
    cur.execute("""
        SELECT id
        FROM midterm_overall_comments
        WHERE student_id=%s AND term=%s AND year=%s
        LIMIT 1
    """, (student_id, term, year))
    row = cur.fetchone()

    if row:
        cur.execute("""
            UPDATE midterm_overall_comments
            SET teacher_comment=%s,
                head_comment=%s,
                special_communication=%s
            WHERE id=%s
        """, (teacher_comment, head_comment, special_comm, row["id"]))
    else:
        cur.execute("""
            INSERT INTO midterm_overall_comments
                (student_id, term, year,
                 teacher_comment, head_comment, special_communication)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (student_id, term, year,
              teacher_comment, head_comment, special_comm))

    conn.commit()
    cur.close()
    conn.close()

    # decide where to go after saving
    target_id = student_id
    if action == "save_next" and next_id:
        target_id = next_id

    return redirect(url_for("midterm_report",
                            student_id=target_id,
                            term=term,
                            year=year))



@app.route("/midterm_report/<int:student_id>/<term>/<int:year>")
@require_role('admin', 'headteacher', 'bursar', 'dos', 'teacher')
def midterm_report(student_id, term, year):
    conn = get_db_connection()
    payload = build_midterm_payload(conn, student_id, term, year)

    if not payload:
        conn.close()
        flash("Learner not found.", "warning")
        return redirect(url_for("reports_hub", term=term, year=year))

    # navigation within same class
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id
        FROM students
        WHERE class_name=%s AND archived=0
        ORDER BY last_name, first_name
    """, (payload["student"]["class_name"],))
    ids = [r["id"] for r in (cur.fetchall() or [])]
    
    # Comment group
    teacher_lib, head_lib = load_comment_library_groups(conn)
    cur.close()
    conn.close()

    prev_id = next_id = None
    if student_id in ids:
        idx = ids.index(student_id)
        if idx > 0:
            prev_id = ids[idx - 1]
        if idx < len(ids) - 1:
            next_id = ids[idx + 1]

    return render_template(
        "midterm_report_citizen.html",
        school=payload["school"],
        student=payload["student"],
        term=payload["term"],
        year=payload["year"],
        rows=payload["rows"],
        aggregate=payload["aggregate"],
        division=payload["division"],
        total_mid_sum=payload["total_mid_sum"],
        avg_mid_overall=payload["avg_mid_overall"],
        comments=payload["comments"],
        grading=payload["grading"],
        show_oth=payload["show_oth"],
        show_bot=payload["show_bot"],
        show_mid=payload["show_mid"],
        prev_id=prev_id,
        next_id=next_id,
        teacher_library=teacher_lib,
        head_library=head_lib,
    )


@app.route("/midterm_report/print_batch", methods=["POST"])
@require_role('admin', 'headteacher', 'bursar', 'dos', 'teacher')
def midterm_report_print_batch():
    class_name = (request.form.get("class_name") or "").strip()
    term = (request.form.get("term") or "").strip()
    year = request.form.get("year", type=int) or datetime.now().year

    try:
        ids = [int(x) for x in request.form.getlist("selected_ids")
               if str(x).strip().isdigit()]
    except Exception:
        ids = []

    if not ids:
        flash("Select at least one student to print (mid-term).", "warning")
        return redirect(url_for("reports_hub",
                                class_name=class_name, term=term, year=year))

    conn = get_db_connection()
    reports = []
    for sid in ids:
        p = build_midterm_payload(conn, sid, term, year)
        if p:
            reports.append(p)
    conn.close()

    if not reports:
        flash("No mid-term data found for the selected students.", "warning")
        return redirect(url_for("reports_hub",
                                class_name=class_name, term=term, year=year))

    # ---------- PDF GENERATION (layout similar to EOT, no position) ----------
    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    PAGE_W, PAGE_H = A4
    M = 10 * mm
    CONTENT_W = PAGE_W - 2 * M

    PAR_FONT_SIZE = 10.5
    LINE_SPACING_MULT = 1.5
    PAR_LEADING = PAR_FONT_SIZE * LINE_SPACING_MULT
    PAR_SPACING_AFTER = 6
    BOTTOM_SPACER = 10

    # column widths
    OTH_W = 16 * mm
    BOT_W = 16 * mm
    MID_W = 16 * mm
    AVG_W = 26 * mm
    GRD_W = 18 * mm
    INIT_W = 18 * mm

    cell_style = ParagraphStyle(
        name="cell", fontName="Helvetica", fontSize=9,
        leading=12, alignment=TA_LEFT
    )
    hdr_style = ParagraphStyle(
        name="hdr", fontName="Helvetica-Bold", fontSize=10,
        leading=13, alignment=TA_LEFT
    )

    SCHOOL_LOGO_PATH = os.path.join(current_app.static_folder, "logo.jpg")

    def draw_header_and_info(rep):
        student = rep["student"]
        school = rep["school"]
        width, height = PAGE_W, PAGE_H
        left = M
        right_margin = M

        COL_NAVY = colors.white
        COL_BLUE = colors.white
        COL_BLUE2 = colors.white

        banner_h = 30 * mm
        banner_y = height - 35 * mm
        left_margin, right_margin_local = left, left
        strip_w = width - left_margin - right_margin_local
        navy_w = strip_w * 0.62
        blue_w = strip_w - navy_w

        c.saveState()
        c.setFillColor(COL_NAVY)
        c.rect(left_margin, banner_y, navy_w, banner_h, stroke=0, fill=1)
        c.setFillColor(COL_BLUE)
        c.rect(left_margin + navy_w, banner_y, blue_w, banner_h, stroke=0, fill=1)

        fold_depth = 11 * mm
        fold_lip = 6 * mm
        c.setFillColor(COL_BLUE2)
        ps = c.beginPath()
        ps.moveTo(left_margin + navy_w, banner_y)
        ps.lineTo(left_margin + navy_w + fold_depth, banner_y + banner_h)
        ps.lineTo(left_margin + navy_w + fold_depth + 2*mm, banner_y + banner_h)
        ps.lineTo(left_margin + navy_w + 2*mm, banner_y)
        ps.close()
        c.drawPath(ps, stroke=0, fill=1)

        flap_col = colors.HexColor("#3a86e0")
        c.setFillColor(flap_col)
        pf = c.beginPath()
        pf.moveTo(left_margin + navy_w - fold_lip, banner_y)
        pf.lineTo(left_margin + navy_w, banner_y)
        pf.lineTo(left_margin + navy_w + fold_depth, banner_y + banner_h)
        pf.lineTo(left_margin + navy_w - fold_lip, banner_y + banner_h)
        pf.close()
        c.drawPath(pf, stroke=0, fill=1)

        logo_box = 24 * mm
        logo_x = left_margin + 6 * mm
        logo_y = banner_y + (banner_h - logo_box) / 2
        if os.path.exists(SCHOOL_LOGO_PATH):
            try:
                c.drawImage(
                    SCHOOL_LOGO_PATH,
                    logo_x, logo_y,
                    width=logo_box, height=logo_box,
                    preserveAspectRatio=True, mask="auto",
                )
            except Exception:
                pass

        SCHOOL_NAME = school.get("name") or ""
        SCHOOL_SUB = school.get("tagline") or school.get("motto") or ""
        name_left = logo_x + logo_box + 6 * mm
        name_right = left_margin + navy_w - 6 * mm
        name_box_w = max(10, name_right - name_left)
        name_fs = 18
        while name_fs >= 10 and c.stringWidth(
            SCHOOL_NAME, "Helvetica-Bold", name_fs
        ) > name_box_w:
            name_fs -= 1
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", name_fs)
        name_y = banner_y + banner_h - (banner_h * 0.30)
        c.drawString(name_left, name_y, SCHOOL_NAME)
        if SCHOOL_SUB:
            sub_fs = 12
            while sub_fs >= 8 and c.stringWidth(
                SCHOOL_SUB, "Helvetica-Bold", sub_fs
            ) > name_box_w:
                sub_fs -= 1
            c.setFont("Helvetica-Bold", sub_fs)
            c.drawString(name_left, name_y - (name_fs * 0.9), SCHOOL_SUB)

        phones_raw = (school.get("phones") or "").strip()
        SCHOOL_PHONE_LINES = [p.strip() for p in phones_raw.split(",") if p.strip()]
        SCHOOL_ADDRESS = school.get("pobox") or ""
        SCHOOL_EMAIL = current_app.config.get("SCHOOL_EMAIL", "")

        c.setFillColor(colors.white)
        c.setFont("Helvetica", 9)
        right_pad = 6 * mm
        text_right = left_margin + strip_w - right_pad
        line_gap = 5.5 * mm
        y_cursor = banner_y + banner_h - 8 * mm
        for ph in SCHOOL_PHONE_LINES:
            c.drawRightString(text_right, y_cursor, ph)
            y_cursor -= line_gap
        if SCHOOL_ADDRESS:
            y_cursor -= 2.5 * mm
            c.drawRightString(text_right, y_cursor, SCHOOL_ADDRESS)
            y_cursor -= line_gap
        if SCHOOL_EMAIL:
            c.drawRightString(text_right, y_cursor, SCHOOL_EMAIL)
        c.restoreState()

        # learner info
        info_top = banner_y - 6 * mm
        info_left = left
        info_width = width - left - right_margin - (40 * mm)

        styles = getSampleStyleSheet()
        lab = ParagraphStyle(
            "lab", parent=styles["Normal"],
            fontName="Helvetica-Bold", fontSize=9, leading=11,
            textColor=colors.black,
        )
        val = ParagraphStyle(
            "val", parent=styles["Normal"],
            fontName="Helvetica", fontSize=9, leading=11,
        )

        full_name = f"{student.get('first_name','')} " \
                    f"{student.get('middle_name','')} " \
                    f"{student.get('last_name','')}".strip()

        info_rows = [
            [Paragraph("Learner's Name:", lab),
             Paragraph(full_name or "-", val)],
            [Paragraph("Student No.:", lab),
             Paragraph(student.get("student_number") or "-", val)],
            [Paragraph("Class / Stream:", lab),
             Paragraph(f"{student.get('class_name','-')} {student.get('stream') or ''}", val)],
            [Paragraph("Term / Year:", lab),
             Paragraph(f"{rep['term']} / {rep['year']}", val)],
        ]
        info_tbl = Table(
            info_rows,
            colWidths=[35 * mm, info_width - 35 * mm],
            hAlign="LEFT",
        )
        info_tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ]))
        w_info, h_info = info_tbl.wrapOn(c, info_width, 9999)
        info_tbl.drawOn(c, info_left, info_top - h_info)


# ---- PHOTO BLOCK (file path only – blob later if you want) ----
        box_w = box_h = 32 * mm
        photo_x = width - right_margin - box_w
        photo_y = info_top - (h_info - box_h) / 2

        photo_path = (student.get("photo") or "").strip()
        if photo_path:
            # normalise: remove any leading slash or accidental "static/"
            photo_path = photo_path.lstrip("/")
            if photo_path.startswith("static/"):
                photo_path = photo_path[len("static/"):]

            full_path = os.path.join(current_app.static_folder, photo_path)
            if os.path.exists(full_path):
                try:
                    img_reader = ImageReader(full_path)
                    c.drawImage(
                        img_reader,
                        photo_x + 2, photo_y - box_h + 2,
                        box_w - 4, box_h - 4,
                        preserveAspectRatio=True,
                        mask="auto",
                    )
                except Exception:
                    pass
            else:
                # file missing – draw placeholder
                c.setStrokeColor(colors.grey)
                c.rect(photo_x, photo_y - box_h, box_w, box_h)
                c.setFont("Helvetica", 7)
                c.drawCentredString(photo_x + box_w/2, photo_y - box_h/2, "Photo")
        else:
            # no path in DB – placeholder
            c.setStrokeColor(colors.grey)
            c.rect(photo_x, photo_y - box_h, box_w, box_h)
            c.setFont("Helvetica", 7)
            c.drawCentredString(photo_x + box_w/2, photo_y - box_h/2, "Photo")


        # title
        title_y = (info_top - h_info) - 4 * mm
        c.setFont("Helvetica-Bold", 13)
        c.setFillColor(colors.black)
        c.drawCentredString(width / 2.0, title_y, "Mid Term Report")
        table_top = title_y - 6 * mm
        return table_top

    def make_table_data_and_widths(rep):
        fixed_total = GRD_W + INIT_W + AVG_W
        if rep["show_oth"]:
            fixed_total += OTH_W
        if rep["show_bot"]:
            fixed_total += BOT_W
        if rep["show_mid"]:
            fixed_total += MID_W
        comment_w = 60 * mm
        subject_w = CONTENT_W - fixed_total - comment_w
        if subject_w < 55 * mm:
            deficit = (55 * mm) - subject_w
            comment_w = max(40 * mm, comment_w - deficit)
            subject_w = CONTENT_W - fixed_total - comment_w

        hdr = ["SUBJECT"]
        if rep["show_oth"]:
            hdr.append("OTH")
        if rep["show_bot"]:
            hdr.append("BOT")
        if rep["show_mid"]:
            hdr.append("MID")
        hdr += ["MID TERM (%)", "GRADE", "COMMENT", "INITIALS"]
        data = [hdr]

        for r in rep["rows"]:
            row = [
                Paragraph(
                    f"{r['name']}" + (
                        f" <font size='8' color='#666666'>({r['code']})</font>"
                        if r.get('code') else ""
                    ),
                    cell_style
                )
            ]
            if rep["show_oth"]:
                row.append("" if r["OTH"] is None else f"{float(r['OTH']):.0f}")
            if rep["show_bot"]:
                row.append("" if r["BOT"] is None else f"{float(r['BOT']):.0f}")
            if rep["show_mid"]:
                row.append("" if r["MID"] is None else f"{float(r['MID']):.0f}")
            row.append("" if r["AVG"] is None else f"{r['AVG']:.1f}")
            row.append(r["grade"] or "")
            row.append(Paragraph(r["comment"] or "", cell_style))
            row.append(r["initials"] or "")
            data.append(row)

        totals_row = [Paragraph("<b>TOTAL</b>", hdr_style)]
        if rep["show_oth"]:
            totals_row.append("")
        if rep["show_bot"]:
            totals_row.append("")
        if rep["show_mid"]:
            totals_row.append("")
        totals_row.append("" if rep["total_mid_sum"] is None
                          else f"{rep['total_mid_sum']:.1f}")
        totals_row.append("")
        totals_row.append("")
        totals_row.append("")
        data.append(totals_row)

        widths = [subject_w]
        if rep["show_oth"]:
            widths.append(OTH_W)
        if rep["show_bot"]:
            widths.append(BOT_W)
        if rep["show_mid"]:
            widths.append(MID_W)
        widths += [AVG_W, GRD_W, comment_w, INIT_W]
        return data, widths

    def draw_table(rep, y):
        data, widths = make_table_data_and_widths(rep)
        row_h = 16
        header_h = 18
        t = Table(data, colWidths=widths, repeatRows=1)
        style = TableStyle([
            ("FONT", (0,0), (-1,0), "Helvetica-Bold", 10),
            ("FONT", (0,1), (-1,-1), "Helvetica", 9),
            ("GRID", (0,0), (-1,-1), 0.6, colors.black),
            ("ALIGN", (1,0), (-2,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("ROWHEIGHT", (0,0), (-1,0), header_h),
        ])
        for r_idx in range(1, len(data)):
            style.add("ROWHEIGHT", (0, r_idx), (-1, r_idx), row_h)
        t.setStyle(style)
        w, h = t.wrapOn(c, CONTENT_W, PAGE_H)
        if y - h < M + BOTTOM_SPACER:
            c.showPage()
            y = draw_header_and_info(rep)
        t.drawOn(c, M, y - h)
        return y - h - 8 * mm

    def draw_summary_and_footer(rep, y):
        c.setFont("Helvetica-Bold", PAR_FONT_SIZE)
        avg_val = rep.get("avg_mid_overall")
        avg_text = "" if avg_val is None else f"{avg_val:.1f}"
        agg_text = "" if rep.get("aggregate") is None else str(rep["aggregate"])
        div_text = rep.get("division") or "NG"
        c.drawString(M, y, f"Average (Mid): {avg_text}")
        c.drawRightString(PAGE_W - M, y,
                          f"Aggregate: {agg_text} Division: {div_text}")
        y -= PAR_LEADING

        ps = ParagraphStyle(
            name="p", fontName="Helvetica", fontSize=PAR_FONT_SIZE,
            leading=PAR_LEADING, spaceAfter=PAR_SPACING_AFTER
        )

        def ensure_space(min_h):
            nonlocal y
            if y - min_h < M + BOTTOM_SPACER:
                c.showPage()
                y = draw_header_and_info(rep)

        # comments (manual already merged, including special communication)
        for label, key in (
            ("Class Teachers’s Remarks:", "teacher_comment"),
            ("Head Teacher’s Remarks:", "head_comment"),
            ("Special Communication:", "special_communication"),
        ):
            txt = f"<b>{label}</b> <i>{(rep['comments'] or {}).get(key,'')}</i>"
            para = Paragraph(txt, ps)
            w, h = para.wrapOn(c, CONTENT_W, PAGE_H)
            ensure_space(h)
            para.drawOn(c, M, y - h)
            y -= (h + PAR_SPACING_AFTER)
        # signatures
        SIGNING_CLEAR_SPACE = max(8 * mm, PAR_LEADING * 1.5)
        ensure_space(SIGNING_CLEAR_SPACE + PAR_LEADING * 2)
        y -= SIGNING_CLEAR_SPACE
        sig_y = y
        seg = (CONTENT_W - 2*mm) / 2.0
        x1 = M
        c.setLineWidth(0.8)
        c.line(x1, sig_y, x1 + seg - 16*mm, sig_y)
        c.line(x1 + seg + 2*mm, sig_y, x1 + 2*seg - 16*mm, sig_y)
        c.setFont("Helvetica", PAR_FONT_SIZE - 1)
        drop = (PAR_FONT_SIZE * 0.95)
        c.drawString(x1, sig_y - drop, "Class Teacher’s Signature")
        c.drawString(x1 + seg + 2*mm, sig_y - drop,
                     "Headteacher’s Signature & Stamp")
        y = sig_y - (PAR_LEADING)

        # grading legend
        grading = rep.get("grading") or []
        if grading:
            parts = []
            for g in grading:
                low, up, gr = g.get("lower_limit"), g.get("upper_limit"), g.get("grade")
                if low is not None and up is not None and gr:
                    parts.append(f"{int(low)}–{int(up)}: {gr}")
            if parts:
                ps2 = ParagraphStyle(
                    name="legend",
                    fontName="Helvetica",
                    fontSize=PAR_FONT_SIZE - 1,
                    leading=(PAR_FONT_SIZE - 1) * LINE_SPACING_MULT,
                    spaceAfter=PAR_SPACING_AFTER
                )
                par = Paragraph("Marks → " + ", ".join(parts), ps2)
                w, h = par.wrapOn(c, CONTENT_W, PAGE_H)
                ensure_space(h)
                par.drawOn(c, M, y - h)
                y -= (h + PAR_SPACING_AFTER)

        # footer
        c.setFont("Helvetica-Oblique", PAR_FONT_SIZE - 1)
        c.drawString(M, y, "Invalid without school Stamp")
        y -= (PAR_FONT_SIZE * LINE_SPACING_MULT)
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.grey)
        c.drawRightString(PAGE_W - M, M + 6,
                          f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        c.setFillColor(colors.black)
        return y

    # render all pages
    for rep in reports:
        y = draw_header_and_info(rep)
        y = draw_table(rep, y)
        y = draw_summary_and_footer(rep, y)
        c.showPage()

    c.save()
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=f"Midterm_{class_name}_{term}_{year}.pdf",
        mimetype="application/pdf",
    )



@app.route('/academic_years', methods=['GET', 'POST'])
@require_role('admin')
def academic_years():
    conn = get_db_connection()
    try:
        if request.method == 'POST':

            # Add year
            if 'add' in request.form:
                year_raw = (request.form.get('year') or '').strip()
                try:
                    y = int(year_raw)
                except Exception:
                    flash("Year must be a number (e.g., 2026).", "warning")
                    return redirect(url_for('academic_years'))

                try:
                    cur = conn.cursor(dictionary=True)
                    cur.execute("INSERT INTO academic_years (year) VALUES (%s)", (y,))
                    conn.commit()
                    cur.close()
                    flash("Academic year added.", "success")
                except Exception:
                    conn.rollback()
                    flash("Year already exists", "warning")

            # Activate year (prevent activating locked year)
            elif 'activate' in request.form:
                year_raw = (request.form.get('year') or '').strip()
                try:
                    y = int(year_raw)
                except Exception:
                    flash("Invalid year.", "warning")
                    return redirect(url_for('academic_years'))

                cur = conn.cursor(dictionary=True)
                cur.execute("SELECT is_locked FROM academic_years WHERE year=%s LIMIT 1", (y,))
                row = cur.fetchone() or {}
                cur.close()

                if int(row.get("is_locked", 0)) == 1:
                    flash(f"Year {y} is locked. Unlock it before activating.", "warning")
                    return redirect(url_for('academic_years'))

                cur = conn.cursor(dictionary=True)
                cur.execute("UPDATE academic_years SET is_active=0")
                cur.execute("UPDATE academic_years SET is_active=1 WHERE year=%s", (y,))
                conn.commit()
                cur.close()
                flash(f"Year {y} activated.", "success")

            # Set term for active year (block if active year locked)
            elif 'set_term' in request.form:
                term = (request.form.get('term') or '').strip()
                if term not in TERMS:
                    flash("Invalid term.", "warning")
                    return redirect(url_for('academic_years'))

                cur = conn.cursor(dictionary=True)
                cur.execute("SELECT year, is_locked FROM academic_years WHERE is_active=1 LIMIT 1")
                active = cur.fetchone() or {}
                cur.close()

                if int(active.get("is_locked", 0)) == 1:
                    flash(f"Active year {active.get('year')} is locked. Unlock to change term.", "warning")
                    return redirect(url_for('academic_years'))

                cur = conn.cursor(dictionary=True)
                cur.execute("UPDATE academic_years SET current_term=%s WHERE is_active=1", (term,))
                conn.commit()
                cur.close()
                flash("Term updated.", "success")

            # Lock year (do not lock active year)
            elif 'lock' in request.form:
                year_raw = (request.form.get('year') or '').strip()
                try:
                    y = int(year_raw)
                except Exception:
                    flash("Invalid year.", "warning")
                    return redirect(url_for('academic_years'))

                cur = conn.cursor(dictionary=True)
                cur.execute("SELECT is_active FROM academic_years WHERE year=%s LIMIT 1", (y,))
                row = cur.fetchone() or {}
                cur.close()

                if int(row.get("is_active", 0)) == 1:
                    flash("You cannot lock the active year.", "warning")
                    return redirect(url_for('academic_years'))

                cur = conn.cursor(dictionary=True)
                cur.execute("UPDATE academic_years SET is_locked=1, locked_on=NOW() WHERE year=%s", (y,))
                conn.commit()
                cur.close()
                flash(f"Year {y} locked.", "success")

            # Unlock year
            elif 'unlock' in request.form:
                year_raw = (request.form.get('year') or '').strip()
                try:
                    y = int(year_raw)
                except Exception:
                    flash("Invalid year.", "warning")
                    return redirect(url_for('academic_years'))

                cur = conn.cursor(dictionary=True)
                cur.execute("UPDATE academic_years SET is_locked=0, locked_on=NULL WHERE year=%s", (y,))
                conn.commit()
                cur.close()
                flash(f"Year {y} unlocked.", "success")

        # Load table
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM academic_years ORDER BY year DESC")
        years = cur.fetchall() or []
        cur.execute("SELECT year, current_term, is_locked FROM academic_years WHERE is_active=1 LIMIT 1")
        active_year = cur.fetchone()
        cur.close()

        return render_template('academic_years.html', years=years, active_year=active_year)

    finally:
        conn.close()


@app.route('/fees/setup', methods=['POST'])
def setup_class_fees():
    data = request.form
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute('''
        INSERT INTO class_fees (class_name, level, section, amount)
        VALUES (%s, %s, %s, %s)
    ''', (
        data['class_name'],
        data['level'],
        data['section'],
        float(data['amount'])
    ))
    conn.commit()
    cur.close()
    conn.close()
    return f"Fees for class_name {data['class_name']} ({data['section']}) set at UGX {data['amount']}."


@app.route('/students/retrieve/<int:student_id>', methods=['POST'])
def retrieve_archived_student(student_id):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        'UPDATE students SET archived = 0, status = "active" WHERE id = %s', (student_id,))
    conn.commit()
    cur.close()
    conn.close()
    return f"Student {student_id} retrieved from archive."


@app.route('/students/drop/<int:student_id>', methods=['POST'])
def drop_student(student_id):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        'UPDATE students SET status = "left" WHERE id = %s', (student_id,))
    conn.commit()
    cur.close()
    conn.close()
    return f"Student {student_id} marked as left."


@app.route('/students/archive/clear', methods=['POST'])
def clear_archived_students():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute('DELETE FROM students WHERE archived = 1')
    conn.commit()
    cur.close()
    conn.close()
    return "All archived students cleared permanently."


@app.route('/expense/categories/init')
def init_expense_categories():
    categories = ['Staff Pay', 'Transport', 'Uniforms',
                  'Secretarial', 'Service Providers', 'Others']
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    for name in categories:
        cur.execute(
            'INSERT IGNORE INTO expense_categories (name) VALUES (%s)', (name,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'message': 'Default categories initialized'})


@app.route('/add_expense', methods=['GET', 'POST'])
@require_role('admin', 'bursar')
def add_expense():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute(
        "SELECT year, current_term FROM academic_years WHERE is_active = 1")
    current = cur.fetchone()
    cur.close()
    if not current:
        flash("No active academic year. Please activate one first.", "warning")
        return redirect(url_for('academic_years'))

    active_year = current['year']
    current_term = current['current_term']
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, name from expense_categories")
    categories = cur.fetchall()
    cur.close()

    if request.method == 'POST':
        description = request.form['description']
        amount = request.form['amount']
        category_id = request.form['category_id']
        type_ = request.form.get('type', 'other')
        date_spent = request.form.get(
            'date_spent') or datetime.now().strftime('%Y-%m-%d')
        recorded_by = session.get('username', 'System')

        try:
            cur = conn.cursor(dictionary=True)
            amount = float(amount)
            cur.execute('''
                INSERT INTO expenses (
                    description, amount, date_spent, category_id,
                    type, recorded_by, term, year
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ''', (description, amount, date_spent, category_id, type_, recorded_by, current_term, active_year))
            conn.commit()
            cur.close()
            flash("Expense added successfully.", "success")
            return redirect(url_for('expenditure_report'))
        except Exception as e:
            flash(f"Error: {e}", "danger")

    conn.close()
    return render_template('add_expense.html', categories=categories)


@app.route('/expenses')
def list_expenses():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute('''
        SELECT e.id, e.title, e.description, e.amount, e.date,
               ec.name AS category, e.recorded_by
        FROM expenses e
        JOIN expense_categories ec ON e.category_id = ec.id
        ORDER BY e.date DESC
    ''')
    expenses = [dict(row) for row in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify(expenses)


@app.route("/export_expenses")
@require_role("admin", "headteacher", "dos", "bursar")
def export_expenses():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT e.id, e.date, e.category, e.amount, e.description,
               u.username AS entered_by
        FROM expenses e
        LEFT JOIN users u ON u.id = e.user_id
        ORDER BY e.date DESC
    """)
    row = cur.fetchall()
    cur.close()
    conn.close()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Headers
    headers = ["ID", "Date", "Category", "Amount", "Description", "Entered By"]
    ws.append(headers)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    # Data rows
    for r in rows:
        ws.append([r["id"], r["date"], r["category"],
                  r["amount"], r["description"], r["entered_by"]])

    # Auto column width
    for column_cells in ws.columns:
        length = max(len(str(cell.value))
                     if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(
            length + 2, 40)

    # Stream to browser
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="expenses.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/expenses/filter', methods=['GET'])
def filter_expenses_by_date():
    start_date = request.args.get('start')
    end_date = request.args.get('end')

    if not start_date or not end_date:
        return jsonify({'error': 'Start and end dates are required'}), 400

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute('''
        SELECT e.id, e.title, e.description, e.amount, e.date,
               ec.name AS category, e.recorded_by
        FROM expenses e
        JOIN expense_categories ec ON e.category_id = ec.id
        WHERE e.date BETWEEN %s AND %s
        ORDER BY e.date DESC
    ''', (start_date, end_date))

    results = [dict(row) for row in cur.fetchall()]
    conn.close()
    return jsonify(results)

@app.route('/pay_teacher', methods=['GET', 'POST'])
@require_role('admin', 'bursar')
def pay_teacher():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    teachers = []
    searched = False
    payment_history = []
    current_year = datetime.now().year

    cur.execute(
        "SELECT current_term FROM academic_years WHERE is_active = 1"
    )
    term_row = cur.fetchone()
    current_term = term_row['current_term'] if term_row else 'N/A'

    if request.method == 'POST':
        if 'search' in request.form:
            search_term = request.form['search_term']
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                SELECT * FROM teachers
                WHERE id LIKE %s OR last_name LIKE %s
            """, (f"%{search_term}%", f"%{search_term}%"))
            teachers = cur.fetchall()
            cur.close()
            searched = True

        elif 'pay_teacher' in request.form:
            teacher_id = request.form['teacher_id']
            amount = float(request.form['amount_paid'])
            term = request.form['term']
            year = int(request.form['year'])
            recorded_by = session.get('full_name', 'Unknown')

            cur.execute("""
                INSERT INTO expenses (
                    description, amount, term, year,
                    category_id, recorded_by, type
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                f"Payment to Teacher ID {teacher_id}",
                amount,
                term,
                year,
                1,  # Assume 1 is the staff payment category
                recorded_by,
                'staff_pay'
            ))
            conn.commit()
            flash("Payment recorded successfully.", "success")
            return redirect(url_for('pay_teacher'))

    cur.execute("""
        SELECT date_spent, amount, term, year, recorded_by,
               REPLACE(description, 'Payment to ', '') AS teacher_name
        FROM expenses
        WHERE type = 'staff_pay'
        ORDER BY date_spent DESC
        LIMIT 10
    """)
    payment_history = cur.fetchall()

    cur.close()
    conn.close()
    return render_template(
        'pay_teacher.html',
        teachers=teachers,
        searched=searched,
        payment_history=payment_history,
        current_term=current_term,
        current_year=current_year
    )


@app.route('/export_teacher_payments')
@require_role('admin', 'bursar')
def export_teacher_payments():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT date_spent, amount, term, year, recorded_by,
               REPLACE(description, 'Payment to ', '') AS teacher_name
        FROM expenses
        WHERE type = 'staff_pay'
    """)
    data = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Teacher Payments"

    headers = ['Date', 'Teacher', 'Amount', 'Term', 'Year', 'Recorded By']
    ws.append(headers)

    for row in data:
        ws.append([
            row['date_spent'],
            row['teacher_name'],
            row['amount'],
            row['term'],
            row['year'],
            row['recorded_by']
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="teacher_payments.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ---------- TEACHERS MODULE ----------
# Put this near your other routes (after get_db_connection, require_role, etc.)


@app.route("/teachers", methods=["GET", "POST"])
@require_role("admin", "director", "headteacher", "dos", "deputyheadteacher")
def teachers_hub():
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)

        # CREATE NEW TEACHER
        if request.method == "POST" and request.form.get("action") == "create":
            try:
                employee_id = int(request.form["employee_id"])
            except Exception:
                cur.close()
                conn.close()
                flash("Select a valid employee.", "warning")
                return redirect(url_for("teachers_hub"))

            initials = (request.form.get("initials") or "").strip().upper()
            if not initials:
                cur.close()
                conn.close()
                flash("Initials are required.", "warning")
                return redirect(url_for("teachers_hub"))

            # Insert teacher
            cur.execute("""
                INSERT INTO teachers (employee_id, initials, status)
                VALUES (%s, %s, 'active')
            """, (employee_id, initials))
            teacher_id = cur.lastrowid

            # Assign subjects/classes
            sel_subjects = request.form.getlist("subjects[]")
            sel_classes = request.form.getlist("classes[]")
            pairs = [(teacher_id, int(sid), cn)
                     for sid in sel_subjects for cn in sel_classes]

            if pairs:
                cur.executemany("""
                    INSERT INTO teacher_subjects (teacher_id, subject_id, class_name)
                    VALUES (%s, %s, %s)
                """, pairs)

            conn.commit()
            cur.close()
            conn.close()
            flash("Teacher created & assignments saved.", "success")
            return redirect(url_for("teachers_hub"))

        # LIST ALL TEACHERS
        cur.execute("""
            SELECT t.id, t.initials, t.status,
                   e.first_name, e.middle_name, e.last_name, e.designation
            FROM teachers t
            JOIN employees e ON e.id = t.employee_id
            ORDER BY (t.status='active') DESC, e.last_name, e.first_name
        """)
        rows = cur.fetchall()

        employees_free = _get_employees_without_teacher(conn)
        subjects = _get_all_subjects(conn)
        classes = _get_all_classes(conn)

        cur.close()
        conn.close()
        return render_template(
            "teachers.html",
            teachers=rows,
            employees_free=employees_free,
            subjects=subjects,
            classes=[r["class_name"] for r in classes]
        )
    except Exception:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()
        raise


@app.route("/teachers/<int:tid>/edit", methods=["GET", "POST"])
@require_role("admin", "director", "headteacher", "dos", "deputyheadteacher")
def edit_teacher(tid):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT t.*, e.first_name, e.middle_name, e.last_name, e.designation
        FROM teachers t
        JOIN employees e ON e.id = t.employee_id
        WHERE t.id=%s
    """, (tid,))
    teacher = cur.fetchone()
    if not teacher:
        cur.close()
        conn.close()
        flash("Teacher not found.", "warning")
        return redirect(url_for("teachers_hub"))

    if request.method == "POST":
        initials = (request.form.get("initials") or "").strip().upper()
        status = request.form.get("status") or "active"

        cur.execute("UPDATE teachers SET initials=%s, status=%s WHERE id=%s",
                    (initials, status, tid))

        # Replace assignments
        cur.execute("DELETE FROM teacher_subjects WHERE teacher_id=%s", (tid,))
        sel_subjects = request.form.getlist("subjects[]")
        sel_classes = request.form.getlist("classes[]")
        pairs = [(tid, int(sid), cn)
                 for sid in sel_subjects for cn in sel_classes]
        if pairs:
            cur.executemany("""
                INSERT INTO teacher_subjects (teacher_id, subject_id, class_name)
                VALUES (%s, %s, %s)
            """, pairs)

        conn.commit()
        cur.close()
        conn.close()
        flash("Teacher updated.", "success")
        return redirect(url_for("teachers_hub"))

    # Preload selected subjects/classes
    cur.execute(
        "SELECT subject_id, class_name FROM teacher_subjects WHERE teacher_id=%s", (tid,))
    rows = cur.fetchall()
    chosen_subjects = {r["subject_id"] for r in rows}
    chosen_classes = {r["class_name"] for r in rows}

    subjects = _get_all_subjects(conn)
    classes = [r["class_name"] for r in _get_all_classes(conn)]

    cur.close()
    conn.close()
    return render_template(
        "teacher_edit.html",
        teacher=teacher,
        subjects=subjects,
        classes=classes,
        chosen_subjects=chosen_subjects,
        chosen_classes=chosen_classes
    )


@app.route("/teachers/<int:tid>/delete", methods=["POST"])
@require_role("admin", "director", "headteacher")
def delete_teacher(tid):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("DELETE FROM teacher_subjects WHERE teacher_id=%s", (tid,))
    cur.execute("DELETE FROM teachers WHERE id=%s", (tid,))
    conn.commit()
    cur.close()
    conn.close()
    flash("Teacher removed.", "info")
    return redirect(url_for("teachers_hub"))


@app.route("/teachers/load", methods=["GET"])
@require_role("admin", "director", "headteacher", "dos", "bursar", "deputyheadteacher")
def teachers_load():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    q = (request.args.get("q") or "").strip()
    class_f = request.args.get("class_name") or ""
    subj_f = request.args.get("subject_id") or ""
    status_f = request.args.get("status") or ""

    where, args = [], []
    if q:
        where.append(
            "(e.first_name LIKE %s OR e.middle_name LIKE %s OR e.last_name LIKE %s)")
        args += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if class_f:
        where.append("ts.class_name = %s")
        args.append(class_f)
    if subj_f:
        where.append("ts.subject_id = %s")
        args.append(subj_f)
    if status_f:
        where.append("t.status = %s")
        args.append(status_f)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    cur.execute(f"""
        SELECT
          t.id AS teacher_id,
          t.initials,
          t.status,
          e.first_name, e.middle_name, e.last_name, e.designation,
          GROUP_CONCAT(
            DISTINCT CONCAT(
              s.name,
              COALESCE(CONCAT(' (', ts.class_name, ')'), '')
            )
            ORDER BY s.name, ts.class_name
            SEPARATOR ', '
          ) AS load_list
        FROM teachers t
        JOIN employees e ON e.id = t.employee_id
        LEFT JOIN teacher_subjects ts ON ts.teacher_id = t.id
        LEFT JOIN subjects s ON s.id = ts.subject_id
        {where_sql}
        GROUP BY t.id, t.initials, t.status, e.first_name, e.middle_name, e.last_name, e.designation
        ORDER BY (t.status='active') DESC, e.last_name, e.first_name
    """, args)
    rows = cur.fetchall()

    cur.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name")
    classes = [r["class_name"] for r in cur.fetchall()]
    cur.execute(
        "SELECT id, name, COALESCE(code,'') AS code FROM subjects ORDER BY name")
    subjects = cur.fetchall()

    cur.close()
    conn.close()
    return render_template(
        "teachers_load.html",
        rows=rows, classes=classes, subjects=subjects,
        q=q, class_f=class_f, subj_f=subj_f, status_f=status_f
    )


@app.route("/teachers/load/export")
@require_role("admin", "director", "headteacher", "dos", "bursar", "deputyheadteacher")
def teachers_load_export():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    q = (request.args.get("q") or "").strip()
    class_f = request.args.get("class_name") or ""
    subj_f = request.args.get("subject_id") or ""
    status_f = request.args.get("status") or ""

    where, args = [], []
    if q:
        where.append(
            "(e.first_name LIKE %s OR e.middle_name LIKE %s OR e.last_name LIKE %s)")
        args += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if class_f:
        where.append("ts.class_name = %s")
        args.append(class_f)
    if subj_f:
        where.append("ts.subject_id = %s")
        args.append(subj_f)
    if status_f:
        where.append("t.status = %s")
        args.append(status_f)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    cur.execute(f"""
        SELECT
          CONCAT(e.last_name, ', ', e.first_name, COALESCE(CONCAT(' ', e.middle_name), '')) AS teacher_name,
          t.initials,
          e.designation,
          t.status,
          GROUP_CONCAT(
            DISTINCT CONCAT(
              s.name,
              COALESCE(CONCAT(' (', ts.class_name, ')'), '')
            )
            ORDER BY s.name, ts.class_name
            SEPARATOR ', '
          ) AS load_list
        FROM teachers t
        JOIN employees e ON e.id = t.employee_id
        LEFT JOIN teacher_subjects ts ON ts.teacher_id = t.id
        LEFT JOIN subjects s ON s.id = ts.subject_id
        {where_sql}
        GROUP BY t.id, teacher_name, t.initials, e.designation, t.status
        ORDER BY (t.status='active') DESC, e.last_name, e.first_name
    """, args)
    data = cur.fetchall()
    cur.close()
    conn.close()

    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(["Teacher", "Initials", "Designation",
                    "Status", "Subjects (Class)"])
    for r in data:
        writer.writerow([
            r["teacher_name"] or "",
            r["initials"] or "",
            r["designation"] or "",
            r["status"] or "",
            r["load_list"] or "",
        ])

    return Response(
        si.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=teacher_load.csv"}
    )


@app.route('/expenses/total')
def total_expense():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute('SELECT SUM(amount) AS total_expense FROM expenses')
    row = cur.fetchone()
    cur.close()
    conn.close()
    return jsonify({'total_expense': row['total_expense'] or 0})


@app.route('/grading_scale', methods=['GET', 'POST'])
@require_role('admin', 'headteacher', 'dos', 'classmanager', 'deputyheadteacher')
def grading_scale():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    if request.method == 'POST':
        grade = request.form['grade']
        lower = int(request.form['lower_limit'])
        upper = int(request.form['upper_limit'])
        comment = request.form.get('comment', '')

        # Check for overlapping ranges
        cur.execute('''
            SELECT * FROM grading_scale
            WHERE NOT (%s < lower_limit OR %s > upper_limit)
        ''', (upper, lower))
        overlapping = cur.fetchall()

        if overlapping:
            flash("Grade range overlaps with existing entries.", "danger")
        else:
            cur.execute('''
                INSERT INTO grading_scale (grade, lower_limit, upper_limit, comment)
                VALUES (%s, %s, %s, %s)
            ''', (grade, lower, upper, comment))
            conn.commit()
            flash("Grade added successfully", "success")

    cur.execute('SELECT * FROM grading_scale ORDER BY lower_limit ASC')
    grades = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('grading_scale.html', grades=grades)


@app.route('/edit_grade/<int:grade_id>', methods=['GET', 'POST'])
@require_role('admin', 'headteacher', 'dos', 'deputyheadteacher')
def edit_grade(grade_id):
    conn = get_db_connection()

    if request.method == 'POST':
        grade = request.form['grade']
        lower = int(request.form['lower_limit'])
        upper = int(request.form['upper_limit'])
        comment = request.form.get('comment', '')

        cur = conn.cursor(dictionary=True)
        cur.execute('''
            UPDATE grading_scale SET grade = %s, lower_limit = %s, upper_limit = %s, comment = %s
            WHERE id = %s
        ''', (grade, lower, upper, comment, grade_id))
        conn.commit()
        cur.close()
        conn.close()
        flash("Grade updated successfully.", "success")
        return redirect(url_for('grading_scale'))

    cur = conn.cursor(dictionary=True)
    cur.execute('SELECT * FROM grading_scale WHERE id = %s', (grade_id,))
    grade_data = cur.fetchone()
    cur.close()
    conn.close()
    return render_template('edit_grade.html', grade=grade_data)


@app.route('/delete_grade/<int:grade_id>')
@require_role('admin')
def delete_grade(grade_id):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute('DELETE FROM grading_scale WHERE id = %s', (grade_id,))
    conn.commit()
    cur.close()
    conn.close()
    flash("Grade deleted.", "info")
    return redirect(url_for('grading_scale'))


@app.route('/export_grading_scale')
@require_role('admin', 'headteacher', 'dos', 'deputyheadteacher')
def export_grading_scale():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM grading_scale ORDER BY lower_limit ASC")
    data = cur.fetchall()
    cur.close()
    conn.close()

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Grading Scale')

    headers = ['Grade', 'Lower Limit', 'Upper Limit', 'Comment']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    for row_num, row in enumerate(data, start=1):
        worksheet.write(row_num, 0, row['grade'])
        worksheet.write(row_num, 1, row['lower_limit'])
        worksheet.write(row_num, 2, row['upper_limit'])
        worksheet.write(row_num, 3, row['comment'] or '')

    workbook.close()
    output.seek(0)

    return send_file(output, download_name='grading_scale.xlsx', as_attachment=True)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ---------- AUDIT TRAIL (admin only) ----------



@app.route('/audit-trail')
@require_role('admin','director')
def audit_trail():
    from math import ceil
    from urllib.parse import urlencode

    q_user = request.args.get('user_id', type=int)
    q_role = (request.args.get('role') or '').strip() or None
    q_act = (request.args.get('action') or '').strip() or None
    q_out = (request.args.get('outcome') or '').strip() or None
    q_route = (request.args.get('route') or '').strip() or None

    page = max(request.args.get('page', 1, type=int), 1)
    per = min(max(request.args.get('per', 50, type=int), 10), 200)
    offset = (page - 1) * per

    conn = get_db_connection()
    ensure_audit_trail_schema(conn)
    cur = conn.cursor(dictionary=True)

    # ---- Build filters ----
    where, params = [], []
    if q_user is not None:
        where.append("a.user_id = %s")
        params.append(q_user)
    if q_role:
        where.append("a.role = %s")
        params.append(q_role)
    if q_act:
        where.append("a.action LIKE %s")
        params.append(f"%{q_act}%")
    if q_out:
        where.append("a.outcome = %s")
        params.append(q_out)
    if q_route:
        where.append("a.route LIKE %s")
        params.append(f"%{q_route}%")

    where_sql = (" WHERE " + " AND ".join(where)) if where else ""

    # ---- Count total ----
    count_sql = f"SELECT COUNT(*) AS n FROM audit_trail a{where_sql}"
    if params:
        cur.execute(count_sql, tuple(params))
    else:
        cur.execute(count_sql)
    total = cur.fetchone()["n"]

    # ---- Select records (no timezone manipulation) ----
    select_sql = f"""
        SELECT
            DATE_FORMAT(a.timestamp, '%Y-%m-%d %H:%i:%S') AS timestamp,
            a.id, a.user_id, a.role, a.action, a.outcome, a.severity,
            a.route, a.method, a.ip_address, a.target_table, a.target_id,
            a.details_json, a.http_status,
            COALESCE(
              NULLIF(TRIM(CONCAT_WS(' ',
                     e.first_name,
                     NULLIF(e.Middle_name, ''),
                     e.last_name
              )), ''),
              u.username, '—'
            ) AS user_name
        FROM audit_trail a
        LEFT JOIN users u ON u.id = a.user_id
        LEFT JOIN employees e ON e.id = u.employee_id
        {where_sql}
        ORDER BY a.timestamp DESC, a.id DESC
        LIMIT {int(per)} OFFSET {int(offset)}
    """

    # ---- Safety check ----
    num_placeholders = select_sql.count("%s")
    if num_placeholders != len(params):
        cur.close(); conn.close()
        raise RuntimeError(
            f"Audit SQL param mismatch: placeholders={num_placeholders}, params={len(params)}; "
            f"SQL={select_sql!r}; params={params!r}"
        )

    if params:
        cur.execute(select_sql, tuple(params))
    else:
        cur.execute(select_sql)
    rows = cur.fetchall()

    cur.close()
    conn.close()

    # ---- Pagination ----
    pages = max(ceil(total / per), 1)
    q = {
        "user_id": q_user,
        "role": q_role,
        "action": q_act,
        "outcome": q_out,
        "route": q_route
    }
    querystring = urlencode({k: v for k, v in q.items() if v not in (None, "")})

    return render_template(
        "audit_trail.html",
        audit_logs=[dict(r) for r in rows],
        total=total, page=page, per=per, pages=pages,
        querystring=querystring, q=q
    )


@app.route('/asset_register', methods=['GET', 'POST'])
@require_role('admin','bursar','headteacher','deputyheadteacher')
def asset_register():
    conn = get_db_connection()
    ensure_assets_schema(conn)

    # Read filters (support GET & POST)
    from_date = (request.values.get('from_date') or '').strip()
    to_date = (request.values.get('to_date') or '').strip()
    name_q = (request.values.get('asset_name') or '').strip()
    cat_q = (request.values.get('category') or '').strip().lower()

    # Add new asset
    if request.method == 'POST' and 'add_asset' in request.form:
        asset_name = (request.form.get('asset_name') or '').strip()
        description = (request.form.get('description') or '').strip()
        model = (request.form.get('model') or '').strip()
        purchase_date = (request.form.get('purchase_date') or '').strip() # YYYY-MM-DD
        asset_condition = (request.form.get('asset_condition') or '').strip()
        value_ = request.form.get('value') or 0
        qty_ = request.form.get('qty') or 0
        location = (request.form.get('location') or '').strip()
        asset_code = (request.form.get('asset_code') or '').strip()
        company_number = (request.form.get('company_number') or '').strip()
        useful_life_years = request.form.get('useful_life_years')

        try: value_ = float(value_)
        except: value_ = 0.0
        try: qty_ = int(qty_)
        except: qty_ = 0
        try: useful_life_years = int(useful_life_years or 0)
        except: useful_life_years = None

        # for legacy: year_purchased mirrors purchase_date (string)
        year_purchased = purchase_date if purchase_date else None

        category = _auto_category(purchase_date, useful_life_years)

        cur = conn.cursor(dictionary=True)
        cur.execute("""
            INSERT INTO assets
              (asset_name, description, model, value, year_purchased, purchase_date,
               asset_condition, qty, location, asset_code, company_number,
               useful_life_years, category)
            VALUES
              (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (asset_name, description, model, value_, year_purchased, purchase_date or None,
              asset_condition, qty_, location, asset_code, company_number,
              useful_life_years, category))
        conn.commit()
        cur.close()
        flash("Asset added.", "success")
        # fall through to listing with same filters

    # Query list
    q, p = _build_asset_query(from_date, to_date, name_q, cat_q)
    cur = conn.cursor(dictionary=True)
    cur.execute(q, p)
    assets = cur.fetchall() or []
    cur.close()

    # Ensure categories are synced
    assets = _refresh_categories(conn, assets)

    conn.close()
    return render_template(
        'asset_register.html',
        assets=assets,
        from_date=from_date,
        to_date=to_date,
        asset_name=name_q,
        category=cat_q
    )



@app.route('/asset/<int:asset_id>/edit', methods=['GET', 'POST'])
@require_role('admin')
def edit_asset(asset_id: int):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Load current asset (for display and for audit snapshots)
    cur.execute("SELECT * FROM assets WHERE id=%s", (asset_id,))
    asset = cur.fetchone()
    if not asset:
        conn.close()
        flash("Asset not found.", "warning")
        # Audit "not found"
        try:
            audit_from_request(
                conn,
                action="asset_edit",
                target_table="assets",
                target_id=asset_id,
                details={"message": "Asset not found for editing"},
                outcome="not_found",
                severity="warning",
            )
        except Exception:
            pass
        return redirect(url_for('asset_register'))

    if request.method != 'POST':
        cur.close(); conn.close()
        return render_template('edit_asset.html', asset=asset)

    # ------------ Parse inputs (mirror your form field names) ------------
    def _f(v, d=0.0):
        try:
            return float(v)
        except:
            return d

    def _i(v, d=0):
        try:
            if v is None or str(v).strip() == "":
                return d
            return int(v)
        except:
            return d

    asset_name = (request.form.get('asset_name') or asset.get("asset_name") or '').strip()
    description = (request.form.get('description') or asset.get("description") or '').strip()
    model = (request.form.get('model') or asset.get("model") or '').strip()
    purchase_date = (request.form.get('purchase_date') or asset.get("purchase_date") or '').strip() or None
    asset_condition = (request.form.get('asset_condition') or asset.get("asset_condition") or '').strip()
    value_ = _f(request.form.get('value') if request.form.get('value') is not None else asset.get("value"))
    qty_ = _i(request.form.get('qty') if request.form.get('qty') is not None else asset.get("qty"))
    location = (request.form.get('location') or asset.get("location") or '').strip()
    asset_code = (request.form.get('asset_code') or asset.get("asset_code") or '').strip()
    company_number = (request.form.get('company_number') or asset.get("company_number") or '').strip()

    uly_raw = request.form.get('useful_life_years')
    if uly_raw is None:
        uly_raw = asset.get("useful_life_years")
    try:
        useful_life_years = int(uly_raw) if (uly_raw not in (None, "")) else None
    except:
        useful_life_years = None

    # Derivations consistent with your existing logic
    year_purchased = purchase_date if purchase_date else None
    try:
        category = _auto_category(purchase_date, useful_life_years) # optional helper you already use
    except Exception:
        # Fallback if helper not present: treat <= 1 year as current else non-current (simple heuristic)
        try:
            yl = int(useful_life_years or 0)
        except:
            yl = 0
        category = "current" if yl <= 1 else "non-current" if yl > 1 else None

    # ------------ BEFORE snapshot (typed) ------------
    def _as_float(x): 
        try: return float(x) if x is not None else 0.0
        except: return 0.0
    def _as_int(x):
        try: return int(x) if x is not None else 0
        except: return 0

    before = {
        "asset_name": asset.get("asset_name"),
        "description": asset.get("description"),
        "model": asset.get("model"),
        "value": _as_float(asset.get("value")),
        "year_purchased": asset.get("year_purchased"),
        "purchase_date": asset.get("purchase_date"),
        "asset_condition": asset.get("asset_condition"),
        "qty": _as_int(asset.get("qty")),
        "location": asset.get("location"),
        "asset_code": asset.get("asset_code"),
        "company_number": asset.get("company_number"),
        "useful_life_years": asset.get("useful_life_years"),
        "category": asset.get("category"),
    }

    # ------------ AFTER payload (what we will write) ------------
    after = {
        "asset_name": asset_name,
        "description": description,
        "model": model,
        "value": float(value_ or 0.0),
        "year_purchased": year_purchased,
        "purchase_date": purchase_date,
        "asset_condition": asset_condition,
        "qty": int(qty_ or 0),
        "location": location,
        "asset_code": asset_code,
        "company_number": company_number,
        "useful_life_years": useful_life_years,
        "category": category,
    }

    changed_only = {
        k: {"before": before.get(k), "after": after.get(k)}
        for k in after.keys() if before.get(k) != after.get(k)
    }

    if not changed_only:
        flash("No changes to save.", "info")
        try:
            audit_from_request(
                conn,
                action="asset_edit",
                target_table="assets",
                target_id=asset_id,
                details={"message": "No field changes"},
                outcome="success",
                severity="info",
            )
        except Exception:
            pass
        cur.close(); conn.close()
        return redirect(url_for('asset_register'))

    # ------------ Update ------------
    try:
        cur.execute("""
            UPDATE assets SET
              asset_name=%s, description=%s, model=%s, value=%s,
              year_purchased=%s, purchase_date=%s,
              asset_condition=%s, qty=%s, location=%s, asset_code=%s,
              company_number=%s, useful_life_years=%s, category=%s
            WHERE id=%s
        """, (after["asset_name"], after["description"], after["model"], after["value"],
              after["year_purchased"], after["purchase_date"],
              after["asset_condition"], after["qty"], after["location"], after["asset_code"],
              after["company_number"], after["useful_life_years"], after["category"],
              asset_id))
        conn.commit()
        cur.close()

        flash("Asset updated.", "success")
        # Audit success with diffs
        audit_from_request(
            conn,
            action="asset_edit",
            target_table="assets",
            target_id=asset_id,
            details={
                "before": before,
                "after": after,
                "changed": changed_only,
                "asset_code": asset.get("asset_code"),
            },
            outcome="success",
            severity="info",
        )
    except Exception as e:
        conn.rollback()
        flash(f"Update failed: {e}", "danger")
        try:
            audit_from_request(
                conn,
                action="asset_edit",
                target_table="assets",
                target_id=asset_id,
                details={"error": str(e), "attempted_changes": changed_only},
                outcome="failure",
                severity="warning",
            )
        except Exception:
            pass
        finally:
            conn.close()
        return redirect(url_for('asset_register'))

    conn.close()
    return redirect(url_for('asset_register'))


@app.route('/asset/<int:asset_id>/delete', methods=['POST'])
@require_role('admin')
def delete_asset(asset_id: int):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Load snapshot for audit
    cur.execute("SELECT * FROM assets WHERE id=%s", (asset_id,))
    asset = cur.fetchone()
    if not asset:
        conn.close()
        flash("Asset not found.", "warning")
        try:
            audit_from_request(
                conn,
                action="asset_delete",
                target_table="assets",
                target_id=asset_id,
                details={"message": "Asset not found for deletion"},
                outcome="not_found",
                severity="warning",
            )
        except Exception:
            pass
        return redirect(url_for('asset_register'))

    try:
        cur.execute("DELETE FROM assets WHERE id=%s", (asset_id,))
        conn.commit()
        cur.close()

        flash("Asset deleted.", "success")
        # Audit success with deleted snapshot
        audit_from_request(
            conn,
            action="asset_delete",
            target_table="assets",
            target_id=asset_id,
            details={"deleted_snapshot": asset},
            outcome="success",
            severity="info",
        )
    except Exception as e:
        conn.rollback()
        flash(f"Delete failed: {e}", "danger")
        try:
            audit_from_request(
                conn,
                action="asset_delete",
                target_table="assets",
                target_id=asset_id,
                details={"error": str(e), "snapshot": asset},
                outcome="failure",
                severity="warning",
            )
        except Exception:
            pass
    finally:
        conn.close()

    return redirect(url_for('asset_register'))


# ---------- Excel Export ----------
@app.route('/assets/export_excel')
@require_role('admin','director','bursar','headteacher', 'deputyheadteacher')
def export_assets():
    from_date = (request.args.get('from_date') or '').strip()
    to_date = (request.args.get('to_date') or '').strip()
    name_q = (request.args.get('asset_name') or '').strip()
    cat_q = (request.args.get('category') or '').strip().lower()

    conn = get_db_connection()
    ensure_assets_schema(conn)
    q, p = _build_asset_query(from_date, to_date, name_q, cat_q)
    cur = conn.cursor(dictionary=True)
    cur.execute(q, p)
    rows = cur.fetchall() or []
    cur.close()
    rows = _refresh_categories(conn, rows)
    conn.close()

    bio = BytesIO()
    wb = xlsxwriter.Workbook(bio, {'in_memory': True})
    ws = wb.add_worksheet("Assets")

    head_fmt = wb.add_format({'bold': True, 'border': 1, 'align': 'center', 'valign':'vcenter'})
    cell_fmt = wb.add_format({'border': 1})
    money_fmt = wb.add_format({'border': 1, 'num_format': '#,##0'})

    headers = [
        "Name", "Description", "Model", "Location", "Code", "Company #",
        "Purchase Date", "Condition", "Qty", "Useful Life (yrs)",
        "Category", "Value (UGX)"
    ]
    for col, h in enumerate(headers):
        ws.write(0, col, h, head_fmt)

    for r, a in enumerate(rows, start=1):
        ws.write(r, 0, a.get("asset_name") or "", cell_fmt)
        ws.write(r, 1, a.get("description") or "", cell_fmt)
        ws.write(r, 2, a.get("model") or "", cell_fmt)
        ws.write(r, 3, a.get("location") or "", cell_fmt)
        ws.write(r, 4, a.get("asset_code") or "", cell_fmt)
        ws.write(r, 5, a.get("company_number") or "", cell_fmt)
        ws.write(r, 6, (a.get("purchase_date") or "") if not a.get("purchase_date") else a["purchase_date"].strftime("%Y-%m-%d"), cell_fmt)
        ws.write(r, 7, a.get("asset_condition") or "", cell_fmt)
        ws.write_number(r, 8, int(a.get("qty") or 0), cell_fmt)
        ws.write_number(r, 9, int(a.get("useful_life_years") or 0), cell_fmt)
        ws.write(r, 10, a.get("category") or "", cell_fmt)
        ws.write_number(r, 11, float(a.get("value") or 0.0), money_fmt)

    ws.set_column(0, 1, 26) # name, description
    ws.set_column(2, 5, 14) # model, location, code, comp #
    ws.set_column(6, 6, 13) # purchase date
    ws.set_column(7, 10, 13) # condition..category
    ws.set_column(11, 11, 14) # value

    wb.close()
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=ASSET_EXCEL_FILENAME,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------- PDF Export (auto-fit + logo + title) ----------
@app.route('/assets/export_pdf')
@require_role('admin','bursar', 'headteacher', 'deputyheadteacher')
def export_assets_pdf():
    from datetime import date, datetime
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    import os

    # ---- filters (optional, used for the DB query only) ----
    from_date = (request.args.get('from_date') or '').strip()
    to_date = (request.args.get('to_date') or '').strip()
    name_q = (request.args.get('asset_name') or '').strip()
    cat_q = (request.args.get('category') or '').strip().lower()

    # ---- fetch ----
    conn = get_db_connection()
    ensure_assets_schema(conn)

    q, p = _build_asset_query(from_date, to_date, name_q, cat_q)
    cur = conn.cursor(dictionary=True)
    cur.execute(q, p)
    rows = cur.fetchall() or []
    cur.close()

    # refresh categories (current vs non-current)
    rows = _refresh_categories(conn, rows)
    conn.close()

    # split sections
    current_rows = [r for r in rows if (r.get("category") or "").lower() == "current"]
    noncurrent_rows = [r for r in rows if (r.get("category") or "").lower() == "non-current"]

    # ---- PDF setup ----
    bio = BytesIO()
    c = rl_canvas.Canvas(bio, pagesize=landscape(A4))
    PAGE_W, PAGE_H = landscape(A4)
    M = 12 * mm
    CONTENT_W = PAGE_W - 2 * M

    # ---- styles ----
    cell_style = ParagraphStyle(
        name="cell", fontName="Helvetica", fontSize=9, leading=12, alignment=TA_LEFT
    )
    hdr_style = ParagraphStyle(
        name="hdr", fontName="Helvetica-Bold", fontSize=10, leading=13, alignment=TA_LEFT
    )
    right_style = ParagraphStyle(
        name="right", fontName="Helvetica", fontSize=9, leading=12, alignment=TA_RIGHT
    )
    center_small = ParagraphStyle(
        name="center_small", fontName="Helvetica", fontSize=9, leading=12, alignment=TA_CENTER
    )

    # ---- columns definition (key, header, relative weight, style) ----
    COLS = [
        ("asset_name", "Name", 1.10, cell_style),
        ("description", "Description", 1.90, cell_style),
        ("model", "Model", 0.80, cell_style),
        ("location", "Location", 0.90, cell_style),
        ("asset_code", "Code", 0.70, center_small),
        ("company_number", "Company #", 0.90, center_small),
        ("purchase_date", "Purchase Date", 0.95, center_small),
        ("asset_condition", "Condition", 0.85, center_small),
        ("qty", "Qty", 0.55, center_small),
        ("useful_life_years","Life(yrs)", 0.65, center_small),
        ("category", "Category", 0.85, center_small),
        ("value", "Value (UGX)", 1.10, right_style),
    ]

    # auto-fit widths from weights
    def col_widths():
        min_w = 18 * mm # never let a col be thinner than this
        weights = [w for _, _, w, _ in COLS]
        W = sum(weights)
        base = CONTENT_W / W
        widths = [max(min_w, base * w) for w in weights]
        # if we exceeded (because of mins), normalize back to CONTENT_W
        total = sum(widths)
        if total != 0:
            factor = CONTENT_W / total
            widths = [w * factor for w in widths]
        return widths

    def fmt_date(v):
        if not v:
            return ""
        if isinstance(v, (datetime, date)):
            return v.strftime("%Y-%m-%d")
        s = str(v)
        # attempt to parse common formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass
        return s # fallback

    def _fmt_money(v):
        try:
            return f"{float(v):,.0f}"
        except Exception:
            return ""

    # ---- header with logo + title ----
    def draw_header(section_title: str):
        y_top = PAGE_H - M
        logo_path = os.path.join(current_app.static_folder, "logo.jpg")

        # left logo (if exists)
        if os.path.exists(logo_path):
            try:
                c.drawImage(logo_path, M, y_top - 14*mm, width=18*mm, height=18*mm,
                            preserveAspectRatio=True, mask='auto')
            except Exception:
                pass

        # center school name + document title
        school = current_app.config.get("SCHOOL_NAME", "DEMO DAY AND BOARDING PRIMARY SCHOOL – MASAJJA")
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(PAGE_W/2, y_top - 2*mm, school)
        c.setFont("Helvetica-Bold", 11.5)
        c.drawCentredString(PAGE_W/2, y_top - 7*mm, "Asset Register")

        # section label (left)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(M, y_top - 20*mm, section_title)

        return y_top - 20*mm # content start Y

    # ---- build a table for a list of rows ----
    def make_table(items):
        # header row
        data = [[h for _, h, _, _ in COLS]]

        # body rows (wrap with Paragraph)
        for a in items:
            row = []
            for key, _, _, sty in COLS:
                if key == "purchase_date":
                    val = fmt_date(a.get(key))
                elif key == "qty" or key == "useful_life_years":
                    v = a.get(key)
                    val = "" if v in (None, "") else str(int(v))
                elif key == "value":
                    val = _fmt_money(a.get("value"))
                else:
                    val = a.get(key) or ""
                row.append(Paragraph(str(val), sty))
            data.append(row)

        t = Table(data, colWidths=col_widths(), repeatRows=1)
        style = TableStyle([
            ("FONT", (0,0), (-1,0), "Helvetica-Bold", 9),
            ("FONT", (0,1), (-1,-1), "Helvetica", 8.5),
            ("GRID", (0,0), (-1,-1), 0.5, colors.black),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN", (0,0), (-1,0), "CENTER"),
            ("LEFTPADDING", (0,0), (-1,-1), 2),
            ("RIGHTPADDING", (0,0), (-1,-1), 2),
            ("ROWHEIGHT", (0,0), (-1,0), 18),
        ])
        # slightly taller body rows for readability
        for i in range(1, len(data)):
            style.add("ROWHEIGHT", (0, i), (-1, i), 15)
        t.setStyle(style)
        return t

    # ---- render Current section ----
    y = draw_header("Current Assets (≤ 1 year remaining life)")
    t = make_table(current_rows)
    w, h = t.wrapOn(c, CONTENT_W, PAGE_H)
    if y - h < M:
        c.showPage()
        y = draw_header("Current Assets (≤ 1 year remaining life)")
    t.drawOn(c, M, y - h)
    y = y - h - 8

    # ---- render Non-current section ----
    t2 = make_table(noncurrent_rows)
    w, h = t2.wrapOn(c, CONTENT_W, PAGE_H)
    if y - h < M:
        c.showPage()
        y = draw_header("Non-current Assets (> 1 year remaining life)")
    else:
        c.setFont("Helvetica-Bold", 11)
        c.drawString(M, y, "Non-current Assets (> 1 year remaining life)")
        y -= 6
    t2.drawOn(c, M, y - h)

    c.showPage()
    c.save()
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="Asset_Register.pdf",
        mimetype="application/pdf",
    )



@app.route('/add_income', methods=['GET', 'POST'])
@require_role('admin', 'bursar')
def add_income():
    preview = None
    if request.method == 'POST':
        if 'preview' in request.form:
            # Just return form data to re-render the page with preview
            preview = {
                'source': request.form['source'],
                'amount': request.form['amount'],
                'term': request.form['term'],
                'year': request.form['year'],
                'description': request.form['description'],
                'date_received': request.form['date_received']
            }
        elif 'confirm' in request.form:
            conn = get_db_connection()
            cur = conn.cursor(dictionary=True)

            source = request.form['source']
            amount = float(request.form['amount'])
            term = request.form['term']
            year = int(request.form['year'])
            description = request.form['description']
            date_received = request.form['date_received']
            recorded_by = session.get('full_name', 'Unknown')

            cur.execute('''
                INSERT INTO other_income (source, amount, term, year, description, recorded_by, date_received)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (source, amount, term, year, description, recorded_by, date_received))

            conn.commit()
            cur.close()
            conn.close()
            flash("Income successfully recorded", "success")
            return redirect(url_for('add_income'))

    return render_template('add_income.html', preview=preview)


@app.route('/income_statement', methods=['GET', 'POST'])
@require_role('admin', 'bursar')
def income_statement():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')

    query_params = ()
    date_filter = ""
    if start_date and end_date:
        date_filter = "AND date_paid BETWEEN %s AND %s"
        query_params = (start_date, end_date)

    # Income from fees
    cur.execute(f'''
        SELECT SUM(amount_paid) as total
        FROM fees
        WHERE payment_type = 'school_fees' {date_filter}
    ''', query_params)
    fee_income = cur.fetchone()['total'] or 0

    # Income from requirements
    cur.execute(f'''
        SELECT SUM(amount_paid) as total
        FROM fees
        WHERE payment_type NOT IN ('school_fees') {date_filter}
    ''', query_params)
    req_income = cur.fetchone()['total'] or 0

    # Other income
    cur.execute(f'''
        SELECT SUM(amount) as total
        FROM other_income
        WHERE date_received IS NOT NULL {f"AND date_received BETWEEN %s AND %s" if date_filter else ""}
    ''', query_params)
    other_income = cur.fetchone()['total'] or 0

    # Expenses
    cur.execute(f'''
        SELECT SUM(amount) as total
        FROM expenses
        WHERE date_spent IS NOT NULL {f"AND date_spent BETWEEN %s AND %s" if date_filter else ""}
    ''', query_params)
    expenses = cur.fetchone()['total'] or 0

    net_income = (fee_income + req_income + other_income) - expenses

    cur.close()
    conn.close()
    return render_template(
        'income_statement.html',
        start_date=start_date,
        end_date=end_date,
        fee_income=fee_income,
        req_income=req_income,
        other_income=other_income,
        total_income=fee_income + req_income + other_income,
        expenses=expenses,
        net_income=net_income
    )


@app.route('/income_statement/export')
@require_role('admin', 'bursar', 'director')
def export_income_statement():
    import io
    import csv
    from flask import make_response, request

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    date_filter = ""
    params = ()
    if start_date and end_date:
        date_filter = "AND date_paid BETWEEN %s AND %s"
        params = (start_date, end_date)

    cur.execute(f'''
        SELECT SUM(amount_paid) FROM fees
        WHERE payment_type = 'school_fees' {date_filter}
    ''', params)
    fee_income = cur.fetchone()[0] or 0

    cur.execute(f'''
        SELECT SUM(amount_paid) FROM fees
        WHERE payment_type NOT IN ('school_fees') {date_filter}
    ''', params)
    req_income = cur.fetchone()[0] or 0

    cur.execute(f'''
        SELECT SUM(amount) FROM other_income
        WHERE date_received IS NOT NULL {f"AND date_received BETWEEN %s AND %s" if date_filter else ""}
    ''', params)
    other_income = cur.fetchone()[0] or 0

    cur.execute(f'''
        SELECT SUM(amount) FROM expenses
        WHERE date_spent IS NOT NULL {f"AND date_spent BETWEEN %s AND %s" if date_filter else ""}
    ''', params)
    expenses = cur.fetchone()[0] or 0

    net = (fee_income + req_income + other_income) - expenses
    cur.close()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Income Statement'])
    writer.writerow(['From:', start_date or '', 'To:', end_date or ''])
    writer.writerow([])
    writer.writerow(['Category', 'Amount (UGX)'])
    writer.writerow(['School Fees', fee_income])
    writer.writerow(['Requirements', req_income])
    writer.writerow(['Other Income', other_income])
    writer.writerow(['Total Income', fee_income + req_income + other_income])
    writer.writerow(['Expenses', expenses])
    writer.writerow(['Net Income', net])

    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=income_statement.csv'
    response.headers['Content-Type'] = 'text/csv'
    return response


@app.route('/income_report/export')
@require_role('admin', 'bursar', 'director')
def export_income_report():
    import io
    import csv
    from flask import make_response, request

    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    term = request.args.get('term', '').strip()

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Fees
    fees_query = '''
        SELECT student_id, amount_paid, date_paid
        FROM fees
        WHERE payment_type = 'school_fees' AND date_paid BETWEEN %s AND %s
    '''
    reqs_query = '''
        SELECT student_id, payment_type, amount_paid, date_paid
        FROM fees
        WHERE payment_type NOT IN ('school_fees') AND date_paid BETWEEN %s AND %s
    '''
    params = (from_date, to_date)

    if term:
        fees_query += ' AND term = %s'
        reqs_query += ' AND term = %s'
        params += (term,)

    cur.execute(fees_query, params)
    fees = cur.fetchall()
    cur.execute(reqs_query, params)
    requirements = cur.fetchall()

    cur.execute('''
        SELECT source, amount, recorded_by, date_received
        FROM other_income
        WHERE date_received BETWEEN %s AND %s
    ''', (from_date, to_date))
    other_income = cur.fetchall()

    cur.close()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Income Report'])
    writer.writerow(['From:', from_date, 'To:',
                    to_date, 'Term:', term or 'All'])
    writer.writerow([])

    writer.writerow(['School Fees'])
    writer.writerow(['Student ID', 'Amount Paid', 'Date'])
    for f in fees:
        writer.writerow([f['student_id'], f['amount_paid'], f['date_paid']])
    writer.writerow([])

    writer.writerow(['Requirement Payments'])
    writer.writerow(['Student ID', 'Item', 'Amount Paid', 'Date'])
    for r in requirements:
        writer.writerow([r['student_id'], r['payment_type'],
                        r['amount_paid'], r['date_paid']])
    writer.writerow([])

    writer.writerow(['Other Income'])
    writer.writerow(['Source', 'Amount', 'Recorded By', 'Date'])
    for o in other_income:
        writer.writerow([o['source'], o['amount'],
                        o['recorded_by'], o['date_received']])

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=income_report.csv"
    response.headers["Content-type"] = "text/csv"
    return response


@app.route('/income_report', methods=['GET', 'POST'])
@require_role('admin', 'bursar', 'director')
def income_report():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    from_date = to_date = datetime.now().strftime('%Y-%m-%d')
    term = ''
    if request.method == 'POST':
        from_date = request.form['from_date']
        to_date = request.form['to_date']
        term = request.form.get('term', '').strip()
    else:
        cur.execute(
            "SELECT current_term FROM academic_years WHERE is_active = 1")
        term_row = cur.fetchone()
        term = term_row['current_term'] if term_row else ''

    filters = {
        'from_date': from_date,
        'to_date': to_date,
        'term': term
    }

    params = [from_date, to_date]
    term_filter = ''
    if term:
        term_filter = 'AND f.term = %s'
        params.append(term)

    cur.execute(f'''
        SELECT f.student_id, s.student_number, s.first_name, s.last_name,
               f.amount_paid, f.date_paid, f.payment_type
        FROM fees f
        JOIN students s ON f.student_id = s.id
        WHERE f.payment_type = 'school_fees' AND f.date_paid BETWEEN %s AND %s {term_filter}
    ''', params)
    fees = cur.fetchall()

    cur.execute(f'''
        SELECT f.student_id, s.student_number, s.first_name, s.last_name,
               f.amount_paid, f.date_paid, f.payment_type
        FROM fees f
        JOIN students s ON f.student_id = s.id
        WHERE f.payment_type != 'school_fees' AND f.date_paid BETWEEN %s AND %s {term_filter}
    ''', params)
    requirements = cur.fetchall()

    cur.execute('''
        SELECT source, amount, recorded_by, date_received
        FROM other_income
        WHERE date_received BETWEEN %s AND %s
    ''', [from_date, to_date])
    other_income = cur.fetchall()

    totals = {
        'fees_total': sum(row['amount_paid'] for row in fees),
        'requirements_total': sum(row['amount_paid'] for row in requirements),
        'other_income_total': sum(row['amount'] for row in other_income),
    }
    totals['overall'] = totals['fees_total'] + \
        totals['requirements_total'] + totals['other_income_total']

    cur.close()
    conn.close()
    return render_template(
        'view_income.html',
        filters=filters,
        fees=fees,
        requirements=requirements,
        other_income=other_income,
        totals=totals
    )

# ---------- Student Hub ----------


@app.route("/students/hub")
@require_role("admin", "headteacher", "dos", "bursar", "deputyheadteacher", "classmanager","teacher")
def students_hub():
    return render_template("students_hub.html")

# ---------- Register Student ----------



@app.route("/register_student", methods=["GET", "POST"])
@require_role("admin", "headteacher", "dos", "bursar", "deputyheadteacher", "classmanager", "teacher")
def register_student():
    ay = get_active_academic_year()
    active_year = int(ay.get("year"))
    active_term = ay.get("current_term") or ay.get("term") or "Term 1"
    today_str = datetime.now().strftime("%Y-%m-%d")
    houses = ["Tiger", "Zebra", "Eagle", "Lion"]

    if request.method == "GET":
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        try:
            cur.execute(
                "SELECT DISTINCT stream FROM classes "
                "WHERE stream IS NOT NULL AND TRIM(stream) <> '' "
                "ORDER BY stream"
            )
            rows = cur.fetchall() or []
            streams = [r["stream"] for r in rows]
        finally:
            cur.close()
            conn.close()

        return render_template(
            "register_student.html",
            active_year=active_year,
            active_term=active_term,
            date_joined=today_str, # for the alert line
            streams=streams,
            houses=houses,
        )

    # ---- POST ----
    f = request.form
    file_photo = request.files.get("photo")

    raw_first = (f.get("first_name") or "").strip()
    raw_middle = (f.get("Middle_name") or f.get("middle_name") or "").strip()
    raw_last = (f.get("last_name") or "").strip()
    raw_sex = (f.get("sex") or "").strip()
    raw_class = (f.get("class_name") or "").strip()
    raw_stream = (f.get("stream") or "").strip()
    raw_section = (f.get("section") or "").strip()

    # ✅ parents + residence + house
    parent1_name = (f.get("parent1_name") or "").strip() or None
    parent1_contact = (f.get("parent1_contact") or "").strip() or None
    parent2_name = (f.get("parent2_name") or "").strip() or None
    parent2_contact = (f.get("parent2_contact") or "").strip() or None
    parent_email = (f.get("parent_email") or "").strip() or None
    residence = (f.get("residence") or "").strip() or None
    house = (f.get("house") or "").strip() or None

    # ✅ allow manual student_number/fees_code, else auto
    manual_student_number = (f.get("student_number") or "").strip()
    manual_fees_code = (f.get("fees_code") or "").strip()

    class_name = norm_class(raw_class)
    stream = norm_stream(raw_stream)
    section = norm_section(raw_section)
    sex = norm_sex(raw_sex)

    # validations (keep yours)
    if not raw_first or not raw_last:
        flash("First and Last name are required.", "danger")
        return redirect(url_for("register_student"))
    if not class_name:
        flash("Please choose a valid Class (Baby, Middle, Top, P1–P7).", "danger")
        return redirect(url_for("register_student"))
    if not section:
        flash("Please choose Section (Day/Boarding).", "danger")
        return redirect(url_for("register_student"))
    if not sex:
        flash("Please choose Sex (M/F).", "danger")
        return redirect(url_for("register_student"))

    year_of_joining = int((f.get("year_of_joining") or active_year))
    term_joined = (f.get("term_joined") or active_term)
    date_joined = (f.get("date_joined") or today_str)

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        student_number = manual_student_number or generate_student_number(conn)
        fees_code = manual_fees_code or generate_fees_code(conn)

        photo_path, photo_blob, photo_mime = save_student_photo(file_photo)

        cur.execute(
            """
            INSERT INTO students (
                first_name, Middle_name, last_name, sex,
                class_name, stream, section,
                student_number, year_of_joining, term_joined, date_joined,
                fees_code,
                residence, house,
                parent_name, parent_contact,
                parent2_name, parent2_contact,
                parent_email,
                photo, photo_blob, photo_mime,
                archived, status
            ) VALUES (
                %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s, %s,
                %s,
                %s, %s,
                %s, %s,
                %s, %s,
                %s,
                %s, %s, %s,
                0, 'active'
            )
            """,
            (
                raw_first, raw_middle, raw_last, sex,
                class_name, stream, section,
                student_number, year_of_joining, term_joined, date_joined,
                fees_code,
                residence, house,
                parent1_name, parent1_contact,
                parent2_name, parent2_contact,
                parent_email,
                photo_path, photo_blob, photo_mime,
            ),
        )
        conn.commit()
        flash(f"Student {raw_first} {raw_last} registered (#{student_number}).", "success")
        return redirect(url_for("register_student"))

    except mysql.connector.Error as e:
        conn.rollback()
        flash(f"Failed to register: {e}", "danger")
        return redirect(url_for("register_student"))
    except Exception as e:
        conn.rollback()
        flash(f"Unexpected error: {e}", "danger")
        return redirect(url_for("register_student"))
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()



# ---------- Edit Student ----------
@app.route("/students/<int:student_id>/edit", methods=["GET", "POST"])
@require_role("admin", "headteacher", "dos", "bursar", "deputyheadteacher", "classmanager", "teacher")
def edit_student(student_id):
    # ---- tiny helpers (scoped here so you don't have to hunt elsewhere) ----
    def _active_year() -> int:
        c = get_db_connection()
        k = c.cursor(dictionary=True)
        try:
            k.execute("SELECT year FROM academic_years WHERE is_active=1 LIMIT 1")
            return int((k.fetchone() or {}).get("year") or 0)
        finally:
            try: k.close()
            except: pass
            c.close()

    def call_proc(sql: str, params: tuple | None = None):
        """Safe CALL wrapper for mysql-connector (buffered dict cursor, no .next_result)."""
        cn = get_db_connection()
        cu = cn.cursor(dictionary=True, buffered=True)
        try:
            cu.execute(sql, params or ())
            while True:
                try:
                    _ = cu.fetchall()
                except Exception:
                    pass
                if not cu.nextset():
                    break
            cn.commit()
        finally:
            try: cu.close()
            except: pass
            cn.close()

    # ---- load dropdowns (unchanged) ----
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name")
    classes = [r["class_name"] for r in cur.fetchall()]
    cur.execute("SELECT DISTINCT stream FROM classes ORDER BY stream")
    streams = [r["stream"] for r in cur.fetchall()]
    houses = ["Tiger", "Zebra", "Eagle", "Lion"]

    if request.method == "POST":
        f = request.form
        file_photo = request.files.get("photo") # NEW

        first_name = (f.get("first_name") or "").strip()
        Middle_name = (f.get("Middle_name") or "").strip()
        last_name = (f.get("last_name") or "").strip()
        sex = norm_sex(f.get("sex"))
        section = norm_section(f.get("section"))
        class_name = norm_class(f.get("class_name"))
        stream = norm_stream(f.get("stream"))
        house = (f.get("house") or "").strip()
        parent_name = (f.get("parent_name") or "").strip()
        parent_contact = (f.get("parent_contact") or "").strip()
        parent2_name = (f.get("parent2_name") or "").strip()
        parent2_contact = (f.get("parent2_contact") or "").strip()
        parent_email = (f.get("parent_email") or "").strip()
        residence = (f.get("residence") or "").strip()
        student_number = (f.get("student_number") or "").strip()
        fees_code = (f.get("fees_code") or "").strip()

        missing = []
        for label, val in [
            ("First name", first_name),
            ("Last name", last_name),
            ("Sex", sex),
            ("Section", section),
            ("Class", class_name),
            ("Stream", stream),
        ]:
            if not val:
                missing.append(label)
        if missing:
            conn.close()
            flash("Missing/invalid fields: " + ", ".join(missing), "warning")
            return redirect(url_for("edit_student", student_id=student_id))

        # --- read OLD class/section + photo before update (no logic change) ---
        cur.execute(
            "SELECT class_name AS old_class, section AS old_section, "
            "photo AS old_photo, photo_blob AS old_blob, photo_mime AS old_mime "
            "FROM students WHERE id=%s",
            (student_id,),
        )
        prev = cur.fetchone() or {}
        old_class = (prev.get("old_class") or "").strip()
        old_section = (prev.get("old_section") or "").strip()
        old_photo = prev.get("old_photo")
        old_blob = prev.get("old_blob")
        old_mime = prev.get("old_mime")

        photo_path, photo_blob, photo_mime = save_student_photo(
            file_photo,
            existing_path=old_photo,
            existing_blob=old_blob,
            existing_mime=old_mime,
        )

        cur.execute(
            """
            UPDATE students SET
              first_name=%s, Middle_name=%s, last_name=%s,
              sex=%s, section=%s, class_name=%s, stream=%s, house=%s,
              parent_name=%s, parent_contact=%s, parent2_name=%s, parent2_contact=%s, parent_email=%s,
              residence=%s, student_number=%s, fees_code=%s,
              photo=%s, photo_blob=%s, photo_mime=%s
            WHERE id=%s
            """,
            (
                first_name,
                Middle_name,
                last_name,
                sex,
                section,
                class_name,
                stream,
                house,
                parent_name,
                parent_contact,
                parent2_name,
                parent2_contact,
                parent_email,
                residence,
                student_number,
                fees_code,
                photo_path,
                photo_blob,
                photo_mime,
                student_id,
            ),
        )
        conn.commit()
        cur.close()
        conn.close()

        # --- recompute ONLY if class or section actually changed ---
        changed = (
            (old_class or "").strip().lower() != (class_name or "").strip().lower()
            or (old_section or "").strip().lower() != (section or "").strip().lower()
        )
        if changed:
            year = _active_year()
            try:
                # fast path: one call that recomputes all 3 terms for this student
                call_proc("CALL recompute_student_all_terms_for_year(%s, %s)", (student_id, year))
            except Exception:
                # fallback: recompute each term individually
                for t in (1, 2, 3):
                    call_proc("CALL recompute_fee_term_summary(%s, %s, %s)", (student_id, year, t))

        flash("Student updated.", "success")
        return redirect(url_for("students"))

    # ---- GET branch (unchanged) ----
    cur.execute("SELECT * FROM students WHERE id=%s", (student_id,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        flash("Student not found.", "warning")
        return redirect(url_for("students"))
    return render_template("edit_student.html", s=row, classes=classes, streams=streams, houses=houses)


# ---------- Import Students (CSV/XLSX) ----------
@app.route("/students/import", methods=["GET", "POST"])
@require_role("admin", "headteacher", "bursar", "deputyheadteacher", "dos")
def students_import():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("Select a CSV or XLSX file.", "warning")
            return redirect(url_for("students_import"))

        filename = file.filename.lower()
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # guiding list only; extras are ignored
        expected = ["first_name", "Middle_name", "last_name", "sex", "section", "class_name", "stream",
                    "house", "parent1_name", "parent1_contact", "parent2_name", "parent2_contact",
                    "parent_email", "residence", "fees_code", "student_number"]

        imported = 0
        try:
            if filename.endswith(".csv"):
                content = io.StringIO(file.stream.read().decode("utf-8"))
                reader = csv.DictReader(content)
                for _row in reader:
                    def g(k): return (_row.get(k) or _row.get(
                        k.upper()) or "").strip()

                    # normalize
                    sex = norm_sex(g("sex"))
                    section = norm_section(g("section"))
                    class_name = norm_class(g("class_name"))
                    stream = norm_stream(g("stream"))

                    # minimal required (normalized values!)
                    if not (g("first_name") and g("last_name") and sex and section and class_name and stream):
                        continue

                    sn = g("student_number") or generate_student_number(
                        conn, class_name)
                    fc = g("fees_code") or generate_fees_code(conn)

                    cur.execute("""
                        INSERT INTO students (
                          first_name, Middle_name, last_name, sex, section, class_name, stream,
                          house, parent_name, parent_contact, parent2_name, parent2_contact,
                          parent_email, residence, student_number, fees_code, archived, status
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0,'active')
                    """, (g("first_name"), g("Middle_name"), g("last_name"), sex, section, class_name, stream,
                          g("house"), g("parent1_name"), g("parent1_contact"), g(
                              "parent2_name"), g("parent2_contact"),
                          g("parent_email"), g("residence"), sn, fc))
                    imported += 1

            elif filename.endswith(".xlsx"):
                from openpyxl import load_workbook
                wb = load_workbook(file, data_only=True)
                ws = wb.active
                headers = [(cell.value or "").strip().lower()
                           for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                idx = {h: i for i, h in enumerate(headers)}

                def val(row, name):
                    i = idx.get(name.lower())
                    if i is None:
                        return ""
                    v = row[i].value
                    return (str(v).strip() if v is not None else "")

                for row in ws.iter_rows(min_row=2):
                    first_name = val(row, "first_name")
                    Middle_name = val(row, "Middle_name") or val(
                        row, "middle_name")
                    last_name = val(row, "last_name")
                    sex = norm_sex(val(row, "sex"))
                    section = norm_section(val(row, "section"))
                    class_name = norm_class(val(row, "class_name"))
                    stream = norm_stream(val(row, "stream"))
                    house = val(row, "house")
                    parent1_name = val(row, "parent1_name")
                    parent1_contact = val(row, "parent1_contact")
                    parent2_name = val(row, "parent2_name")
                    parent2_contact = val(row, "parent2_contact")
                    parent_email = val(row, "parent_email")
                    residence = val(row, "residence")
                    sn = val(row, "student_number") or generate_student_number(
                        conn, class_name)
                    fc = val(row, "fees_code") or generate_fees_code(conn)

                    if not (first_name and last_name and sex and section and class_name and stream):
                        continue

                    cur.execute("""
                        INSERT INTO students (
                          first_name, Middle_name, last_name, sex, section, class_name, stream,
                          house, parent_name, parent_contact, parent2_name, parent2_contact,
                          parent_email, residence, student_number, fees_code, archived, status
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0,'active')
                    """, (first_name, Middle_name, last_name, sex, section, class_name, stream,
                          house, parent1_name, parent1_contact, parent2_name, parent2_contact,
                          parent_email, residence, sn, fc))
                    imported += 1
            else:
                flash("Unsupported file type. Use .csv or .xlsx", "warning")
                conn.close()
                return redirect(url_for("students_import"))

            conn.commit()
            flash(f"Imported {imported} students.", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Import failed: {e}", "danger")
        finally:
            cur.close()
            conn.close()

        return redirect(url_for("students"))

    return render_template("students_import.html")


# ---------- Students list (search) ----------
# If you already have a /students route, keep its name and replace the body.
@app.route("/students")
@require_role("admin", "headteacher", "dos", "bursar", "teacher", "classmanager","deputyheadteacher")
def students():
    q_class = (request.args.get("class_name") or "").strip()
    q_num = (request.args.get("student_number") or "").strip()
    q_last = (request.args.get("last_name") or "").strip()

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name")
    classes = [r["class_name"] for r in cur.fetchall()]

    where = ["archived = 0"]
    params = []
    if q_class:
        where.append("class_name = %s")
        params.append(q_class)
    if q_num:
        where.append("student_number LIKE %s")
        params.append(f"%{q_num}%")
    if q_last:
        where.append("last_name LIKE %s")
        params.append(f"%{q_last}%")

    sql = f"""
      SELECT id, student_number, first_name, Middle_name, last_name,
             sex, section, class_name, stream, house,
             parent_name, parent_contact, parent2_name, parent2_contact,
             parent_email, residence, fees_code, date_joined, term_joined, year_of_joining
      FROM students
      WHERE {' AND '.join(where)}
      ORDER BY class_name, last_name, first_name
    """
    cur.execute(sql, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("students_manage.html",
                           rows=rows, classes=classes,
                           q_class=q_class, q_num=q_num, q_last=q_last)

# ---------- Export Students (CSV) ----------


@app.route("/students/export")
@require_role("admin", "headteacher", "dos", "bursar", "classmanager","deputyheadteacher","teacher")
def students_export():
    # reuse same filters as /students
    q_class = (request.args.get("class_name") or "").strip()
    q_num = (request.args.get("student_number") or "").strip()
    q_last = (request.args.get("last_name") or "").strip()

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    where = ["archived = 0"]
    params = []
    if q_class:
        where.append("class_name = %s")
        params.append(q_class)
    if q_num:
        where.append(
            "student_number LIKE %s")
        params.append(f"%{q_num}%")
    if q_last:
        where.append("last_name LIKE %s")
        params.append(f"%{q_last}%")

    sql = f"""SELECT student_number, first_name, Middle_name, last_name, sex, section,
                     class_name, stream, house, parent_name, parent_contact,
                     parent2_name, parent2_contact, parent_email, residence, fees_code
              FROM students
              WHERE {' AND '.join(where)}
              ORDER BY class_name, last_name, first_name"""
    cur.execute(sql, params)
    rows = cur.fetchall()
    cur.close
    conn.close()

    # build CSV
    output = io.StringIO()
    w = csv.writer(output)
    header = ["student_number", "first_name", "Middle_name", "last_name", "sex", "section",
              "class_name", "stream", "house", "parent1_name", "parent1_contact",
              "parent2_name", "parent2_contact", "parent_email", "residence", "fees_code"]
    w.writerow(header)
    for r in rows:
        w.writerow([r["student_number"], r["first_name"], r["Middle_name"], r["last_name"], r["sex"], r["section"],
                    r["class_name"], r["stream"], r["house"], r["parent_name"], r["parent_contact"],
                    r["parent2_name"], r["parent2_contact"], r["parent_email"], r["residence"], r["fees_code"]])
    csv_data = output.getvalue()
    return Response(csv_data, mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=students.csv"})

# ---------- Parents Directory + Export ----------


@app.route("/parents")
@require_role("admin", "headteacher", "bursar", "dos" "teacher", "classmanager", "deputyheadteacher")
def parents_directory():
    search = (request.args.get("search") or "").strip()

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    if search:
        cur.execute("""
            SELECT student_number,
                   parent_name AS parent1_name, parent_contact AS parent1_contact,
                   parent2_name, parent2_contact,
                   CONCAT_WS(' ', first_name, COALESCE(Middle_name, ''), last_name) AS student_name,
                   class_name, stream
            FROM students
            WHERE archived = 0 AND (
                  parent_name LIKE %s OR parent_contact LIKE %s OR
                  parent2_name LIKE %s OR parent2_contact LIKE %s OR
                  student_number LIKE %s OR last_name LIKE %s
            )
            ORDER BY parent_name, parent2_name
        """, (f"%{search}%",)*6)
        rows = cur.fetchall()
    else:
        cur.execute("""
            SELECT student_number,
                   parent_name AS parent1_name, parent_contact AS parent1_contact,
                   parent2_name, parent2_contact,
                   CONCAT_WS(' ', first_name, COALESCE(Middle_name, ''), last_name) AS student_name,
                   class_name, stream
            FROM students
            WHERE archived = 0
            ORDER BY parent_name, parent2_name
        """)
        rows = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("parents_directory.html", rows=rows, search=search)
# for Jinja templates: has_role('admin') etc.


@app.context_processor
def inject_role_utils():
    def has_role(*roles):
        r = session.get("role")
        return r in roles
    return dict(has_role=has_role)


@app.route("/parents/export")
@require_role("admin", "headteacher", "bursar", "deputyheadteacher")
def parents_export():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT student_number,
               parent_name AS parent1_name, parent_contact AS parent1_contact,
               parent2_name, parent2_contact, parent_email,
               CONCAT_WS(' ', first_name, COALESCE(Middle_name, ''), last_name) AS student_name,
               class_name, stream
        FROM students
        WHERE archived = 0
        ORDER BY parent_name, parent2_name
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    output = io.StringIO()
    w = csv.writer(output)
    header = ["student_number", "student_name", "class_name", "stream",
              "parent1_name", "parent1_contact", "parent2_name", "parent2_contact", "parent_email"]
    w.writerow(header)
    for r in rows:
        w.writerow([r["student_number"], r["student_name"], r["class_name"], r["stream"],
                    r["parent1_name"], r["parent1_contact"], r["parent2_name"], r["parent2_contact"], r["parent_email"]])
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=parents.csv"})


@app.route('/fees/report')
@require_role('admin', 'bursar', 'headteacher', 'deputyheadteacher')
def fees_report():
    # active period
    conn = get_db_connection(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT year, current_term FROM academic_years WHERE is_active=1 LIMIT 1")
    academic = cur.fetchone()
    cur.close(); conn.close()
    if not academic:
        flash("Please activate an academic year.", "warning")
        return redirect(url_for('dashboard'))

    year = int(academic['year'])
    term_label = (academic['current_term'] or '').strip().lower()
    term_no = {'term 1': 1, 'term 2': 2, 'term 3': 3}.get(term_label)
    term = academic['current_term']

    if not term_no:
        flash("Active term is not set correctly (Term 1/2/3).", "warning")
        return redirect(url_for('dashboard'))

    # filters
    class_filter  = (request.args.get('class_name') or '').strip()
    stream_filter = (request.args.get('stream') or '').strip()
    raw_status    = (request.args.get('status') or '').strip().lower().replace(' ', '')
    status_map    = {
        'full':'full','partial':'partial','none':'none',
        'fullypaid':'full','partialpaid':'partial','notpaid':'none',
        'credit':'full','overpaid':'full'
    }
    status_filter = status_map.get(raw_status, '')

    # base where + params
    where = ["s.archived=0"]
    params = [year, term_no]
    if class_filter:
        where.append("s.class_name=%s"); params.append(class_filter)
    if stream_filter:
        where.append("(s.stream=%s OR (s.stream IS NULL AND %s=''))")
        params.extend([stream_filter, stream_filter])
    where_sql = " AND ".join(where)

    conn = get_db_connection(); cur = conn.cursor(dictionary=True)
    cur.execute(f"""
      SELECT
        s.id, s.student_number, s.first_name, s.last_name, s.class_name, s.stream,

        /* expected components */
        COALESCE(fs.expected_fees,0)          AS expected_fees,
        COALESCE(fs.expected_reqs_base,0)     AS expected_reqs_base,
        COALESCE(fs.transport_due_term,0)     AS transport_due_term,
        COALESCE(fs.bursary_current,0)        AS bursary_current,

        /* authoritative from summary */
        COALESCE(fs.fees_expected_net,0)      AS fees_expected_net,
        COALESCE(fs.req_expected_final,0)     AS req_expected_final,
        COALESCE(fs.carry_forward,0)          AS carry_forward,
        COALESCE(fs.overall_expected,0)       AS overall_expected,   /* excludes CF per proc */
        COALESCE(fs.overall_paid,0)           AS overall_paid,

        /* ✅ signed overall balance (can be negative = credit) */
        (COALESCE(fs.carry_forward,0) + (COALESCE(fs.overall_expected,0) - COALESCE(fs.overall_paid,0)))
          AS overall_outstanding_signed

      FROM students s
      LEFT JOIN fee_term_summary fs
        ON fs.student_id = s.id AND fs.year = %s AND fs.term_no = %s
      WHERE {where_sql}
      ORDER BY s.class_name, s.stream, s.last_name, s.first_name
    """, params)
    rows = cur.fetchall() or []
    cur.close(); conn.close()

    EPS = 0.01
    report = {'full': [], 'partial': [], 'none': []}

    for r in rows:
        signed_bal = float(r.get('overall_outstanding_signed') or 0.0)

        expected_display = float(r.get('expected_fees') or 0.0) \
                         + float(r.get('expected_reqs_base') or 0.0) \
                         + float(r.get('transport_due_term') or 0.0)

        bursary_display  = float(r.get('bursary_current') or 0.0)
        carried          = float(r.get('carry_forward') or 0.0)
        paid_total       = float(r.get('overall_paid') or 0.0)

        # classify: credit is treated as "full" but will display negative balance
        if signed_bal <= EPS:
            status = 'full'
        elif (expected_display - bursary_display) > EPS and paid_total <= EPS:
            status = 'none'
        else:
            status = 'partial'

        if status_filter and status != status_filter:
            continue

        final_expected = (float(r.get('fees_expected_net') or 0.0) + float(r.get('req_expected_final') or 0.0)) + carried

        report[status].append({
            'student': {
                'id': r['id'], 'student_number': r['student_number'],
                'first_name': r['first_name'], 'last_name': r['last_name'],
                'class_name': r['class_name'], 'stream': r['stream']
            },
            'expected': expected_display,
            'bursary': bursary_display,
            'carried': carried,
            'final_expected': final_expected,
            'paid': paid_total,
            'balance': signed_bal,   # ✅ will be negative if overpaid
        })

    return render_template(
        'fees_report.html',
        report=report, term=term, year=year,
        class_filter=class_filter, stream_filter=stream_filter,
        status_filter=status_filter
    )



@app.route('/fees/report/export.csv')
@require_role('admin', 'bursar', 'headteacher')
def export_fees_report():
    # Active period
    conn = get_db_connection(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT year, current_term FROM academic_years WHERE is_active=1 LIMIT 1")
    academic = cur.fetchone(); cur.close(); conn.close()
    if not academic:
        flash("Please activate an academic year.", "warning")
        return redirect(url_for('fees_report'))

    year = int(academic['year'])
    term_label = (academic['current_term'] or '').strip().lower()
    term = academic['current_term']
    term_no = {'term 1':1,'term 2':2,'term 3':3}.get(term_label)

    # Filters (match your page)
    class_filter  = (request.args.get('class_name') or '').strip()
    stream_filter = (request.args.get('stream') or '').strip()
    raw_status    = (request.args.get('status') or '').strip().lower().replace(' ', '')
    status_map    = {'full':'full','partial':'partial','none':'none',
                     'fullypaid':'full','partialpaid':'partial','notpaid':'none'}
    status_filter = status_map.get(raw_status, '')

    where = ["s.archived=0"]
    params = [year, term_no]
    if class_filter:
        where.append("s.class_name=%s"); params.append(class_filter)
    if stream_filter:
        where.append("(s.stream=%s OR (s.stream IS NULL AND %s=''))")
        params.extend([stream_filter, stream_filter])
    where_sql = " AND ".join(where)

    # Fetch rows with authoritative numbers from fee_term_summary
    conn = get_db_connection(); cur = conn.cursor(dictionary=True)
    cur.execute(f"""
      SELECT
        s.id, s.student_number, s.first_name, s.last_name, s.class_name, s.stream,
        COALESCE(fs.expected_fees,0)               AS expected_fees,
        COALESCE(fs.expected_reqs_base,0)          AS expected_reqs_base,
        COALESCE(fs.transport_due_term,0)          AS transport_due_term,
        COALESCE(fs.bursary_current,0)             AS bursary_current,
        COALESCE(fs.fees_expected_net,0)           AS fees_expected_net,
        COALESCE(fs.req_expected_final,0)          AS req_expected_final,
        COALESCE(fs.carry_forward,0)               AS carry_forward,
        COALESCE(fs.overall_paid,0)                AS overall_paid,
        COALESCE(fs.overall_outstanding,0)         AS overall_outstanding
      FROM students s
      LEFT JOIN fee_term_summary fs
        ON fs.student_id = s.id AND fs.year = %s AND fs.term_no = %s
      WHERE {where_sql}
      ORDER BY s.class_name, s.stream, s.last_name, s.first_name
    """, params)
    rows = cur.fetchall() or []
    cur.close(); conn.close()

    # Bucket + compute the same display fields
    EPS = 0.01
    export_rows = []
    for r in rows:
        expected_display = float(r['expected_fees'] or 0.0) \
                           + float(r['expected_reqs_base'] or 0.0) \
                           + float(r['transport_due_term'] or 0.0)
        bursary_display = float(r['bursary_current'] or 0.0)
        carried         = float(r['carry_forward'] or 0.0)
        paid_total      = float(r['overall_paid'] or 0.0)
        overall_bal     = float(r['overall_outstanding'] or 0.0)

        final_expected  = max(float(r['fees_expected_net'] or 0.0), 0.0) \
                          + float(r['req_expected_final'] or 0.0) \
                          + carried

        if overall_bal <= EPS:
            status = 'full'
        elif (expected_display - bursary_display) > EPS and paid_total <= EPS:
            status = 'none'
        else:
            status = 'partial'

        # Respect filter
        if status_filter and status != status_filter:
            continue

        export_rows.append({
            "Student Number": r['student_number'],
            "Name": f"{r['first_name']} {r['last_name']}",
            "Class": r['class_name'],
            "Stream": r['stream'],
            "Status": status.capitalize(),
            "Expected (Fees+Req+Transport)": expected_display,
            "Bursary": bursary_display,
            "Carry Forward": carried,
            "Final Expected": final_expected,
            "Paid": paid_total,
            "Balance": overall_bal,
        })

    # CSV
    si = StringIO()
    writer = csv.writer(si)
    headers = [
        "Student Number","Name","Class","Stream","Status",
        "Expected (Fees+Req+Transport)","Bursary","Carry Forward",
        "Final Expected","Paid","Balance"
    ]
    writer.writerow(headers)
    for r in export_rows:
        writer.writerow([r[h] for h in headers])

    si.seek(0)
    filename = f"Fees_Report_{term}_{year}.csv"
    return Response(si.getvalue(),
                    mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


# Download a blank template for import (CSV)
@app.route("/students/import/template")
@require_role("admin", "headteacher", "dos", "bursar", "deputyheadteacher")
def students_import_template():
    header = ["first_name", "Middle_name", "last_name", "sex", "section", "class_name", "stream",
              "house", "parent1_name", "parent1_contact", "parent2_name", "parent2_contact",
              "parent_email", "residence", "fees_code", "student_number"]
    output = io.StringIO()
    csv.writer(output).writerow(header)
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=students_import_template.csv"})


# ---------- FEES EDIT (admin only) ----------

@app.route("/fees/<int:fee_id>/edit", methods=["GET", "POST"])
@require_role("admin")
def fees_edit(fee_id: int):

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT f.*, s.student_number, s.first_name, COALESCE(s.Middle_name,'') AS middle_name,
               s.last_name, s.class_name, s.stream
          FROM fees f
          JOIN students s ON s.id = f.student_id
         WHERE f.id = %s
    """, (fee_id,))
    fee = cur.fetchone()
    if not fee:
        conn.close()
        flash("Transaction not found.", "warning")
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        term = (request.form.get("term") or fee["term"] or "").strip()
        year = int(request.form.get("year") or fee["year"] or 0)
        date_paid = (request.form.get("date_paid")
                     or fee["date_paid"] or "").strip()
        method = (request.form.get("method") or fee["method"] or "").strip()
        ptype_raw = (request.form.get("payment_type")
                     or fee["payment_type"] or "").strip()

        ptype_lc = ptype_raw.lower().replace(" ", "_")
        if ptype_lc in ("school_fees", "schoolfees", "tuition"):
            payment_type = "school_fees"
        elif ptype_lc in ("fees",):
            payment_type = "fees"
        elif ptype_lc in ("requirements", "requirement"):
            payment_type = "requirements"
        elif ptype_lc in ("opening_balance", "opening-balance", "ob"):
            payment_type = "opening_balance"
        else:
            payment_type = ptype_raw

        def _f(v):
            try:
                return float(v)
            except:
                return 0.0

        amount_paid = _f(request.form.get("amount_paid"))
        expected_amount = _f(request.form.get("expected_amount"))
        bursary_amount = _f(request.form.get("bursary_amount"))
        carried_forward = _f(request.form.get("carried_forward"))
        requirement_name = (request.form.get("requirement_name") or "").strip()
        comment = (request.form.get("comment") or "").strip()

        # snapshot BEFORE
        before = {
            "term": fee["term"], "year": int(fee["year"] or 0),
            "date_paid": fee["date_paid"], "method": fee["method"],
            "payment_type": fee["payment_type"],
            "amount_paid": float(fee["amount_paid"] or 0.0),
            "expected_amount": float(fee["expected_amount"] or 0.0),
            "bursary_amount": float(fee["bursary_amount"] or 0.0),
            "carried_forward": float(fee["carried_forward"] or 0.0),
            "requirement_name": fee["requirement_name"],
            "comment": fee["comment"],
        }

        make_void = request.form.get("void_entry") == "on"
        if make_void:
            who = session.get("username") or "admin"
            when = datetime.now().strftime("%Y-%m-%d %H:%M")
            # build comment safely, even if comment is None/empty
            comment = f"{(comment or '').strip()} [VOIDED by {who} on {when}]".strip(
            )
            # zero amounts
            amount_paid = expected_amount = bursary_amount = carried_forward = 0.0

        try:
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                UPDATE fees
                   SET term=%s,
                       year=%s,
                       date_paid=%s,
                       method=%s,
                       payment_type=%s,
                       amount_paid=%s,
                       expected_amount=%s,
                       bursary_amount=%s,
                       carried_forward=%s,
                       requirement_name=%s,
                       comment=%s,
                       processed_on=CURRENT_TIMESTAMP
                 WHERE id=%s
            """, (term, year, date_paid, method, payment_type,
                  amount_paid, expected_amount, bursary_amount, carried_forward,
                  requirement_name, comment, fee_id))
            conn.commit()
            cur.close()
            flash("Transaction updated.", "success")

            # AFTER + changed-only for the audit payload
            after = {
                "term": term, "year": year, "date_paid": date_paid, "method": method,
                "payment_type": payment_type, "amount_paid": amount_paid,
                "expected_amount": expected_amount, "bursary_amount": bursary_amount,
                "carried_forward": carried_forward, "requirement_name": requirement_name,
                "comment": comment,
            }
            changed_only = {
                k: {"before": before.get(k), "after": after.get(k)}
                for k in after.keys() if before.get(k) != after.get(k)
            }

            audit_from_request(
                conn,
                action="fees_edit",
                target_table="fees",
                target_id=fee_id,
                details={
                    "student_id": fee["student_id"],
                    "before": before,
                    "after": after,
                    "changed": changed_only,
                    "voided": make_void
                },
                outcome="success",
                severity="info"
            )
        except Exception as e:
            conn.rollback()
            flash(f"Update failed: {e}", "danger")
            audit_from_request(
                conn,
                action="fees_edit",
                target_table="fees",
                target_id=fee_id,
                details={"error": str(e)},
                outcome="failure",
                severity="warning"
            )
        finally:
            conn.close()

        return redirect(url_for("student_statement_by_id", student_id=fee["student_id"]))

    cur.close()
    conn.close()
    return render_template("fees_edit.html", fee=fee)


# ========= Secure Change Password (hashed) =========
# Requires: from flask import request, render_template, redirect, url_for, flash, session, current_app


@app.route("/change_password", methods=["GET", "POST"])
def change_password():
    username = session.get("username")
    if not username:
        flash("You must be logged in to change your password.", "warning")
        return redirect(url_for("login"))

    if request.method == "POST":
        old_password = (request.form.get("old_password") or "").strip()
        new_password = (request.form.get("new_password") or "").strip()
        new_password2 = (request.form.get("new_password2") or "").strip()

        if not old_password or not new_password or not new_password2:
            flash("Please fill in all password fields.", "warning")
            return redirect(url_for("change_password"))
        if new_password != new_password2:
            flash("New passwords do not match.", "danger")
            return redirect(url_for("change_password"))
        if len(new_password) < 8:
            flash("New password must be at least 8 characters.", "warning")
            return redirect(url_for("change_password"))

        conn, cur = None, None
        try:
            conn = get_db_connection()  # use your standard helper
            cur = conn.cursor(dictionary=True)

            # Figure out password column (fallback to 'password')
            pw_col = "password"
            cur.execute("SHOW COLUMNS FROM users LIKE 'password_hash'")
            if cur.fetchone():
                pw_col = "password_hash"

            cur.execute("SELECT * FROM users WHERE username=%s", (username,))
            user = cur.fetchone()
            if not user:
                flash("User not found.", "danger")
                return redirect(url_for("change_password"))

            stored_hash = user.get(pw_col)
            if not stored_hash:
                flash("Password is not set for this account. Contact admin.", "danger")
                return redirect(url_for("change_password"))

            if not check_password_hash(stored_hash, old_password):
                flash("Old password is incorrect.", "danger")
                return redirect(url_for("change_password"))

            if check_password_hash(stored_hash, new_password):
                flash("New password must be different from the old one.", "warning")
                return redirect(url_for("change_password"))

            new_hash = generate_password_hash(new_password)
            cur.execute(
                f"UPDATE users SET {pw_col}=%s WHERE username=%s", (new_hash, username))
            conn.commit()

            flash("Password changed successfully.", "success")
            return redirect(url_for("dashboard"))

        except Exception as e:
            try:
                current_app.logger.exception("[change_password] failed")
            except Exception:
                pass
            flash(f"Could not change password: {e}", "danger")
            return redirect(url_for("change_password"))
        finally:
            if cur is not None:
                try:
                    cur.close()
                except Exception:
                    pass
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    # GET
    return render_template("change_password.html")


# ---------- small helpers ----------


# --- normalize role every request (belt & braces) ---


# --- Login route ---


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        next_url = request.args.get("next") or request.form.get("next")

        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT * FROM users WHERE username=%s LIMIT 1", (username,)
        )
        user = cur.fetchone()
        cur.close()
        conn.close()

        if not user:
            flash("Invalid credentials.", "danger")
            return render_template("login.html", current_year=datetime.now().year, next=next_url)

        if "status" in user.keys() and user["status"] != "active":
            flash("This account is archived/disabled.", "warning")
            return render_template("login.html", current_year=datetime.now().year, next=next_url)

        if not check_password_hash(user["password_hash"], password):
            flash("Invalid credentials.", "danger")
            return render_template("login.html", current_year=datetime.now().year, next=next_url)

        # success
        session.clear()
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["role"] = _norm_role(user["role"])  # normalized
        session["employee_id"] = user.get("employee_id")
        session["staff_id"] = user.get("employee_id")

        prof = get_user_profile(user["id"])
        session["initials"] = prof.get("initials", "") or ""
        session["full_name"] = (prof.get("full_name")
                                or user["username"]).strip()

        flash("Login successful.", "success")
        if next_url and _is_safe_url(next_url):
            return redirect(next_url)
        return redirect(url_for("dashboard"))

    return render_template("login.html", current_year=datetime.now().year, next=request.args.get("next", ""))

# --- Legacy login_post (fixed) ---



@app.route("/profile", methods=["GET", "POST"])
@require_role("admin", "bursar", "teacher", "headteacher", "director", "clerk", "deputyheadteacher")
def user_profile():
    user_id = get_user_id()
    if not user_id:
        flash("Please login first.", "danger")
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Load user
    cur.execute("SELECT id, username, role, status, employee_id FROM users WHERE id=%s", (user_id,))
    user = cur.fetchone()

    if request.method == "POST":
        # Only allow username change (optional)
        new_username = request.form.get("username", "").strip()

        # Only update if provided
        if new_username:
            cur.execute("""
                UPDATE users SET username=%s WHERE id=%s
            """, (new_username, user_id))
            conn.commit()
            flash("Profile updated.", "success")
        else:
            flash("Nothing to update.", "info")

        cur.close()
        conn.close()
        return redirect(url_for("user_profile"))

    cur.close()
    conn.close()
    return render_template("user_profile.html", user=user)



@app.route('/clearance')
@require_role('admin', 'bursar', 'headteacher')
def clearance():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Get active term and year
    cur.execute(
        "SELECT year, current_term FROM academic_years WHERE is_active = 1")
    current = cur.fetchone()
    if not current:
        flash("No active academic year found.", "warning")
        return redirect(url_for('academic_years'))

    year, term = current['year'], current['current_term']

    # Fetch students who have fully paid (school fees only) for that term and year
    cur.execute('''
        SELECT s.id, s.first_name, s.Middle_name, s.last_name, s.class_name, SUM(f.amount_paid) as total_paid, f.expected_amount
        FROM students s
        JOIN fees f ON s.id = f.student_id
        WHERE f.term = %s AND f.year = %s AND f.payment_type = 'school_fees'
        GROUP BY s.id
        HAVING total_paid >= f.expected_amount
    ''', (term, int(year)))
    cleared_students = cur.fetchall()

    cur.close()
    conn.close()
    return render_template('clearance.html', students=cleared_students, term=term, year=year)


@app.route('/expenditure_report', methods=['GET', 'POST'])
@require_role('admin', 'bursar')
def expenditure_report():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    filters = {"from_date": "", "to_date": ""}
    expenses = []
    summary = []
    type_summary = []
    total = 0

    if request.method == 'POST':
        filters['from_date'] = request.form.get('from_date')
        filters['to_date'] = request.form.get('to_date')

        cur.execute('''
            SELECT e.*, ec.name as category_name
            FROM expenses e
            LEFT JOIN expense_categories ec ON e.category_id = ec.id
            WHERE date_spent BETWEEN %s AND %s
            ORDER BY date_spent DESC
        ''', (filters['from_date'], filters['to_date']))
        expenses = cur.fetchall()

        total = sum([row['amount'] for row in expenses])

        cur.execute('''
            SELECT ec.name as category, SUM(e.amount) as total
            FROM expenses e
            LEFT JOIN expense_categories ec ON e.category_id = ec.id
            WHERE date_spent BETWEEN %s AND %s
            GROUP BY ec.name
            ORDER BY total DESC
        ''', (filters['from_date'], filters['to_date']))
        summary = cur.fetchall()

        cur.execute('''
            SELECT type, SUM(amount) as total
            FROM expenses
            WHERE date_spent BETWEEN %s AND %s
            GROUP BY type
            ORDER BY total DESC
        ''', (filters['from_date'], filters['to_date']))
        type_summary = cur.fetchall()

    cur.close()
    conn.close()
    return render_template('expenditure_report.html',
                           expenses=expenses, total=total,
                           filters=filters, summary=summary,
                           type_summary=type_summary)

    def _list_classes():
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT DISTINCT TRIM(class_name) AS class_name
            FROM students
            WHERE archived = 0 AND class_name IS NOT NULL AND TRIM(class_name) <> ''
            ORDER BY class_name
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return [r["class_name"] for r in rows]

    def _year_choices(base_year: int):
        # collect years seen in HP plus base±1 so the list isn't empty
        years = set()
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        try:
            for r in cur.execute("SELECT DISTINCT year FROM holiday_package ORDER BY year"):
                try:
                    years.add(int(r[0]))
                except Exception:
                    pass
        finally:
            cur.close()
            conn.close()
        if base_year:
            years.update({base_year - 1, base_year, base_year + 1})
        ys = sorted(y for y in years if y)
        # sensible fallback if db is empty
        if not ys:
            ys = [base_year] if base_year else []
        return ys

    def _list_students_for_class(class_name: str):
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, student_number, first_name, Middle_name, last_name,
                   class_name, stream
            FROM students
            WHERE archived = 0 AND LOWER(TRIM(class_name)) = LOWER(TRIM(%s))
            ORDER BY last_name, first_name
        """, (class_name,))
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return rows

    def _fetch_hp_for_class_term_year(class_name: str, term: str, year: int):
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT hp.*
            FROM holiday_package hp
            JOIN students s ON s.id = hp.student_id
            WHERE s.archived = 0
              AND LOWER(TRIM(s.class_name)) = LOWER(TRIM(%s))
              AND LOWER(TRIM(hp.term)) = LOWER(TRIM(%s))
              AND hp.year = %s
        """, (class_name, term, year))
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return {r["student_id"]: r for r in rows}

    def _upsert_hp_row(student_id: int, term: str, year: int, rowvals: dict):
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id FROM holiday_package
            WHERE student_id=%s AND LOWER(TRIM(term))=LOWER(TRIM(%s)) AND year=%s
        """, (student_id, term, year))
        existing = cur.fetchone()

        cols = ["eng", "mat", "sci", "sst", "agg", "total"]
        vals = [rowvals.get(k) for k in cols]

        if existing:
            cur.execute("""
                UPDATE holiday_package
                   SET eng=%s, mat=%s, sci=%s, sst=%s, agg=%s, total=%s
                 WHERE id=%s
            """, (*vals, existing["id"]))
        else:
            cur.execute("""
                INSERT INTO holiday_package (student_id, term, year, eng, mat, sci, sst, agg, total)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (student_id, term, year, *vals))
        conn.commit()
        cur.close()
        conn.close()

    def _to_num(x):
        try:
            if x is None or str(x).strip() == "":
                return None
            return float(str(x).replace(",", "").strip())
        except Exception:
            return None

    # ---------- filters ----------
    class_name = (request.values.get("class_name") or "").strip()
    term = (request.values.get("term") or "").strip()
    try:
        year = int(request.values.get("year") or 0)
    except ValueError:
        year = 0

    # active year for defaults & year options
    try:
        ay = get_active_academic_year()
        active_year = int(ay.get("year") or 0)
    except Exception:
        active_year = 0

    class_options = _list_classes()
    year_options = _year_choices(active_year)

    # ---------- POST: upload ----------
    if request.method == "POST":
        f = request.files.get("file")
        if not (class_name and term and year):
            flash("Select Class, Term and Year first.", "warning")
            return redirect(url_for("holiday_hub", class_name=class_name, term=term, year=year))

        if not f or not f.filename:
            flash("No file selected.", "warning")
            return redirect(url_for("holiday_hub", class_name=class_name, term=term, year=year))

        _, ext = os.path.splitext(f.filename.lower())
        if ext not in ALLOWED_EXTS_HP:
            flash("Unsupported file type. Use .xlsx, .xls, or .csv", "warning")
            return redirect(url_for("holiday_hub", class_name=class_name, term=term, year=year))

        # build map for selected class
        stu_rows = _list_students_for_class(class_name)
        sn_to_id = {r["student_number"]: r["id"] for r in stu_rows}
        saved = skipped = 0

        try:
            if ext == ".csv":
                reader = csv.DictReader(TextIOWrapper(
                    f.stream, encoding="utf-8", errors="ignore"))
                for r in reader:
                    sn = (r.get("student_number") or r.get(
                        "Student Number") or r.get("student") or "").strip()
                    sid = sn_to_id.get(sn)
                    if not sid:
                        skipped += 1
                        continue
                    _upsert_hp_row(sid, term, year, {
                        "eng": _to_num(r.get("eng")),
                        "mat": _to_num(r.get("mat")),
                        "sci": _to_num(r.get("sci")),
                        "sst": _to_num(r.get("sst")),
                        "agg": _to_num(r.get("agg")),
                        "total": _to_num(r.get("total")),
                    })
                    saved += 1
            else:
                if not load_workbook:
                    flash(
                        "openpyxl not installed; cannot read .xlsx/.xls. Use .csv or install openpyxl.", "danger")
                    return redirect(url_for("holiday_hub", class_name=class_name, term=term, year=year))

                wb = load_workbook(f, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(
                    ws.iter_rows(min_row=1, max_row=1))]
                def _idx(name): return headers.index(
                    name) if name in headers else -1
                idx = {
                    "student_number": _idx("student_number"),
                    "eng": _idx("eng"), "mat": _idx("mat"),
                    "sci": _idx("sci"), "sst": _idx("sst"),
                    "agg": _idx("agg"), "total": _idx("total"),
                }
                if idx["student_number"] < 0:
                    flash("Template must include 'student_number' column.", "danger")
                    return redirect(url_for("holiday_hub", class_name=class_name, term=term, year=year))

                for row in ws.iter_rows(min_row=2):
                    sn_cell = row[idx["student_number"]].value
                    sn = (str(sn_cell).strip() if sn_cell is not None else "")
                    if not sn:
                        continue
                    sid = sn_to_id.get(sn)
                    if not sid:
                        skipped += 1
                        continue
                    _upsert_hp_row(sid, term, year, {
                        "eng": _to_num(row[idx["eng"]].value) if idx["eng"] >= 0 else None,
                        "mat": _to_num(row[idx["mat"]].value) if idx["mat"] >= 0 else None,
                        "sci": _to_num(row[idx["sci"]].value) if idx["sci"] >= 0 else None,
                        "sst": _to_num(row[idx["sst"]].value) if idx["sst"] >= 0 else None,
                        "agg": _to_num(row[idx["agg"]].value) if idx["agg"] >= 0 else None,
                        "total": _to_num(row[idx["total"]].value) if idx["total"] >= 0 else None,
                    })
                    saved += 1

            flash(
                f"Holiday Package saved: {saved} rows. Skipped: {skipped}.", "success")
        except Exception as e:
            current_app.logger.exception(f"[HP upload] failed: {e}")
            flash(f"Upload failed: {e}", "danger")

        return redirect(url_for("holiday_hub", class_name=class_name, term=term, year=year))

    # ---------- GET: load view ----------
    students = []
    hp_rows = {}
    if class_name and term and year:
        students = _list_students_for_class(class_name)
        hp_rows = _fetch_hp_for_class_term_year(class_name, term, year)

    # fallbacks so dropdowns aren't empty
    if not class_options:
        class_options = [class_name] if class_name else []
    if not year_options:
        year_options = [year] if year else (
            [active_year] if active_year else [])

    return render_template(
        "holiday_hub.html",
        class_name=class_name,
        term=term,
        year=year or (active_year or 0),
        students=students,
        hp_rows=hp_rows,
        TERMS=TERMS,
        class_options=class_options,
        year_options=year_options,
        ALLOWED_EXTS_HP={".xlsx", ".xls", ".csv"},
    )


# ========================= REQUIREMENTS MANAGEMENT =========================
# Assumes you already have: app, require_role, get_db_connection


@app.route("/admin/requirements", methods=["GET", "POST"])
@require_role("admin", "bursar", "headteacher", "deputyheadteacher")
def admin_requirements():
    conn = get_db_connection()
    ensure_requirements_schema(conn)

    fee_groups = ["Old", "New"]
    active_year, active_term, active_term_no = get_active_year_term()

    # ---------- POST ----------
    if request.method == "POST":
        f = request.form
        rid = (f.get("id") or "").strip()

        class_name = (f.get("class_name") or "").strip()
        name = (f.get("name") or "").strip()

        qty_raw = (f.get("qty") or "").strip()
        amt_raw = (f.get("amount") or "").strip()

        year_raw = (f.get("year") or "").strip()
        term_no_raw = (f.get("term_no") or "").strip()
        fee_group = (f.get("fee_group") or "Old").strip() or "Old"
        section = (f.get("section") or "").strip() or None

        try:
            qty = int(qty_raw)
        except Exception:
            qty = 1
        try:
            amount = float(amt_raw)
        except Exception:
            amount = 0.0

        try:
            year_val = int(year_raw or active_year)
        except Exception:
            year_val = active_year

        try:
            term_no = int(term_no_raw) if term_no_raw else None
        except Exception:
            term_no = None

        if fee_group not in fee_groups:
            fee_group = "Old"

        # term_label stored in requirements.term (generated term_no depends on this)
        term_label = term_no_to_label(term_no) if term_no else None

        if not class_name or not name:
            conn.close()
            flash("Class and item name are required.", "warning")
            return redirect(url_for("admin_requirements"))

        # LOCK CHECK
        if is_academic_year_locked(year_val):
            conn.close()
            flash(f"Year {year_val} is locked. You cannot change requirements for a locked year.", "warning")
            return redirect(url_for("admin_requirements", year=year_val, fee_group=fee_group))

        cur = conn.cursor(dictionary=True)
        try:
            old_class = old_term_no = old_year = old_group = None

            if rid:
                cur.execute("""
                    SELECT class_name, year, term_no, fee_group
                    FROM requirements WHERE id=%s
                """, (rid,))
                old = cur.fetchone() or {}
                old_class = (old.get("class_name") or "").strip()
                old_year = int(old.get("year") or 0)
                old_term_no = old.get("term_no")  # can be NULL
                old_group = (old.get("fee_group") or "").strip()

                cur.execute("""
                    UPDATE requirements
                       SET class_name=%s,
                           name=%s,
                           qty=%s,
                           amount=%s,
                           term=%s,
                           year=%s,
                           fee_group=%s,
                           section=%s
                     WHERE id=%s
                """, (class_name, name, qty, amount, term_label, year_val, fee_group, section, rid))
            else:
                # Manual upsert by (class, name, year, term_no, fee_group)
                cur.execute("""
                    SELECT id
                    FROM requirements
                    WHERE class_name=%s
                      AND name=%s
                      AND year=%s
                      AND fee_group=%s
                      AND (
                           (term_no IS NULL AND %s IS NULL)
                           OR (term_no = %s)
                      )
                    LIMIT 1
                """, (class_name, name, year_val, fee_group, term_no, term_no))
                existing = cur.fetchone()

                if existing:
                    cur.execute("""
                        UPDATE requirements
                           SET qty=%s,
                               amount=%s,
                               term=%s,
                               section=%s
                         WHERE id=%s
                    """, (qty, amount, term_label, section, existing["id"]))
                else:
                    cur.execute("""
                        INSERT INTO requirements (class_name, name, qty, amount, term, year, fee_group, section)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (class_name, name, qty, amount, term_label, year_val, fee_group, section))

            conn.commit()
            flash("Requirement saved.", "success")

            # recompute new side
            _recompute_class(class_name, term_no)

            # recompute old side if moved (only if editing)
            if rid and old_class:
                moved = (
                    old_class != class_name
                    or old_year != int(year_val)
                    or (old_term_no != term_no)  # both can be None
                    or (old_group != fee_group)
                )
                if moved:
                    _recompute_class(old_class, old_term_no)

        except mysql.connector.Error as e:
            conn.rollback()
            if getattr(e, "errno", None) == 1062:
                flash("Duplicate item for this class/term/year/group.", "danger")
            else:
                flash(f"Failed to save: {e}", "danger")
        finally:
            cur.close()
            conn.close()

        return redirect(url_for("admin_requirements", year=year_val, term_no=(term_no or ""), fee_group=fee_group, class_name=class_name))

    # ---------- GET ----------
    q_class = (request.args.get("class_name") or "").strip()
    q_year = (request.args.get("year") or "").strip()
    q_term_no = (request.args.get("term_no") or "").strip()
    q_group = (request.args.get("fee_group") or "").strip()

    try:
        year_filter = int(q_year) if q_year else active_year
    except Exception:
        year_filter = active_year

    # term filter is optional; if empty, show all (including generic)
    try:
        term_no_filter = int(q_term_no) if q_term_no else None
    except Exception:
        term_no_filter = None
    if term_no_filter not in (None, 1, 2, 3):
        term_no_filter = None

    group_filter = q_group if q_group in fee_groups else "Old"
    locked = is_academic_year_locked(year_filter)

    where = ["year=%s", "fee_group=%s"]
    params = [year_filter, group_filter]

    if q_class:
        where.append("class_name=%s")
        params.append(q_class)

    if term_no_filter is None:
        # show all terms + generic
        pass
    else:
        where.append("(term_no=%s OR term_no IS NULL)")
        params.append(term_no_filter)

    cur = conn.cursor(dictionary=True)
    cur.execute(f"""
        SELECT id, class_name, name, qty, amount, term, year, term_no, fee_group, section
          FROM requirements
         WHERE {' AND '.join(where)}
         ORDER BY class_name, COALESCE(term,''), name
    """, params)
    rows = cur.fetchall() or []
    cur.close()

    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT DISTINCT class_name FROM classes ORDER BY 1")
    classes = [r["class_name"] for r in (cur.fetchall() or [])]
    cur.close()
    conn.close()

    return render_template(
        "admin_requirements.html",
        items=rows,
        classes=classes,
        terms=TERMS,
        fee_groups=fee_groups,

        default_year=year_filter,
        default_term_no=(term_no_filter or ""),
        default_fee_group=group_filter,

        q_class=q_class,
        q_year=str(year_filter),
        q_term_no=str(term_no_filter or ""),
        q_fee_group=group_filter,

        year_locked=locked,
        active_year=active_year,
        active_term_no=active_term_no
    )






@app.route("/admin/requirements/<int:rid>/delete", methods=["POST"])
@require_role("admin", "bursar", "headteacher")
def admin_requirements_delete(rid):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT class_name, year, term_no, fee_group
            FROM requirements WHERE id=%s
        """, (rid,))
        row = cur.fetchone() or {}
        cls = row.get("class_name")
        y = int(row.get("year") or 0)
        tno = row.get("term_no")  # can be None
        grp = (row.get("fee_group") or "Old")

        if y and is_academic_year_locked(y):
            flash(f"Year {y} is locked. Cannot delete requirements.", "warning")
            return redirect(url_for("admin_requirements", year=y, term_no=(tno or ""), fee_group=grp, class_name=cls))

        cur.execute("DELETE FROM requirements WHERE id=%s", (rid,))
        conn.commit()

        if cls:
            _recompute_class(cls, tno)

        flash("Requirement deleted.", "success")
        return redirect(url_for("admin_requirements", year=y, term_no=(tno or ""), fee_group=grp, class_name=cls))

    except Exception as e:
        conn.rollback()
        flash(f"Delete failed: {e}", "danger")
        return redirect(url_for("admin_requirements"))
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()

# ===================== PROMOTIONS: HISTORY + UNDO + BATCH =====================

@app.route("/promotions/hub", methods=["GET"])
@require_role("admin", "headteacher", "deputyheadteacher" , "dos", "classmanager", "teacher")
def promotions_hub():
    conn = get_db_connection()
    classes = _distinct_classes(conn)
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT DISTINCT stream FROM students WHERE stream IS NOT NULL AND TRIM(stream)<>'' ORDER BY stream"
    )
    streams = [r["stream"] for r in cur.fetchall()]

    f_class = (request.args.get("class") or "").strip()
    f_stream = (request.args.get("stream") or "").strip()
    f_status = (request.args.get("status")
                or "active").strip()  # default 'active'
    q = (request.args.get("q") or "").strip()

    sql = """
      SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS Middle_name,
             last_name, class_name, stream, archived
      FROM students
      WHERE 1=1
    """
    params = []
    if f_class:
        sql += " AND class_name=%s"
        params.append(f_class)
    if f_stream:
        sql += " AND stream=%s"
        params.append(f_stream)
    if f_status == "active":
        sql += " AND archived=0"
    elif f_status == "archived":
        sql += " AND archived=1"
    if q:
        sql += " AND (student_number = %s OR CONCAT_WS(' ', first_name, COALESCE(Middle_name,''), last_name) LIKE %s)"
        params += [q, f"%{q}%"]

    sql += " ORDER BY class_name, last_name, first_name"
    cur.execute(sql, params)
    students = cur.fetchall()
    cur.close()
    conn.close()

    return render_template("promotions_hub.html",
                           classes=classes, streams=streams,
                           preview_students=None, source_class=None, target_class=None,
                           students=students, f_class=f_class, f_stream=f_stream, f_status=f_status, q=q)


@app.route("/promotions/preview", methods=["POST"])
@require_role("admin", "headteacher", "dos", "classmanager", "deputyheadteacher", "teacher")
def promotions_preview():
    source_class = (request.form.get("source_class") or "").strip()
    target_class = (request.form.get("target_class") or "").strip()
    if not source_class or not target_class:
        flash("Please choose both Source and Target classes.", "warning")
        return redirect(url_for("promotions_hub"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, student_number,
               first_name, COALESCE(Middle_name,'') AS middle_name, last_name,
               class_name, COALESCE(stream,'') AS stream
        FROM students
        WHERE archived=0 AND class_name=%s
        ORDER BY last_name, first_name
    """, (source_class,))
    rows = cur.fetchall()
    classes = _distinct_classes(conn)
    cur.close()
    conn.close()

    return render_template(
        "promotions_hub.html",
        classes=classes,
        preview_students=rows,
        source_class=source_class,
        target_class=target_class,
        promoted_students=None,
        p7_archived_count=0
    )

# ========================= PROMOTIONS: COMMIT =========================


@app.route("/promotions/commit", methods=["POST"])
@require_role("admin", "headteacher", "dos", "deputyheadteacher")
def promotions_commit():
    conn = None
    cur = None
    try:
        # 1) Open connection first, then ensure schema
        conn = get_db_connection()
        ensure_promotions_log_schema(conn)

        # 2) Read inputs
        source_class = (request.form.get("source_class") or "").strip()
        target_class = (request.form.get("target_class") or "").strip()
        if not source_class or not target_class:
            flash("Missing source/target class.", "warning")
            return redirect(url_for("promotions_hub"))

        cur = conn.cursor(dictionary=True)

        # 3) Fetch students in source class
        cur.execute("""
            SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS Middle_name,
                   last_name, class_name, COALESCE(stream,'') AS stream
            FROM students
            WHERE archived=0 AND class_name=%s
            ORDER BY last_name, first_name
        """, (source_class,))
        students = cur.fetchall()

        if not students:
            flash(f"No students found in {source_class}.", "info")
            return redirect(url_for("promotions_hub"))

        batch_id = f"BATCH-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        actor = session.get("username") or session.get("role") or "system"

        # 4) Archive P7 leavers, or promote others
        if source_class == "P7":
            count = 0
            for s in students:
                _archive_student(
                    conn, s["id"], new_status="completed", stage="P7 Leaver")
                cur.execute("""
                    INSERT INTO promotions_log (student_id, from_class, to_class, actor, batch_id, reversed)
                    VALUES (%s, %s, %s, %s, %s, 0)
                """, (s["id"], "P7", "ARCHIVED", actor, batch_id))
                count += 1
            conn.commit()
            flash(f"Archived {count} P7 leaver(s) as completed.", "success")
            promoted_students = []
        else:
            # promote within primary
            cur.execute(
                "UPDATE students SET class_name=%s WHERE archived=0 AND class_name=%s",
                (target_class, source_class)
            )
            cur.executemany("""
                INSERT INTO promotions_log (student_id, from_class, to_class, actor, batch_id, reversed)
                VALUES (%s, %s, %s, %s, %s, 0)
            """, [(s["id"], source_class, target_class, actor, batch_id) for s in students])
            conn.commit()

            flash(
                f"Promoted {len(students)} student(s): {source_class} \u2192 {target_class}.", "success")

            # Show who’s now in the target class
            cur.execute("""
                SELECT student_number,
                       CONCAT_WS(' ', first_name, COALESCE(Middle_name, ''), last_name) AS full_name,
                       class_name, COALESCE(stream,'') AS stream
                FROM students
                WHERE archived=0 AND class_name=%s
                ORDER BY last_name, first_name
            """, (target_class,))
            promoted_students = cur.fetchall()

        # 5) Refresh class list for the page
        classes = _distinct_classes(conn)

        return render_template(
            "promotions_hub.html",
            classes=classes,
            preview_students=None,
            source_class=source_class,
            target_class=target_class,
            promoted_students=promoted_students,
            p7_archived_count=(len(students) if source_class == "P7" else 0)
        )

    except Exception as e:
        if conn:
            conn.rollback()
        current_app.logger.exception("[promotions_commit] failed")
        flash(f"Promotion failed: {e}", "danger")
        return redirect(url_for("promotions_hub"))

    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


@app.route("/promotions/batch_adjacent", methods=["POST"], endpoint="promotions_bulk_adjacent")
@require_role("admin", "headteacher", "dos", "deputyheadteacher")
def promotions_bulk_adjacent():
    src = (request.form.get("source_class") or "").strip()
    direction = (request.form.get("direction") or "").strip().lower()
    if not src or direction not in ("up", "down"):
        flash("Invalid batch move inputs.", "warning")
        return redirect(url_for("promotions_hub"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        if src == "P7" and direction == "up":
            cur.execute(
                "SELECT id FROM students WHERE archived=0 AND class_name=%s", (
                    src,)
            )
            ids = [r["id"] for r in cur.fetchall()]
            for sid in ids:
                _archive_student(
                    conn, sid, new_status="completed", stage="P7 Leaver")
            flash(f"Archived {len(ids)} P7 leaver(s).", "success")
        else:
            target = next_class_name(
                src) if direction == "up" else prev_class_name(src)
            if not target:
                flash(f"No adjacent target for '{src}'.", "warning")
                conn.close()
                return redirect(url_for("promotions_hub"))
            cur.execute(
                "UPDATE students SET class_name=%s WHERE archived=0 AND class_name=%s", (target, src))
            conn.commit()
            flash(f"Batch moved students: {src} → {target}.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Batch move failed: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("promotions_hub"))


@app.route("/promotions/schoolwide", methods=["POST"])
@require_role("admin", "headteacher", "dos", "deputyheadteacher")
def promotions_schoolwide():
    """Run a single atomic school-wide promotion (P7 archive, others move up)."""
    from datetime import datetime

    # open connection first, then ensure schema
    conn = get_db_connection()
    ensure_promotions_log_schema(conn)
    cur = conn.cursor(dictionary=True)

    try:
        # academic year + batch guard
        ay = (get_active_academic_year() or {})
        try:
            year = int(ay.get("year") or ay.get(
                "active_year") or datetime.now().year)
        except Exception:
            year = datetime.now().year

        batch_id = f"SCHOOLWIDE-{year}"
        actor = session.get("username") or session.get("role") or "system"

        ORDER = ["Baby", "Middle", "Top", "P1",
                 "P2", "P3", "P4", "P5", "P6", "P7"]

        # check if this batch already executed
        cur.execute(
            "SELECT 1 FROM promotions_log WHERE batch_id=%s LIMIT 1", (batch_id,))
        if cur.fetchone():
            flash(
                f"School-wide promotion already completed for {year}.", "info")
            return redirect(url_for("promotions_hub"))

        # begin transaction
        cur.execute("START TRANSACTION")

        # (1) Archive all P7 leavers
        cur.execute(
            "SELECT id FROM students WHERE archived=0 AND class_name='P7'")
        p7_ids = [r["id"] for r in cur.fetchall()]
        for sid in p7_ids:
            _archive_student(
                conn, sid, new_status="completed", stage="P7 Leaver")
            cur.execute("""
                INSERT INTO promotions_log (student_id, from_class, to_class, actor, batch_id, reversed)
                VALUES (%s, 'P7', 'ARCHIVED', %s, %s, 0)
            """, (sid, actor, batch_id))

        # (2) Promote lower classes upwards (P6→P7, ..., Baby→Top)
        ladder = ORDER[:-1][::-1]  # ['P6','P5','P4',...,'Baby']
        for src in ladder:
            idx = ORDER.index(src)
            target = ORDER[idx + 1]

            cur.execute(
                "SELECT id FROM students WHERE archived=0 AND class_name=%s", (src,))
            ids = [r["id"] for r in cur.fetchall()]
            if not ids:
                continue

            cur.execute(
                "UPDATE students SET class_name=%s WHERE archived=0 AND class_name=%s", (target, src))
            cur.executemany("""
                INSERT INTO promotions_log (student_id, from_class, to_class, actor, batch_id, reversed)
                VALUES (%s, %s, %s, %s, %s, 0)
            """, [(sid, src, target, actor, batch_id) for sid in ids])

        conn.commit()
        flash(
            f"School-wide promotion completed for {year}. P7 archived: {len(p7_ids)}.", "success")

    except Exception as e:
        conn.rollback()
        current_app.logger.exception("[promotions_schoolwide] failed")
        flash(f"School-wide promotion failed: {e}", "danger")
    finally:
        try:
            cur.close()
            conn.close()
        except Exception:
            pass

    return redirect(url_for("promotions_hub"))


# ========================= PROMOTIONS: HISTORY & REVERSE =========================

@app.route("/promotions/history", methods=["GET"])
@require_role("admin", "headteacher", "dos", "bursar", "deputyheadteacher", "teacher", "classmanager")
def promotions_history():
    # open connection first, then ensure schema
    conn = get_db_connection()
    ensure_promotions_log_schema(conn)

    q_student = (request.args.get("q") or "").strip()
    q_from = (request.args.get("from_class") or "").strip()
    q_to = (request.args.get("to_class") or "").strip()
    q_batch = (request.args.get("batch_id") or "").strip()

    cur = conn.cursor(dictionary=True)
    classes = _distinct_classes(conn)

    sql = """
      SELECT p.*,
             s.student_number,
             CONCAT_WS(' ', s.first_name, COALESCE(s.Middle_name, ''), s.last_name) AS full_name
      FROM promotions_log p
      JOIN students s ON s.id = p.student_id
      WHERE 1=1
    """
    args = []

    if q_student:
        sql += """
          AND (
                s.student_number = %s
             OR CONCAT_WS(' ', s.first_name, COALESCE(s.Middle_name, ''), s.last_name) LIKE %s
          )
        """
        args += [q_student, f"%{q_student}%"]

    if q_from:
        sql += " AND p.from_class = %s"
        args.append(q_from)

    if q_to:
        sql += " AND p.to_class = %s"
        args.append(q_to)

    if q_batch:
        sql += " AND p.batch_id = %s"
        args.append(q_batch)

    sql += " ORDER BY p.created_at DESC LIMIT 500"

    cur.execute(sql, args)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return render_template(
        "promotions_history.html",
        rows=rows,
        classes=classes,
        q_student=q_student, q_from=q_from, q_to=q_to, q_batch=q_batch
    )



@app.route("/promotions/demote/<int:student_id>", methods=["POST"])
@require_role("admin", "headteacher", "dos", "classmanager", "deputyheadteacher")
def promotions_demote(student_id: int):
    conn = get_db_connection() # create conn first
    ensure_promotions_log_schema(conn) # then ensure schema

    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT id, from_class, to_class
            FROM promotions_log
            WHERE student_id=%s
            ORDER BY created_at DESC
            LIMIT 1
        """, (student_id,))
        last = cur.fetchone()

        if not last:
            flash("No promotion history for this student.", "info")
            return redirect(request.referrer or url_for("promotions_history"))

        raw_from = last["from_class"] or ""
        raw_to = last["to_class"] or ""

        # Normalise to safe ENUM values
        from_enum = normalize_class_enum(raw_from)
        to_enum = normalize_class_enum(raw_to)

        if raw_to == "ARCHIVED":
            # Unarchive P7 leaver back to their former class
            if not from_enum:
                flash(
                    f"Cannot demote: invalid previous class '{raw_from}' stored in log.",
                    "danger",
                )
                return redirect(request.referrer or url_for("promotions_history"))

            _unarchive_student(conn, student_id)
            cur.execute(
                "UPDATE students SET class_name=%s, status='active' WHERE id=%s",
                (from_enum, student_id),
            )
        else:
            # Step back one class
            if not from_enum:
                flash(
                    f"Cannot demote: invalid previous class '{raw_from}' stored in log.",
                    "danger",
                )
                return redirect(request.referrer or url_for("promotions_history"))

            cur.execute(
                "UPDATE students SET class_name=%s, status='active' WHERE id=%s",
                (from_enum, student_id),
            )

        actor = session.get("username") or session.get("role") or "system"

        # Log the UNDO using clean enums where possible
        log_from = raw_to if raw_to == "ARCHIVED" else (to_enum or raw_to)
        log_to = from_enum or raw_from

        cur.execute("""
            INSERT INTO promotions_log (student_id, from_class, to_class, actor, batch_id, reversed)
            VALUES (%s, %s, %s, %s, %s, 1)
        """, (student_id, log_from, log_to, actor, f"UNDO-{last['id']}"))

        conn.commit()
        flash(f"Reverted. Student returned to {log_to}.", "success")

    except Exception as e:
        conn.rollback()
        current_app.logger.exception("[promotions_demote] failed")
        flash(f"Demotion failed: {e}", "danger")
    finally:
        cur.close()
        conn.close()

    return redirect(request.referrer or url_for("promotions_history"))


# =================== ARCHIVE / UNARCHIVE ROUTES ===================


@app.route("/students/<int:student_id>/archive", methods=["POST"])
@require_role("admin", "headteacher","bursar")
def archive_student(student_id: int):
    """Archive a single student (any class)."""
    conn = get_db_connection()
    try:
        changed = _archive_student(
            conn,
            student_id,
            new_status="completed",
            stage="Manual Archive"
        )
        flash("Student archived." if changed else "No change.",
              "success" if changed else "info")

        # ---- AUDIT ----
        audit_from_request(
            conn,
            action="student_archive",
            outcome="success" if changed else "warning",
            severity="info" if changed else "warning",
            target_table="students",
            target_id=student_id,
            details={"changed": bool(changed), "stage": "Manual Archive"}
        )
    except Exception as e:
        conn.rollback()
        current_app.logger.exception("[archive_student] failed")
        flash(f"Archive failed: {e}", "danger")
        audit_from_request(
            conn,
            action="student_archive",
            outcome="failure",
            severity="warning",
            target_table="students",
            target_id=student_id,
            details={"error": str(e)}
        )
    finally:
        conn.close()
    return redirect(request.referrer or url_for("archive_hub"))


@app.route("/students/<int:student_id>/unarchive", methods=["POST"])
@require_role("admin", "headteacher", "bursar")
def unarchive_student(student_id: int):
    conn = get_db_connection()
    try:
        ensure_archived_students_table(conn)  # safe no-op if exists
        _unarchive_student(conn, student_id, remove_archive_rows=True)
        flash("Student restored.", "success")

        # ---- AUDIT ----
        audit_from_request(
            conn,
            action="student_unarchive",
            target_table="students",
            target_id=student_id,
            details={"remove_archive_rows": True}
        )
    except Exception as e:
        conn.rollback()
        flash(f"Unarchive failed: {e}", "danger")
        audit_from_request(
            conn,
            action="student_unarchive",
            outcome="failure",
            severity="warning",
            target_table="students",
            target_id=student_id,
            details={"error": str(e)}
        )
    finally:
        conn.close()
    return redirect(request.referrer or url_for("archive_hub"))


@app.route("/students/<int:student_id>/delete", methods=["POST"])
@require_role("admin", "headteacher", "bursar")
def delete_student(student_id: int):
    """Treat 'delete' as soft-archive to avoid FK errors and fill archive list."""
    conn = get_db_connection()
    try:
        _archive_student(conn, student_id,
                         new_status="completed", stage="Manual Archive")
        flash("Student archived (soft-deleted).", "success")
    except Exception as e:
        conn.rollback()
        current_app.logger.exception("[delete_student] failed")
        flash(f"Archive failed: {e}", "danger")
    finally:
        conn.close()
    return redirect(request.referrer or url_for("archive_hub"))


@app.route("/archive", methods=["GET"])
@require_role("admin", "headteacher", "dos")
def archive_hub():
    conn = get_db_connection()
    ensure_archived_students_table(conn)

    q_class = (request.args.get("class") or "").strip()
    q_sn = (request.args.get("student_number") or "").strip()
    q_year = (request.args.get("year") or "").strip()
    q_lname = (request.args.get("last_name") or "").strip()

    sql = "SELECT * FROM archived_students WHERE 1=1"
    params = []
    if q_class:
        sql += " AND class_name = %s"
        params.append(q_class)
    if q_sn:
        sql += " AND student_number = %s"
        params.append(q_sn)
    if q_year:
        try:
            sql += " AND year_completed = %s"
            params.append(int(q_year))
        except ValueError:
            pass
    if q_lname:
        sql += " AND full_name LIKE %s"
        params.append(f"%{q_lname}%")

    sql += " ORDER BY year_completed DESC, class_name, full_name LIMIT 1000"

    cur = conn.cursor(dictionary=True)
    cur.execute(sql, params)
    rows = cur.fetchall()

    # dropdowns
    try:
        classes = ORDER[:]  # if global ORDER exists
    except NameError:
        classes = []

    # ALWAYS compute years
    years = []
    cur.execute(
        "SELECT DISTINCT class_name FROM archived_students ORDER BY class_name")
    if not classes:
        classes = [r["class_name"] for r in cur.fetchall()]
    else:
        cur.fetchall()  # drain cursor if needed

    cur.execute(
        "SELECT DISTINCT year_completed FROM archived_students ORDER BY year_completed DESC")
    years = [r["year_completed"] for r in cur.fetchall()]

    cur.close()
    conn.close()

    return render_template(
        "archive_hub.html",
        rows=rows, classes=classes, years=years,
        q_class=q_class, q_sn=q_sn, q_year=q_year, q_lname=q_lname
    )


@app.route("/admin/class_fees", methods=["GET", "POST"])
@require_role("admin", "headteacher", "bursar")
def admin_class_fees():
    conn = get_db_connection()
    try:
        ensure_class_fees_schema(conn)

        class_options = ["Baby", "Middle", "Top", "P1", "P2", "P3", "P4", "P5", "P6", "P7"]
        section_options = ["Day", "Boarding"]
        fee_groups = ["Old", "New"]

        active_year, active_term, active_term_no = get_active_year_term()

        # ---------- POST ----------
        if request.method == "POST":
            f = request.form

            raw_class = (f.get("class_name") or "").strip()
            raw_section = (f.get("section") or "").strip()
            raw_level = (f.get("level") or "").strip() or None
            raw_amount = (f.get("amount") or "").strip()

            year_raw = (f.get("year") or "").strip()
            term_no_raw = (f.get("term_no") or "").strip()
            fee_group = (f.get("fee_group") or "Old").strip() or "Old"

            class_name = norm_class(raw_class) or (raw_class or "").title()
            section = norm_section(raw_section)
            level = raw_level

            try:
                amount = float(raw_amount)
            except Exception:
                amount = None

            try:
                year_val = int(year_raw or active_year)
            except Exception:
                year_val = active_year

            try:
                term_no = int(term_no_raw or active_term_no)
            except Exception:
                term_no = active_term_no
            if term_no not in (1, 2, 3):
                term_no = active_term_no

            if fee_group not in fee_groups:
                fee_group = "Old"

            # LOCK CHECK
            if is_academic_year_locked(year_val):
                flash(f"Year {year_val} is locked. You cannot change fees for a locked year.", "warning")
                return redirect(url_for("admin_class_fees", year=year_val, term_no=term_no, fee_group=fee_group))

            if not class_name or not section or amount is None:
                flash("Please provide Class, Section, Amount, Year, Term, and Fee Group.", "warning")
                return redirect(url_for("admin_class_fees", year=year_val, term_no=term_no, fee_group=fee_group))

            cur = conn.cursor(dictionary=True)
            try:
                # Manual upsert by (class,section,year,term_no,group)
                cur.execute("""
                    SELECT id
                    FROM class_fees
                    WHERE class_name=%s AND section=%s AND year=%s AND term_no=%s AND fee_group=%s
                    LIMIT 1
                """, (class_name, section, year_val, term_no, fee_group))
                existing = cur.fetchone()

                if existing:
                    cur.execute("""
                        UPDATE class_fees
                           SET level=%s, amount=%s
                         WHERE id=%s
                    """, (level, amount, existing["id"]))
                else:
                    cur.execute("""
                        INSERT INTO class_fees (class_name, section, level, amount, year, term_no, fee_group)
                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                    """, (class_name, section, level, amount, year_val, term_no, fee_group))

                conn.commit()
                flash("Class fee saved.", "success")

                # recompute affected class/term only
                _recompute_class(class_name, term_no)

            except mysql.connector.Error as e:
                conn.rollback()
                flash(f"Failed to save: {e}", "danger")
            finally:
                cur.close()

            return redirect(url_for("admin_class_fees", year=year_val, term_no=term_no, fee_group=fee_group))

        # ---------- GET ----------
        q_year = (request.args.get("year") or "").strip()
        q_term_no = (request.args.get("term_no") or "").strip()
        q_group = (request.args.get("fee_group") or "").strip()

        try:
            year_filter = int(q_year) if q_year else active_year
        except Exception:
            year_filter = active_year

        try:
            term_no_filter = int(q_term_no) if q_term_no else active_term_no
        except Exception:
            term_no_filter = active_term_no
        if term_no_filter not in (1, 2, 3):
            term_no_filter = active_term_no

        fee_group_filter = q_group if q_group in fee_groups else "Old"

        locked = is_academic_year_locked(year_filter)

        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, class_name, section, level, amount, year, term_no, fee_group
            FROM class_fees
            WHERE year=%s AND term_no=%s AND fee_group=%s
            ORDER BY
              CASE class_name
                WHEN 'Baby' THEN 0 WHEN 'Middle' THEN 1 WHEN 'Top' THEN 2
                WHEN 'P1' THEN 3 WHEN 'P2' THEN 4 WHEN 'P3' THEN 5
                WHEN 'P4' THEN 6 WHEN 'P5' THEN 7 WHEN 'P6' THEN 8 WHEN 'P7' THEN 9
                ELSE 99
              END, section
        """, (year_filter, term_no_filter, fee_group_filter))
        fees = cur.fetchall() or []
        cur.close()

        return render_template(
            "admin_class_fees.html",
            class_options=class_options,
            section_options=section_options,
            fee_groups=fee_groups,
            fees=fees,

            default_year=year_filter,
            default_term_no=term_no_filter,
            default_fee_group=fee_group_filter,

            q_year=str(year_filter),
            q_term_no=str(term_no_filter),
            q_fee_group=fee_group_filter,

            year_locked=locked,
            active_year=active_year,
            active_term_no=active_term_no
        )
    finally:
        conn.close()







@app.route("/admin/class_fees/<int:fee_id>/delete", methods=["POST"])
@require_role("admin", "headteacher", "bursar")
def delete_class_fee(fee_id):
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT class_name, year, term_no, fee_group FROM class_fees WHERE id=%s", (fee_id,))
        row = cur.fetchone() or {}
        cls = row.get("class_name")
        y = int(row.get("year") or 0)
        tno = int(row.get("term_no") or 1)
        grp = (row.get("fee_group") or "Old")

        if y and is_academic_year_locked(y):
            flash(f"Year {y} is locked. Cannot delete fees.", "warning")
            return redirect(url_for("admin_class_fees", year=y, term_no=tno, fee_group=grp))

        cur.execute("DELETE FROM class_fees WHERE id=%s", (fee_id,))
        conn.commit()
        cur.close()

        if cls:
            _recompute_class(cls, tno)

        flash("Fee deleted.", "success")
        return redirect(url_for("admin_class_fees", year=y, term_no=tno, fee_group=grp))

    except Exception as e:
        conn.rollback()
        flash(f"Failed to delete: {e}", "danger")
        return redirect(url_for("admin_class_fees"))
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


# ---- Bursaries: list/add/export/import/edit/delete ----


@app.route("/admin/bursaries", methods=["GET", "POST"])
@require_role("admin", "bursar", "headteacher")
def bursaries():
    # ---------- POST: create/update ----------
    if request.method == "POST":
        conn = get_db_connection()
        try:
            ensure_bursaries_schema(conn)

            f = request.form
            # find student (by SN or last name)
            student_number = (f.get("student_number") or "").strip()
            last_name = (f.get("last_name") or "").strip()
            student = _find_student_by_sn_or_ln(student_number, last_name)
            if not student:
                flash("Student not found. Use Student Number or Last Name.", "warning")
                return redirect(url_for("bursaries"))

            # NOT NULL sponsor (empty string is fine and will collide for upsert)
            sponsor_name = (f.get("sponsor_name") or "").strip()

            # year & amount
            ay = get_active_academic_year() or {}
            default_year = int(ay.get("year") or ay.get(
                "active_year") or datetime.now().year)
            try:
                year_val = int(f.get("year") or default_year)
            except ValueError:
                year_val = default_year

            amount_raw = (f.get("amount") or "").strip()
            try:
                amount = float(amount_raw)
            except ValueError:
                flash("Amount must be numeric.", "danger")
                return redirect(url_for("bursaries"))

            # which term(s)
            apply_to = (f.get("apply_to") or "one").lower()
            if apply_to == "year":
                terms_to_apply = TERMS[:]  # all three
            elif apply_to == "two":
                pair_key = (f.get("term_pair") or "").strip()
                pair_map = {
                    "t1_t2": ["Term 1", "Term 2"],
                    "t1_t3": ["Term 1", "Term 3"],
                    "t2_t3": ["Term 2", "Term 3"],
                }
                terms_to_apply = pair_map.get(pair_key, [])
                if not terms_to_apply:
                    flash("Choose a valid two-term pair.", "warning")
                    return redirect(url_for("bursaries"))
            else:
                term_one = (f.get("term_one") or "").strip()
                if term_one in TERMS:
                    terms_to_apply = [term_one]
                else:
                    flash("Choose a valid term.", "warning")
                    return redirect(url_for("bursaries"))

            # Upsert for each term
            cur = conn.cursor(dictionary=True)
            for t in terms_to_apply:
                cur.execute("""
                    INSERT INTO bursaries (student_id, sponsor_name, amount, term, year)
                    VALUES (%s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                        amount = VALUES(amount),
                        sponsor_name = VALUES(sponsor_name)
                """, (student["id"], sponsor_name, amount, t, year_val))
            conn.commit()
            cur.close()
            flash(f"Bursary saved for {', '.join(terms_to_apply)}.", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Failed to save bursary: {e}", "danger")
        finally:
            conn.close()

        return redirect(url_for("bursaries"))

    # ---------- GET: list + filters ----------
    q_sn = (request.args.get("student_number") or "").strip()
    q_ln = (request.args.get("last_name") or "").strip()
    q_year = (request.args.get("year") or "").strip()
    q_term = (request.args.get("term") or "").strip()

    where, params = ["1=1"], []
    if q_year:
        where.append("b.year = %s")
        params.append(int(q_year))
    if q_term:
        where.append("b.term = %s")
        params.append(q_term)
    if q_sn:
        where.append("s.student_number = %s")
        params.append(q_sn)
    if q_ln:
        where.append("s.last_name LIKE %s")
        params.append(f"%{q_ln}%")

    conn = get_db_connection()
    try:
        ensure_bursaries_schema(conn)
        cur = conn.cursor(dictionary=True)
        cur.execute(f"""
            SELECT b.id, b.student_id, b.sponsor_name, b.amount, b.term, b.year,
                   s.student_number, s.first_name, s.Middle_name, s.last_name, s.class_name, s.stream
            FROM bursaries b
            JOIN students s ON s.id = b.student_id
            WHERE {' AND '.join(where)}
            ORDER BY b.year DESC,
                     CASE b.term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 99 END,
                     s.last_name, s.first_name
        """, params)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    # year default for form
    ay = get_active_academic_year() or {}
    default_year = int(ay.get("year") or ay.get(
        "active_year") or datetime.now().year)

    return render_template(
        "admin_bursaries.html",
        bursaries=rows,
        terms=TERMS,
        default_year=default_year,
        q_sn=q_sn, q_ln=q_ln, q_year=q_year, q_term=q_term
    )


@app.route("/admin/bursaries/<int:bid>/delete", methods=["POST"])
@require_role("admin", "bursar", "headteacher")
def delete_bursary(bid):
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("DELETE FROM bursaries WHERE id=%s", (bid,))
        conn.commit()
        cur.close()
        flash("Bursary entry deleted.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Delete failed: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("bursaries"))


@app.route("/admin/bursaries/<int:bid>/update", methods=["POST"])
@require_role("admin", "bursar", "headteacher")
def update_bursary(bid):
    sponsor_name = (request.form.get("sponsor_name") or "").strip() or None
    term = (request.form.get("term") or "").strip()
    year_raw = (request.form.get("year") or "").strip()
    amount_raw = (request.form.get("amount") or "").strip()

    if term not in TERMS:
        flash("Choose a valid term.", "warning")
        return redirect(url_for("bursaries"))
    try:
        year = int(year_raw)
        amount = float(amount_raw)
    except ValueError:
        flash("Year and amount must be numeric.", "warning")
        return redirect(url_for("bursaries"))

    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            UPDATE bursaries
            SET sponsor_name = %s, amount = %s, term = %s, year = %s
            WHERE id = %s
        """, (sponsor_name, amount, term, year, bid))
        conn.commit()
        cur.close()
        flash("Bursary updated.", "success")
    except mysql.connector.Error:
        flash("That change duplicates an existing bursary for the student/term/year/sponsor.", "danger")
    except Exception as e:
        conn.rollback()
        flash(f"Update failed: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("bursaries"))

# ---- Export / Sample / Import ----


@app.route("/admin/bursaries/export")
@require_role("admin", "bursar", "headteacher")
def bursaries_export():
    q_sn = (request.args.get("student_number") or "").strip()
    q_ln = (request.args.get("last_name") or "").strip()
    q_year = (request.args.get("year") or "").strip()
    q_term = (request.args.get("term") or "").strip()

    where, params = ["1=1"], []
    if q_year:
        where.append("b.year=%s")
        params.append(int(q_year))
    if q_term:
        where.append("b.term=%s")
        params.append(q_term)
    if q_sn:
        where.append("s.student_number=%s")
        params.append(q_sn)
    if q_ln:
        where.append("s.last_name LIKE %s")
        params.append(f"%{q_ln}%")

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(f"""
        SELECT s.student_number, s.first_name, s.Middle_name, s.last_name,
               s.class_name, s.stream,
               b.term, b.year, b.sponsor_name, b.amount
        FROM bursaries b
        JOIN students s ON s.id=b.student_id
        WHERE {' AND '.join(where)}
        ORDER BY b.year DESC,
                 CASE b.term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 99 END,
                 s.last_name, s.first_name
    """, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["student_number", "first_name", "Middle_name", "last_name",
                     "class_name", "stream", "term", "year", "sponsor_name", "amount"])
    for r in rows:
        writer.writerow([
            r["student_number"], r["first_name"], r["Middle_name"], r["last_name"],
            r["class_name"], r["stream"], r["term"], r["year"], r["sponsor_name"] or "", r["amount"] or 0
        ])
    mem = io.BytesIO(out.getvalue().encode("utf-8"))
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(mem, mimetype="text/csv", as_attachment=True,
                     download_name=f"bursaries_{ts}.csv")


@app.route("/admin/bursaries/sample")
@require_role("admin", "bursar", "headteacher")
def bursaries_sample():
    headers = ["student_number", "last_name", "year", "amount",
               "sponsor_name", "terms", "apply_to", "term_one", "term_two"]
    sample = [
        ["P7-2025-0001", "", 2025, 150000, "Sponsor A", "Term 1, Term 2", "", "", ""],
        ["", "Okello", 2025, 120000, "PTA", "", "year", "", ""],
        ["P6-2025-0003", "", 2025, 80000, "", "", "two", "Term 1", "Term 3"],
    ]
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(headers)
    writer.writerows(sample)
    mem = io.BytesIO(out.getvalue().encode("utf-8"))
    return send_file(mem, mimetype="text/csv", as_attachment=True,
                     download_name="bursaries_sample.csv")


try:
    import pandas as pd
except Exception:
    pd = None

@app.route("/admin/bursaries/import", methods=["POST"])
@require_role("admin", "bursar", "headteacher")
def bursaries_import():
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("Choose a CSV or Excel file.", "warning")
        return redirect(url_for("bursaries"))

    filename = file.filename.lower()
    rows = []
    try:
        if filename.endswith(".csv"):
            text = file.read().decode("utf-8", errors="ignore")
            rows = list(csv.DictReader(io.StringIO(text)))
        elif (filename.endswith(".xlsx") or filename.endswith(".xls")) and pd is not None:
            df = pd.read_excel(file)
            rows = df.to_dict(orient="records")
        else:
            text = file.read().decode("utf-8", errors="ignore")
            rows = list(csv.DictReader(io.StringIO(text)))
    except Exception as e:
        flash(f"Could not read file: {e}", "danger")
        return redirect(url_for("bursaries"))

    errors, processed = [], 0
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    def _expand_terms_from_row(row: dict):
        def norm(t):
            t = (t or "").strip()
            return t if t in TERMS else None
        terms_mult = row.get("terms") or row.get(
            "Terms") or row.get("TERMS") or ""
        if terms_mult:
            parts = [p.strip()
                     for p in terms_mult.replace(";", ",").split(",")]
            expanded = [p for p in (norm(p) for p in parts) if p]
            if expanded:
                return expanded
        single = norm(row.get("term") or row.get("Term") or row.get("TERM"))
        if single:
            return [single]
        apply_to = (row.get("apply_to") or row.get(
            "Apply_To") or "").lower().strip()
        if apply_to == "year":
            return TERMS[:]
        if apply_to == "two":
            t1 = norm(row.get("term_one") or row.get("Term_One"))
            t2 = norm(row.get("term_two") or row.get("Term_Two"))
            return [t for t in (t1, t2) if t]
        return []

    for idx, row in enumerate(rows, start=2):
        sn = (row.get("student_number") or "").strip()
        ln = (row.get("last_name") or "").strip()
        sponsor = (row.get("sponsor_name") or "").strip()

        try:
            year = int(str(row.get("year") or "").strip())
            amount = float(str(row.get("amount") or "").strip())
        except Exception:
            errors.append(f"Row {idx}: invalid year/amount")
            continue

        student = _find_student_by_sn_or_ln(sn, ln)
        if not student:
            errors.append(
                f"Row {idx}: student not found (SN='{sn}' LN='{ln}')")
            continue

        terms_to_apply = _expand_terms_from_row(row)
        if not terms_to_apply:
            errors.append(f"Row {idx}: no valid term(s)")
            continue

        try:
            for t in terms_to_apply:
                cur.execute("""
                    INSERT INTO bursaries (student_id, sponsor_name, amount, term, year)
                    VALUES (%s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                    amount = VALUES(amount),
                    sponsor_name=VALUES(sponsor_name) 
                """, (student["id"], sponsor, amount, t, year))
                processed += 1
        except Exception as e:
            errors.append(f"Row {idx}: {e}")

    try:
        conn.commit()
    except Exception as e:
        conn.rollback()
        errors.append(f"Commit failed: {e}")
    finally:
        cur.close()
        conn.close()

    msg = f"Import complete; Processed: {processed} entries."
    if errors:
        msg += (" Some issues: " +
                "; ".join(errors[:6]) + ("..." if len(errors) > 6 else ""))
        flash(msg, "warning")
    else:
        flash(msg, "success")
    return redirect(url_for("bursaries")) 


@app.route("/admin/fix_fees", methods=["POST"])
@require_role("admin", "headteacher", "bursar", "clerk")
def run_fix_fees():
    conn = get_db_connection()
    try:
        # get current term/year from your helper
        # must return {'term': 'Term 3', 'year': 2025} (names you use)
        ay = get_active_academic_year()
        term = ay.get("term") or ay.get("term_name") or "Term 1"
        year = int(ay.get("year") or ay.get("year_number"))

        # 1) create missing fee rows
        created = ensure_fee_rows_for_all(conn, term, year)

        # 2) recalc all fees (school_fees/fees)
        updated = _recalc_all_fees(conn)

        flash(
            f"Fee records refreshed: {updated} updated; {created} created.", "success")
    except Exception as e:
        conn.rollback()
        app.logger.exception("run_fix_fees()")
        flash(f"Fix failed: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("admin_class_fees"))


@app.route("/finance", methods=["GET", "POST"])
@require_role("admin", "bursar", "headteacher", "director")
def finance_hub():
    ay = get_active_academic_year()
    f = _parse_finance_filters(request, ay)
    fees_rows, req_rows, other_rows, exp_rows, totals = _fetch_finance_data(f)
    snapshot = _balance_sheet_snapshot(f)
    return render_template(
        "finance_hub.html",
        terms=TERMS,
        filters=f,
        fees_rows=fees_rows,
        req_rows=req_rows,
        other_rows=other_rows,
        exp_rows=exp_rows,
        totals=totals,
        snapshot=snapshot
    )


# ========================= FINANCE EXPORT =========================

@app.route("/finance/export", methods=["GET"])
@require_role("admin", "bursar", "headteacher", "director")
def finance_export():
    import openpyxl
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    ay = get_active_academic_year()
    f = _parse_finance_filters(request, ay)
    view = (request.args.get("view") or "fees").strip().lower()

    # Reuse the same datasets as the hub
    fees_rows, req_rows, other_rows, exp_rows, totals = _fetch_finance_data(f)
    snapshot = _balance_sheet_snapshot(f)

    wb = openpyxl.Workbook()
    ws = wb.active

    def autosize(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    def sheet(title, headers, rows):
        nonlocal ws
        ws.title = title[:31] or "Sheet1"
        ws.append(headers)
        for r in rows:
            ws.append(r)
        autosize(ws)

    # ---------------- Sheets ----------------
    if view == "fees":
        headers = ["Date", "Student No.", "Full Name", "Method", "Term", "Year", "Amount (UGX)"]
        data = [
            [r["date_paid"], r["student_number"], r["full_name"], r["method"], r["term"], r["year"], r["amount_paid"]]
            for r in fees_rows
        ]
        sheet("Fees", headers, data)

    elif view == "requirements":
        headers = ["Date", "Student No.", "Full Name", "Item", "Method", "Term", "Year", "Amount (UGX)"]
        data = [
            [r["date_paid"], r["student_number"], r["full_name"], r["requirement_name"],
             r["method"], r["term"], r["year"], r["amount_paid"]]
            for r in req_rows
        ]
        sheet("Requirements", headers, data)

    elif view == "other":
        headers = ["Date", "Source", "Description", "Recorded By", "Term", "Year", "Amount (UGX)"]
        data = [
            [r["date_received"], r["source"], r["description"], r["recorded_by"],
             r["term"], r["year"], r["amount"]]
            for r in other_rows
        ]
        sheet("Other Income", headers, data)

    elif view == "expenses":
        headers = ["Date", "Category", "Description", "Type", "Recorded By", "Term", "Year", "Amount (UGX)"]
        data = [
            [r["date_spent"], r["category"], r["description"], r["type"], r["recorded_by"],
             r["term"], r["year"], r["amount"]]
            for r in exp_rows
        ]
        sheet("Expenses", headers, data)

    elif view == "income_statement":
        headers = ["Account", "Amount (UGX)"]
        data = [
            ["Fees Income", totals["fees_total"]],
            ["Requirements Income", totals["requirements_total"]],
            ["Other Income", totals["other_income_total"]],
            ["Total Income", totals["income_total"]],
            ["Total Expenses", totals["expenses_total"]],
            ["Net (Income - Expenses)", totals["net_total"]],
        ]
        sheet("Income Statement", headers, data)

    elif view == "balance_sheet":
        headers = ["Account", "Amount (UGX)"]
        data = [
            ["Cash & Cash Equivalents (period, ALL)", snapshot["cash_in_all"]],
            ["Cash & Cash Equivalents (period, Active)", snapshot["cash_in_active"]],
            ["Expenses (period)", -snapshot["expenses"]],
            ["Accounts Receivable (ALL-Time, ALL)", snapshot["receivables_all"]],
            ["Accounts Receivable (ALL-Time, Active)", snapshot["receivables_active"]],
            ["Net Position (ALL)", snapshot["net_all"]],
            ["Net Position (Active)", snapshot["net_active"]],
        ]
        sheet("Balance Sheet", headers, data)

    # ---------------- Download ----------------
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    label = f["use_dates"] and f'{f["from_date"]}_to_{f["to_date"]}' or f'{f["term"]}_{f["year"]}'
    filename = f"{view}_{label}.xlsx".replace(" ", "").replace(":", "-")
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )



# ---- EMPLOYEES HUB ----
@app.route("/employees", methods=["GET", "POST"], endpoint="employees_hub")
@require_role("admin", "director", "headteacher", "bursar", "dos", "deputyheadteacher")
def employees_hub():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Add new employee
    if request.method == "POST":
        data = {
            "first_name": (request.form.get("first_name") or "").strip(),
            "middle_name": (request.form.get("middle_name") or "").strip(),
            "last_name": (request.form.get("last_name") or "").strip(),
            "gender": request.form.get("gender"),
            "contact": request.form.get("contact"),
            "email": request.form.get("email"),
            "residence": request.form.get("residence"),
            "department": request.form.get("department"),
            "designation": request.form.get("designation"),
            "hire_date": request.form.get("hire_date"),
            "status": request.form.get("status") or "active",
            "base_salary": float(request.form.get("base_salary") or 0),
            "allowance": float(request.form.get("allowance") or 0),
            "bonus": float(request.form.get("bonus") or 0),
            "pay_cycle": request.form.get("pay_cycle") or "monthly",
            "bank_name": request.form.get("bank_name"),
            "bank_account": request.form.get("bank_account"),
            "tin": request.form.get("tin"),
            "notes": request.form.get("notes"),
        }
        if not data["first_name"] or not data["last_name"]:
            conn.close()
            flash("First and last name are required.", "warning")
            return redirect(url_for("employees_hub"))

        cur.execute("""
            INSERT INTO employees (
              first_name, middle_name, last_name, gender, contact, email, residence,
              department, designation, hire_date, status, base_salary, allowance, bonus,
              pay_cycle, bank_name, bank_account, tin, notes
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, tuple(data.values()))
        conn.commit()
        conn.close()
        flash("Employee added.", "success")
        return redirect(url_for("employees_hub"))

    # Search/list
    q = (request.args.get("q") or "").strip()
    if q:
        cur.execute("""
            SELECT * FROM employees
            WHERE first_name LIKE %s OR last_name LIKE %s
               OR middle_name LIKE %s OR designation LIKE %s
               OR department LIKE %s
            ORDER BY last_name, first_name
        """, (f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"))
        rows = cur.fetchall()
    else:
        cur.execute("""
            SELECT * FROM employees
            ORDER BY (status='active') DESC, last_name, first_name
        """)
        rows = cur.fetchall()

    cur.close()
    conn.close()
    return render_template("employees_hub.html", employees=rows, q=q)


# ---- EDIT EMPLOYEE ----
@app.route("/employees/<int:eid>/edit", methods=["GET", "POST"], endpoint="employee_edit")
@require_role("admin", "director", "bursar","deputyheadteacher")
def employee_edit(eid):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM employees WHERE id=%s", (eid,))
    emp = cur.fetchone()
    if not emp:
        conn.close()
        flash("Employee not found.", "warning")
        return redirect(url_for("employees_hub"))

    if request.method == "POST":
        fields = [
            "first_name", "middle_name", "last_name", "gender", "contact", "email", "residence",
            "department", "designation", "hire_date", "status", "base_salary", "allowance", "bonus",
            "pay_cycle", "bank_name", "bank_account", "tin", "notes"
        ]
        values = [request.form.get(k) for k in fields]
        # numeric cleanup
        for i, k in enumerate(fields):
            if k in ("base_salary", "allowance", "bonus"):
                try:
                    values[i] = float(values[i] or 0)
                except:
                    values[i] = 0.0

        set_clause = ", ".join([f"{k}=%s" for k in fields])
        cur.execute(
            f"UPDATE employees SET {set_clause} WHERE id=%s", (*values, eid))
        conn.commit()
        conn.close()
        flash("Employee updated.", "success")
        return redirect(url_for("employees_hub"))

    cur.close()
    conn.close()
    return render_template("employee_edit.html", emp=emp)


# ---- (Optional) DELETE ----
@app.route("/employees/<int:eid>/delete", methods=["POST"], endpoint="employee_delete")
@require_role("admin", "director","bursar")
def employee_delete(eid):
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)
        # First try a hard delete
        cur.execute("DELETE FROM employees WHERE id=%s", (eid,))
        conn.commit()
        cur.close()
        flash("Employee deleted.", "info")
    except IntegrityError as e:
        # FK violation? fallback to archive if supported
        if getattr(e, "errno", None) == 1451:
            cur = conn.cursor(dictionary=True)
            # Check if 'archived' column exists
            cur.execute("SHOW COLUMNS FROM employees LIKE 'archived'")
            has_archived = bool(cur.fetchone())
            if has_archived:
                cur.execute("UPDATE employees SET archived=1 WHERE id=%s", (eid,))
                conn.commit()
                cur.close()
                flash("Employee has related payroll records. Marked as archived instead.", "warning")
            else:
                cur.close()
                flash("Cannot delete: employee has related payroll records. Remove those entries first.", "danger")
        else:
            # Other MySQL integrity errors
            flash(f"Delete failed: {e}", "danger")
    except Exception as e:
        flash(f"Delete failed: {e}", "danger")
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return redirect(url_for("employees_hub"))


# ---------- report page ----------


@app.route("/employees/report", methods=["GET"], endpoint="employees_report")
@require_role("admin", "director", "headteacher", "bursar","deputyheadteacher")
def employees_report():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    department = (request.args.get("department") or "").strip()
    designation = (request.args.get("designation") or "").strip()

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    sql, params = _employees_query_and_params(
        q, status, department, designation)
    cur.execute(sql, params)
    rows = cur.fetchall()

    # quick summary
    summary = {
        "count_all": len(rows),
        "count_active": sum(1 for r in rows if (r["status"] or "") == "active"),
        "count_archived": sum(1 for r in rows if (r["status"] or "") == "archived"),
        "total_base_salary": sum(float(r["base_salary"] or 0) for r in rows),
        "total_allowance": sum(float(r["allowance"] or 0) for r in rows),
        "total_bonus": sum(float(r["bonus"] or 0) for r in rows),
    }
    cur.close()
    conn.close()

    return render_template(
        "employees_report.html",
        rows=rows,
        q=q, status=status, department=department, designation=designation,
        summary=summary,
        today=datetime.now().strftime("%Y-%m-%d")
    )

# ---------- export (CSV / Excel) ----------


@app.route("/employees/export", methods=["GET"], endpoint="employees_export")
@require_role("admin", "director", "headteacher", "bursar","deputyheadteacher")
def employees_export():
    fmt = (request.args.get("format") or "csv").lower()  # 'csv' or 'xlsx'
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    department = (request.args.get("department") or "").strip()
    designation = (request.args.get("designation") or "").strip()

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    sql, params = _employees_query_and_params(
        q, status, department, designation)
    cur.execute(sql, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # Normalize rows -> list of dicts
    data = []
    for r in rows:
        data.append({
            "ID": r["id"],
            "First Name": r["first_name"],
            "Middle Name": r.get("middle_name") or r.get("Middle_name") or "",
            "Last Name": r["last_name"],
            "Gender": r["gender"],
            "Contact": r["contact"],
            "Email": r["email"],
            "Residence": r["residence"],
            "Department": r["department"],
            "Designation": r["designation"],
            "Hire Date": r["hire_date"],
            "Status": r["status"],
            "Base Salary": r["base_salary"],
            "Allowance": r["allowance"],
            "Bonus": r["bonus"],
            "Pay Cycle": r["pay_cycle"],
            "Bank": r["bank_name"],
            "Bank Account": r["bank_account"],
            "TIN": r["tin"],
            "Notes": r["notes"],
            "Created": r["created_at"],
        })

    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_filename = f"staff_report_{stamp}"

    if fmt == "xlsx":
        # Try to create a real Excel file if pandas is available; otherwise fall back to CSV
        try:
            import pandas as pd
            import tempfile
            import os
            df = pd.DataFrame(data)
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp.close()
            with pd.ExcelWriter(tmp.name, engine="xlsxwriter") as writer:
                df.to_excel(writer, index=False, sheet_name="Staff")
            return send_file(
                tmp.name,
                as_attachment=True,
                download_name=f"{base_filename}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception:
            # fall through to CSV
            pass

    # CSV export (works everywhere; Excel opens it fine)
    si = io.StringIO()
    writer = csv.DictWriter(si, fieldnames=list(data[0].keys()) if data else [
        "ID", "First Name", "Middle Name", "Last Name", "Gender", "Contact", "Email",
        "Residence", "Department", "Designation", "Hire Date", "Status", "Base Salary",
        "Allowance", "Bonus", "Pay Cycle", "Bank", "Bank Account", "TIN", "Notes", "Created"
    ])
    writer.writeheader()
    for row in data:
        writer.writerow(row)

    mem = io.BytesIO()
    mem.write(si.getvalue().encode("utf-8-sig"))  # BOM for Excel
    mem.seek(0)
    return send_file(
        mem,
        as_attachment=True,
        download_name=f"{base_filename}.csv",
        mimetype="text/csv"
    )


# ===================== PAYROLL =====================


@app.route("/payroll", methods=["GET", "POST"])
@require_role("admin", "bursar", "director", "headteacher")
def payroll_hub():
    from datetime import datetime

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Employees for dropdowns
    cur.execute("""
        SELECT id, first_name, Middle_name, last_name, designation, status
        FROM employees
        ORDER BY (status='active') DESC, last_name, first_name
    """)
    employees = cur.fetchall()

    # Create a payroll row
    if request.method == "POST" and request.form.get("action") == "create":
        try:
            employee_id = int(request.form["employee_id"])
            term = request.form["term"]
            year = int(request.form["year"])
            expected_salary = float(request.form.get("expected_salary") or 0)
            bonus = float(request.form.get("bonus") or 0)
            allowance = float(request.form.get("allowance") or 0)
        except Exception:
            conn.close()
            flash("Invalid input for payroll creation.", "danger")
            return redirect(url_for("payroll_hub"))

        total = expected_salary + bonus + allowance
        status = _payroll_status(total, 0)

        cur.execute("""
            INSERT INTO payroll
                (employee_id, term, year, expected_salary, bonus,
                 allowance, total, paid_amount, status, date_paid)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s, NOW())
        """, (employee_id, term, year, expected_salary, bonus, allowance, total, 0.0, status))
        conn.commit()
        conn.close()
        flash("Payroll row created.", "success")
        return redirect(url_for("payroll_hub"))

    # Filters
    sel_term = request.args.get("term") or ""
    sel_year = request.args.get("year") or ""
    sel_emp = request.args.get("employee_id") or ""
    q_sql, q_args = [], []
    if sel_term:
        q_sql.append("p.term = %s")
        q_args.append(sel_term)
    if sel_year:
        q_sql.append("p.year = %s")
        q_args.append(sel_year)
    if sel_emp:
        q_sql.append("p.employee_id = %s")
        q_args.append(sel_emp)
    where = ("WHERE " + " AND ".join(q_sql)) if q_sql else ""

    # Rows
    cur.execute(f"""
        SELECT p.*, e.first_name, e.Middle_name, e.last_name, e.designation
        FROM payroll p
        LEFT JOIN employees e ON e.id = p.employee_id
        {where}
        ORDER BY p.year DESC,
                 CASE p.term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 9 END,
                 e.last_name, e.first_name
    """, q_args)
    rows = cur.fetchall()

    # Summary
    cur.execute(f"""
        SELECT
          IFNULL(SUM(p.total), 0) AS total_expected,
          IFNULL(SUM(p.paid_amount), 0) AS total_paid,
          IFNULL(SUM(p.total - p.paid_amount), 0) AS total_outstanding,
          SUM(CASE WHEN p.status='fully_paid' THEN 1 ELSE 0 END) AS cnt_fully,
          SUM(CASE WHEN p.status='partially_paid' THEN 1 ELSE 0 END) AS cnt_partial,
          SUM(CASE WHEN p.status='not_paid' THEN 1 ELSE 0 END) AS cnt_none,
          COUNT(*) AS row_count
        FROM payroll p
        {where}
    """, q_args)
    summary = cur.fetchone()

    cur.close()
    conn.close()
    default_year = datetime.now().year

    return render_template(
        "payroll.html",
        employees=employees,
        rows=rows,
        TERMS=TERMS,
        sel_term=sel_term,
        sel_year=sel_year,
        sel_emp=sel_emp,
        default_year=default_year,
        summary=summary
    )


@app.route("/payroll/pay/<int:pid>", methods=["POST"])
@require_role("admin", "bursar", "director")
def payroll_add_payment(pid):
    """Add a payment to payroll AND mirror it into expenses (Salaries)."""
    # Amount
    try:
        amount = float(request.form.get("amount") or 0)
        if amount <= 0:
            raise ValueError
    except Exception:
        flash("Invalid payment amount.", "warning")
        return redirect(url_for("payroll_hub"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Load row + employee (for description)
    cur.execute("""
        SELECT p.id, p.employee_id, p.term, p.year, p.total, p.paid_amount,
               e.first_name, e.Middle_name, e.last_name, e.designation
        FROM payroll p
        LEFT JOIN employees e ON e.id = p.employee_id
        WHERE p.id=%s
    """, (pid,))
    row = cur.fetchone()

    if not row:
        conn.close()
        flash("Payroll row not found.", "warning")
        return redirect(url_for("payroll_hub"))

    # Update payroll amounts + status
    new_paid = (row["paid_amount"] or 0) + amount
    status = _payroll_status(row["total"], new_paid)
    cur.execute("""
        UPDATE payroll
           SET paid_amount = %s, status = %s, date_paid = NOW()
         WHERE id = %s
    """, (new_paid, status, pid))

    # Ensure "Salaries" category exists, then insert into expenses
    cat_id = get_or_create_expense_category(conn, "Salaries")
    emp_name = f"{(row['last_name'] or '').strip()}, {(row['first_name'] or '').strip()} {(row['Middle_name'] or '' ).strip()}".strip(
    ).strip(',')
    description = f"Salary payment - {emp_name} — {row['term']} {row['year']}"
    recorded_by = session.get("username", "system")

    cur.execute("""
        INSERT INTO expenses (description, amount, term, year, date_spent, category_id, recorded_by, type)
        VALUES (%s, %s, %s, %s, NOW(), %s, %s, 'staff_pay')
    """, (description, amount, row["term"], row["year"], cat_id, recorded_by))

    conn.commit()
    cur.close()
    conn.close()
    flash("Payment recorded and posted to expenses (Salaries).", "success")
    return redirect(url_for("payroll_hub"))


@app.route("/payroll/edit/<int:pid>", methods=["POST"])
@require_role("admin", "bursar", "director")
def payroll_edit(pid):
    """Edit expected/pay components; recompute total and status."""
    try:
        expected_salary = float(request.form.get("expected_salary") or 0)
        bonus = float(request.form.get("bonus") or 0)
        allowance = float(request.form.get("allowance") or 0)
    except Exception:
        flash("Invalid amounts for edit.", "danger")
        return redirect(url_for("payroll_hub"))

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT total, paid_amount FROM payroll WHERE id=%s", (pid,))
    row = cur.fetchone()
    if not row:
        conn.close()
        flash("Payroll row not found.", "warning")
        return redirect(url_for("payroll_hub"))

    new_total = expected_salary + bonus + allowance
    status = _payroll_status(new_total, row["paid_amount"])

    cur.execute("""
        UPDATE payroll
           SET expected_salary=%s, bonus=%s, allowance=%s, total=%s, status=%s
         WHERE id=%s
    """, (expected_salary, bonus, allowance, new_total, status, pid))
    conn.commit()
    cur.close()
    conn.close()
    flash("Payroll updated.", "success")
    return redirect(url_for("payroll_hub"))


@app.route("/payroll/delete/<int:pid>", methods=["POST"])
@require_role("admin", "director")
def payroll_delete(pid):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("DELETE FROM payroll WHERE id=%s", (pid,))
    conn.commit()
    cur.close()
    conn.close()
    flash("Payroll row deleted.", "info")
    return redirect(url_for("payroll_hub"))


@app.route("/payroll/export")
@require_role("admin", "bursar", "director")
def payroll_export():
    """Export current filtered view to CSV."""
    import csv
    import io
    sel_term = request.args.get("term") or ""
    sel_year = request.args.get("year") or ""
    sel_emp = request.args.get("employee_id") or ""

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    q_sql, q_args = [], []
    if sel_term:
        q_sql.append("p.term = %s")
        q_args.append(sel_term)
    if sel_year:
        q_sql.append("p.year = %s")
        q_args.append(sel_year)
    if sel_emp:
        q_sql.append("p.employee_id = %s")
        q_args.append(sel_emp)
    where = ("WHERE " + " AND ".join(q_sql)) if q_sql else ""

    cur.execute(f"""
        SELECT p.*, e.first_name, e.Middle_name, e.last_name, e.designation
        FROM payroll p
        LEFT JOIN employees e ON e.id = p.employee_id
        {where}
        ORDER BY p.year DESC,
                 CASE p.term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 9 END,
                 e.last_name, e.first_name
    """, q_args)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Employee", "Designation", "Term", "Year",
                     "Expected Salary", "Bonus", "Allowance", "Total",
                     "Paid Amount", "Status", "Last Paid"])
    for r in rows:
        fullname = f"{(r['last_name'] or '')}, {(r['first_name'] or '')} {r['Middle_name'] or ''}".strip()
        writer.writerow([
            fullname, r["designation"] or "",
            r["term"], r["year"],
            r["expected_salary"] or 0, r["bonus"] or 0, r["allowance"] or 0,
            r["total"] or 0, r["paid_amount"] or 0, r["status"], r["date_paid"] or ""
        ])
    resp = Response(buf.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = "attachment; filename=payroll.csv"
    return resp


# ----- Payroll Report (read-only) -----
@app.route("/reports/payroll", methods=["GET"])
@require_role("admin", "bursar", "director", "headteacher")
def payroll_report():
    from datetime import datetime

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Reuse if you already define TERMS globally; else fallback:
    TERMS_LOCAL = globals().get("TERMS", ["Term 1", "Term 2", "Term 3"])

    # Employees for filter dropdown
    cur.execute("""
        SELECT id, first_name, Middle_name, last_name, designation, status
        FROM employees
        ORDER BY (status='active') DESC, last_name, first_name
    """)
    employees = cur.fetchall()

    # Filters
    sel_term = request.args.get("term") or ""
    sel_year = request.args.get("year") or ""
    sel_emp = request.args.get("employee_id") or ""

    q_sql, q_args = [], []
    if sel_term:
        q_sql.append("p.term = %s")
        q_args.append(sel_term)
    if sel_year:
        q_sql.append("p.year = %s")
        q_args.append(sel_year)
    if sel_emp:
        q_sql.append("p.employee_id = %s")
        q_args.append(sel_emp)
    where = ("WHERE " + " AND ".join(q_sql)) if q_sql else ""

    cur.execute(f"""
        SELECT p.*, e.first_name, e.Middle_name, e.last_name, e.designation
        FROM payroll p
        LEFT JOIN employees e ON e.id = p.employee_id
        {where}
        ORDER BY p.year DESC,
                 CASE p.term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 9 END,
                 e.last_name, e.first_name
    """, q_args)
    rows = cur.fetchall()

    cur.execute(f"""
        SELECT
          IFNULL(SUM(p.total), 0) AS total_expected,
          IFNULL(SUM(p.paid_amount), 0) AS total_paid,
          IFNULL(SUM(p.total - p.paid_amount), 0) AS total_outstanding,
          SUM(CASE WHEN p.status='fully_paid' THEN 1 ELSE 0 END) AS cnt_fully,
          SUM(CASE WHEN p.status='partially_paid' THEN 1 ELSE 0 END) AS cnt_partial,
          SUM(CASE WHEN p.status='not_paid' THEN 1 ELSE 0 END) AS cnt_none,
          COUNT(*) AS row_count
        FROM payroll p
        {where}
    """, q_args)
    summary = cur.fetchone()

    cur.close()
    conn.close()

    return render_template(
        "payroll_report.html",
        employees=employees,
        rows=rows,
        TERMS=TERMS_LOCAL,
        sel_term=sel_term,
        sel_year=sel_year,
        sel_emp=sel_emp,
        summary=summary,
        default_year=datetime.now().year
    )


@app.route("/comment_rules", methods=["GET", "POST"])
@require_role("admin", "dos","headteacher","deputyheadteacher","bursar","classmanager")
def comment_rules():
    conn = get_db_connection()
    ensure_comment_rules_schema(conn)

    # -------- Dropdown helpers ----------
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT DISTINCT class_name
        FROM students
        WHERE class_name IS NOT NULL
        ORDER BY class_name
    """)
    classes = [r["class_name"] for r in cur.fetchall()]
    cur.close()

    terms = ["Term 1", "Term 2", "Term 3"]

    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT DISTINCT grade FROM grading_scale ORDER BY 1")
    grades = [r["grade"] for r in cur.fetchall()]
    cur.close()

    divisions = [1, 2, 3, 4, 5, 6, 7, 8, 9]

    # ---- NEW: student list for optional per-learner rules ----
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, first_name, last_name, student_number
        FROM students
        ORDER BY first_name, last_name
    """)
    students = cur.fetchall()
    cur.close()

    # -------- Insert new rule OR new library comment ----------
    if request.method == "POST":
        form_kind = request.form.get("form_kind") or "rule"

        # --- NEW: add comment into comment_library ---
        if form_kind == "library":
            try:
                cat = (request.form.get("lib_category") or "good").lower()
                if cat not in ("good", "moderate", "poor"):
                    cat = "good"

                lib_role = request.form.get("lib_role") or "teacher"
                lib_scope = request.form.get("lib_scope") or "overall"
                lib_text = (request.form.get("lib_text") or "").strip()

                if lib_text:
                    cur = conn.cursor()
                    cur.execute(
                        """
                        INSERT INTO comment_library
                            (category, text, role, scope, uses)
                        VALUES (%s,%s,%s,%s,0)
                        """,
                        (cat, lib_text, lib_role, lib_scope),
                    )
                    conn.commit()
                    cur.close()
                    flash("Library comment added.", "success")
                else:
                    flash("Comment text cannot be empty.", "warning")
            except Exception as e:
                app.logger.error(f"Add library comment failed: {e}", exc_info=True)
                conn.rollback()
                flash("Failed to add library comment.", "danger")

        # --- EXISTING: add comment rule ---
        else:
            data = request.form
            try:
                template = (data.get("template_text") or "").strip()

                # Ensure NO {name} is stored in rules
                template = template.replace("{name}", "").strip()
                template = template.lstrip(",. ") # optional cleanup

                cur = conn.cursor(dictionary=True)
                cur.execute(
                    """
                    INSERT INTO comment_rules
                        (role, scope, match_type,
                         grade, division, lower_limit, upper_limit,
                         class_name, level, term,
                         student_id,
                         template_text, priority, active)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        data.get("role"),
                        data.get("scope"),
                        data.get("match_type"),
                        (data.get("grade") or None),
                        (int(data.get("division")) if data.get("division") else None),
                        (float(data.get("lower_limit")) if data.get("lower_limit") else None),
                        (float(data.get("upper_limit")) if data.get("upper_limit") else None),
                        (data.get("class_name") or None),
                        (data.get("level") or None),
                        (data.get("term") or None),
                        (int(data.get("student_id")) if data.get("student_id") else None),
                        template,
                        int(data.get("priority") or 100),
                        int(data.get("active") or 1),
                    ),
                )
                conn.commit()
                cur.close()
                flash("Rule added.", "success")
            except Exception as e:
                app.logger.error(f"Add comment rule failed: {e}", exc_info=True)
                conn.rollback()
                flash("Failed to add rule.", "danger")

    # -------- Load rules table ----------
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM comment_rules ORDER BY role, scope, priority, id")
    rows = cur.fetchall()
    cur.close()

    # -------- NOW load comment library from DB (for presets) ----------
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT id, category, text, role, scope, uses
        FROM comment_library
        WHERE role = 'teacher' AND scope = 'overall'
        ORDER BY category, id
        """
    )
    lib_rows = cur.fetchall() or []
    cur.close()

    library_groups = {
        "excellent": [], # category 'good'
        "moderate": [], # category 'moderate'
        "poor": [] # category 'poor'
    }
    for r in lib_rows:
        cat = (r.get("category") or "").lower()
        if cat == "good":
            key = "excellent"
        elif cat == "moderate":
            key = "moderate"
        elif cat == "poor":
            key = "poor"
        else:
            continue

        library_groups[key].append(
            {
                "id": r["id"],
                "text": r["text"],
                "role": r["role"],
                "scope": r["scope"],
                "uses": r["uses"],
            }
        )

    conn.close()

    # -------- Render page ----------
    return render_template(
        "comment_rules.html",
        rows=rows,
        classes=classes,
        terms=terms,
        grades=grades,
        divisions=divisions,
        students=students,
        library_groups=library_groups,
    )


@app.route("/comment_rules/<int:rid>/delete", methods=["POST"])
@require_role("admin", "dos", "classmanager", "headteacher", "deputyheadteacher")
def delete_comment_rule(rid):
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("DELETE FROM comment_rules WHERE id=%s", (rid,))
        conn.commit()
        cur.close()
        flash("Rule deleted.", "success")
    except Exception as e:
        app.logger.error(f"Delete comment rule failed: {e}", exc_info=True)
        flash("Failed to delete rule.", "danger")
    finally:
        conn.close()
    return redirect(url_for("comment_rules"))


@app.route("/students/template")
@require_role("admin", "headteacher", "deputyheadteacher", "teacher", "classmanager")  # adjust roles
def download_students_template():
    from io import BytesIO
    import pandas as pd

    # Define expected columns
    cols = [
        "first_name", "Middle_name", "last_name",
        "sex", "section", "class_name", "stream",
        "parent_name", "parent_contact",
        "parent2_name", "parent2_contact",
        "student_number", "fees_code"
    ]
    df = pd.DataFrame(columns=cols)

    # Write empty sheet with headers
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Students")
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="students_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/admin/check_ensures")
def admin_check_ensures():
    names = [
        "ensure_expenses_schema",
        "ensure_other_income_schema",
    ]
    visible = {n: callable(globals().get(n)) for n in names}
    return jsonify(visible)


# ================= ESC/POS RAW PRINT HELPERS (Windows) =================

# Try Pillow for logo support (gracefully skip if missing)
try:
    from PIL import Image
except Exception:
    Image = None

# --- ESC/POS command bytes ---
ESC_INIT = b"\x1b\x40"  # Initialize
TXT_BOLD_ON = b"\x1b\x45\x01"
TXT_BOLD_OFF = b"\x1b\x45\x00"
ALIGN_LEFT = b"\x1b\x61\x00"  # Left
ALIGN_CTR = b"\x1b\x61\x01"  # Center
ALIGN_RGT = b"\x1b\x61\x02"  # Right
CUT_FULL = b"\x1d\x56\x00"  # Full cut (some models prefer \x1d\x56\x42\x00)
FEED_6 = b"\n" * 6


# ---- 5) (Optional) Tiny admin utility to add routes -------------------------------
@app.route("/transport/routes/save", methods=["POST"])
@require_role("admin", "bursar", "headteacher")
def transport_route_save():
    ensure_transport_as_requirement_schema()
    name = (request.form.get("route_name") or "").strip()
    fare = float(request.form.get("fare_per_term") or 0)
    if not name:
        flash("Route name required.", "warning")
        return redirect(request.referrer or url_for("start_payment"))
    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
                INSERT INTO transport_routes (name, fare_per_term)
                VALUES (%s, %s) ON DUPLICATE KEY UPDATE
                fare_per_term = VALUES(fare_per_term)""",
                    (name, fare)
                    )
        conn.commit()
        cur.close()
        conn.close()
        flash("Route saved.", "success")
    except Exception as e:
        flash(f"Failed to save route: {e}", "danger")
    return redirect(request.referrer or url_for("start_payment"))




# ---------- 3) Student lookup API (for auto-fill + guard hint) ----------


@app.route("/transport/subscribe", methods=["POST"])
@require_role("admin", "bursar")
def transport_simple_subscribe():
    try:
        student_number = (request.form.get("student_number") or "").strip()
        route_id = int(request.form.get("route_id") or 0)
        term = (request.form.get("term") or "").strip()
        year = int(request.form.get("year") or 0)

        if not student_number or not route_id or not term or not year:
            flash("Missing inputs for subscribe.", "warning")
            return redirect(url_for("start_payment", student_number=student_number))

        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT id FROM students WHERE student_number=%s AND archived=0", (student_number,))
        stu = cur.fetchone()
        cur.close()
        conn.close()
        if not stu:
            flash("Student not found or archived.", "warning")
            return redirect(url_for("start_payment", student_number=student_number))

        transport_subscribe(stu["id"], route_id, term, year)
        flash(
            "Subscribed: transport requirement will now appear in Requirements.", "success")
        return redirect(url_for("start_payment", student_number=student_number, term=term))
    except Exception as e:
        flash(f"Subscribe failed: {e}", "danger")
        return redirect(url_for("start_payment"))


@app.route("/transport/unsubscribe", methods=["POST"])
@require_role("admin", "bursar")
def transport_simple_unsubscribe():
    try:
        student_number = (request.form.get("student_number") or "").strip()
        route_id = int(request.form.get("route_id") or 0)
        term = (request.form.get("term") or "").strip()

        if not student_number or not route_id:
            flash("Missing inputs for unsubscribe.", "warning")
            return redirect(url_for("start_payment", student_number=student_number, term=term))

        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT id FROM students WHERE student_number=%s", (student_number,))
        stu = cur.fetchone()
        cur.close()
        conn.close()
        if not stu:
            flash("Student not found.", "warning")
            return redirect(url_for("start_payment", student_number=student_number, term=term))

        transport_unsubscribe(stu["id"])
        flash("Unsubscribed: transport requirement removed for this term.", "info")
        return redirect(url_for("start_payment", student_number=student_number, term=term))
    except Exception as e:
        flash(f"Unsubscribe failed: {e}", "danger")
        return redirect(url_for("start_payment"))

# ---------- 4) CSV template (unchanged) ----------


@app.route("/transport/template")
@require_role("admin", "bursar", "headteacher")
def transport_template_download():
    csv_text = (
        "student_number,route_name,action,amount,method\n"
        "STU-0001,City Route,SUBSCRIBE,,\n"
        "STU-0002,City Route,PAY,50000,Cash\n"
        "STU-0003,City Route,UNSUBSCRIBE,,\n"
    )
    resp = make_response(csv_text)
    resp.headers["Content-Type"] = "text/csv"
    resp.headers["Content-Disposition"] = "attachment; filename=transport_template.csv"
    return resp

# ---------- 5) Hub (kept, with pay-guard + lookup auto-fill) ----------


@app.route("/transport/hub", methods=["GET", "POST"])
@require_role("admin", "bursar", "headteacher")
def transport_hub():
    conn = get_db_connection()
    try:
        ensure_transport_schema(conn)
    finally:
        conn.close()
    ay = get_active_academic_year()
    active_term = ay.get("current_term") or ay.get("term") or "Term 1"
    active_year = int(ay.get("year"))

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        # --- Add/Update route ---
        if action == "add_route":
            name = (request.form.get("route_name") or "").strip()
            fare = float(request.form.get("fare_per_term") or 0)
            if not name:
                flash("Route name is required.", "warning")
            else:
                try:
                    conn = get_db_connection()
                    cur = conn.cursor(dictionary=True)
                    cur.execute("""
                        INSERT INTO transport_routes (name, fare_per_term)
                        VALUES (%s, %s) ON DUPLICATE KEY UPDATE
                        fare_per_term = VALUES(fare_per_term)""",
                                (name, fare)
                                )
                    conn.commit()
                    cur.close()
                    conn.close()
                    flash("Route saved.", "success")
                except Exception as e:
                    flash(f"Failed to save route: {e}", "danger")
            return redirect(url_for("transport_hub"))

        # --- Delete route ---
        if action == "delete_route":
            rid = request.form.get("route_id")
            try:
                conn = get_db_connection()
                cur = conn.cursor(dictionary=True)
                cur.execute("DELETE FROM transport_routes WHERE id=%s", (rid,))
                conn.commit()
                cur.close()
                conn.close()
                flash("Route deleted.", "info")
            except Exception as e:
                flash(f"Delete failed: {e}", "danger")
            return redirect(url_for("transport_hub"))

        # --- Subscribe ---
        if action == "subscribe":
            student_number = (request.form.get("student_number") or "").strip()
            route_id = int(request.form.get("route_id") or 0)
            term = (request.form.get("term") or active_term)
            year = int(request.form.get("year") or active_year)

            if not student_number or not route_id:
                flash("Student number and route are required.", "warning")
                return redirect(url_for("transport_hub"))

            conn = get_db_connection()
            cur = conn.cursor(dictionary=True)
            cur.execute(
                "SELECT id FROM students WHERE archived=0 AND student_number=%s", (student_number,))
            stu = cur.fetchone()
            cur.close()
            conn.close()
            if not stu:
                flash("Student not found or archived.", "warning")
                return redirect(url_for("transport_hub"))

            conn = get_db_connection()
            if transport_is_already_subscribed(conn, stu["id"], route_id):
                conn.close()
                flash("Already subscribed to this route.", "info")
                return redirect(url_for("transport_hub"))
            conn.close()

            transport_subscribe(stu["id"], route_id, term, year)
            flash("Subscribed.", "success")
            return redirect(url_for("transport_hub"))

        # --- Unsubscribe ---
        if action == "unsubscribe":
            student_number = (request.form.get("student_number") or "").strip()
            route_id = int(request.form.get("route_id") or 0)

            if not student_number or not route_id:
                flash("Student number and route are required.", "warning")
                return redirect(url_for("transport_hub"))

            conn = get_db_connection()
            cur = conn.cursor(dictionary=True)
            cur.execute(
                "SELECT id FROM students WHERE student_number=%s", (student_number,))
            stu = cur.fetchone()
            cur.close()
            conn.close()
            if not stu:
                flash("Student not found.", "warning")
                return redirect(url_for("transport_hub"))

            transport_unsubscribe(stu["id"], route_id)
            flash("Unsubscribed.", "success")
            return redirect(url_for("transport_hub"))

        # --- Pay (into other_income) with GUARD ---
        if action == "pay":
            student_number = (request.form.get("student_number") or "").strip()
            route_id = int(request.form.get("route_id") or 0)
            amount = float(request.form.get("amount") or 0)
            method = (request.form.get("method") or "Cash").strip()
            term = (request.form.get("term") or active_term)
            year = int(request.form.get("year") or active_year)

            if not student_number or amount <= 0:
                flash("Student number and a positive amount are required.", "warning")
                return redirect(url_for("transport_hub"))

            # Resolve student + subscription check
            conn = get_db_connection()
            cur = conn.cursor(dictionary=True)
            cur.execute(
                "SELECT id FROM students WHERE archived=0 AND student_number=%s", (student_number,))
            stu = cur.fetchone()
            cur.close()
            conn.close()
            if not stu:
                flash("Student not found or archived.", "warning")
                return redirect(url_for("transport_hub"))

            if not transport_has_active_subscription(get_db_connection(), stu["id"], term, year, (route_id or None)):
                flash(
                    "Payment rejected: student is NOT an active transport subscriber for this term/route.", "danger")
                return redirect(url_for("transport_hub"))

            # Pretty route name (optional)
            route_name = ""
            if route_id:
                conn = get_db_connection()
                cur = conn.cursor(dictionary=True)
                cur.execute(
                    "SELECT name FROM transport_routes WHERE id=%s", (route_id,))
                row = cur.fetchone()
                cur.close()
                conn.close()
                route_name = row["name"] if row else ""

            recorded_by = session.get("full_name") or session.get(
                "username") or session.get("role") or "system"
            try:
                transport_record_payment_for_student_number(
                    student_number, route_name, term, year, amount, method, recorded_by)
                flash("Transport payment recorded (Other Income).", "success")
            except Exception as e:
                current_app.logger.exception("[transport pay] insert failed")
                flash(f"Payment failed: {e}", "danger")

            return redirect(url_for("transport_hub"))

        # --- Bulk CSV (unchanged logic) ---
        if action == "upload_csv":
            file = request.files.get("file")
            if not file:
                flash("No file provided.", "warning")
                return redirect(url_for("transport_hub"))

            import pandas as pd
            df = pd.read_csv(file) if file.filename.lower().endswith(
                ".csv") else pd.read_excel(file)
            ok, skip = 0, 0

            for _, r in df.iterrows():
                try:
                    sn = str(r.get("student_number") or "").strip()
                    rname = str(r.get("route_name") or "").strip()
                    act = str(r.get("action") or "").strip().upper()

                    # resolve route id (if provided)
                    route_id = None
                    if rname:
                        conn = get_db_connection()
                        cur = conn.cursor(dictionary=True)
                        cur.execute(
                            "SELECT id FROM transport_routes WHERE name=%s", (rname,))
                        row = cur.fetchone()
                        cur.close()
                        conn.close()
                        route_id = row["id"] if row else None

                    if act == "SUBSCRIBE":
                        if not route_id:
                            skip += 1
                            continue
                        conn = get_db_connection()
                        cur = conn.cursor(dictionary=True)
                        cur.execute(
                            "SELECT id FROM students WHERE archived=0 AND student_number=%s", (sn,))
                        stu = cur.fetchone()
                        cur.close()
                        conn.close()
                        if not stu:
                            skip += 1
                            continue
                        transport_subscribe(
                            stu["id"], route_id, active_term, active_year)
                        ok += 1

                    elif act == "UNSUBSCRIBE":
                        if not route_id:
                            skip += 1
                            continue
                        conn = get_db_connection()
                        cur = conn.cursor(dictionary=True)
                        cur.execute(
                            "SELECT id FROM students WHERE student_number=%s", (sn,))
                        stu = cur.fetchone()
                        cur.close()
                        conn.close()
                        if not stu:
                            skip += 1
                            continue
                        transport_unsubscribe(stu["id"], route_id)
                        ok += 1

                    elif act == "PAY":
                        amt = float(r.get("amount") or 0)
                        if amt <= 0:
                            skip += 1
                            continue
                        method = str(r.get("method") or "Cash")

                        # optional guard here too: require active sub (any route)
                        conn = get_db_connection()
                        cur = conn.cursor(dictionary=True)
                        cur.execute(
                            "SELECT id FROM students WHERE archived=0 AND student_number=%s", (sn,))
                        stu = cur.fetchone()
                        cur.close()
                        conn.close()
                        if not stu or not transport_has_active_subscription(get_db_connection(), stu["id"], active_term, active_year, route_id):
                            skip += 1
                            continue

                        transport_record_payment_for_student_number(sn, rname or "", active_term, active_year, amt, method,
                                                                    session.get("full_name") or session.get("username") or "system")
                        ok += 1
                    else:
                        skip += 1
                except Exception:
                    skip += 1

            flash(f"Processed: {ok}, Skipped: {skip}.", "info")
            return redirect(url_for("transport_hub"))

    # GET: show lists/balances
    conn = get_db_connection()
    routes = transport_get_routes(conn)

    route_filter = request.args.get("route_id")
    try:
        route_filter_id = int(route_filter) if route_filter else None
    except Exception:
        route_filter_id = None

    subs = transport_get_active_subscribers(
        conn, active_term, active_year, route_filter_id)

    balances = []
    for row in subs:
        due = float(row["fare_per_term"] or 0.0)
        paid = transport_paid_total_for_sn(
            conn, row["student_number"], active_term, active_year)
        balances.append({
            "subscription_id": row["id"],
            "student_number": row["student_number"],
            "full_name": f"{row['first_name']} {row['Middle_name'] or ''} {row['last_name']}".replace(" ", " ").strip(),
            "class_name": row["class_name"],
            "stream": row["stream"],
            "route_name": row["route_name"],
            "fare_per_term": due,
            "paid": paid,
            "balance": max(due - paid, 0.0),
        })

    conn.close()

    return render_template(
        "transport_hub.html",
        routes=routes,
        balances=balances,
        active_term=active_term,
        active_year=active_year,
        TERMS=TERMS,
        route_filter=(route_filter_id or "")
    )


@app.route("/api/student/by_number")
@require_role("admin", "bursar", "headteacher")
def api_student_by_number():
    """
    Returns basic student info (active only) and current transport route (if subscribed).
    Query: %ssn=STU-0001
    Response:
      { success: true,
        data: { id, student_number, full_name, class_name, stream,
                route_id, route_name } }
      or { success: false, message: "..." }
    """
    sn = (request.args.get("sn") or "").strip()
    if not sn:
        return jsonify({"success": False, "message": "Missing student number."}), 400

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS mid, last_name,
               class_name, stream
        FROM students
        WHERE archived=0 AND student_number=%s
        LIMIT 1
    """, (sn,))
    stu = cur.fetchone()

    if not stu:
        conn.close()
        return jsonify({"success": False, "message": "Student not found or archived."}), 404

    # try find an active transport subscription (any route)
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT ts.route_id, tr.name AS route_name
        FROM transport_subscriptions ts
        JOIN transport_routes tr ON tr.id=ts.route_id
        WHERE ts.active=1 AND ts.student_id=%s
        ORDER BY ts.created_at DESC
        LIMIT 1
    """, (stu["id"],))
    sub = cur.fetchone()
    cur.close()
    conn.close()

    full_name = f"{stu['first_name']} {stu['mid']} {stu['last_name']}".replace(
        " ", " ").strip()
    data = {
        "id": stu["id"],
        "student_number": stu["student_number"],
        "full_name": full_name,
        "class_name": stu["class_name"],
        "stream": stu["stream"],
        "route_id": (sub["route_id"] if sub else None),
        "route_name": (sub["route_name"] if sub else None),
    }
    return jsonify({"success": True, "data": data})
# --------------------------------------------------------------------

# ===================== /TRANSPORT (as Other Income) =====================
# ================= /TRANSPORT as REQUIREMENT (Single Start Payment) ===================


# 2) /fees/pay (unchanged UX; after commit we print via the wrapper)

@app.route("/fees/pay", methods=["POST"])
@require_role("admin", "bursar")
def pay_fees():
    f = request.form
    student_id = int(f.get("student_id") or 0)
    amount_paid = float(f.get("amount_paid") or 0)
    method = (f.get("method") or "Cash").strip()
    payment_type = (f.get("payment_type") or "fees").strip()

    if not student_id or amount_paid <= 0:
        flash("Student and a positive amount are required.", "warning")
        return redirect(request.referrer or url_for("dashboard"))

    ay = get_active_academic_year()
    term = f.get("term") or ay.get(
        "current_term") or ay.get("term") or "Term 1"
    year = int(f.get("year") or ay.get("year"))

    conn = get_db_connection()

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            INSERT INTO fees (
                student_id, term, year, amount_paid, method, payment_type, date_paid, recorded_by
            ) VALUES (%s, %s, %s, %s, %s, %s, NOW(), %s)
        """, (student_id, term, year, amount_paid, method, payment_type, (session.get("username") or "system")))
        fee_id = c.lastrowid
        conn.commit()
        cur.close()

        ok = handle_payment_and_print(fee_id)
        if ok:
            flash("Payment saved and sent to printer.", "success")
        else:
            flash(
                "Payment saved. Printer not confirmed — open the receipt and try Reprint.", "warning")

        return redirect(url_for("receipt_view", payment_id=fee_id))

    except Exception as e:
        conn.rollback()
        current_app.logger.exception(f"[fees/pay] insert failed: {e}")
        flash(f"Failed to save payment: {e}", "danger")
        return redirect(request.referrer or url_for("dashboard"))
    finally:
        conn.close()


# ========================= START PAYMENT (with Transport-as-Requirement) =========================
# Assumes you pasted earlier helpers:
# ensure_transport_as_requirement_schema, get_student_requirements
# If not, paste those first (from my previous message).


# --- Put these near your other helpers (once) -------------------------------


@app.route("/fees/<int:fee_id>/print")
@require_role("admin", "clerk", "headteacher", "bursar")
def print_fee_receipt(fee_id):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM fees WHERE id=%s", (fee_id,))
    fee = cur.fetchone()
    if not fee:
        conn.close()
        flash("Payment not found.", "danger")
        return redirect(url_for("start_payment"))

    cur.execute("SELECT * FROM students WHERE id=%s", (fee["student_id"],))
    stu = cur.fetchone()
    cur.close()
    conn.close()

    ok = send_receipt_to_printer(
        fee, stu,
        printer_name=current_app.config["RECEIPT_PRINTER_NAME"],
        school_name=current_app.config.get("SCHOOL_NAME", "School"),
        logo_path=current_app.config.get("RECEIPT_LOGO_PATH"),
        paper_width_dots=current_app.config.get("RECEIPT_PAPER_DOTS", 576)
    )
    flash("Receipt sent to printer." if ok else "Could not print receipt.",
          "success" if ok else "warning")
    return redirect(request.referrer or url_for("start_payment"))


@app.route("/receipt/reprint/<int:fee_id>")
@require_role("admin", "bursar")
def reprint_receipt(fee_id):
    """Reprint an existing receipt by fee_id."""
    try:
        ok = handle_payment_and_print(fee_id)
        if ok:
            flash(f"Receipt {fee_id} sent to printer.", "success")
        else:
            flash(f"Receipt {fee_id} could not be printed.", "warning")
    except Exception as e:
        current_app.logger.exception(f"[REPRINT] failed fee_id={fee_id}: {e}")
        flash("Unexpected error while reprinting.", "danger")
    return redirect(request.referrer or url_for("student_statement"))





@app.route("/start_payment", methods=["GET", "POST"])
@require_role("admin", "bursar")
def start_payment():
    ay = get_active_academic_year()
    current_term = ay.get("current_term") or ay.get("term") or "Term 1"
    current_year = int(ay.get("year"))
    sel_term = (request.values.get("term") or current_term).strip()

    q_student_id = request.values.get("student_id", type=int)
    q_student_number = (request.values.get("student_number") or "").strip()

    # ---------- helpers ----------
    def _get_student_by_id(stid: int):
        if not stid:
            return None
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS middle_name,
                   last_name, class_name, stream, section
              FROM students
             WHERE id=%s AND archived=0
        """, (stid,))
        row = cur.fetchone()
        cur.close(); conn.close()
        return row

    def _get_student_by_number(snum: str):
        if not snum:
            return None
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS middle_name,
                   last_name, class_name, stream, section
              FROM students
             WHERE TRIM(LOWER(student_number))=TRIM(LOWER(%s)) AND archived=0
             LIMIT 1
        """, (snum,))
        row = cur.fetchone()
        cur.close(); conn.close()
        return row

    def _all_students_for_dropdown():
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS m,
                   last_name, class_name, COALESCE(stream,'') AS stream
              FROM students
             WHERE archived=0
             ORDER BY last_name, first_name
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        return rows

    # ---------- POST: record payment (flow unchanged, INSERTS FIXED) ----------
    if request.method == "POST":
        sid = request.form.get("sid", type=int)
        if not sid:
            flash("Select a student before recording payment.", "warning")
            return redirect(url_for("start_payment", student_id=q_student_id, term=sel_term))

        payment_type = (request.form.get("payment_type") or "school_fees").strip()
        method = (request.form.get("method") or "cash").strip()
        term = (request.form.get("term") or current_term).strip()
        year = int(request.form.get("year") or current_year)
        comment = (request.form.get("comment") or request.form.get("payment_item") or "").strip()
        recorded_by = session.get("full_name") or session.get("username") or session.get("role") or "system"
        today = datetime.now().strftime("%Y-%m-%d")

        conn = get_db_connection()
        try:
            if payment_type == "requirements":
                ids = request.form.getlist("req_id[]")
                names = request.form.getlist("req_name[]")
                amts = request.form.getlist("req_amount[]")

                total = 0.0
                for rid, rname, ramt in zip(ids, names, amts):
                    try:
                        amt = float(ramt)
                    except ValueError:
                        amt = 0.0
                    if amt <= 0:
                        continue

                    total += amt
                    cur = conn.cursor(dictionary=True)
                    cur.execute("""
                        INSERT INTO fees (
                            student_id, term, year,
                            amount_paid,
                            requirement_name, req_term,
                            date_paid, method, comment,
                            payment_type, recorded_by,
                            payment_item
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'requirements',%s,'transport')
                    """, (
                        sid, term, year,
                        amt,
                        rname, term,
                        today, method, comment,
                        recorded_by
                    ))
                    cur.close()

                conn.commit()
                flash(f"Requirements payment recorded (UGX {total:,.0f}).", "success")

            else:
                raw = (request.form.get("amount_paid") or "0").strip()
                try:
                    amount_paid = float(raw)
                except ValueError:
                    amount_paid = 0.0

                if amount_paid <= 0:
                    flash("Amount must be greater than zero.", "warning")
                    return redirect(url_for("start_payment", student_id=q_student_id, term=term))

                cur = conn.cursor(dictionary=True)
                cur.execute("""
                    INSERT INTO fees (
                        student_id, term, year,
                        amount_paid,
                        date_paid, method, comment,
                        payment_type, recorded_by
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,'school_fees',%s)
                """, (
                    sid, term, year,
                    amount_paid,
                    today, method, comment,
                    recorded_by
                ))
                fee_id = cur.lastrowid
                conn.commit()
                cur.close()

                flash("Fees payment recorded.", "success")
                try:
                    current_app.logger.info(f"[PRINT] start_payment -> printing id={fee_id}")
                    handle_payment_and_print(fee_id)
                except Exception as e:
                    current_app.logger.exception(f"[PRINT] failed: {e}")

        except Exception as e:
            conn.rollback()
            flash(f"Payment failed: {e}", "danger")
        finally:
            conn.close()

        return redirect(url_for("start_payment", student_id=sid, term=term))

    # ---------- GET ----------
    student = None
    if q_student_id:
        student = _get_student_by_id(q_student_id)
    if not student and q_student_number:
        student = _get_student_by_number(q_student_number)
        if student:
            q_student_id = student["id"]

    students_list = _all_students_for_dropdown()
    reqs, fin, transport = [], None, None
    try:
        routes = transport_get_routes()
    except Exception:
        routes = []

    if student:
        reqs = get_class_requirements(student["class_name"], sel_term) or []

        term_no = term_to_no(sel_term) or 1
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT TRIM(LOWER(requirement_name)) AS nm, COALESCE(SUM(amount_paid),0) AS t
              FROM fees
             WHERE student_id=%s
               AND year=%s AND term_no=%s
               AND payment_type_norm IN('requirements','requirement')
               AND (comment IS NULL OR LOWER(comment) NOT LIKE '%void%')
             GROUP BY TRIM(LOWER(requirement_name))
        """, (student["id"], current_year, term_no))
        paid_map = {r["nm"]: float(r["t"] or 0.0) for r in (cur.fetchall() or [])}
        cur.close(); conn.close()

        for r in reqs:
            nm = ((r.get("name") or "").strip().lower())
            amt = float(r.get("amount") or 0.0)
            r["amount"] = amt
            r["is_paid"] = paid_map.get(nm, 0.0) >= amt - 0.0001
            r["paid_amount"] = paid_map.get(nm, 0.0)

        # balances summary (unchanged call)
        fin = compute_student_financials(student["id"], student["class_name"], sel_term, current_year)

        try:
            tinfo = transport_subscription_info(student["id"], sel_term, current_year)
        except Exception:
            tinfo = None

        if tinfo and float(tinfo.get("fare_per_term") or 0) > 0:
            conn = get_db_connection()
            try:
                tp_paid = transport_paid_via_requirements(conn, student["id"], sel_term, current_year)
            finally:
                conn.close()

            fare = float(tinfo["fare_per_term"] or 0.0)
            is_paid = float(tp_paid or 0.0) >= fare - 0.0001
            transport = {
                "route_name": tinfo["route_name"],
                "fare_per_term": fare,
                "paid": float(tp_paid or 0.0),
                "balance": max(fare - float(tp_paid or 0.0), 0.0),
                "is_paid": is_paid,
            }
            reqs.append({
                "id": "transport",
                "name": f"Transport ({tinfo['route_name']})",
                "qty": 1,
                "amount": fare,
                "is_paid": is_paid,
                "paid_amount": float(tp_paid or 0.0),
            })

    return render_template(
        "start_payment.html",
        terms=TERMS,
        current_term=sel_term or current_term,
        current_year=current_year,
        student=student,
        students_list=students_list,
        reqs=reqs,
        fin=fin,
        transport=transport,
        routes=routes,
        student_number=q_student_number,
    )

@app.route("/payments/confirm", methods=["POST"])
@require_role("admin", "bursar", "clerk", "headteacher")
def payments_confirm():
    f = request.form
    ay = get_active_academic_year()
    term = (f.get("term") or ay.get("current_term")
            or ay.get("term") or "Term 1").strip()
    year = int(f.get("year") or ay.get("year") or datetime.now().year)

    # Resolve student
    student_id = f.get("student_id")
    student_number = (f.get("student_number") or "").strip()
    if not student_id and student_number:
        conn = get_db_connection()
        try:
            cur = conn.cursor(dictionary=True)
            cur.execute(
                "SELECT id FROM students WHERE student_number=%s AND archived=0",
                (student_number,)
            )
            r = cur.fetchone()
            if not r:
                flash("Student not found.", "warning")
                return redirect(url_for("dashboard"))
            student_id = r["id"]
        finally:
            cur.close()
            conn.close()

    try:
        student_id = int(student_id)
    except (TypeError, ValueError):
        flash("Invalid or missing student.", "danger")
        return redirect(url_for("dashboard"))

    # Amount + basics
    try:
        amount_paid = float(f.get("amount_paid") or 0)
        if amount_paid <= 0:
            raise ValueError
    except Exception:
        flash("Enter a valid positive amount.", "danger")
        return redirect(url_for("dashboard"))

    method = (f.get("method") or "N/A").strip()
    payment_type = (f.get("payment_type") or "fees").strip().lower()
    payment_item = (f.get("payment_item") or "").strip()
    recorded_by = session.get("username") or session.get("role") or "System"

    # Save
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            INSERT INTO fees (
              student_id, term, year, amount_paid,
              payment_item, bursary_amount, carried_forward, expected_amount,
              date_paid, method, payment_type, recorded_by
            ) VALUES (%s, %s, %s, %s, %s, 0, 0, 0, NOW(), %s, %s, %s)
        """, (student_id, term, year, amount_paid, payment_item, method, payment_type, recorded_by))
        fee_id = cur.lastrowid
        conn.commit()
        cur.close()
    except Exception as e:
        conn.rollback()
        current_app.logger.exception("Failed saving payment")
        flash(f"Failed to save payment: {e}", "danger")
        return redirect(url_for("dashboard"))
    finally:
        conn.close()

    # Auto print
    ok = handle_payment_and_print(fee_id)
    if ok:
        flash("Payment saved and sent to printer.", "success")
    else:
        flash("Payment saved. Printer not confirmed — open the receipt and use Reprint.", "warning")

    return redirect(url_for("receipt_view", payment_id=fee_id))


@app.route("/receipt/reprint/<int:payment_id>", methods=["POST"])
@require_role("admin", "bursar")
def receipt_reprint(payment_id: int):
    ok = handle_payment_and_print(payment_id)
    if ok:
        flash("Receipt sent to printer.", "success")
    else:
        flash("Printer not confirmed. Check connection and try again.", "warning")
    return redirect(url_for("receipt_view", payment_id=payment_id))
# ========================================================================================


# ---------- Opening Balance: BULK UPLOAD ----------


@app.route("/balances/opening-balance/bulk", methods=["GET", "POST"])
@require_role("admin", "bursar", "clerk")
def opening_balance_bulk():
    """
    Upload a CSV/XLSX with columns:
      - student_number (required)
      - amount (required)
      - asof_year (optional; falls back to active year)

    Saves one row per student into fees table using payment_type='opening_balance'.
    """
    ay = get_active_academic_year()
    active_year = int(ay.get("year"))

    # GET: show page
    if request.method == "GET":
        sample = [
            {"row": 1, "student_number": "STD-2025-001",
                "amount": 150000, "asof_year": active_year},
            {"row": 2, "student_number": "STD-2025-002",
                "amount": 90000, "asof_year": active_year},
        ]
        return render_template("opening_balance_bulk.html",
                               active_year=active_year, results=[],
                               sample_rows=sample)

    # POST: process upload
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Please choose a CSV or XLSX file.", "warning")
        return redirect(url_for("opening_balance_bulk"))

    # --- Load rows (CSV/XLSX) ---
    try:
        fname = file.filename.lower()
        if fname.endswith(".csv"):
            df = pd.read_csv(file)
        elif fname.endswith(".xlsx") or fname.endswith(".xls"):
            df = pd.read_excel(file)
        else:
            flash("Only .csv, .xlsx are supported.", "danger")
            return redirect(url_for("opening_balance_bulk"))
    except Exception as e:
        flash(f"Could not read file: {e}", "danger")
        return redirect(url_for("opening_balance_bulk"))

    # Normalize column names
    cols = {c.strip().lower(): c for c in df.columns}
    need = {"student_number", "amount"}
    if not need.issubset(cols.keys()):
        flash("Required columns: student_number, amount (optional: asof_year).", "danger")
        return redirect(url_for("opening_balance_bulk"))

    # Prepare iteration
    results = []
    ok = bad = 0

    # Open one DB connection for all rows
    conn = get_db_connection()

    # Iterate rows
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        sn = str(row.get(cols.get("student_number"), "")).strip()
        amt = row.get(cols.get("amount"))
        asof = row.get(cols.get("asof_year"), active_year)

        # Validate basics
        if not sn or pd.isna(amt):
            bad += 1
            results.append({
                "row": i, "student_number": sn, "amount": amt,
                "asof_year": asof, "status": "ERROR",
                "message": "Missing student_number or amount"
            })
            continue

        # Resolve student_id from student_number
        sid = resolve_student_id(conn, student_number=sn)
        if not sid:
            bad += 1
            results.append({
                "row": i, "student_number": sn, "amount": amt,
                "asof_year": asof, "status": "ERROR",
                "message": "Student not found"
            })
            continue

        # Coerce types (any failure -> error row)
        try:
            amount = float(amt)
            asof_year = int(asof) if str(asof).strip() else active_year
        except Exception:
            bad += 1
            results.append({
                "row": i, "student_number": sn, "amount": amt,
                "asof_year": asof, "status": "ERROR",
                "message": "Amount or asof_year is invalid"
            })
            continue

        # ✅ Save opening balance (no surrounding try/except; let hard errors surface)
        set_opening_balance(conn, sid, amount, asof_year,
                            note=f"opening_balance import row {i}")
        ok += 1
        results.append({
            "row": i, "student_number": sn, "amount": amount,
            "asof_year": asof_year, "status": "OK", "message": "saved"
        })

    # Commit once for the batch
    conn.commit()
    conn.close()

    if ok and not bad:
        flash(f"{ok} opening balances saved.", "success")
    elif ok and bad:
        flash(f"{ok} saved; {bad} failed. See details below.", "warning")
    else:
        flash("No rows saved. Check the errors below.", "danger")

    return render_template("opening_balance_bulk.html",
                           active_year=active_year,
                           results=results,
                           sample_rows=[])


@app.route("/balances/opening/template.csv")
@require_role("admin", "bursar", "headteacher")
def opening_balance_template_csv():
    """
    Download a CSV template: student_number,amount,asof_year
    """
    csv_text = "student_number,amount,asof_year\nSTD-2025-001,250000,2024\n"
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=opening_balance_template.csv"}
    )


@app.route("/balances/opening/template.xlsx")
@require_role("admin", "bursar")
def opening_balance_template_xlsx():
    """
    Download an XLSX template with the same columns.
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "opening_balance"
    ws.append(["student_number", "amount", "asof_year"])
    ws.append(["STD-2025-001", 250000, 2024])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="opening_balance_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------- (NEW) CSV template download ----------


@app.route("/holiday/template.csv")
@require_role("admin", "teacher", "headteacher", "dos")
def holiday_template_csv():
    _hp_ensure()
    ay = get_active_academic_year()
    term = ay.get("current_term") or ay.get("term") or "Term 1"
    year = int(ay.get("year"))

    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["student_number", "subject_name", "assessment_type",
               "score", "max_score", "weight", "term", "year"])
    w.writerow(["STD-2025-001", "Holiday Package", "Test",
               "85", "100", "", "%s" % term, "%d" % year])
    data = output.getvalue()
    output.close()
    return Response(
        data,
        mimetype="text/csv",
        headers={
            "Content-Disposition": 'attachment; filename="holiday_package_template.csv"'}
    )

# ---------- (NEW) subjects that actually have HP marks in class/period ----------


def _hp_subjects_for_class_term_year(conn, class_name: str, term: str, year: int):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
      SELECT DISTINCT h.subject_id,
             (SELECT name FROM subjects s WHERE s.id=h.subject_id) AS name
      FROM holiday_package_scores h
      JOIN students st ON st.id = h.student_id
      WHERE st.class_name = %s AND h.term = %s AND h.year = %s
      ORDER BY name
    """, (class_name, term, year))
    rows = cur.fetchall()
    cur.close()
    return [{"id": r["subject_id"], "name": r["name"]} for r in rows]

# 6) Route (original logic kept; preview & template link added)


@app.route("/holiday/hub", methods=["GET", "POST"])
@require_role("admin", "teacher", "headteacher", "dos")
def holiday_hub():
    _hp_ensure()
    ay = get_active_academic_year()
    active_term = ay.get("current_term") or ay.get("term") or "Term 1"
    active_year = int(ay.get("year"))

    # -------- POST: ORIGINAL ACTIONS (unchanged) --------
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        if action == "add_one":
            sn = (request.form.get("student_number") or "").strip()
            sid = request.form.get("student_id")
            if not sid and sn:
                conn = get_db_connection()
                cur = conn.cursor(dictionary=True)
                cur.execute(
                    "SELECT id FROM students WHERE student_number=%s AND archived=0", (sn,))
                r = cur.fetchone()
                cur.close()
                conn.close()
                if not r:
                    flash("Student not found.", "warning")
                    return redirect(url_for("holiday_hub"))
                sid = r["id"]

            subject_id = request.form.get("subject_id")
            subj_name = (request.form.get("subject_name") or "").strip()
            if not subject_id and subj_name:
                subject_id = get_or_create_subject_by_name(subj_name)

            try:
                hp_add_score(
                    int(sid),
                    int(subject_id),
                    request.form.get("assessment_type") or "Test",
                    float(request.form.get("score") or 0),
                    request.form.get("term") or active_term,
                    int(request.form.get("year") or active_year),
                    float(request.form.get("max_score") or 100),
                    request.form.get("weight")
                )
                flash("Holiday score recorded.", "success")
            except Exception as e:
                flash(f"Failed: {e}", "danger")
            return redirect(url_for("holiday_hub"))

        if action == "upload_csv":
            file = request.files.get("file")
            if not file:
                flash("No file", "warning")
                return redirect(url_for("holiday_hub"))

            if file.filename.lower().endswith(".csv"):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)

            ok, skip = 0, 0
            conn = get_db_connection()
            for _, r in df.iterrows():
                try:
                    cur = conn.cursor(dictionary=True)
                    cur.execute(
                        "SELECT id FROM students WHERE student_number=%s",
                        (str(r["student_number"]),)
                    )
                    rs = cur.fetchone()
                    cur.close()
                    if not rs:
                        skip += 1
                        continue
                    sid = rs["id"]
                    subj = get_or_create_subject_by_name(
                        str(r["subject_name"]))
                    hp_add_score(
                        sid, subj,
                        str(r.get("assessment_type", "Test")),
                        float(r["score"]),
                        str(r.get("term", active_term)),
                        int(r.get("year", active_year)),
                        float(r.get("max_score", 100)),
                        r.get("weight", None)
                    )
                    ok += 1
                except Exception:
                    skip += 1
            conn.close()
            flash(f"Uploaded {ok}, skipped {skip}.", "info")
            return redirect(url_for("holiday_hub"))

        if action == "sync":
            class_name = (request.form.get("class_name") or "").strip()
            try:
                n = hp_sync_into_record_score(
                    class_name, active_term, active_year, session.get("initials") or "HP")
                flash(f"Synced {n} scores", "success")
            except Exception as e:
                flash(f"Sync failed: {e}", "danger")
            return redirect(url_for("holiday_hub"))

    # -------- GET: options + (NEW) preview --------
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name")
    classes = [r["classes"] for r in cur.fetchall()]
    cur.close()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id,name FROM subjects ORDER BY name")
    subjects = cur.fetchall()
    cur.close()

    # Preview params (optional)
    preview_class = (request.args.get("class_name") or "").strip()
    preview_term = (request.args.get("term") or active_term).strip()
    preview_year = int(request.args.get("year") or active_year)
    do_preview = request.args.get("preview") == "1"

    preview = None
    if do_preview and preview_class:
        subj_list = _hp_subjects_for_class_term_year(
            conn, preview_class, preview_term, preview_year)
        cur = conn.cursor(dictionary=True)
        cur.execute("""
          SELECT id, student_number, first_name, Middle_name, last_name
          FROM students
          WHERE archived=0 AND class_name=%s
          ORDER BY last_name, first_name
        """, (preview_class,))
        students = cur.fetchall()
        cur.close()

        rows = []
        for s in students:
            row = {
                "student_id": s["id"],
                "student_number": s["student_number"],
                "name": f'{s["first_name"]} {s["Middle_name"] or ""} {s["last_name"]}'.replace(" ", " ").strip(),
                "subjects": {}
            }
            for sj in subj_list:
                avg = hp_aggregate_student_subject(
                    conn, s["id"], sj["id"], preview_term, preview_year)
                row["subjects"][sj["id"]] = avg


# ========================= STUDENTS FINANCE REPORT =========================

@app.route("/reports/students_finance", methods=["GET"])
@require_role("admin", "bursar", "headteacher", "director")
def students_finance_report():
    f = _filters_from_request()
    term, year = (f.get("term") or "").strip(), f.get("year")

    if not term or not year:
        ay, tno = _active_year_term()  # returns (year, term_no)
        rev = {1: "Term 1", 2: "Term 2", 3: "Term 3"}
        if not year:
            year = ay
        if not term:
            term = rev.get(tno, "Term 1")
        f["year"], f["term"] = year, term

    term_no = {'term 1': 1, 'term 2': 2, 'term 3': 3}.get(term.strip().lower(), 1)

    # students base
    where = ["s.archived=0"]
    params = [year, term_no]

    if f.get("student_number"):
        where.append("s.student_number=%s"); params.append(f["student_number"])
    if f.get("last_name"):
        where.append("s.last_name LIKE %s"); params.append(f"%{f['last_name']}%")
    if f.get("class_name"):
        where.append("s.class_name=%s"); params.append(f["class_name"])

    where_sql = " AND ".join(where)

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute(f"""
        SELECT
          s.student_number,
          CONCAT_WS(' ', s.first_name, s.last_name) AS full_name,
          s.class_name, s.stream,

          '' AS transport_route,  /* placeholder; your template expects it */

          /* Fees */
          COALESCE(fs.expected_fees,0)        AS fees_expected,
          COALESCE(fs.paid_fees_nonvoid,0)    AS fees_paid,
          GREATEST(COALESCE(fs.fees_expected_net,0) - COALESCE(fs.paid_fees_nonvoid,0), 0) AS fees_balance,

          /* Requirements + Transport */
          (COALESCE(fs.expected_reqs_base,0) + COALESCE(fs.transport_due_term,0)) AS req_expected,
          COALESCE(fs.paid_reqs_nonvoid,0)    AS req_paid,
          GREATEST(
            (COALESCE(fs.expected_reqs_base,0) + COALESCE(fs.transport_due_term,0))
            - COALESCE(fs.paid_reqs_nonvoid,0), 0
          ) AS req_balance,

          COALESCE(fs.transport_due_term,0)   AS tr_due,
          COALESCE(fs.transport_paid_term,0)  AS tr_paid,

          /* Carry forward (can be negative = credit) */
          COALESCE(fs.carry_forward,0)        AS carried_forward,

          /* Overall (exclude CF in expected, add CF in outstanding) */
          (COALESCE(fs.fees_expected_net,0) + COALESCE(fs.req_expected_final,0)) AS overall_expected,
          COALESCE(fs.overall_paid,0)         AS overall_paid,

          /* ✅ SIGNED overall outstanding (credit allowed) */
          (COALESCE(fs.carry_forward,0) + (
              (COALESCE(fs.fees_expected_net,0) + COALESCE(fs.req_expected_final,0))
              - COALESCE(fs.overall_paid,0)
          )) AS overall_outstanding

        FROM students s
        LEFT JOIN fee_term_summary fs
          ON fs.student_id = s.id AND fs.year = %s AND fs.term_no = %s
        WHERE {where_sql}
        ORDER BY s.class_name, s.stream, s.last_name, s.first_name
    """, params)

    rows = cur.fetchall() or []
    cur.close()
    conn.close()

    totals = {
        "fees_expected": 0.0, "fees_paid": 0.0, "fees_balance": 0.0,
        "req_expected": 0.0, "req_paid": 0.0, "req_balance": 0.0,
        "tr_due": 0.0, "tr_paid": 0.0, "tr_balance": 0.0,
        "carried_forward": 0.0,
        "overall_expected": 0.0, "overall_paid": 0.0, "overall_outstanding": 0.0
    }

    for r in rows:
        # transport balance for template
        tr_due = float(r.get("tr_due") or 0.0)
        tr_paid = float(r.get("tr_paid") or 0.0)
        r["tr_balance"] = max(tr_due - tr_paid, 0.0)

        # roll totals (must match template keys)
        for k in totals.keys():
            totals[k] += float(r.get(k, 0) or 0)

    return render_template(
        "students_finance_report.html",
        rows=rows, totals=totals, terms=TERMS, filters=f,
        class_options=_class_options()
    )

@app.route("/reports/students_finance/export.csv")
@require_role("admin", "bursar", "headteacher", "director")
def export_students_finance_csv():
    f = _filters_from_request()
    term, year = f.get("term"), f.get("year")
    if not term or not year:
        ay, at = _active_year_term(); year = year or ay; term = term or at
        f["year"], f["term"] = year, term
    term_no = {'term 1':1,'term 2':2,'term 3':3}[term.strip().lower()]

    # Base filters (same as screen)
    where = ["s.archived=0"]
    params = [year, term_no]
    if f.get("student_number"):
        where.append("s.student_number=%s"); params.append(f["student_number"])
    if f.get("last_name"):
        where.append("s.last_name LIKE %s"); params.append(f"%{f['last_name']}%")
    if f.get("class_name"):
        where.append("s.class_name=%s"); params.append(f["class_name"])
    where_sql = " AND ".join(where)

    conn = get_db_connection(); cur = conn.cursor(dictionary=True)
    cur.execute(f"""
        SELECT
          s.student_number,
          CONCAT_WS(' ', s.first_name, s.last_name) AS full_name,
          s.class_name, s.stream,

          COALESCE(fs.expected_fees,0)             AS fees_expected,
          COALESCE(fs.paid_fees_nonvoid,0)         AS fees_paid,
          GREATEST(COALESCE(fs.fees_expected_net,0) - COALESCE(fs.paid_fees_nonvoid,0), 0) AS fees_balance,

          COALESCE(fs.expected_reqs_base,0)        AS req_expected_base,
          COALESCE(fs.transport_due_term,0)        AS tr_due,
          COALESCE(fs.paid_reqs_nonvoid,0)         AS req_paid,
          COALESCE(fs.transport_paid_term,0)       AS tr_paid,

          COALESCE(fs.expected_reqs_base,0) + COALESCE(fs.transport_due_term,0) AS req_expected,
          GREATEST(
            (COALESCE(fs.expected_reqs_base,0) + COALESCE(fs.transport_due_term,0))
            - COALESCE(fs.paid_reqs_nonvoid,0), 0
          ) AS req_balance,

          COALESCE(fs.carry_forward,0)             AS carried_forward,

          COALESCE(fs.fees_expected_net,0) + COALESCE(fs.req_expected_final,0) AS overall_expected,
          COALESCE(fs.overall_paid,0)             AS overall_paid,
          COALESCE(fs.overall_outstanding,0)      AS overall_outstanding
        FROM students s
        LEFT JOIN fee_term_summary fs
          ON fs.student_id = s.id AND fs.year = %s AND fs.term_no = %s
        WHERE {where_sql}
        ORDER BY s.class_name, s.stream, s.last_name, s.first_name
    """, params)
    rows = cur.fetchall() or []
    cur.close(); conn.close()

    # (Optional) route name lookup — keep fast by skipping, or enable if needed:
    # If you want route names included, uncomment this block (requires transport_subscription_info):
    # def _route_name_for(sid, term, year):
    #     try:
    #         t = transport_subscription_info(sid=None, term=term, year=year)  # adapt to your signature if needed
    #     except Exception:
    #         t = None
    #     return (t or {}).get("route_name") or ""
    # and then add "transport_route" column in export with the value.

    # Build CSV
    si = StringIO(); writer = csv.writer(si)
    headers = [
        "Student Number","Name","Class","Stream",
        "Fees Expected","Fees Paid","Fees Balance",
        "Req Expected","Req Paid","Req Balance",
        "Transport Due","Transport Paid","Transport Balance",
        "Carry Forward",
        "Overall Expected","Overall Paid","Overall Outstanding"
    ]
    writer.writerow(headers)
    for r in rows:
        tr_bal = max(float(r["tr_due"] or 0) - float(r["tr_paid"] or 0), 0.0)
        writer.writerow([
            r["student_number"], r["full_name"], r["class_name"], r["stream"],
            r["fees_expected"], r["fees_paid"], r["fees_balance"],
            r["req_expected"], r["req_paid"], r["req_balance"],
            r["tr_due"], r["tr_paid"], tr_bal,
            r["carried_forward"],
            r["overall_expected"], r["overall_paid"], r["overall_outstanding"]
        ])

    si.seek(0)
    filename = f"Students_Finance_Report_{term}_{year}.csv"
    return Response(si.getvalue(),
                    mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


# ---------- STUDENT STATEMENT ----------

# Use this name (or keep your renamed 'student_statement') consistently in url_for(...)

@app.route("/student-statement", methods=["GET"])
@require_role("admin", "director", "bursar", "headteacher", "deputyheadteacher")
def student_statement():
    """
    If no query is provided, render a simple search form.
    If student_number or last_name is provided, find the student and redirect to their statement.
    """
    sn = (request.args.get("student_number") or "").strip()
    ln = (request.args.get("last_name") or "").strip()

    # No input → show search page (do NOT redirect)
    if not sn and not ln:
        # Optional: show the info message on the page instead of flashing
        return render_template("student_statement_search.html")

    # With input → do lookup
    conn = get_db_connection()
    try:
        if sn:
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                SELECT id, student_number, first_name,
                       COALESCE(Middle_name,'') AS middle_name, last_name,
                       class_name, stream, section, sex, parent_name, parent_contact
                FROM students
                WHERE student_number=%s AND archived=0
                LIMIT 1
            """, (sn,))
            stu = cur.fetchone()
            cur.close()
        else:
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                SELECT id, student_number, first_name,
                       COALESCE(Middle_name,'') AS middle_name, last_name,
                       class_name, stream, section, sex, parent_name, parent_contact
                FROM students
                WHERE last_name LIKE %s AND archived=0
                ORDER BY last_name, first_name
                LIMIT 1
            """, (f"%{ln}%",))
            stu = cur.fetchone()
            cur.close()
    finally:
        conn.close()

    if not stu:
        flash("Student not found or archived.", "warning")
        return redirect(url_for("student_statement"))

    return redirect(url_for("student_statement_by_id", student_id=stu["id"]))


@app.route("/student-statement/<int:student_id>", methods=["GET"])
@require_role("admin", "bursar", "headteacher", "director", "deputyheadteacher")
def student_statement_by_id(student_id: int):
    """
    Printable Student Statement (transactions + summary).
    - Uses compute_student_financials() for active term/year summary (credit-aware).
    - Keeps receipt_no backfill logic.
    """
    # ensure schema has the receipt_no column
    try:
        ensure_fees_has_receipt_no()
    except Exception:
        pass

    ay = get_active_academic_year() or {}
    term = (ay.get("current_term") or ay.get("term") or "Term 1")
    try:
        year = int(ay.get("year") or ay.get("active_year") or datetime.now().year)
    except Exception:
        year = datetime.now().year

    conn = get_db_connection()
    try:
        # student
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, student_number, first_name,
                   COALESCE(Middle_name,'') AS middle_name, last_name,
                   class_name, stream, section, sex, parent_name, parent_contact, residence, photo
            FROM students
            WHERE id=%s LIMIT 1
        """, (student_id,))
        stu = cur.fetchone()
        cur.close()

        if not stu:
            flash("Student not found.", "warning")
            return redirect(url_for("register_student"))

        # transactions (full history)
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, term, year, date_paid, method, payment_type,
                   amount_paid, expected_amount, bursary_amount, carried_forward,
                   COALESCE(requirement_name, '') AS requirement_name,
                   COALESCE(comment, '') AS comment,
                   receipt_no
            FROM fees
            WHERE student_id = %s
            ORDER BY year, term_order, id
        """, (student_id,))
        tx = cur.fetchall() or []
        cur.close()

        # backfill missing receipt numbers
        missing_ids = [row["id"] for row in tx if not row.get("receipt_no")]
        if missing_ids:
            for fee_id in missing_ids:
                rcpt = generate_receipt_no(conn, fee_id)
                cur = conn.cursor(dictionary=True)
                try:
                    cur.execute("UPDATE fees SET receipt_no=%s WHERE id=%s", (rcpt, fee_id))
                except mysql.connector.Error:
                    rcpt2 = f"{rcpt}-{fee_id%1000:03d}"
                    cur.execute("UPDATE fees SET receipt_no=%s WHERE id=%s", (rcpt2, fee_id))
                cur.close()
            conn.commit()

            # re-fetch (so template shows filled receipts)
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                SELECT id, term, year, date_paid, method, payment_type,
                       amount_paid, expected_amount, bursary_amount, carried_forward,
                       COALESCE(requirement_name, '') AS requirement_name,
                       COALESCE(comment, '') AS comment,
                       receipt_no
                FROM fees
                WHERE student_id = %s
                ORDER BY year, term_order, id
            """, (student_id,))
            tx = cur.fetchall() or []
            cur.close()

    finally:
        conn.close()

    # summary for ACTIVE term/year (this now reflects OB→CF via procedure)
    fin = compute_student_financials(student_id, stu["class_name"], term, year)

    grouped = defaultdict(lambda: defaultdict(list))
    for r in tx:
        grouped[r["year"]][r["term"]].append(r)

    school = {
        "name": current_app.config.get("SCHOOL_NAME", "DEMO DAY AND BOARDING PRIMARY SCHOOL – KAMPALA"),
        "address": current_app.config.get("SCHOOL_ADDRESS", "P.O Box 1X1X1 Kampala"),
        "phone": current_app.config.get("SCHOOL_PHONE", "+256778878411, +256759685640, +256773589232, +256750347624"),
    }

    return render_template(
        "student_statement.html",
        school=school,
        student=stu,
        grouped=grouped,
        fin=fin,
        active_term=term,
        active_year=year,
        today=datetime.now().strftime("%d %b %Y %H:%M"),
    )
    
    

@app.route("/student-statement/<int:student_id>/pdf", methods=["GET"])
@require_role("admin", "bursar", "headteacher", "director", "deputyheadteacher")
def student_statement_pdf(student_id: int):
    """
    Student Statement (PDF)
    - Header with logo on the left, centered school info
    - Summary table (two narrow columns: label/value; will generate multiple rows)
    - Transactions table (Receipt, Date, Term, Type, Method, Item, Paid, Comment)
    - Auto page-breaks; no overlapping; compact but readable
    """
    # ----- active session -----
    ay = get_active_academic_year() or {}
    term = (ay.get("current_term") or ay.get("term") or "Term 1")
    try:
        year = int(ay.get("year") or datetime.now().year)
    except Exception:
        year = datetime.now().year

    # ----- data: student + transactions -----
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS middle_name, last_name,
                   class_name, stream, section, parent_name, parent_contact, photo
            FROM students WHERE id=%s LIMIT 1
        """, (student_id,))
        stu = cur.fetchone()
        cur.close()
        if not stu:
            flash("Student not found.", "warning")
            return redirect(url_for("student_statement"))
        
        
        # NEW: resolve student photo path (if any)
        photo_path = None
        rel = (stu.get("photo") or "").strip()
        if rel:
            # if stored like 'static/uploads/...' use app root
            if rel.startswith("static/"):
                photo_path = os.path.join(current_app.root_path, rel)
            else:
                # fallback: treat as inside static folder
                photo_path = os.path.join(current_app.static_folder, rel)
            if not os.path.exists(photo_path):
                photo_path = None


        # ensure receipt numbers exist
        try:
            ensure_fees_has_receipt_no()
        except Exception:
            pass

        # full transaction history
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, date_paid, term, year, method, payment_type,
                   amount_paid, COALESCE(requirement_name,'') AS requirement_name,
                   COALESCE(comment,'') AS comment, receipt_no
            FROM fees
            WHERE student_id=%s
            ORDER BY year,
                     CASE LOWER(term)
                       WHEN 'term 1' THEN 1 WHEN 'term 2' THEN 2 WHEN 'term 3' THEN 3 ELSE 99 END,
                     id
        """, (student_id,))
        tx = cur.fetchall() or []
        cur.close()

        # backfill missing receipt numbers (rare)
        missing = [r["id"] for r in tx if not r.get("receipt_no")]
        if missing:
            for fee_id in missing:
                try:
                    cur = conn.cursor(dictionary=True)
                    rcpt = generate_receipt_no(conn, fee_id)
                    cur.execute("UPDATE fees SET receipt_no=%s WHERE id=%s", (rcpt, fee_id))
                    cur.close()
                except Exception:
                    pass
            conn.commit()
            # re-fetch to include backfilled numbers
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                SELECT id, date_paid, term, year, method, payment_type,
                       amount_paid, COALESCE(requirement_name,'') AS requirement_name,
                       COALESCE(comment,'') AS comment, receipt_no
                FROM fees
                WHERE student_id=%s
                ORDER BY year,
                         CASE LOWER(term)
                           WHEN 'term 1' THEN 1 WHEN 'term 2' THEN 2 WHEN 'term 3' THEN 3 ELSE 99 END,
                         id
            """, (student_id,))
            tx = cur.fetchall() or []
            cur.close()
    finally:
        conn.close()

    # summary (current active term/year)
    fin = compute_student_financials(student_id, stu["class_name"], term, year)

    # group transactions (not strictly needed for PDF, but handy if you want to section later)
    grouped = defaultdict(list)
    for r in tx:
        grouped[(r["year"], r["term"])].append(r)

    # ----- school meta & logo path -----
    school = {
        "name": current_app.config.get("SCHOOL_NAME", "DEMO DAY AND BOARDING PRIMARY SCHOOL – KAMPALA"),
        "address": current_app.config.get("SCHOOL_ADDRESS", "P.O Box 1X1X1 Kampala"),
        "phone": current_app.config.get("SCHOOL_PHONE", "+256778878411, +256759685640, +256773589232, +256750347624"),
        }
    logo_path = os.path.join(current_app.static_folder, "logo.jpg")
    # ----- styles -----
    styles = getSampleStyleSheet()
    base_font = 9
    styles.add(ParagraphStyle(
        name="Body",
        fontName="Helvetica",
        fontSize=base_font,
        leading=int(base_font * 1.5), # ~1.5 line spacing
    ))
    styles.add(ParagraphStyle(
        name="BodyMono",
        fontName="Courier",
        fontSize=base_font-1,
        leading=int((base_font-1) * 1.4),
    ))
    p = lambda t: Paragraph(str(t), styles["Body"])
    pmono = lambda t: Paragraph(str(t), styles["BodyMono"])

    # ----- header (canvas) -----
    def draw_header(canvas, doc):
        canvas.saveState()
        left = doc.leftMargin
        right = doc.pagesize[0] - doc.rightMargin
        top_y = doc.pagesize[1] - 10*mm # header baseline

        # blue line
        canvas.setStrokeColorRGB(0.05, 0.43, 0.99)
        canvas.setLineWidth(2)
        canvas.line(left, top_y - 16*mm, right, top_y - 16*mm)

        # logo left
        try:
            canvas.drawImage(logo_path, left, top_y - 16*mm, width=20*mm, height=20*mm,
                             preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

        # centered text
        cx = (left + right) / 2.0
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawCentredString(cx, top_y - 4*mm, school["name"])
        canvas.setFont("Helvetica", 9)
        canvas.drawCentredString(cx, top_y - 9*mm, "The early bird catches the worm")
        canvas.setFont("Helvetica", 8.5)
        canvas.drawCentredString(cx, top_y - 13*mm, f"Tel: {school['phone']} | {school['address']}")

        canvas.restoreState()

    # ----- document -----
    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=32*mm, # leave space under header line
        bottomMargin=12*mm,
    )

    story = []
    story.append(Spacer(1, 4*mm))

    # ---- STUDENT INFO (4 cols + photo col) ----
    name = f"{stu['first_name']} {stu.get('middle_name') or ''} {stu['last_name']}".strip()

    # Build the image element (or empty if none)
    photo_el = ""
    if photo_path:
        try:
            photo_el = RLImage(photo_path, width=22*mm, height=26*mm)
        except Exception:
            photo_el = ""

    info_rows = [
        # row 0: student + stud no + photo
        [
            p("<b>Student</b>"), p(name),
            p("<b>Stud No.</b>"), p(stu["student_number"]),
            photo_el,
        ],
        # row 1: class/stream + section
        [
            p("<b>Class/Str</b>"),
            p(f"{stu['class_name']} {stu.get('stream') or ''}".strip()),
            p("<b>Sect.</b>"),
            p(stu.get("section") or "-"),
            "",
        ],
        # row 2: parent + contact
        [
            p("<b>Parent</b>"),
            p(stu.get("parent_name") or "-"),
            p("<b>Contact</b>"),
            p(stu.get("parent_contact") or "-"),
            "",
        ],
        # row 3: statement + active session
        [
            p("<b>Statement</b>"),
            p(f"Generated: {datetime.now():%d %b %Y %H:%M}"),
            p("<b>Session</b>"),
            p(f"{term} {year}"),
            "",
        ],
    ]

    info_tbl = Table(
        info_rows,
        colWidths=[28*mm, 50*mm, 25*mm, 45*mm, 24*mm],
        hAlign="LEFT",
    )

    info_style = [
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("BACKGROUND", (2, 0), (2, -1), colors.whitesmoke),
        ("BACKGROUND", (1, 0), (1, -1), colors.white),
        ("BACKGROUND", (3, 0), (3, -1), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("FONTSIZE", (0, 0), (-1, -1), base_font),
    ]

    # If we actually have a photo, span it across all rows in the last column
    if photo_el:
        info_style.append(("SPAN", (4, 0), (4, 3)))
        info_style.append(("VALIGN", (4, 0), (4, 3), "MIDDLE"))

    info_tbl.setStyle(TableStyle(info_style))
    story.append(info_tbl)
    story.append(Spacer(1, 4*mm))
    
    # small helpers you likely already have
    def fmt_amt(x): 
        try: 
            return f"UGX {float(x):,.0f}"
        except Exception:
            return "UGX 0"

    # Build the flat list of (label, value) first
    sum_rows = [
        ("Expected Fees (Term)", fmt_amt(fin.get("expected_fees", 0))),
        ("Expected Requirements", fmt_amt(fin.get("expected_requirements", 0))),
        ("Bursary (Term)", fmt_amt(fin.get("bursary_current", 0))),
        ("Paid Fees (Term)", fmt_amt(fin.get("paid_fees", 0))),
        ("Paid Requirements (Term)", fmt_amt(fin.get("paid_requirements", 0))),
        ("Carry Forward (incl. OB)", fmt_amt(fin.get("carry_forward", 0))),
        ("Total Due (This Term)", fmt_amt(fin.get("total_due_this_term", 0))),
        ("Balance (This Term)", fmt_amt(fin.get("balance_this_term", 0))),
        ("Overall Outstanding", fmt_amt(fin.get("overall_balance", 0))),
    ]

    # Convert to 4-column rows: [label1, value1, label2, value2]
    sum_rows_4col = []
    for i in range(0, len(sum_rows), 2):
        left = sum_rows[i]
        right = sum_rows[i+1] if i+1 < len(sum_rows) else ("", "")
        sum_rows_4col.append([
            p(f"<b>{left[0]}</b>"), p(left[1]),
            p(f"<b>{right[0]}</b>"), p(right[1]),
        ])

    sum_tbl = Table(
        sum_rows_4col,
        colWidths=[40*mm, 45*mm, 40*mm, None], # balanced across the page
        hAlign="LEFT",
    )
    sum_tbl.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("BACKGROUND", (0,0), (0,-1), colors.whitesmoke),
        ("BACKGROUND", (2,0), (2,-1), colors.whitesmoke),
        ("BACKGROUND", (1,0), (1,-1), colors.white),
        ("BACKGROUND", (3,0), (3,-1), colors.white),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING",(0,0), (-1,-1), 4),
        ("FONTSIZE", (0,0), (-1,-1), base_font),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 5*mm))


    # --- Transactions table (compact, no overlap). Keep columns minimal. ---
    # Columns: Receipt, Date, Term, Type, Method, Item, Paid, Comment
    tx_header = ["Receipt", "Date", "Term", "Type", "Method", "Item", "Paid (UGX)", "Comment"]
    rows = [tx_header]
    for r in tx:
        # Normalize date
        d = r.get("date_paid")
        if isinstance(d, (datetime, date)):
            d = d.strftime("%Y-%m-%d")
        else:
            d = (str(d) or "")[:10]

        # shorten receipt visually but preserve value (monospace, can wrap)
        receipt = r.get("receipt_no") or "-"
        rows.append([
            pmono(receipt),
            p(d),
            p(r.get("term") or ""),
            p((r.get("payment_type") or "").capitalize()),
            p(r.get("method") or ""),
            p(r.get("requirement_name") or "—"),
            p(f"{float(r.get('amount_paid') or 0):,.0f}"),
            p(r.get("comment") or ""),
        ])

    # widths tuned to avoid overlap; table will flow to multiple pages as needed
    col_widths = [30*mm, 20*mm, 18*mm, 20*mm, 22*mm, 48*mm, 24*mm, None]
    tx_tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tx_tbl.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f0f4ff")),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold", base_font),
        ("FONT", (0,1), (-1,-1), "Helvetica", base_font-0.5),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ALIGN", (6,1), (6,-1), "RIGHT"), # Paid column right-aligned
    ]))
    story.append(tx_tbl)

    # footer note
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph('<font size="8" color="#777777">This is a system-generated statement.</font>', styles["Body"]))

    # ----- build -----
    def _hdr(c, d): # wrapper to pass header data
        draw_header(c, d)

    doc.build(story, onFirstPage=_hdr, onLaterPages=_hdr)

    bio.seek(0)
    fname = f"Statement_{stu['student_number']}.pdf"
    return send_file(bio, as_attachment=True, download_name=fname, mimetype="application/pdf")
# ---------- /STUDENT STATEMENT ----------


@app.route("/mark_sheet", methods=["GET"])
@require_role("admin", "teacher", "headteacher", "dos", "classmanager", "deputyheadteacher")
def mark_sheet():
    # ---- filters ----
    ay = get_active_academic_year()
    class_name = (request.args.get("class_name") or "P7").strip()
    stream = (request.args.get("stream") or "").strip()
    term = (request.args.get("term") or ay["current_term"]).strip()
    year = int(request.args.get("year") or ay["year"])
    school_title = current_app.config.get("SCHOOL_NAME", "Your School")

    conn = get_db_connection()

    # subjects that actually have marks, cores first
    subjects = subjects_with_marks(conn, class_name, stream, term, year)
    subjects = sorted(
        subjects,
        key=lambda s: (0 if ((s.get("code") or s["name"]).strip().upper() in CORE_CODES) else 1, s["name"])
    )

    
    # ---- fetch available streams for this class ----
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT DISTINCT COALESCE(stream, '') AS stream
        FROM students
        WHERE class_name = %s AND archived = 0
        ORDER BY stream
    """, (class_name,))
    stream_rows = cur.fetchall() or []
    streams = [r["stream"] for r in stream_rows]
    

    # students
    cur = conn.cursor(dictionary=True)

    params = [class_name]
    extra_stream_where = ""

    # if a specific stream was chosen, filter by it; if stream == "" we use ALL
    if stream:
        extra_stream_where = " AND COALESCE(stream,'') = %s"
        params.append(stream)

    cur.execute(f"""
        SELECT id, student_number,
               first_name, COALESCE(Middle_name,'') AS middle_name, last_name,
               class_name, COALESCE(stream,'') AS stream
        FROM students
        WHERE archived = 0
          AND class_name = %s
          {extra_stream_where}
        ORDER BY last_name, first_name
    """, params)
    students = cur.fetchall() or []


    # raw scores (components + average_mark) – for the students we already selected
    cur = conn.cursor(dictionary=True)
    student_ids = [s["id"] for s in students]
    score_rows = []

    if student_ids:
        fmt_ids = ",".join(["%s"] * len(student_ids))
        cur.execute(f"""
            SELECT rs.student_id, rs.subject_id,
                   rs.other_mark, rs.holiday_mark, rs.bot_mark, rs.midterm_mark, rs.eot_mark, rs.ca_mark,
                   rs.average_mark
              FROM record_score rs
             WHERE rs.term = %s AND rs.year = %s
               AND rs.student_id IN ({fmt_ids})
        """, [term, year, *student_ids])
        score_rows = cur.fetchall() or []
    else:
        score_rows = []

    # === grading scale (once) ===
    gcur = conn.cursor(dictionary=True)
    gcur.execute("""
        SELECT grade, lower_limit, upper_limit
          FROM grading_scale
         ORDER BY lower_limit DESC
    """)
    SCALE = gcur.fetchall() or []
    gcur.close()

    def grade_from_scale(score: float) -> str:
        try:
            s = float(score)
        except Exception:
            return "NG"
        for r in SCALE:
            lo, hi = float(r["lower_limit"]), float(r["upper_limit"])
            if lo <= s <= hi:
                return (r["grade"] or "").strip()
        return "NG"

    def pick_average_row_score(r: dict):
        if r.get("average_mark") is not None:
            return float(r["average_mark"])
        return _mean_nonnull([r.get(k) for k in COMPONENT_FIELDS])

    per = defaultdict(dict)
    for r in score_rows:
        sc = pick_average_row_score(r)
        if sc is not None:
            per[r["student_id"]][r["subject_id"]] = sc
            

    # ✅ IMPORTANT: remove subjects that have NO usable marks in this class/stream/term/year
    if student_ids and subjects:
        subj_keep = []
        for subj in subjects:
            sid_subj = subj["id"]
            has_any = any(per.get(stid, {}).get(sid_subj) is not None for stid in student_ids)
            if has_any:
                subj_keep.append(subj)
        subjects = subj_keep


    # core detection for subjects list
    def _is_core(subj: dict) -> bool:
        code = (subj.get("code") or "").strip().upper()
        name = (subj.get("name") or "").strip().lower()
        if code in CORE_CODES: return True
        if name.startswith("eng"): return True
        if name.startswith(("mat", "math")): return True
        if name.startswith("sci"): return True
        if name in {"sst", "soc. studies", "social studies", "social std", "socialstudies"}: return True
        return False

    rows = []
    for idx, s in enumerate(students, start=1):
        sid = s["id"]
        cells = []
        total = 0.0
        n = 0
        agg_sum = 0
        agg_cnt = 0

        for subj in subjects:
            m = per.get(sid, {}).get(subj["id"])
            if m is None:
                cells.append({"text": ""})
                continue
            g = grade_from_scale(m) or ""
            cells.append({"text": f"{m:.0f} ({g})"})
            total += m; n += 1
            if _is_core(subj) and g in AGG_MAP:
                agg_sum += AGG_MAP[g]; agg_cnt += 1

        ave = round(total / n, 1) if n else 0.0
        agg = agg_sum if agg_cnt == 4 else None
        if agg is not None:
            div = division_from_aggregate(agg)
        else:
            div = "NG" # consistent when cores missing

        rows.append({
            "no": idx,
            "student_number": s["student_number"],
            "full_name": f"{s['first_name']} {s['middle_name']} {s['last_name']}".replace(" "," ").strip(),
            "cells": cells,
            "total": round(total, 0),
            "ave": ave,
            "agg": agg if agg is not None else "",
            "div": div,
        })

    # positions by total (ties share rank)
    ranked = sorted(rows, key=lambda r: (-r["total"], r["full_name"]))
    pos = 0; seen = 0; prev = None
    for r in ranked:
        seen += 1
        if r["total"] != prev:
            pos = seen; prev = r["total"]
        r["pos"] = pos
    pos_by_stuno = {r["student_number"]: r["pos"] for r in ranked}
    for r in rows:
        r["pos"] = pos_by_stuno.get(r["student_number"], "")

    conn.close()

    return render_template(
        "mark_sheet.html",
        school_title=school_title,
        class_name=class_name, stream=stream, term=term, year=year,
        subjects=subjects,
        rows=rows,
        streams=streams
    )



@app.route("/mark_sheet/pdf_reportlab")
@require_role("admin", "teacher", "headteacher", "dos", "classmanager", "deputyheadteacher")
def mark_sheet_pdf_reportlab():
    import os
    from io import BytesIO
    from collections import defaultdict
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm

    # ---- filters ----
    ay = get_active_academic_year()
    class_name = (request.args.get("class_name") or "P7").strip()
    stream = (request.args.get("stream") or "").strip()
    term = (request.args.get("term") or ay["current_term"]).strip()
    year = int(request.args.get("year") or ay["year"])
    school_title = current_app.config.get("SCHOOL_NAME", "Your School")

    # ---- data (same logic as HTML route) ----
    conn = get_db_connection()

    subjects = subjects_with_marks(conn, class_name, stream, term, year)
    subjects = sorted(
        subjects,
        key=lambda s: (0 if ((s.get("code") or s["name"]).strip().upper() in CORE_CODES) else 1, s["name"])
    )

    cur = conn.cursor(dictionary=True)
    params = [class_name]
    extra_stream_where = ""

    # if a specific stream was chosen, filter by it; if stream == "" we use ALL
    if stream:
        extra_stream_where = " AND COALESCE(stream,'') = %s"
        params.append(stream)

    cur.execute(f"""
        SELECT id, student_number,
               first_name, COALESCE(Middle_name,'') AS middle_name, last_name,
               class_name, COALESCE(stream,'') AS stream
        FROM students
        WHERE archived = 0
          AND class_name = %s
          {extra_stream_where}
        ORDER BY last_name, first_name
    """, params)
    students = cur.fetchall() or []

    cur = conn.cursor(dictionary=True)
    student_ids = [s["id"] for s in students]
    score_rows = []

    if student_ids:
        fmt_ids = ",".join(["%s"] * len(student_ids))
        cur.execute(f"""
            SELECT rs.student_id, rs.subject_id,
                   rs.other_mark, rs.holiday_mark, rs.bot_mark, rs.midterm_mark, rs.eot_mark, rs.ca_mark,
                   rs.average_mark
              FROM record_score rs
             WHERE rs.term = %s AND rs.year = %s
               AND rs.student_id IN ({fmt_ids})
        """, [term, year, *student_ids])
        score_rows = cur.fetchall() or []
    else:
        score_rows = []

    # grading scale (once)
    gcur = conn.cursor(dictionary=True)
    gcur.execute("""
        SELECT grade, lower_limit, upper_limit
          FROM grading_scale
         ORDER BY lower_limit DESC
    """)
    SCALE = gcur.fetchall() or []
    gcur.close()

    def grade_from_scale(score: float) -> str:
        try:
            s = float(score)
        except Exception:
            return "NG"
        for r in SCALE:
            lo, hi = float(r["lower_limit"]), float(r["upper_limit"])
            if lo <= s <= hi:
                return (r["grade"] or "").strip()
        return "NG"

    def pick_average_row_score(r: dict):
        if r.get("average_mark") is not None:
            return float(r["average_mark"])
        return _mean_nonnull([r.get(k) for k in COMPONENT_FIELDS])

    per = defaultdict(dict)
    for r in score_rows:
        sc = pick_average_row_score(r)
        if sc is not None:
            per[r["student_id"]][r["subject_id"]] = sc
            

    # ✅ IMPORTANT: keep only subjects that have at least one usable mark
    if student_ids and subjects:
        subj_keep = []
        for subj in subjects:
            sid_subj = subj["id"]
            has_any = any(per.get(stid, {}).get(sid_subj) is not None for stid in student_ids)
            if has_any:
                subj_keep.append(subj)
        subjects = subj_keep

    def _is_core(subj: dict) -> bool:
        code = (subj.get("code") or "").strip().upper()
        name = (subj.get("name") or "").strip().lower()
        if code in CORE_CODES: return True
        if name.startswith("eng"): return True
        if name.startswith(("mat", "math")): return True
        if name.startswith("sci"): return True
        if name in {"sst", "soc. studies", "social studies", "social std", "socialstudies"}: return True
        return False

    rows = []
    for i, s in enumerate(students, start=1):
        sid = s["id"]
        cells = []; total = 0.0; n = 0; agg_sum = 0; agg_cnt = 0
        for subj in subjects:
            m = per.get(sid, {}).get(subj["id"])
            if m is None:
                cells.append("")
                continue
            g = grade_from_scale(m) or ""
            cells.append(f"{m:.0f} ({g})")
            total += m; n += 1
            if _is_core(subj) and g in AGG_MAP:
                agg_sum += AGG_MAP[g]; agg_cnt += 1
        ave = round(total / n, 1) if n else 0.0
        agg = agg_sum if agg_cnt == 4 else None
        div = division_from_aggregate(agg) if agg is not None else "NG"
        rows.append({
            "no": i,
            "student_number": s["student_number"],
            "full_name": f"{s['first_name']} {s['middle_name']} {s['last_name']}".replace(" ", " ").strip(),
            "cells": cells,
            "total": round(total, 0),
            "ave": ave,
            "agg": agg if agg is not None else "",
            "div": div,
        })

    # positions (ties share)
    ranked = sorted(rows, key=lambda r: (-r["total"], r["full_name"]))
    pos = 0; seen = 0; prev = None
    for r in ranked:
        seen += 1
        if r["total"] != prev:
            pos = seen; prev = r["total"]
        r["pos"] = pos
    pos_by_stuno = {r["student_number"]: r["pos"] for r in ranked}
    for r in rows:
        r["pos"] = pos_by_stuno.get(r["student_number"], "")

    cur.close(); conn.close()

    # ---- PDF render (auto-fit, zebra rows, 1.5 spacing, consistent gap below header line) ----
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=landscape(A4))
    width, height = landscape(A4)

    # header titles
    y = height - 1.5 * cm
    try:
        from reportlab.lib.utils import ImageReader
        logo_path = os.path.join(current_app.static_folder, "logo.jpg")
        if os.path.exists(logo_path):
            c.drawImage(ImageReader(logo_path), 1.0*cm, y-0.7*cm, width=0.8*cm, height=0.8*cm,
                        preserveAspectRatio=True, mask='auto')
    except Exception:
        pass

    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(width/2, y, school_title)
    y -= 0.55 * cm
    c.setFont("Helvetica", 11)
    c.drawCentredString(width/2, y, f"Mark Sheet {class_name} {stream} — {term} {year}")
    y -= 0.45 * cm

    # --- Auto-fit column widths ---
    left_margin = 0.8 * cm
    right_margin = 0.8 * cm
    usable = width - left_margin - right_margin

    base_cols = [
        ("No", 1.0 * cm),
        ("Student No.", 3.0 * cm),
        ("Name", 6.0 * cm),
    ] + [((s.get("code") or s["name"]), 2.6 * cm) for s in subjects] + [
        ("Total", 2.2 * cm),
        ("AVE", 1.8 * cm),
        ("AGG", 1.6 * cm),
        ("DIV", 1.8 * cm),
        ("POS", 1.8 * cm),
    ]
    base_total = sum(w for _, w in base_cols)
    scale = usable / base_total if base_total > 0 else 1.0
    cols = [(t, w * scale) for (t, w) in base_cols]

    x_edges = [left_margin]
    for _, w in cols:
        x_edges.append(x_edges[-1] + w)

    # body font + line height (1.5x)
    font_body = 9
    c.setFont("Helvetica", font_body)
    row_height = (font_body * 1.5 / 28.3465) * cm # points -> cm

    # header spacing config (line tight to titles, gap below)
    HEADER_LINE_OFFSET = 0.42 * cm
    GAP_BELOW_HEADER = 0.80 * cm

    def _draw_table_header():
        nonlocal y
        c.setFont("Helvetica-Bold", 9)
        for (title, _), x0, x1 in zip(cols, x_edges[:-1], x_edges[1:]):
            cx = (x0 + x1) / 2.0
            if title == "Name":
                c.drawString(x0 + 2, y, str(title))
            else:
                c.drawCentredString(cx, y, str(title))
        y -= HEADER_LINE_OFFSET
        c.line(left_margin, y, x_edges[-1], y) # line close to titles
        y -= GAP_BELOW_HEADER
        c.setFont("Helvetica", font_body)

    def _new_page():
        nonlocal y
        c.showPage()
        y = height - 1.2 * cm
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width/2, y, f"Mark Sheet {class_name} {stream} — {term} {year}")
        y -= 0.5 * cm
        _draw_table_header()

    # print header (first page)
    _draw_table_header()

    # draw rows with zebra background
    for idx, row in enumerate(rows, start=1):
        if y < 1.7 * cm:
            _new_page()

        # zebra stripe
        if idx % 2 == 0:
            c.saveState()
            c.setFillGray(0.93)
            c.rect(left_margin, y - (row_height - 0.05 * cm),
                   x_edges[-1] - left_margin, row_height, fill=1, stroke=0)
            c.restoreState()

        flat = [
            row["no"],
            row["student_number"],
            row["full_name"],
            *row["cells"],
            f"{row['total']:.0f}",
            f"{row['ave']:.1f}",
            (row["agg"] if row["agg"] != "" else ""),
            row["div"],
            row["pos"],
        ]
        for val, (title, _), x0, x1 in zip(flat, cols, x_edges[:-1], x_edges[1:]):
            cx = (x0 + x1) / 2.0
            if title == "Name":
                c.drawString(x0 + 2, y, str(val))
            else:
                c.drawCentredString(cx, y, str(val))

        y -= row_height

    c.showPage()
    c.save()
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"Mark_Sheet_{class_name}_{stream}_{term}_{year}.pdf",
        mimetype="application/pdf",
    )
    
    # ------------------ ROUTE: HTML FORM ------------------ #

@app.route("/reports/competency_checklist/<int:student_id>", methods=["GET", "POST"])
@require_role("admin", "headteacher", "bursar", "teacher", "classmanager", "deputyheadteacher")
def competency_checklist(student_id):
    ay = get_active_academic_year() or {}
    current_term = ay.get("current_term") or ay.get("term") or "Term 1"
    current_year = int(ay.get("year") or datetime.now().year)

    term = (request.values.get("term") or current_term).strip()
    try:
        year = int(request.values.get("year") or current_year)
    except ValueError:
        year = current_year

    conn = get_db_connection()
    ensure_robotics_checklist_schema(conn)
    ensure_robotics_checklist_meta_schema(conn)
    cur = conn.cursor(dictionary=True)

    # ---------- student ----------
    cur.execute(
        """
        SELECT id,
               student_number,
               first_name,
               COALESCE(Middle_name,'') AS Middle_name,
               last_name,
               class_name,
               stream,
               section
        FROM students
        WHERE id=%s
        """,
        (student_id,),
    )
    student = cur.fetchone()
    if not student:
        cur.close()
        conn.close()
        flash("Student not found.", "warning")
        return redirect(url_for("students"))

    # ---------- navigation within same class ----------
    prev_id = next_id = None
    if student.get("class_name"):
        cur.execute(
            """
            SELECT id
            FROM students
            WHERE class_name=%s AND archived=0
            ORDER BY last_name, first_name, id
            """,
            (student["class_name"],),
        )
        id_rows = cur.fetchall() or []
        all_ids = [r["id"] for r in id_rows]
        if student_id in all_ids:
            idx = all_ids.index(student_id)
            if idx > 0:
                prev_id = all_ids[idx - 1]
            if idx < len(all_ids) - 1:
                next_id = all_ids[idx + 1]

    # ---------- checklist items ----------
    items = _fetch_checklist_items()

    # existing checklist rows
    cur.execute(
        """
        SELECT area, area_code, section, label, competence, tick, remark
        FROM robotics_checklist
        WHERE student_id=%s AND term=%s AND year=%s
        """,
        (student_id, term, year),
    )
    saved_rows = cur.fetchall() or []
    saved_map = {
        (r["area"], r["section"], r["label"], r["competence"]): r
        for r in saved_rows
    }

    # existing meta for this learner/period
    meta = _fetch_checklist_meta(student_id, term, year)

    # ---- auto-fill next term dates from term_dates if missing ---- # <-- added
    try:
        ensure_term_dates_schema(conn)
    except Exception:
        pass # if helper fails, just keep existing meta

    cur.execute(
        """
        SELECT next_term_date, next_term_end_date
        FROM term_dates
        WHERE year=%s AND term=%s
        LIMIT 1
        """,
        (year, term),
    )
    td = cur.fetchone() or {}

    from datetime import datetime as _dt

    if not meta.get("next_term_begin") and td.get("next_term_date"):
        try:
            meta["next_term_begin"] = _dt.strptime(
                td["next_term_date"], "%Y-%m-%d"
            ).date()
        except Exception:
            # if it's already a date, just keep it
            meta["next_term_begin"] = td["next_term_date"]

    if not meta.get("next_term_end") and td.get("next_term_end_date"):
        try:
            meta["next_term_end"] = _dt.strptime(
                td["next_term_end_date"], "%Y-%m-%d"
            ).date()
        except Exception:
            meta["next_term_end"] = td["next_term_end_date"]
    # ---- end auto-fill block ----

    # ---------- dropdown options for overall remark (keep your existing logic) ----------
    cur.execute(
        """
        SELECT DISTINCT overall_remark
        FROM robotics_checklist_meta
        WHERE overall_remark IS NOT NULL
          AND overall_remark <> ''
        ORDER BY overall_remark
        """
    )
    extra_overall = [
        r["overall_remark"]
        for r in (cur.fetchall() or [])
        if r["overall_remark"] not in RECOMMENDATION_OPTIONS
    ]
    overall_options = RECOMMENDATION_OPTIONS + extra_overall

    # ---------- load comment library groups (teacher / headteacher) ----------
    teacher_library, head_library = load_comment_library_groups(conn)

    if request.method == "POST":
        # ---------- save checklist rows ----------
        cur.execute(
            """
            DELETE FROM robotics_checklist
            WHERE student_id=%s AND term=%s AND year=%s
            """,
            (student_id, term, year),
        )
        for idx, (area, acode, section, label, comp) in enumerate(items):
            tick_val = 1 if request.form.get(f"tick_{idx}") == "on" else 0
            remark_val = (request.form.get(f"remark_{idx}") or "").strip()
            cur.execute(
                """
                INSERT INTO robotics_checklist
                    (student_id, term, year,
                     area, area_code, section, label, competence,
                     tick, remark)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    student_id,
                    term,
                    year,
                    area,
                    acode,
                    section,
                    label,
                    comp,
                    tick_val,
                    remark_val,
                ),
            )

        # ---------- save meta (overall remark, comms, dates, fees) ----------
        def _parse_date(val: str):
            try:
                return datetime.strptime(val, "%Y-%m-%d").date()
            except Exception:
                return None

        overall = (
            (request.form.get("overall_custom") or "").strip()
            or (request.form.get("overall_remark") or "").strip()
        )
        special = (request.form.get("special_communication") or "").strip()
        nb = _parse_date((request.form.get("next_term_begin") or "").strip())
        ne = _parse_date((request.form.get("next_term_end") or "").strip())
        fees = (request.form.get("school_fees") or "").strip()
        fees_dc = (request.form.get("school_fees_daycare") or "").strip()

        cur.execute(
            """
            DELETE FROM robotics_checklist_meta
            WHERE student_id=%s AND term=%s AND year=%s
            """,
            (student_id, term, year),
        )
        cur.execute(
            """
            INSERT INTO robotics_checklist_meta
                (student_id, term, year,
                 overall_remark, special_communication,
                 next_term_begin, next_term_end,
                 school_fees, school_fees_daycare)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                student_id,
                term,
                year,
                overall or None,
                special or None,
                nb,
                ne,
                fees or None,
                fees_dc or None,
            ),
        )

        conn.commit()
        cur.close()
        conn.close()
        flash("Checklist saved.", "success")
        return redirect(
            url_for(
                "competency_checklist",
                student_id=student_id,
                term=term,
                year=year,
            )
        )

    cur.close()
    conn.close()

    # format for the <input type="date"> fields
    if meta.get("next_term_begin"):
        meta["next_term_begin"] = meta["next_term_begin"].strftime("%Y-%m-%d")
    if meta.get("next_term_end"):
        meta["next_term_end"] = meta["next_term_end"].strftime("%Y-%m-%d")

    return render_template(
        "competency_checklist_form.html",
        student=student,
        term=term,
        year=year,
        items=items,
        saved_map=saved_map,
        remark_options=REMARK_OPTIONS,
        recommendation_options=RECOMMENDATION_OPTIONS,
        overall_options=overall_options,
        meta=meta,
        prev_id=prev_id,
        next_id=next_id,
        teacher_library=teacher_library,
        head_library=head_library,
    )




# ------------------ ROUTE: KINDERGARTEN PDF REPORT ------------------ #


@app.route("/reports/competency_checklist_pdf/<int:student_id>")
@require_role("admin", "headteacher", "teacher", "class_teacher", "bursar", "classmanager")
def competency_checklist_pdf(student_id):
    # ---- resolve term/year ----
    ay = get_active_academic_year() or {}
    term = (
        request.args.get("term")
        or ay.get("current_term")
        or ay.get("term")
        or "Term 1"
    ).strip()
    year = int(
        request.args.get("year")
        or ay.get("year")
        or ay.get("active_year")
        or datetime.now().year
    )

    # ---- student ----
    student = _fetch_student_for_checklist(student_id)
    if not student:
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)
        c.drawString(40, 800, "Student not found.")
        c.save()
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="checklist_missing.pdf",
        )
    
    # ---- items & saved ticks/remarks ----
    items = _fetch_checklist_items()
    saved_map = _fetch_saved_checklist_map(student_id, term, year)
    # key: (area, section, label, competence) -> {'tick': bool, 'remark': text}
    

    # ---- load meta (overall remark, special comms, dates, fees) ----
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # 1) Existing meta (remarks, fees, and possibly dates)
    cur.execute(
        """
        SELECT overall_remark, special_communication,
               next_term_begin, next_term_end,
               school_fees, school_fees_daycare
        FROM robotics_checklist_meta
        WHERE student_id=%s AND term=%s AND year=%s
        """,
        (student_id, term, year),
    )
    meta = cur.fetchone() or {}

    # 2) Auto-pick dates from term_dates if missing
    # (does NOT override existing meta dates)
    try:
        ensure_term_dates_schema(conn)
    except Exception:
        pass # if helper fails, just fall back to whatever is in meta

    cur.execute(
        """
        SELECT next_term_date, next_term_end_date
        FROM term_dates
        WHERE year=%s AND term=%s
        LIMIT 1
        """,
        (year, term),
    )
    td = cur.fetchone() or {}

    from datetime import datetime as _dt

    if not meta.get("next_term_begin") and td.get("next_term_date"):
        try:
            meta["next_term_begin"] = _dt.strptime(
                td["next_term_date"], "%Y-%m-%d"
            ).date()
        except Exception:
            # if parsing fails, leave it out; footer logic stays unchanged
            pass

    if not meta.get("next_term_end") and td.get("next_term_end_date"):
        try:
            meta["next_term_end"] = _dt.strptime(
                td["next_term_end_date"], "%Y-%m-%d"
            ).date()
        except Exception:
            pass

    cur.close()
    conn.close()


    # ---- PDF + HEADER (only page 1) ----
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    left = 40
    top = height - 40
    bottom_margin = 60 # space for footer/remarks

    # =================== CLEAN BRANDED HEADER (NO ICONS) ===================
    banner_h = 30 * mm
    banner_y = height - 35 * mm
    left_margin, right_margin = left, left
    strip_w = width - left_margin - right_margin
    navy_w = strip_w * 0.73
    blue_w = strip_w - navy_w

    c.saveState()
    c.setFillColor(COL_NAVY)
    c.rect(left_margin, banner_y, navy_w, banner_h, stroke=0, fill=1)
    c.setFillColor(COL_BLUE)
    c.rect(left_margin + navy_w, banner_y, blue_w, banner_h, stroke=0, fill=1)

    fold_depth = 11 * mm
    fold_lip = 6 * mm
    c.setFillColor(COL_BLUE2)
    ps = c.beginPath()
    ps.moveTo(left_margin + navy_w, banner_y)
    ps.lineTo(left_margin + navy_w + fold_depth, banner_y + banner_h)
    ps.lineTo(left_margin + navy_w + fold_depth + 2 * mm, banner_y + banner_h)
    ps.lineTo(left_margin + navy_w + 2 * mm, banner_y)
    ps.close()
    c.drawPath(ps, stroke=0, fill=1)
    flap_col = colors.HexColor("#3a86e0")
    c.setFillColor(flap_col)
    pf = c.beginPath()
    pf.moveTo(left_margin + navy_w - fold_lip, banner_y)
    pf.lineTo(left_margin + navy_w, banner_y)
    pf.lineTo(left_margin + navy_w + fold_depth, banner_y + banner_h)
    pf.lineTo(left_margin + navy_w - fold_lip, banner_y + banner_h)
    pf.close()
    c.drawPath(pf, stroke=0, fill=1)

    SCHOOL_LOGO_PATH = os.path.join(current_app.static_folder, "logo.jpg")
    logo_box = 24 * mm
    logo_x = left_margin + 6 * mm
    logo_y = banner_y + (banner_h - logo_box) / 2
    if os.path.exists(SCHOOL_LOGO_PATH):
        try:
            c.drawImage(
                SCHOOL_LOGO_PATH,
                logo_x,
                logo_y,
                width=logo_box,
                height=logo_box,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass

    name_left = logo_x + logo_box + 6 * mm
    name_right = left_margin + navy_w - 6 * mm
    name_box_w = max(10, name_right - name_left)

    # centre of the text block (between logo-right & navy-right)
    center_x = (name_left + name_right) / 2.0

    name_text = SCHOOL_NAME or ""
    name_fs = 18
    while (
        name_fs >= 10
        and c.stringWidth(name_text, "Helvetica-Bold", name_fs) > name_box_w
    ):
        name_fs -= 1

    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", name_fs)
    # a bit lower + more room for lines below
    name_y = banner_y + banner_h - 5 * mm
    c.drawCentredString(center_x, name_y, name_text)

    sub_text = SCHOOL_SUB or ""
    addr_text = SCHOOL_ADDRESS or "" # P.O Box line

    # tagline line
    if sub_text:
        sub_fs = 12
        while (
            sub_fs >= 8
            and c.stringWidth(sub_text, "Helvetica-Bold", sub_fs) > name_box_w
        ):
            sub_fs -= 1
        c.setFont("Helvetica-Bold", sub_fs)
        tagline_y = name_y - (name_fs * 1.15) # <- bigger = more spacing
        c.drawCentredString(center_x, tagline_y, sub_text)
    else:
        tagline_y = name_y

    # P.O Box line
    if addr_text:
        addr_fs = max(8, (sub_fs - 1) if sub_text else 10)
        c.setFont("Helvetica-Bold", addr_fs)
        addr_y = tagline_y - (addr_fs * 1.1) # extra spacing again
        c.drawCentredString(center_x, addr_y, addr_text)
    
    # --- phones: ensure one number per line (3 rows) ---

    c.setFillColor(colors.white)
    c.setFont("Helvetica", 9)
    right_pad = 6 * mm
    text_right = left_margin + strip_w - right_pad
    line_gap = 5.5 * mm
    y_cursor = banner_y + banner_h - 8 * mm

    # -------- Contacts block (RIGHT SIDE, one per line) --------
    raw = SCHOOL_PHONE_LINES or ""

    # Normalize to one comma-separated string first
    if isinstance(raw, (list, tuple)):
        combined = ", ".join(str(p) for p in raw)
    else:
        combined = str(raw)

    # Now definitely split into separate phone numbers
    phone_lines = [p.strip() for p in combined.split(",") if p.strip()]

    # draw each phone on its own line
    for ph in phone_lines:
        c.drawRightString(text_right, y_cursor, ph)
        y_cursor -= line_gap

    # small extra gap, then email if present
    if SCHOOL_EMAIL:
        y_cursor -= 2.5 * mm
        c.drawRightString(text_right, y_cursor, SCHOOL_EMAIL)


    c.restoreState()

    # ========== LEARNER INFO TABLE (two columns) ==========
    info_top = banner_y - 6 * mm
    info_left = left_margin
    info_width = width - left_margin - right_margin - (40 * mm) # room for photo

    styles = getSampleStyleSheet()
    lab = ParagraphStyle(
        "lab",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=colors.black,
    )
    val = ParagraphStyle(
        "val",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=11,
    )

    full_name = f"{student.get('first_name','')} {student.get('middle_name','')} {student.get('last_name','')}".strip()

    info_rows = [
        [Paragraph("Learner's Name:", lab), Paragraph(full_name or "-", val)],
        [Paragraph("Student No.:", lab), Paragraph(student.get("student_number") or "-", val)],
        [
            Paragraph("Class / Stream:", lab),
            Paragraph(f"{student.get('class_name','-')} {student.get('stream','') or ''}", val),
        ],
        [Paragraph("Term / Year:", lab), Paragraph(f"{term} / {year}", val)],
    ]

    info_tbl = Table(
        info_rows,
        colWidths=[35 * mm, info_width - 35 * mm],
        hAlign="LEFT",
    )
    info_tbl.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ]
        )
    )
    w_info, h_info = info_tbl.wrapOn(c, info_width, 9999)
    info_tbl.drawOn(c, info_left, info_top - h_info)

    # ---- PHOTO BLOCK (works with DB blob or file path) ----
    photo_blob = student.get("photo_blob")
    photo_path = student.get("photo")

    # Your existing box position
    box_w = box_h = 32 * mm
    photo_x = width - right_margin - box_w
    photo_y = info_top - (h_info - box_h) / 2

    if photo_blob:
        try:
            img_reader = ImageReader(io.BytesIO(photo_blob))
            c.drawImage(
                img_reader,
                photo_x + 2, photo_y - box_h + 2,
                box_w - 4, box_h - 4,
                preserveAspectRatio=True,
                mask="auto"
            )
        except:
            pass

    elif photo_path:
        full_path = os.path.join(app.root_path, photo_path)
        if os.path.exists(full_path):
            try:
                img_reader = ImageReader(full_path)
                c.drawImage(
                    img_reader,
                    photo_x + 2, photo_y - box_h + 2,
                    box_w - 4, box_h - 4,
                    preserveAspectRatio=True,
                    mask="auto"
                )
            except:
                pass

    else:
        # placeholder
        c.setStrokeColor(colors.grey)
        c.rect(photo_x, photo_y - box_h, box_w, box_h)
        c.setFont("Helvetica", 7)
        c.drawCentredString(photo_x + box_w/2, photo_y - box_h/2, "Photo")


    # ====== Big title above competence table ======
    title_y = (info_top - h_info) - 6 * mm
    c.setFont("Helvetica-Bold", 13)
    c.setFillColor(colors.black)
    c.drawString(left, title_y, "Learner's Competency Checklist [Early Childhood]")

    table_top = title_y - 6 * mm

    # ---------- TABLE DATA ----------
    col_area_w = 26 * mm
    col_skill_w = 32 * mm
    col_tick_w = 8 * mm
    col_rem_w = 36 * mm
    col_comp_w = width - left * 2 - col_area_w - col_skill_w - col_tick_w - col_rem_w

    styles = getSampleStyleSheet()
    p_head = ParagraphStyle(
        "head",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        alignment=1,
        textColor=colors.white,
    )
    p_skill = ParagraphStyle(
        "skill",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=9,
    )
    p_section = ParagraphStyle(
        "section",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
    )
    p_comp = ParagraphStyle(
        "comp",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=9,
    )
    p_rem = ParagraphStyle(
        "rem",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7,
        leading=8,
    )

    data = [
        [
            Paragraph("Area", p_head),
            Paragraph("Competence", p_head),
            Paragraph("Skill", p_head),
            Paragraph("✓", p_head),
            Paragraph("Remarks", p_head),
        ]
    ]
    row_meta = [{"type": "header"}]
    collected_remarks = []
    last_area = None
    last_section = None
    area_dividers = []

    for (area, acode, section, label, comp) in items:
        key = (area, section, label, comp)
        saved = saved_map.get(key, {})
        tick = bool(saved.get("tick"))
        remark_txt = (saved.get("remark") or "").strip()

        if remark_txt:
            collected_remarks.append(
                f"{section} - {label}: {remark_txt}"
                if section
                else f"{label}: {remark_txt}"
            )

        # area marker row
        if area != last_area:
            data.append(["", "", "", "", ""])
            row_meta.append({"type": "area", "area": area})
            area_dividers.append(len(data) - 1)
            last_area = area
            last_section = None

        # section row
        if section and section != last_section:
            data.append(["", Paragraph(section, p_section), "", "", ""])
            row_meta.append(
                {
                    "type": "section",
                    "area": area,
                    "section": section,
                }
            )
            last_section = section

        # skill row
        data.append(
            [
                "",
                Paragraph(label, p_skill),
                Paragraph(comp, p_comp),
                "✔" if tick else "",
                Paragraph(remark_txt, p_rem) if remark_txt else "",
            ]
        )
        row_meta.append(
            {
                "type": "skill",
                "area": area,
                "section": section,
                "label": label,
            }
        )

    # ---------- BUILD FULL TABLE ONCE ----------
    base_table = Table(
        data,
        colWidths=[col_area_w, col_skill_w, col_comp_w, col_tick_w, col_rem_w],
        repeatRows=0,
    )
    ts = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.0, 0.45, 0.80)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("BOX", (0, 0), (-1, -1), 0.75, colors.lightgrey),
            ("LINEBELOW", (0, 0), (-1, 0), 0.75, colors.lightgrey),
            ("GRID", (1, 1), (-1, -1), 0.25, colors.lightgrey),
            ("LINEAFTER", (0, 0), (0, -1), 0.75, colors.lightgrey),
            ("ALIGN", (3, 1), (3, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]
    )

    # bold section labels
    for i in range(1, len(data)):
        if row_meta[i]["type"] == "section":
            ts.add("FONT", (1, i), (1, i), "Helvetica-Bold", 8)

    base_table.setStyle(ts)
    avail_height_first = table_top - bottom_margin
    base_table.wrapOn(c, width - left * 2, avail_height_first)
    full_rows = base_table._cellvalues
    full_heights = list(base_table._rowHeights)

    # ---------- SPLIT INTO PAGES ----------
    pages = []
    header_row = full_rows[0]
    header_meta = row_meta[0]

    max_height = avail_height_first
    cur_rows = [header_row]
    cur_meta = [header_meta]
    cur_height = full_heights[0]

    i = 1
    while i < len(full_rows):
        meta_i = row_meta[i]
        # If this row starts a new AREA block, grab the whole block
        if meta_i.get("type") == "area":
            start = i
            j = i
            # area block continues until the next 'area' row or end
            while j + 1 < len(full_rows) and row_meta[j + 1].get("type") != "area":
                j += 1
        else:
            # safety: treat single row as its own tiny block
            start = i
            j = i

        block_rows = full_rows[start : j + 1]
        block_meta = row_meta[start : j + 1]
        block_height = sum(full_heights[k] for k in range(start, j + 1))

        # If the block can't fit on this page (and we already have more
        # than just the header), push current page and start a new one
        if cur_height + block_height > max_height and len(cur_rows) > 1:
            pages.append((cur_rows, cur_meta))
            cur_rows = [header_row]
            cur_meta = [header_meta]
            cur_height = full_heights[0]
            max_height = height - 80 # subsequent pages

        # Add the whole block to the current page
        cur_rows.extend(block_rows)
        cur_meta.extend(block_meta)
        cur_height += block_height
        i = j + 1

    pages.append((cur_rows, cur_meta))


    # ---------- DRAW PAGES ----------
    last_table_y = table_top
    carry_area = None

    try:
        base_cmds = list(ts.getCommands())
    except AttributeError:
        base_cmds = list(ts._cmds)

    for page_index, (page_rows, page_meta) in enumerate(pages):
        if page_index == 0:
            top_y = table_top
        else:
            c.showPage()
            top_y = height - 60
            c.setFont("Helvetica-Bold", 10)
            c.setFillColor(colors.black)
            c.drawString(left, top_y, "Learner's Competency Checklist (continued)")
            top_y -= 16

        page_table = Table(
            page_rows,
            colWidths=[col_area_w, col_skill_w, col_comp_w, col_tick_w, col_rem_w],
            repeatRows=1,
        )

        extra_cmds = []
        for i, meta_row in enumerate(page_meta):
            if meta_row.get("type") == "area":
                extra_cmds.append(
                    ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#bfbfbf"))
                )
                extra_cmds.append(
                    ("TEXTCOLOR", (0, i), (-1, i), colors.white)
                )
                extra_cmds.append(
                    ("LINEABOVE", (0, i), (-1, i), 0.9, colors.black)
                )
                extra_cmds.append(
                    ("LINEBELOW", (0, i), (-1, i), 0.9, colors.black)
                )

        ts_page = TableStyle(base_cmds + extra_cmds)
        page_table.setStyle(ts_page)

        avail_h = top_y - bottom_margin
        w, h = page_table.wrapOn(c, width - left * 2, avail_h)
        table_y = top_y - h

        row_heights = list(page_table._rowHeights)

        area_ranges = _compute_area_ranges_from_meta(page_meta, previous_area=carry_area)

        _draw_area_icons_vertical(
            c=c,
            left=left,
            table_y=table_y,
            row_heights=row_heights,
            col_area_w=col_area_w,
            area_ranges=area_ranges,
        )

        page_table.drawOn(c, left, table_y)
        last_table_y = table_y

        for m in page_meta:
            if m.get("type") == "area":
                carry_area = m.get("area")

    # ---------- REMARKS + FOOTER ----------
    remarks_y = last_table_y - 24
    if remarks_y < 120:
        c.showPage()
        remarks_y = height - 120

    raw_name = (session.get("full_name") or session.get("username") or "").strip()
    if raw_name:
        parts = raw_name.split()
        if len(parts) >= 2:
            raw_name = f"{parts[0]} {parts[-1]}"
    prepared_name = f"Tr. {raw_name}" if raw_name else "Tr. __________________"
    today_str = datetime.now().strftime("%d-%b-%Y")

    overall = (meta.get("overall_remark") or "").strip()
    # convert \n to <br/> so each role appears on its own line
    overall_html = (overall or " ").replace("\n", "<br/>")

    sc = (meta.get("special_communication") or "").strip()
    ntb = meta.get("next_term_begin")
    nte = meta.get("next_term_end")
    #fees = (meta.get("school_fees") or "").strip()
    #fees_dc = (meta.get("school_fees_daycare") or "").strip()

    ntb_str = ntb.strftime("%d/%m/%y") if ntb else "__________"
    nte_str = nte.strftime("%d/%m/%y") if nte else "__________"

    styles = getSampleStyleSheet()
    h_style = ParagraphStyle(
        "bottom_head",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.white,
        alignment=0,
    )
    label_style = ParagraphStyle(
        "bottom_label",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
    )
    text_style = ParagraphStyle(
        "bottom_text",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
    )

    bottom_data = [
        [Paragraph("Remarks & Recommendations to the learner:", h_style), ""],
        [Paragraph(overall_html, text_style), ""],
        [Paragraph("Special Communication:", label_style), Paragraph(sc or " ", text_style)],
        [Paragraph("Next term begins on:", label_style),
         Paragraph(f"{ntb_str} will end on: {nte_str}", text_style)],
        #[Paragraph("School fees:", label_style), Paragraph(fees or " ", text_style)],
        #[Paragraph("School fees + daycare:", label_style), Paragraph(fees_dc or " ", text_style)],
        [Paragraph("Prepared by:", label_style), Paragraph(prepared_name, text_style)],
        [Paragraph("Date:", label_style), Paragraph(today_str, text_style)],
    ]

    bottom_table = Table(
        bottom_data,
        colWidths=[(width - 2 * left) * 0.28, (width - 2 * left) * 0.72],
    )
    bt = TableStyle(
        [
            ("BOX", (0, 0), (-1, -1), 0.75, colors.lightgrey),
            ("GRID", (0, 2), (-1, -1), 0.25, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.0, 0.45, 0.80)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ]
    )
    bt.add("SPAN", (0, 0), (-1, 0))
    bt.add("SPAN", (0, 1), (-1, 1))
    bottom_table.setStyle(bt)

    bw, bh = bottom_table.wrapOn(c, width - 2 * left, remarks_y - 40)
    bottom_y = remarks_y - bh
    bottom_table.drawOn(c, left, bottom_y)

    c.showPage()
    c.save()
    buf.seek(0)

    filename = f"Checklist_{student.get('student_number','')}_{term}_{year}.pdf".replace(" ", "_")
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )





@app.route("/reports/competency_special_comm", methods=["GET", "POST"])
@require_role("admin", "headteacher", "teacher", "class_teacher", "bursar", "classmanager", "dos", "deputyheadteacher")
def competency_special_comm():
    ay = get_active_academic_year() or {}
    current_term = ay.get("current_term") or ay.get("term") or "Term 1"
    current_year = int(ay.get("year") or datetime.now().year)

    term = (request.values.get("term") or current_term).strip()
    try:
        year = int(request.values.get("year") or current_year)
    except ValueError:
        year = current_year

    conn = get_db_connection()
    ensure_robotics_checklist_meta_schema(conn)

    # We don't need full list of classes – template will show Baby/Middle/Top
    classes = KG_CLASSES[:]

    affected = 0

    if request.method == "POST":
        scope = request.form.get("scope") or "class"
        class_name = (request.form.get("class_name") or "").strip()
        special = (request.form.get("special_communication") or "").strip()
        override = bool(request.form.get("override_existing"))

        if not special:
            flash("Special communication text cannot be empty.", "warning")
        else:
            cur = conn.cursor(dictionary=True)

            # decide target learners
            if scope == "all_kindergarten":
                # ALL KG classes: Baby, Middle, Top
                cur.execute(
                    """
                    SELECT id
                    FROM students
                    WHERE class_name IN (%s,%s,%s) AND archived = 0
                    """,
                    tuple(KG_CLASSES),
                )
            else:
                if not class_name:
                    flash(
                        "Please choose a class when scope is 'Single class'.",
                        "warning",
                    )
                    cur.close()
                    conn.close()
                    return redirect(
                        url_for(
                            "competency_special_comm",
                            term=term,
                            year=year,
                        )
                    )
                cur.execute(
                    """
                    SELECT id
                    FROM students
                    WHERE class_name=%s AND archived = 0
                    """,
                    (class_name,),
                )

            stu_rows = cur.fetchall() or []
            cur.close()

            if not stu_rows:
                flash("No learners found for the selected scope.", "warning")
            else:
                cur = conn.cursor(dictionary=True)
                for row in stu_rows:
                    sid = row["id"]
                    cur.execute(
                        """
                        SELECT id, special_communication
                        FROM robotics_checklist_meta
                        WHERE student_id=%s AND term=%s AND year=%s
                        LIMIT 1
                        """,
                        (sid, term, year),
                    )
                    meta_row = cur.fetchone()
                    if meta_row:
                        existing = (meta_row.get("special_communication") or "").strip()
                        if override or not existing:
                            cur.execute(
                                """
                                UPDATE robotics_checklist_meta
                                SET special_communication=%s
                                WHERE id=%s
                                """,
                                (special, meta_row["id"]),
                            )
                            affected += 1
                    else:
                        cur.execute(
                            """
                            INSERT INTO robotics_checklist_meta
                                (student_id, term, year, special_communication)
                            VALUES (%s,%s,%s,%s)
                            """,
                            (sid, term, year, special),
                        )
                        affected += 1

                conn.commit()
                cur.close()

                flash(
                    f"Special communication applied to {affected} learner(s) for {term} {year}.",
                    "success",
                )

        conn.close()
        return redirect(
            url_for(
                "competency_special_comm",
                term=term,
                year=year,
            )
        )

    conn.close()
    return render_template(
        "competency_special_comm.html",
        term=term,
        year=year,
        classes=classes, # not strictly needed now, but harmless
    )


 
   
@app.route("/reports/competency_checklist_batch_pdf", methods=["POST"])
@require_role('admin', 'headteacher', 'bursar', 'dos', 'deputyheadteacher', 'classmanager', 'teacher')
def competency_checklist_batch_pdf():
    ay = get_active_academic_year() or {}

    # accept both term/year and termsel/yearsel from reports_hub
    raw_term = (
        request.form.get("term")
        or request.form.get("termsel")
        or ay.get("current_term")
        or ay.get("term")
        or "Term 1"
    )
    term = (raw_term or "Term 1").strip()

    raw_year = (
        request.form.get("year")
        or request.form.get("yearsel")
        or ay.get("year")
        or ay.get("active_year")
        or datetime.now().year
    )
    try:
        year = int(raw_year)
    except (TypeError, ValueError):
        year = int(ay.get("year") or datetime.now().year)

    # ids posted from reports_hub checkboxes
    raw_ids = request.form.getlist("selected_ids")
    student_ids = []
    for sid in raw_ids:
        sid = (sid or "").strip()
        if sid.isdigit():
            student_ids.append(int(sid))

    if not student_ids:
        flash("Please select at least one learner first.", "warning")
        return redirect(url_for(
            "reports_hub",
            class_name=request.form.get("class_name", ""),
            term=term,
            year=year,
        ))

    # ---- single combined PDF buffer & canvas ----
    import io
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    # ========= LOOP THROUGH EACH STUDENT =========
    for idx, student_id in enumerate(student_ids, start=1):
        # ---- student ----
        student = _fetch_student_for_checklist(student_id)
        if not student:
            # skip missing learners in batch mode
            continue

        # ---- items & saved ticks/remarks ----
        items = _fetch_checklist_items()
        saved_map = _fetch_saved_checklist_map(student_id, term, year)
        # key: (area, section, label, competence) -> {'tick': bool, 'remark': text}

        # load meta (overall remark, special comms, dates, fees)
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT overall_remark, special_communication,
                   next_term_begin, next_term_end,
                   school_fees, school_fees_daycare
            FROM robotics_checklist_meta
            WHERE student_id=%s AND term=%s AND year=%s
            """,
            (student_id, term, year),
        )
        meta = cur.fetchone() or {}
        cur.close()
        conn.close()


        # ---- PDF + HEADER (only page 1) ----

        left = 40
        top = height - 40
        bottom_margin = 60 # space for footer/remarks

        # =================== CLEAN BRANDED HEADER (NO ICONS) ===================
        banner_h = 30 * mm
        banner_y = height - 35 * mm
        left_margin, right_margin = left, left
        strip_w = width - left_margin - right_margin
        navy_w = strip_w * 0.73
        blue_w = strip_w - navy_w

        c.saveState()
        c.setFillColor(COL_NAVY)
        c.rect(left_margin, banner_y, navy_w, banner_h, stroke=0, fill=1)
        c.setFillColor(COL_BLUE)
        c.rect(left_margin + navy_w, banner_y, blue_w, banner_h, stroke=0, fill=1)

        fold_depth = 11 * mm
        fold_lip = 6 * mm
        c.setFillColor(COL_BLUE2)
        ps = c.beginPath()
        ps.moveTo(left_margin + navy_w, banner_y)
        ps.lineTo(left_margin + navy_w + fold_depth, banner_y + banner_h)
        ps.lineTo(left_margin + navy_w + fold_depth + 2 * mm, banner_y + banner_h)
        ps.lineTo(left_margin + navy_w + 2 * mm, banner_y)
        ps.close()
        c.drawPath(ps, stroke=0, fill=1)
        flap_col = colors.HexColor("#3a86e0")
        c.setFillColor(flap_col)
        pf = c.beginPath()
        pf.moveTo(left_margin + navy_w - fold_lip, banner_y)
        pf.lineTo(left_margin + navy_w, banner_y)
        pf.lineTo(left_margin + navy_w + fold_depth, banner_y + banner_h)
        pf.lineTo(left_margin + navy_w - fold_lip, banner_y + banner_h)
        pf.close()
        c.drawPath(pf, stroke=0, fill=1)

        SCHOOL_LOGO_PATH = os.path.join(current_app.static_folder, "logo.jpg")
        logo_box = 24 * mm
        logo_x = left_margin + 6 * mm
        logo_y = banner_y + (banner_h - logo_box) / 2
        if os.path.exists(SCHOOL_LOGO_PATH):
            try:
                c.drawImage(
                    SCHOOL_LOGO_PATH,
                    logo_x,
                    logo_y,
                    width=logo_box,
                    height=logo_box,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass
        name_left = logo_x + logo_box + 6 * mm
        name_right = left_margin + navy_w - 6 * mm
        name_box_w = max(10, name_right - name_left)

        # centre of the text block (between logo-right & navy-right)
        center_x = (name_left + name_right) / 2.0

        name_text = SCHOOL_NAME or ""
        name_fs = 18
        while (
            name_fs >= 10
            and c.stringWidth(name_text, "Helvetica-Bold", name_fs) > name_box_w
        ):
            name_fs -= 1

        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", name_fs)
        # a bit lower + more room for lines below
        name_y = banner_y + banner_h - 5 * mm
        c.drawCentredString(center_x, name_y, name_text)

        sub_text = SCHOOL_SUB or ""
        addr_text = SCHOOL_ADDRESS or "" # P.O Box line

        # tagline line
        if sub_text:
            sub_fs = 12
            while (
                sub_fs >= 8
                and c.stringWidth(sub_text, "Helvetica-Bold", sub_fs) > name_box_w
            ):
                sub_fs -= 1
            c.setFont("Helvetica-Bold", sub_fs)
            tagline_y = name_y - (name_fs * 1.15) # <- bigger = more spacing
            c.drawCentredString(center_x, tagline_y, sub_text)
        else:
            tagline_y = name_y

        # P.O Box line
        if addr_text:
            addr_fs = max(8, (sub_fs - 1) if sub_text else 10)
            c.setFont("Helvetica-Bold", addr_fs)
            addr_y = tagline_y - (addr_fs * 1.1) # extra spacing again
            c.drawCentredString(center_x, addr_y, addr_text)
        
        # --------- Contacts block (RIGHT SIDE) ----------

        c.setFillColor(colors.white)
        c.setFont("Helvetica", 9)
        right_pad = 6 * mm
        text_right = left_margin + strip_w - right_pad
        line_gap = 5.5 * mm
        y_cursor = banner_y + banner_h - 8 * mm

        # -------- Contacts block (RIGHT SIDE, one per line) --------
        raw = SCHOOL_PHONE_LINES or ""

        # Normalize to one comma-separated string first
        if isinstance(raw, (list, tuple)):
            combined = ", ".join(str(p) for p in raw)
        else:
            combined = str(raw)

        # Now definitely split into separate phone numbers
        phone_lines = [p.strip() for p in combined.split(",") if p.strip()]

        # draw each phone on its own line
        for ph in phone_lines:
            c.drawRightString(text_right, y_cursor, ph)
            y_cursor -= line_gap

        # small extra gap, then email if present
        if SCHOOL_EMAIL:
            y_cursor -= 2.5 * mm
            c.drawRightString(text_right, y_cursor, SCHOOL_EMAIL)


        c.restoreState()

        # ========== LEARNER INFO TABLE (two columns) ==========
        info_top = banner_y - 6 * mm
        info_left = left_margin
        info_width = width - left_margin - right_margin - (40 * mm) # room for photo

        styles = getSampleStyleSheet()
        lab = ParagraphStyle(
            "lab",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=11,
            textColor=colors.black,
        )
        val = ParagraphStyle(
            "val",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=11,
        )

        full_name = f"{student.get('first_name','')} {student.get('middle_name','')} {student.get('last_name','')}".strip()

        info_rows = [
            [Paragraph("Learner's Name:", lab), Paragraph(full_name or "-", val)],
            [Paragraph("Student No.:", lab), Paragraph(student.get("student_number") or "-", val)],
            [
                Paragraph("Class / Stream:", lab),
                Paragraph(f"{student.get('class_name','-')} {student.get('stream','') or ''}", val),
            ],
            [Paragraph("Term / Year:", lab), Paragraph(f"{term} / {year}", val)],
        ]

        info_tbl = Table(
            info_rows,
            colWidths=[35 * mm, info_width - 35 * mm],
            hAlign="LEFT",
        )
        info_tbl.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ]
            )
        )
        w_info, h_info = info_tbl.wrapOn(c, info_width, 9999)
        info_tbl.drawOn(c, info_left, info_top - h_info)

        # learner photo — right side
        box_w = box_h = 32 * mm
        photo_x = width - right_margin - box_w
        photo_y = info_top - (h_info - box_h) / 2
        c.setStrokeColor(colors.grey)
        c.rect(photo_x, photo_y - box_h, box_w, box_h, stroke=1, fill=0)
        c.setFont("Helvetica", 7)
        c.drawCentredString(photo_x + box_w / 2, photo_y - box_h / 2, "Photo")

        photo_path = (student.get("photo") or "").strip()
        if photo_path and os.path.exists(photo_path):
            try:
                c.drawImage(
                    photo_path,
                    photo_x + 2,
                    photo_y - box_h + 2,
                    box_w - 4,
                    box_h - 4,
                    preserveAspectRatio=True,
                    anchor="c",
                    mask="auto",
                )
            except Exception:
                pass

        # ====== Big title above competence table ======
        title_y = (info_top - h_info) - 6 * mm
        c.setFont("Helvetica-Bold", 13)
        c.setFillColor(colors.black)
        c.drawString(left, title_y, "Learner's Competency Checklist [Early Childhood]")

        table_top = title_y - 6 * mm

        # ---------- TABLE DATA ----------
        col_area_w = 26 * mm
        col_skill_w = 32 * mm
        col_tick_w = 8 * mm
        col_rem_w = 36 * mm
        col_comp_w = width - left * 2 - col_area_w - col_skill_w - col_tick_w - col_rem_w

        styles = getSampleStyleSheet()
        p_head = ParagraphStyle(
            "head",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            alignment=1,
            textColor=colors.white,
        )
        p_skill = ParagraphStyle(
            "skill",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=9,
        )
        p_section = ParagraphStyle(
            "section",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=9,
        )
        p_comp = ParagraphStyle(
            "comp",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=9,
        )
        p_rem = ParagraphStyle(
            "rem",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7,
            leading=8,
        )

        data = [
            [
                Paragraph("Area", p_head),
                Paragraph("Skill", p_head),
                Paragraph("Competence", p_head),
                Paragraph("✓", p_head),
                Paragraph("Remarks", p_head),
            ]
        ]
        row_meta = [{"type": "header"}]
        collected_remarks = []
        last_area = None
        last_section = None
        area_dividers = []

        for (area, acode, section, label, comp) in items:
            key = (area, section, label, comp)
            saved = saved_map.get(key, {})
            tick = bool(saved.get("tick"))
            remark_txt = (saved.get("remark") or "").strip()

            if remark_txt:
                collected_remarks.append(
                    f"{section} - {label}: {remark_txt}"
                    if section
                    else f"{label}: {remark_txt}"
                )

            # area marker row
            if area != last_area:
                data.append(["", "", "", "", ""])
                row_meta.append({"type": "area", "area": area})
                area_dividers.append(len(data) - 1)
                last_area = area
                last_section = None

            # section row
            if section and section != last_section:
                data.append(["", Paragraph(section, p_section), "", "", ""])
                row_meta.append(
                    {
                        "type": "section",
                        "area": area,
                        "section": section,
                    }
                )
                last_section = section

            # skill row
            data.append(
                [
                    "",
                    Paragraph(label, p_skill),
                    Paragraph(comp, p_comp),
                    "✔" if tick else "",
                    Paragraph(remark_txt, p_rem) if remark_txt else "",
                ]
            )
            row_meta.append(
                {
                    "type": "skill",
                    "area": area,
                    "section": section,
                    "label": label,
                }
            )

        # ---------- BUILD FULL TABLE ONCE ----------
        base_table = Table(
            data,
            colWidths=[col_area_w, col_skill_w, col_comp_w, col_tick_w, col_rem_w],
            repeatRows=0,
        )
        ts = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.0, 0.45, 0.80)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("BOX", (0, 0), (-1, -1), 0.75, colors.lightgrey),
                ("LINEBELOW", (0, 0), (-1, 0), 0.75, colors.lightgrey),
                ("GRID", (1, 1), (-1, -1), 0.25, colors.lightgrey),
                ("LINEAFTER", (0, 0), (0, -1), 0.75, colors.lightgrey),
                ("ALIGN", (3, 1), (3, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ]
        )

        # bold section labels
        for i in range(1, len(data)):
            if row_meta[i]["type"] == "section":
                ts.add("FONT", (1, i), (1, i), "Helvetica-Bold", 8)

        base_table.setStyle(ts)
        avail_height_first = table_top - bottom_margin
        base_table.wrapOn(c, width - left * 2, avail_height_first)
        full_rows = base_table._cellvalues
        full_heights = list(base_table._rowHeights)

        # ---------- SPLIT INTO PAGES ----------

        pages = []
        header_row = full_rows[0]
        header_meta = row_meta[0]

        max_height = avail_height_first
        cur_rows = [header_row]
        cur_meta = [header_meta]
        cur_height = full_heights[0]

        i = 1
        while i < len(full_rows):
            meta_i = row_meta[i]
            # If this row starts a new AREA block, grab the whole block
            if meta_i.get("type") == "area":
                start = i
                j = i
                # area block continues until the next 'area' row or end
                while j + 1 < len(full_rows) and row_meta[j + 1].get("type") != "area":
                    j += 1
            else:
                # safety: treat single row as its own tiny block
                start = i
                j = i

            block_rows = full_rows[start : j + 1]
            block_meta = row_meta[start : j + 1]
            block_height = sum(full_heights[k] for k in range(start, j + 1))

            # If the block can't fit on this page (and we already have more
            # than just the header), push current page and start a new one
            if cur_height + block_height > max_height and len(cur_rows) > 1:
                pages.append((cur_rows, cur_meta))
                cur_rows = [header_row]
                cur_meta = [header_meta]
                cur_height = full_heights[0]
                max_height = height - 80 # subsequent pages

            # Add the whole block to the current page
            cur_rows.extend(block_rows)
            cur_meta.extend(block_meta)
            cur_height += block_height
            i = j + 1

        pages.append((cur_rows, cur_meta))


        # ---------- DRAW PAGES ----------
        last_table_y = table_top
        carry_area = None

        try:
            base_cmds = list(ts.getCommands())
        except AttributeError:
            base_cmds = list(ts._cmds)

        for page_index, (page_rows, page_meta) in enumerate(pages):
            if page_index == 0:
                top_y = table_top
            else:
                c.showPage()
                top_y = height - 60
                c.setFont("Helvetica-Bold", 10)
                c.setFillColor(colors.black)
                c.drawString(left, top_y, "Learner's Competency Checklist (continued)")
                top_y -= 16

            page_table = Table(
                page_rows,
                colWidths=[col_area_w, col_skill_w, col_comp_w, col_tick_w, col_rem_w],
                repeatRows=1,
            )

            extra_cmds = []
            for i, meta_row in enumerate(page_meta):
                if meta_row.get("type") == "area":
                    extra_cmds.append(
                        ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#bfbfbf"))
                    )
                    extra_cmds.append(
                        ("TEXTCOLOR", (0, i), (-1, i), colors.white)
                    )
                    extra_cmds.append(
                        ("LINEABOVE", (0, i), (-1, i), 0.9, colors.black)
                    )
                    extra_cmds.append(
                        ("LINEBELOW", (0, i), (-1, i), 0.9, colors.black)
                    )

            ts_page = TableStyle(base_cmds + extra_cmds)
            page_table.setStyle(ts_page)

            avail_h = top_y - bottom_margin
            w, h = page_table.wrapOn(c, width - left * 2, avail_h)
            table_y = top_y - h

            row_heights = list(page_table._rowHeights)

            area_ranges = _compute_area_ranges_from_meta(page_meta, previous_area=carry_area)

            _draw_area_icons_vertical(
                c=c,
                left=left,
                table_y=table_y,
                row_heights=row_heights,
                col_area_w=col_area_w,
                area_ranges=area_ranges,
            )

            page_table.drawOn(c, left, table_y)
            last_table_y = table_y

            for m in page_meta:
                if m.get("type") == "area":
                    carry_area = m.get("area")

        # ---------- REMARKS + FOOTER ----------
        remarks_y = last_table_y - 24
        if remarks_y < 120:
            c.showPage()
            remarks_y = height - 120

        raw_name = (session.get("full_name") or session.get("username") or "").strip()
        if raw_name:
            parts = raw_name.split()
            if len(parts) >= 2:
                raw_name = f"{parts[0]} {parts[-1]}"
        prepared_name = f"Tr. {raw_name}" if raw_name else "Tr. __________________"
        today_str = datetime.now().strftime("%d-%b-%Y")

        overall = (meta.get("overall_remark") or "").strip()
        # convert \n to <br/> so each role appears on its own line
        overall_html = (overall or " ").replace("\n", "<br/>")

        sc = (meta.get("special_communication") or "").strip()
        ntb = meta.get("next_term_begin")
        nte = meta.get("next_term_end")
        #fees = (meta.get("school_fees") or "").strip()
        #fees_dc = (meta.get("school_fees_daycare") or "").strip()

        ntb_str = ntb.strftime("%d/%m/%y") if ntb else "__________"
        nte_str = nte.strftime("%d/%m/%y") if nte else "__________"

        styles = getSampleStyleSheet()
        h_style = ParagraphStyle(
            "bottom_head",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            textColor=colors.white,
            alignment=0,
        )
        label_style = ParagraphStyle(
            "bottom_label",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
        )
        text_style = ParagraphStyle(
            "bottom_text",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
        )

        bottom_data = [
            [Paragraph("Remarks & Recommendations to the learner:", h_style), ""],
            [Paragraph(overall_html, text_style), ""],
            [Paragraph("Special Communication:", label_style), Paragraph(sc or " ", text_style)],
            [Paragraph("Next term begins on:", label_style),
             Paragraph(f"{ntb_str} will end on: {nte_str}", text_style)],
            #[Paragraph("School fees:", label_style), Paragraph(fees or " ", text_style)],
            #[Paragraph("School fees + Requirements:", label_style), Paragraph(fees_dc or " ", text_style)],
            [Paragraph("Prepared by:", label_style), Paragraph(prepared_name, text_style)],
            [Paragraph("Date:", label_style), Paragraph(today_str, text_style)],
        ]

        bottom_table = Table(
            bottom_data,
            colWidths=[(width - 2 * left) * 0.28, (width - 2 * left) * 0.72],
        )
        bt = TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.75, colors.lightgrey),
                ("GRID", (0, 2), (-1, -1), 0.25, colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.0, 0.45, 0.80)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ]
        )
        bt.add("SPAN", (0, 0), (-1, 0))
        bt.add("SPAN", (0, 1), (-1, 1))
        bottom_table.setStyle(bt)

        bw, bh = bottom_table.wrapOn(c, width - 2 * left, remarks_y - 40)
        bottom_y = remarks_y - bh
        bottom_table.drawOn(c, left, bottom_y)
        c.showPage()

    # ========= END LOOP =========

    c.save()
    buf.seek(0)
    filename = f"Checklist_Batch_{term}_{year}.pdf".replace(" ", "_")
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


#=========================X-TER ASSESSMENT========================
@app.route("/character_form", methods=["GET", "POST"])
@require_role("admin", "headteacher", "teacher", "class_teacher", "deputyheadteacher", "dos", "bursar", "classmanager")
def character_form():
    ay = get_active_academic_year() or {}

    sid_raw = request.args.get("student_id") or request.form.get("student_id")
    if not sid_raw or not str(sid_raw).isdigit():
        flash("Missing or invalid learner.", "danger")
        return redirect(url_for("students"))
    student_id = int(sid_raw)

    raw_term = (request.values.get("term") or ay.get("current_term") or ay.get("term") or "Term 1")
    term = (raw_term or "Term 1").strip()

    raw_year = (request.values.get("year") or ay.get("year") or ay.get("active_year") or datetime.now().year)
    try:
        year = int(raw_year)
    except (TypeError, ValueError):
        year = int(ay.get("year") or datetime.now().year)

    # ✅ FETCH STUDENT EARLY (so it's always defined)
    student = _fetch_student_basic(student_id)
    if not student:
        flash("Learner not found.", "danger")
        return redirect(url_for("students"))

    # ✅ ITEMS EARLY TOO
    items = _fetch_character_items()
    if not items:
        flash("No character assessment items defined yet.", "warning")

    # ================= POST =================
    if request.method == "POST":
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        for item in items:
            key = f"level_{item['id']}"
            level = (request.form.get(key) or "").strip()
            if not level:
                cur.execute("""
                    DELETE FROM character_scores
                    WHERE student_id=%s AND term=%s AND year=%s AND item_id=%s
                """, (student_id, term, year, item["id"]))
            else:
                cur.execute("""
                    INSERT INTO character_scores (student_id, term, year, item_id, level)
                    VALUES (%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE level=VALUES(level)
                """, (student_id, term, year, item["id"], level))

        class_comment = (request.form.get("class_teacher_comment") or "").strip()
        head_comment = (request.form.get("head_teacher_comment") or "").strip()
        overall_custom = (request.form.get("overall_custom") or "").strip()

        if overall_custom:
            lines = [l.strip() for l in overall_custom.splitlines() if l.strip()]
            for l in lines:
                low = l.lower()
                if low.startswith("class teacher:"):
                    class_comment = l.split(":", 1)[1].strip()
                elif low.startswith("headteacher:"):
                    head_comment = l.split(":", 1)[1].strip()

        next_term_begin = request.form.get("next_term_begin") or None
        next_term_end = request.form.get("next_term_end") or None
        special = (request.form.get("special_communication") or "").strip()

        cur.execute("""
            INSERT INTO character_meta
              (student_id, term, year, class_teacher_comment, head_teacher_comment,
               next_term_begin, next_term_end, special_communication)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
              class_teacher_comment=VALUES(class_teacher_comment),
              head_teacher_comment=VALUES(head_teacher_comment),
              next_term_begin=VALUES(next_term_begin),
              next_term_end=VALUES(next_term_end),
              special_communication=VALUES(special_communication)
        """, (student_id, term, year, class_comment, head_comment, next_term_begin, next_term_end, special))

        conn.commit()
        cur.close()
        conn.close()

        flash("Character assessment saved.", "success")
        return redirect(url_for("character_form", student_id=student_id, term=term, year=year))

    # ================= GET =================
    scores_map = _fetch_character_scores_map(student_id, term, year)
    meta = _fetch_character_meta(student_id, term, year) or {}

    # normalize dates
    for k in ("next_term_begin", "next_term_end"):
        if meta.get(k) and hasattr(meta[k], "strftime"):
            meta[k] = meta[k].strftime("%Y-%m-%d")

    # any override logic here is now safe because student exists

    level_options = list(REMARK_OPTIONS)

    conn = get_db_connection()
    teacher_library, head_library = load_comment_library_groups(conn)
    conn.close()

    # prev/next
    prev_id = None
    next_id = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT id FROM students
            WHERE class_name=%s AND stream=%s AND id < %s
            ORDER BY id DESC LIMIT 1
        """, (student.get("class_name"), student.get("stream"), student_id))
        row = cur.fetchone()
        if row: prev_id = row[0]

        cur.execute("""
            SELECT id FROM students
            WHERE class_name=%s AND stream=%s AND id > %s
            ORDER BY id ASC LIMIT 1
        """, (student.get("class_name"), student.get("stream"), student_id))
        row = cur.fetchone()
        if row: next_id = row[0]

        cur.close()
        conn.close()
    except Exception:
        prev_id = None
        next_id = None

    return render_template(
        "character_form.html",
        student=student,
        term=term,
        year=year,
        items=items,
        scores_map=scores_map,
        meta=meta,
        level_options=level_options,
        prev_id=prev_id,
        next_id=next_id,
        teacher_library=teacher_library,
        head_library=head_library,
    )





@app.route("/character_pdf", methods=["GET"])
@require_role("admin", "headteacher", "teacher", "class_teacher", "deputyheadteacher", "dos", "bursar", "classmanager")
def character_pdf():
    sid_raw = request.args.get("student_id")
    if not sid_raw or not str(sid_raw).isdigit():
        flash("Missing or invalid learner.", "danger")
        return redirect(url_for("students"))
    student_id = int(sid_raw)

    ay = get_active_academic_year() or {}
    raw_term = (
        request.args.get("term")
        or ay.get("current_term")
        or ay.get("term")
        or "Term 1"
    )
    term = (raw_term or "Term 1").strip()
    raw_year = (
        request.args.get("year")
        or ay.get("year")
        or ay.get("active_year")
        or datetime.now().year
    )
    try:
        year = int(raw_year)
    except (TypeError, ValueError):
        year = int(ay.get("year") or datetime.now().year)

    student = _fetch_student_basic(student_id)
    if not student:
        flash("Learner not found.", "danger")
        return redirect(url_for("students"))

    items = _fetch_character_items()
    scores_map = _fetch_character_scores_map(student_id, term, year)
    meta = _fetch_character_meta(student_id, term, year)

    import io
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _draw_character_report_page(c, student, term, year, items, scores_map, meta)
    c.save()
    buf.seek(0)

    filename = f"Character_{student.get('student_number','')}_{term}_{year}.pdf".replace(" ", "_")
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


@app.route("/reports/character_batch_pdf", methods=["POST"])
@require_role("admin", "headteacher", "teacher", "class_teacher", "deputyheadteacher", "dos", "bursar", "classmanager")
def character_batch_pdf():
    ay = get_active_academic_year() or {}
    raw_term = (
        request.form.get("term")
        or request.form.get("termsel")
        or ay.get("current_term")
        or ay.get("term")
        or "Term 1"
    )
    term = (raw_term or "Term 1").strip()
    raw_year = (
        request.form.get("year")
        or request.form.get("yearsel")
        or ay.get("year")
        or ay.get("active_year")
        or datetime.now().year
    )
    try:
        year = int(raw_year)
    except (TypeError, ValueError):
        year = int(ay.get("year") or datetime.now().year)

    raw_ids = request.form.getlist("selected_ids")
    student_ids = [int(s.strip()) for s in raw_ids if s and s.strip().isdigit()]
    if not student_ids:
        flash("Please select at least one learner first.", "warning")
        return redirect(
            url_for(
                "reports_hub",
                class_name=request.form.get("class_name", ""),
                term=term,
                year=year,
            )
        )

    items = _fetch_character_items()

    import io
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    for sid in student_ids:
        student = _fetch_student_basic(sid)
        if not student:
            continue
        scores_map = _fetch_character_scores_map(sid, term, year)
        meta = _fetch_character_meta(sid, term, year)
        _draw_character_report_page(c, student, term, year, items, scores_map, meta)

    c.save()
    buf.seek(0)
    filename = f"Character_Batch_{term}_{year}.pdf".replace(" ", "_")
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )
    
    
@app.route("/admin/character_items", methods=["GET", "POST"])
@require_role("admin", "headteacher")
def admin_character_items():
    term = (request.values.get("term") or "Term 1").strip()
    term_no = _term_to_no(term)
    class_name = (request.values.get("class_name") or "").strip() or None

    if request.method == "POST":
        action = request.form.get("action")

        conn = get_db_connection()
        cur = conn.cursor()

        if action == "add":
            area = (request.form.get("area") or "").strip()
            section = (request.form.get("section") or "").strip() or None
            skill_label = (request.form.get("skill_label") or "").strip()
            description = (request.form.get("description") or "").strip()
            sort_order = int(request.form.get("sort_order") or 0)
            is_active = 1 if request.form.get("is_active") else 0
            # "ALL" option => NULL in DB
            term_no_in = request.form.get("term_no")
            term_no_db = None if term_no_in == "ALL" else int(term_no_in)
            class_in = (request.form.get("class_name") or "").strip()
            class_db = None if class_in == "ALL" or class_in == "" else class_in

            cur.execute(
                """
                INSERT INTO character_items (area, section, term_no, class_name, skill_label, description, sort_order, is_active)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (area, section, term_no_db, class_db, skill_label, description, sort_order, is_active),
            )

        elif action == "update":
            item_id = int(request.form["id"])
            area = (request.form.get("area") or "").strip()
            section = (request.form.get("section") or "").strip() or None
            skill_label = (request.form.get("skill_label") or "").strip()
            description = (request.form.get("description") or "").strip()
            sort_order = int(request.form.get("sort_order") or 0)
            is_active = 1 if request.form.get("is_active") else 0

            term_no_in = request.form.get("term_no")
            term_no_db = None if term_no_in == "ALL" else int(term_no_in)
            class_in = (request.form.get("class_name") or "").strip()
            class_db = None if class_in == "ALL" or class_in == "" else class_in

            cur.execute(
                """
                UPDATE character_items
                SET area=%s, section=%s, term_no=%s, class_name=%s,
                    skill_label=%s, description=%s, sort_order=%s, is_active=%s
                WHERE id=%s
                """,
                (area, section, term_no_db, class_db, skill_label, description, sort_order, is_active, item_id),
            )

        elif action == "delete":
            item_id = int(request.form["id"])
            # safer than hard delete:
            cur.execute("UPDATE character_items SET is_active=0 WHERE id=%s", (item_id,))

        conn.commit()
        cur.close()
        conn.close()

        flash("Saved.", "success")
        return redirect(url_for("admin_character_items", term=term, class_name=class_name or ""))

    # GET: list items for filter
    items = _fetch_character_items(term, class_name)

    return render_template(
        "admin_character_items.html",
        items=items,
        term=term,
        term_no=term_no,
        class_name=class_name or "",
    )
    
    
@app.get("/api/character_items")
@require_role("admin","headteacher","teacher","class_teacher","deputyheadteacher","dos","bursar","classmanager")
def api_character_items():
    term = (request.args.get("term") or "Term 1").strip()
    class_name = (request.args.get("class_name") or "").strip() or None
    items = _fetch_character_items(term, class_name)
    return {"items": items}

    


#======================X-TER ASSESSMENT ENDS=============================================


# ===================== TIMETABLES =====================

@app.route("/timetables", methods=["GET", "POST"])
@require_role("admin", "headteacher", "deputyheadteacher", "dos", "teacher", "classmanager")
def timetables():
    """
    - Admin / DOS / Head / Deputy: can upload timetables.
    - All allowed roles: can view list and open/download the file.
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    role = session.get("role") or ""
    staff_id = session.get("staff_id")

    # ----- lists for dropdowns -----
    cur.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name")
    class_rows = cur.fetchall() or []
    classes = [r["class_name"] for r in class_rows]

    cur.execute(
        """
        SELECT DISTINCT stream
        FROM classes
        WHERE stream IS NOT NULL AND TRIM(stream) <> ''
        ORDER BY stream
        """
    )
    stream_rows = cur.fetchall() or []
    streams = [r["stream"] for r in stream_rows]

    # ---------- UPLOAD ----------
    if request.method == "POST" and role in {"admin", "dos", "headteacher", "deputyheadteacher"}:
        if not staff_id:
            flash("Your account is not linked to an employee record. Cannot save uploader.", "danger")
            cur.close()
            conn.close()
            return redirect(url_for("timetables"))

        class_name = (request.form.get("class_name") or "").strip()
        stream = (request.form.get("stream") or "").strip() or None
        term = (request.form.get("term") or "").strip() or "Term 1"
        year_raw = request.form.get("year") or datetime.now().year
        title = (request.form.get("title") or "").strip()

        if not title:
            title = f"{class_name} {stream or ''} {term} {year_raw}".strip()

        try:
            year = int(year_raw)
        except Exception:
            year = datetime.now().year

        file = request.files.get("timetable_file")

        try:
            file_name, file_mime, file_blob = extract_upload_blob(file)
        except ValueError as e:
            flash(str(e), "danger")
            cur.close()
            conn.close()
            return redirect(url_for("timetables"))

        cur.execute(
            """
            INSERT INTO timetables
                (class_name, stream, term, year, title,
                 file_name, file_mime, file_blob, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (class_name, stream, term, year, title,
             file_name, file_mime, file_blob, staff_id),
        )
        conn.commit()
        flash("Timetable uploaded.", "success")
        cur.close()
        conn.close()
        return redirect(url_for("timetables"))

    # ---------- LIST ----------
    cur.execute(
        """
        SELECT t.id, t.class_name, t.stream, t.term, t.year,
               t.title, t.created_at,
               e.first_name, e.last_name
        FROM timetables t
        LEFT JOIN employees e ON e.id = t.created_by
        ORDER BY t.year DESC, t.term DESC, t.class_name, t.stream
        """
    )
    rows = cur.fetchall() or []

    # build display name for template
    for r in rows:
        if r.get("first_name") or r.get("last_name"):
            r["uploaded_by"] = f"{r.get('first_name') or ''} {r.get('last_name') or ''}".strip()
        else:
            r["uploaded_by"] = "Unknown"

    cur.close()
    conn.close()

    return render_template(
        "timetables.html",
        rows=rows,
        role=role,
        classes=classes,
        streams=streams,
    )


@app.route("/timetables/<int:tt_id>/file")
@require_role("teacher", "classmanager", "dos", "headteacher", "deputyheadteacher", "admin", "classteacher")
def timetable_file(tt_id):
    """
    ?dl=1 -> download
    else -> open in browser
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT file_name, file_mime, file_blob FROM timetables WHERE id=%s",
        (tt_id,),
    )
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row or not row["file_blob"]:
        abort(404)

    download = request.args.get("dl") == "1"

    buf = io.BytesIO(row["file_blob"])
    buf.seek(0)

    return send_file(
        buf,
        mimetype=row["file_mime"] or "application/octet-stream",
        as_attachment=download,
        download_name=row["file_name"],
    )


@app.route("/timetables/<int:tt_id>/delete", methods=["POST"])
@require_role("admin", "headteacher", "deputyheadteacher", "dos")
def timetable_delete(tt_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM timetables WHERE id=%s", (tt_id,))
    conn.commit()
    cur.close()
    conn.close()
    flash("Timetable deleted.", "success")
    return redirect(url_for("timetables"))


# ===================== NOTES (UPLOAD & MANAGE) =====================


@app.route("/notes/manage", methods=["GET", "POST"])
@require_role("admin", "dos", "headteacher")
def notes_manage():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    staff_id = session.get("employee_id") or session.get("staff_id")

    # subjects for dropdown
    cur.execute("SELECT id, code, name FROM subjects ORDER BY name")
    subjects = cur.fetchall() or []

    # classes & streams for dropdowns
    cur.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name")
    class_rows = cur.fetchall() or []
    classes = [r["class_name"] for r in class_rows]

    cur.execute(
        """
        SELECT DISTINCT stream
        FROM classes
        WHERE stream IS NOT NULL AND TRIM(stream) <> ''
        ORDER BY stream
        """
    )
    stream_rows = cur.fetchall() or []
    streams = [r["stream"] for r in stream_rows]

    # ---------- UPLOAD ----------
    if request.method == "POST":
        if not staff_id:
            flash("Your account is not linked to an employee record. Cannot save uploader.", "danger")
            cur.close()
            conn.close()
            return redirect(url_for("notes_manage"))

        subject_code = (request.form.get("subject_code") or "").strip()
        class_name = (request.form.get("class_name") or "").strip() or None
        stream = (request.form.get("stream") or "").strip() or None
        term = (request.form.get("term") or "").strip() or "Term 1"
        year_raw = request.form.get("year") or datetime.now().year
        title = (request.form.get("title") or "").strip()
        desc = (request.form.get("description") or "").strip()

        if not subject_code:
            flash("Please choose a subject.", "danger")
            cur.close()
            conn.close()
            return redirect(url_for("notes_manage"))

        try:
            year = int(year_raw)
        except Exception:
            year = datetime.now().year


        file = request.files.get("note_file")

        if not staff_id:
            flash("Your account is not linked to an employee record. Ask the admin to link you under Users.", "warning")
            return redirect(url_for("notes_manage"))

        try:
            file_name, file_mime, file_blob = extract_upload_blob(file)
        except ValueError as e:
            flash(str(e), "danger")
            return redirect(url_for("notes_manage"))

        cur.execute(
            """
            INSERT INTO subject_notes
                (subject_code, class_name, stream, term, year,
                 title, description, file_name, file_mime, file_blob, uploaded_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (subject_code, class_name, stream, term, year,
             title, desc, file_name, file_mime, file_blob, staff_id),
        )

        conn.commit()
        flash("Notes uploaded.", "success")
        cur.close()
        conn.close()
        return redirect(url_for("notes_manage"))

    # ---------- LIST ALL NOTES ----------
    cur.execute(
        """
        SELECT n.id, n.subject_code, n.class_name, n.stream,
               n.term, n.year, n.title, n.created_at,
               e.first_name, e.last_name
        FROM subject_notes n
        LEFT JOIN employees e ON e.id = n.uploaded_by
        ORDER BY n.year DESC, n.term DESC, n.subject_code, n.class_name
        """
    )
    notes = cur.fetchall() or []

    for r in notes:
        if r.get("first_name") or r.get("last_name"):
            r["uploaded_by"] = f"{r.get('first_name') or ''} {r.get('last_name') or ''}".strip()
        else:
            r["uploaded_by"] = "Unknown"

    cur.close()
    conn.close()

    return render_template(
        "notes_manage.html",
        subjects=subjects,
        classes=classes,
        streams=streams,
        notes=notes,
    )
    
@app.route("/notes/<int:note_id>/delete", methods=["POST"])
@require_role("admin", "dos", "headteacher")
def notes_delete(note_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM subject_notes WHERE id=%s", (note_id,))
    conn.commit()
    cur.close()
    conn.close()
    flash("Notes entry deleted.", "success")
    return redirect(url_for("notes_manage"))

# ===================== NOTES (TEACHER VIEW) =====================

@app.route("/notes/my")
@require_role("teacher", "classmanager", "dos", "headteacher", "deputyheadteacher", "admin", "classteacher")
def notes_my():
    """
    Teachers see only notes for subjects they handle.
    DOS/Head can see everything from here as well.
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    role = session.get("role") or ""
    staff_id = session.get("staff_id")

    # If admin-type roles come here, just show all
    if role in {"dos", "headteacher", "deputyheadteacher"}:
        cur.execute(
            """
            SELECT n.id, n.subject_code, n.class_name, n.stream,
                   n.term, n.year, n.title, n.created_at,
                   e.first_name, e.last_name
            FROM subject_notes n
            LEFT JOIN employees e ON e.id = n.uploaded_by
            ORDER BY n.year DESC, n.term DESC, n.subject_code, n.class_name
            """
        )
        notes = cur.fetchall() or []
        cur.close()
        conn.close()
        return render_template("notes_my.html", notes=notes)

    # ---------- Normal teacher: limit by subjects they teach ----------
    # staff_id here is employees.id
    cur.execute(
        """
        SELECT DISTINCT s.code AS subject_code
        FROM teacher_subjects ts
        JOIN teachers t ON t.id = ts.teacher_id
        JOIN subjects s ON s.id = ts.subject_id
        WHERE t.employee_id = %s
        """,
        (staff_id,),
    )
    subject_rows = cur.fetchall() or []
    subjects = [row["subject_code"] for row in subject_rows]

    if not subjects:
        cur.close()
        conn.close()
        return render_template("notes_my.html", notes=[])

    placeholders = ",".join(["%s"] * len(subjects))
    sql = f"""
        SELECT n.id, n.subject_code, n.class_name, n.stream,
               n.term, n.year, n.title, n.created_at,
               e.first_name, e.last_name
        FROM subject_notes n
        LEFT JOIN employees e ON e.id = n.uploaded_by
        WHERE n.subject_code IN ({placeholders})
        ORDER BY n.year DESC, n.term DESC, n.class_name, n.subject_code
    """
    cur.execute(sql, subjects)
    notes = cur.fetchall() or []
    cur.close()
    conn.close()

    return render_template("notes_my.html", notes=notes)


# ===================== NOTES (DOWNLOAD SINGLE FILE) =====================


@app.route("/notes/<int:note_id>/file")
@require_role("teacher", "classmanager", "dos", "headteacher", "deputyheadteacher", "admin", "classteacher")
def notes_file(note_id):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute(
        "SELECT subject_code, file_name, file_mime, file_blob FROM subject_notes WHERE id=%s",
        (note_id,),
    )
    note = cur.fetchone()
    if not note:
        cur.close()
        conn.close()
        abort(404)

    role = session.get("role") or ""
    staff_id = session.get("staff_id")

    # Admin-type roles: allowed
    allowed = role in {"admin", "headteacher", "deputyheadteacher", "dos"}

    if not allowed:
        # Check if this teacher teaches the subject
        cur.execute(
            """
            SELECT 1 FROM teacher_subjects ts
            JOIN subjects s ON s.id = ts.subject_id
            WHERE ts.teacher_id=%s AND subject_code=%s
            LIMIT 1
            """,
            (staff_id, note["subject_code"]),
        )
        allowed = bool(cur.fetchone())

    cur.close()
    conn.close()

    if not allowed:
        flash("You are not allowed to access these notes.", "danger")
        return redirect(url_for("notes_my"))

    if not note["file_blob"]:
        abort(404)

    download = request.args.get("dl") == "1"

    buf = io.BytesIO(note["file_blob"])
    buf.seek(0)

    return send_file(
        buf,
        mimetype=note["file_mime"] or "application/pdf",
        as_attachment=download,
        download_name=note["file_name"],
    )

#================WORKPLANS========================================================================
@app.route("/workplans", methods=["GET", "POST"])
@require_role(
    "teacher", "classmanager", "classteacher",
    "dos", "headteacher", "deputyheadteacher", "director", "admin"
)
def workplans():
    """
    Teacher view:
    - Add / edit / delete own work plans.
    - See list of own plans.
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # ----- dropdown data: classes, streams -----
    cur.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name")
    class_rows = cur.fetchall() or []
    classes = [r["class_name"] for r in class_rows]

    cur.execute(
        """
        SELECT DISTINCT stream FROM classes
        WHERE stream IS NOT NULL AND TRIM(stream) <> ''
        ORDER BY stream
        """
    )
    stream_rows = cur.fetchall() or []
    streams = [r["stream"] for r in stream_rows]

    employee_id = session.get("employee_id") # from users.employee_id
    if not employee_id:
        cur.close()
        conn.close()
        flash("Your account is not linked to an employee record. Cannot use work plans.", "warning")
        return render_template(
            "workplans.html",
            classes=classes,
            streams=streams,
            terms=["Term 1", "Term 2", "Term 3"],
            weeks=WEEKS,
            plans=[],
            editing=None,
            today=date.today().isoformat(),
        )

    # ---------- POST: SAVE or DELETE ----------
    if request.method == "POST":
        action = request.form.get("action") or "save"
        plan_id = request.form.get("plan_id")

        # DELETE branch
        if action == "delete" and plan_id:
            cur.execute(
                "DELETE FROM teacher_work_plans WHERE id=%s AND teacher_id=%s",
                (plan_id, employee_id),
            )
            conn.commit()
            cur.close()
            conn.close()
            flash("Work plan deleted.", "success")
            return redirect(url_for("workplans"))

        # SAVE / UPDATE branch
        plan_date = request.form.get("plan_date") or ""
        class_name = (request.form.get("class_name") or "").strip() or None
        stream = (request.form.get("stream") or "").strip() or None
        term = (request.form.get("term") or "").strip() or "Term 1"
        week_label = (request.form.get("week_label") or "").strip() or None

        period_from = request.form.get("period_from") or ""
        period_to = request.form.get("period_to") or ""
        planned_activities = (request.form.get("planned_activities") or "").strip()
        done_activities = (request.form.get("done_activities") or "").strip() or None
        comment = (request.form.get("comment") or "").strip() or None
        status = (request.form.get("status") or "pending").strip()

        if not plan_date or not period_from or not period_to or not planned_activities:
            flash("Date, period and planned activities are required.", "danger")
            cur.close()
            conn.close()
            return redirect(url_for("workplans"))

        if plan_id: # UPDATE
            cur.execute(
                """
                UPDATE teacher_work_plans
                SET plan_date=%s,
                    class_name=%s,
                    stream=%s,
                    term=%s,
                    week_label=%s,
                    period_from=%s,
                    period_to=%s,
                    activities=%s,
                    done_activities=%s,
                    status=%s,
                    comment=%s
                WHERE id=%s AND teacher_id=%s
                """,
                (
                    plan_date, class_name, stream, term, week_label,
                    period_from, period_to,
                    planned_activities, done_activities,
                    status, comment,
                    plan_id, employee_id,
                ),
            )
            flash("Work plan updated.", "success")
        else: # INSERT
            cur.execute(
                """
                INSERT INTO teacher_work_plans
                    (teacher_id, class_name, stream, term, week_label,
                     plan_date, period_from, period_to,
                     activities, done_activities, status, comment)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    employee_id, class_name, stream, term, week_label,
                    plan_date, period_from, period_to,
                    planned_activities, done_activities,
                    status, comment,
                ),
            )
            flash("Work plan saved.", "success")

        conn.commit()
        cur.close()
        conn.close()
        return redirect(url_for("workplans"))

    # ---------- EDIT MODE (optional) ----------
    edit_id = request.args.get("edit", type=int)
    editing = None
    if edit_id:
        cur.execute(
            """
            SELECT *
            FROM teacher_work_plans
            WHERE id=%s AND teacher_id=%s
            """,
            (edit_id, employee_id),
        )
        editing = cur.fetchone()

    # ---------- LIST OWN PLANS ----------
    cur.execute(
        """
        SELECT
            id,
            plan_date,
            class_name,
            stream,
            term,
            week_label,
            period_from,
            period_to,
            activities AS planned_activities,
            done_activities,
            status,
            comment
        FROM teacher_work_plans
        WHERE teacher_id=%s
        ORDER BY plan_date DESC, period_from
        """,
        (employee_id,),
    )
    plans = cur.fetchall() or []

    cur.close()
    conn.close()

    return render_template(
        "workplans.html",
        classes=classes,
        streams=streams,
        terms=["Term 1", "Term 2", "Term 3"],
        weeks=WEEKS,
        plans=plans,
        editing=editing,
        today=date.today().isoformat(),
    )



@app.route("/workplans/<int:plan_id>/delete", methods=["POST"])
@require_role("teacher", "classteacher", "classmanager",
              "dos", "headteacher", "deputyheadteacher",
              "director", "admin")
def workplans_delete(plan_id):
    """
    Delete a work plan. Only owner or supervisor can delete.
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    role = (session.get("role") or "").lower()
    emp_id = current_employee_id()

    if not emp_id:
        flash("Your account is not linked to an employee record.", "warning")
        cur.close()
        conn.close()
        return redirect(url_for("workplans"))

    if role in SUPERVISOR_ROLES:
        cur.execute("DELETE FROM teacher_work_plans WHERE id=%s", (plan_id,))
    else:
        cur.execute(
            "DELETE FROM teacher_work_plans WHERE id=%s AND teacher_id=%s",
            (plan_id, emp_id),
        )
    conn.commit()
    cur.close()
    conn.close()

    flash("Work plan deleted.", "success")
    return redirect(url_for("workplans"))



@app.route("/workplans/view/<int:plan_id>")
@require_role(
    "teacher", "classmanager", "classteacher",
    "dos", "headteacher", "deputyheadteacher", "director", "admin"
)

def workplan_view(plan_id):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    role = session.get("role") or ""
    employee_id = session.get("employee_id")

    cur.execute(
        """
        SELECT
            w.*,
            e.first_name,
            e.last_name,
            e.department
        FROM teacher_work_plans w
        LEFT JOIN employees e ON e.id = w.teacher_id
        WHERE w.id = %s
        """,
        (plan_id,),
    )
    plan = cur.fetchone()
    cur.close()
    conn.close()

    if not plan:
        abort(404)

    # permission check: non-supervisors can only see their own plans
    supervisors = {"dos", "headteacher", "deputyheadteacher", "director", "admin"}
    if role not in supervisors and plan["teacher_id"] != employee_id:
        flash("You are not allowed to view this work plan.", "danger")
        return redirect(url_for("workplans"))

    return render_template("workplan_view.html", plan=plan)



@app.route("/workplans/export/pdf")
@require_role("teacher", "classteacher", "classmanager",
              "dos", "headteacher", "deputyheadteacher",
              "director", "admin")
def workplans_export_pdf():
    """
    Export work plans to PDF.
    - Teachers: their own plans.
    - Supervisors: all or filtered by ?teacher_id=&from=&to=
    """
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    role = (session.get("role") or "").lower()
    emp_id = current_employee_id()
    if not emp_id:
        flash("Your account is not linked to an employee record.", "warning")
        cur.close()
        conn.close()
        return redirect(url_for("workplans"))

    is_supervisor = role in SUPERVISOR_ROLES

    where = []
    params = []

    if is_supervisor:
        teacher_filter = (request.args.get("teacher_id") or "").strip()
        if teacher_filter:
            where.append("wp.teacher_id = %s")
            params.append(teacher_filter)
    else:
        where.append("wp.teacher_id = %s")
        params.append(emp_id)

    date_from = (request.args.get("from") or "").strip()
    date_to = (request.args.get("to") or "").strip()
    if date_from:
        where.append("wp.plan_date >= %s")
        params.append(date_from)
    if date_to:
        where.append("wp.plan_date <= %s")
        params.append(date_to)

    where_sql = ""
    if where:
        where_sql = "WHERE " + " AND ".join(where)

    sql = f"""
        SELECT wp.plan_date, wp.period_from, wp.period_to,
               wp.activities, wp.status, wp.comment,
               e.first_name, e.last_name
        FROM teacher_work_plans wp
        JOIN employees e ON e.id = wp.teacher_id
        {where_sql}
        ORDER BY wp.plan_date DESC, wp.period_from
    """
    cur.execute(sql, params)
    rows = cur.fetchall() or []
    cur.close()
    conn.close()

    # ------- Build PDF -------
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    top = height - 20 * mm
    x_margin = 15 * mm
    y = top

    c.setFont("Helvetica-Bold", 14)
    c.drawString(x_margin, y, "Teacher Work Plans")
    y -= 8 * mm

    c.setFont("Helvetica", 9)

    for row in rows:
        if y < 25 * mm: # new page
            c.showPage()
            y = top
            c.setFont("Helvetica-Bold", 14)
            c.drawString(x_margin, y, "Teacher Work Plans (cont.)")
            y -= 8 * mm
            c.setFont("Helvetica", 9)

        teacher_name = f"{row['first_name']} {row['last_name']}"
        line1 = f"{row['plan_date']} {row['period_from']} - {row['period_to']} [{row['status']}]"
        line2 = f"Teacher: {teacher_name}"
        c.drawString(x_margin, y, line1)
        y -= 5 * mm
        c.drawString(x_margin, y, line2)
        y -= 5 * mm

        # activities
        act = (row["activities"] or "").replace("\r", "")
        for para in act.split("\n"):
            for chunk in _wrap_text(para, 95):
                if y < 25 * mm:
                    c.showPage()
                    y = top
                    c.setFont("Helvetica", 9)
                c.drawString(x_margin + 5 * mm, y, f"- {chunk}")
                y -= 4 * mm

        # comment
        comment = (row["comment"] or "").replace("\r", "")
        if comment:
            if y < 25 * mm:
                c.showPage()
                y = top
                c.setFont("Helvetica", 9)
            c.drawString(x_margin + 5 * mm, y, "Comment:")
            y -= 4 * mm
            for para in comment.split("\n"):
                for chunk in _wrap_text(para, 95):
                    if y < 25 * mm:
                        c.showPage()
                        y = top
                        c.setFont("Helvetica", 9)
                    c.drawString(x_margin + 10 * mm, y, chunk)
                    y -= 4 * mm

        y -= 6 * mm # gap between plans

    c.showPage()
    c.save()
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="teacher_work_plans.pdf",
    )




@app.route("/workplans/all")
@require_role("admin", "dos", "headteacher", "deputyheadteacher", "director")
def workplans_all():
    """
    Supervisors' view: see ALL teachers' work plans.
    Allows filtering by teacher / class / term / week and export to PDF.
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    teacher_id = request.args.get("teacher_id", type=int)
    class_name = (request.args.get("class_name") or "").strip() or None
    term = (request.args.get("term") or "").strip() or None
    week = (request.args.get("week") or "").strip() or None

    where = []
    params = []

    if teacher_id:
        where.append("w.teacher_id = %s")
        params.append(teacher_id)
    if class_name and class_name != "ALL":
        where.append("w.class_name = %s")
        params.append(class_name)
    if term and term != "ALL":
        where.append("w.term = %s")
        params.append(term)
    if week and week != "ALL":
        where.append("w.week_label = %s")
        params.append(week)

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    sql = f"""
        SELECT
            w.id,
            w.teacher_id,
            w.plan_date,
            w.period_from,
            w.period_to,
            w.class_name,
            w.stream,
            w.term,
            w.week_label,
            w.activities AS planned_activities,
            w.status,
            w.comment,
            e.first_name,
            e.last_name,
            e.department
        FROM teacher_work_plans w
        LEFT JOIN employees e ON e.id = w.teacher_id
        {where_sql}
        ORDER BY w.plan_date DESC,
                 e.last_name,
                 e.first_name,
                 w.period_from
    """
    cur.execute(sql, params)
    rows = cur.fetchall() or []

    # for filters drop-downs
    cur.execute("SELECT id, first_name, last_name FROM employees ORDER BY first_name, last_name")
    teachers = cur.fetchall() or []

    cur.execute(
        "SELECT DISTINCT class_name FROM classes WHERE class_name IS NOT NULL ORDER BY class_name"
    )
    class_rows = cur.fetchall() or []
    classes = [r["class_name"] for r in class_rows]

    cur.close()
    conn.close()

    # If you later add a /workplans/all/export route, this button will hit it.
    return render_template(
        "workplans_all.html",
        rows=rows,
        teachers=teachers,
        classes=classes,
        current_teacher_id=teacher_id,
        current_class=class_name or "ALL",
        current_term=term or "ALL",
        current_week=week or "ALL",
    )


# with app.app_context():
 #   bootstrap()
if __name__ == "__main__":
    configure_logging(app)
    # with app.app_context():
    # bootstrap()
    # populate_default_expense_categories()
    # apply_schema_guards(app)
    # upsert_admin_user()
    # add_created_at_if_missing()
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "8080")), debug=True)
